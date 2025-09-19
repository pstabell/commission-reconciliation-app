#!/usr/bin/env python3
"""Analyze the exported Excel file using openpyxl."""

try:
    from openpyxl import load_workbook
    
    # Load the workbook
    wb = load_workbook('./PRIVATE mode contacts_export_20250919_134134.xlsx')
    
    print("=== EXCEL FILE ANALYSIS ===\n")
    print(f"Sheet names: {wb.sheetnames}\n")
    
    # Check Commission Rules sheet
    if 'Commission Rules' in wb.sheetnames:
        ws = wb['Commission Rules']
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"Commission Rules headers: {headers}")
        
        # Find column indices
        carrier_idx = headers.index('carrier_name') if 'carrier_name' in headers else None
        mga_idx = headers.index('mga_name') if 'mga_name' in headers else None
        
        print(f"\nLooking for Johnson and Johnson associations...")
        
        jj_rules = []
        jj_carriers = set()
        
        # Scan all rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if mga_idx is not None and row[mga_idx] == 'Johnson and Johnson':
                jj_rules.append(row)
                if carrier_idx is not None:
                    jj_carriers.add(row[carrier_idx])
        
        print(f"\nFound {len(jj_rules)} rules with Johnson and Johnson")
        print(f"Carriers associated with Johnson and Johnson: {sorted(jj_carriers)}")
        
        # Show the rules
        if jj_rules and len(headers) > 0:
            print("\nJohnson and Johnson rules details:")
            for rule in jj_rules[:10]:  # Show first 10
                print(f"  Carrier: {rule[carrier_idx] if carrier_idx is not None else 'N/A'}")
                for i, header in enumerate(headers):
                    if i != carrier_idx and i != mga_idx and rule[i] is not None:
                        print(f"    {header}: {rule[i]}")
                print()
    
    # Check MGAs sheet
    if 'MGAs' in wb.sheetnames:
        ws = wb['MGAs']
        headers = [cell.value for cell in ws[1]]
        
        print("\nMGAs sheet:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and 'johnson' in str(row[0]).lower():
                print(f"  Found: {row}")
    
    wb.close()
    
except ImportError:
    print("openpyxl not installed. Trying alternative method...")
    
    # Try to at least check file size and basic info
    import os
    import zipfile
    
    file_path = './PRIVATE mode contacts_export_20250919_134134.xlsx'
    print(f"File size: {os.path.getsize(file_path)} bytes")
    
    # Excel files are zip archives, try to peek inside
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            print(f"\nFiles in Excel archive: {zip_file.namelist()[:10]}")
    except:
        print("Could not read Excel file structure")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()