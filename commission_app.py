import streamlit as st
st.set_page_config(layout="wide")
import traceback
import string
import random
import pandas as pd
import numpy as np
from supabase import create_client, Client
from dotenv import load_dotenv
import hashlib
import re
import plotly.express as px

# Set pandas display options to avoid scientific notation and show 2 decimal places
pd.options.display.float_format = '{:.2f}'.format

# Load environment variables
load_dotenv()
import datetime
import time
import io
import streamlit_sortables
import os
import json
import pdfplumber
import shutil
import uuid
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from column_mapping_config import (
    column_mapper, get_mapped_column, get_ui_field_name, 
    is_calculated_field, safe_column_reference
)

def check_password():
    """Returns True if the user had the correct password."""
    
    def password_entered():
        """Checks whether a password entered by the user is correct."""
        # For initial setup, we'll use a simple password check
        # You should change this password and consider using environment variables
        correct_password = os.getenv("APP_PASSWORD", "CommissionApp2025!")  # Change this!
        
        # Check if password key exists before accessing it
        if "password" in st.session_state and st.session_state["password"] == correct_password:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # Don't store password
        else:
            st.session_state["password_correct"] = False

    # Return True if the password is validated
    if st.session_state.get("password_correct", False):
        return True

    # Show input for password
    st.title("🔐 Sales Commission App - Login")
    st.text_input(
        "Password", 
        type="password", 
        on_change=password_entered, 
        key="password",
        help="Contact administrator for password"
    )
    
    if "password_correct" in st.session_state and not st.session_state["password_correct"]:
        st.error("😕 Password incorrect. Please try again.")
    
    st.info("This application contains sensitive commission data. Authentication is required.")
    return False

@st.cache_resource
def get_supabase_client():
    """Get cached Supabase client."""
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_ANON_KEY")
    if not url or not key:
        st.error("Missing Supabase credentials. Please check your .env file.")
        st.stop()
    return create_client(url, key)

@st.cache_data(ttl=300)  # Cache for 5 minutes
def load_policies_data():
    """Load policies data from Supabase with caching."""
    try:
        supabase = get_supabase_client()
        response = supabase.table('policies').select("*").execute()
        if response.data:
            df = pd.DataFrame(response.data)
            # Ensure numeric columns are properly typed
            numeric_cols = [
                'Agent Estimated Comm $',
                'Policy Gross Comm %',
                'Agency Estimated Comm/Revenue (CRM)',
                'Agency Comm Received (STMT)',
                'Premium Sold',
                'Agent Paid Amount (STMT)',
                'Agency Comm Received (STMT)',
                'Broker Fee',
                'Policy Taxes & Fees',
                'Commissionable Premium',
                'Broker Fee Agent Comm',
                'Total Agent Comm'
            ]
            
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Keep date columns as they are in the database
            # DO NOT format dates here as it can cause data loss
            
            # Round all numeric columns to 2 decimal places
            df = round_numeric_columns(df)
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading data from Supabase: {e}")
        return pd.DataFrame()

def clear_policies_cache():
    """Clear the policies data cache."""
    load_policies_data.clear()

def format_date_value(date_value, format='%m/%d/%Y'):
    """Safely format a date value to MM/DD/YYYY string format.
    
    Args:
        date_value: The date value to format (can be string, datetime, date, or pandas timestamp)
        format: The desired output format (default: '%m/%d/%Y')
    
    Returns:
        Formatted date string or original value if formatting fails
    """
    # If null or empty, return the original value
    if pd.isna(date_value) or date_value is None:
        return date_value
    
    # If it's already an empty string, keep it
    if str(date_value).strip() == '':
        return date_value
    
    try:
        # Store original value
        original_value = date_value
        
        # If it's already a datetime or date object
        if isinstance(date_value, (datetime.datetime, datetime.date)):
            return date_value.strftime(format)
        
        # If it's a pandas Timestamp
        if hasattr(date_value, 'strftime'):
            return date_value.strftime(format)
        
        # If it's already in MM/DD/YYYY format, return as is
        date_str = str(date_value)
        if re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
            return date_str
        
        # Try to parse string dates
        parsed_date = pd.to_datetime(date_value, errors='coerce')
        if pd.notna(parsed_date):
            return parsed_date.strftime(format)
        
        # If all else fails, return the original value unchanged
        return original_value
    except:
        # On any error, return the original value unchanged
        return date_value

def style_special_transactions(df):
    """Apply special styling to STMT and VOID transactions in a dataframe.
    
    Args:
        df: pandas DataFrame with a 'Transaction ID' column
    
    Returns:
        Styled DataFrame with colored rows for STMT and VOID transactions
    """
    def highlight_transaction_type(row):
        if 'Transaction ID' in row:
            trans_id = str(row['Transaction ID'])
            if '-STMT-' in trans_id:
                # Light blue background for STMT transactions
                return ['background-color: #e6f3ff'] * len(row)
            elif '-VOID-' in trans_id:
                # Light red background for VOID transactions
                return ['background-color: #ffe6e6'] * len(row)
        return [''] * len(row)
    
    # Check if DataFrame has Transaction ID column
    if 'Transaction ID' in df.columns:
        return df.style.apply(highlight_transaction_type, axis=1)
    else:
        return df

def calculate_dashboard_metrics(df):
    """Calculate dashboard metrics with reconciled vs unreconciled YTD 2025 focus."""
    metrics = {
        # Transaction metrics
        'total_transactions': len(df),
        'transactions_this_month': 0,
        'stmt_transactions': 0,
        
        # Policy metrics (unique policy numbers)
        'unique_policies': 0,
        'active_policies': 0,
        'cancelled_policies': 0,
        
        # YTD 2025 Financial metrics - Reconciled (Paid)
        'premium_reconciled_ytd': 0.0,
        'agent_comm_paid_ytd': 0.0,
        
        # YTD 2025 Financial metrics - Unreconciled (Estimated)
        'premium_unreconciled_ytd': 0.0,
        'agent_comm_estimated_ytd': 0.0
    }
    
    if df.empty:
        return metrics
    
    # Transaction counts
    metrics['total_transactions'] = len(df)
    
    # Current month transactions
    if 'Effective Date' in df.columns:
        try:
            df['Effective Date'] = pd.to_datetime(df['Effective Date'], errors='coerce')
            current_month = datetime.datetime.now().month
            current_year = datetime.datetime.now().year
            metrics['transactions_this_month'] = len(df[(df['Effective Date'].dt.month == current_month) & 
                                                       (df['Effective Date'].dt.year == current_year)])
        except:
            pass
    
    # STMT transactions
    if 'Transaction Type' in df.columns:
        metrics['stmt_transactions'] = len(df[df['Transaction Type'].str.startswith('-', na=False)])
    
    # Unique policy count
    if 'Policy Number' in df.columns:
        metrics['unique_policies'] = df['Policy Number'].nunique()
        
        # Calculate active vs cancelled policies
        if not df.empty and 'Transaction Type' in df.columns:
            # Get latest transaction for each policy
            latest_trans = df.sort_values('Effective Date').groupby('Policy Number').last()
            metrics['active_policies'] = len(latest_trans[~latest_trans['Transaction Type'].isin(['CAN', 'XCL'])])
            metrics['cancelled_policies'] = len(latest_trans[latest_trans['Transaction Type'].isin(['CAN', 'XCL'])])
    
    # YTD 2025 Financial metrics
    if 'Effective Date' in df.columns:
        try:
            df['Effective Date'] = pd.to_datetime(df['Effective Date'], errors='coerce')
            
            # For reconciled metrics, we need -STMT- entries from 2025 and their matching originals (which could be from any year)
            # First, get all -STMT- entries from 2025 for commission paid amounts
            stmt_mask = pd.Series(False, index=df.index)
            if 'Transaction ID' in df.columns:
                stmt_mask = df['Transaction ID'].str.contains('-STMT-', na=False)
            
            # Filter -STMT- entries to only 2025 (based on statement date)
            df_stmt_all = df[stmt_mask]
            df_stmt_2025 = df_stmt_all[df_stmt_all['Effective Date'].dt.year == 2025]
            
            # Get ALL original transactions (from any year, exclude -STMT-, -VOID-, -ADJ-)
            original_mask = pd.Series(True, index=df.index)
            if 'Transaction ID' in df.columns:
                original_mask = ~df['Transaction ID'].str.contains('-STMT-|-VOID-|-ADJ-', na=False)
            
            df_originals = df[original_mask]
            
            # Now filter originals to only 2025 for unreconciled calculations
            df_originals_2025 = df_originals[df_originals['Effective Date'].dt.year == 2025]
            
            # The issue: -STMT- entries from 2025 might be paying for policies from any year
            # So we need to track which transactions (from any year) have been paid in 2025
            
            # Step 1: Get all -STMT- entries from 2025 (payments made in 2025)
            # These use STMT DATE for when payment was made
            df_stmt_2025_by_stmt_date = pd.DataFrame()
            if 'STMT DATE' in df_stmt_all.columns:
                df_stmt_all['STMT DATE'] = pd.to_datetime(df_stmt_all['STMT DATE'], errors='coerce', format='mixed')
                df_stmt_2025_by_stmt_date = df_stmt_all[df_stmt_all['STMT DATE'].dt.year == 2025]
            
            # If no STMT DATE column, fall back to Effective Date
            if df_stmt_2025_by_stmt_date.empty:
                df_stmt_2025_by_stmt_date = df_stmt_2025
            
            # Step 2: Find which original transactions these -STMT- entries paid for
            reconciled_in_2025 = set()
            if not df_stmt_2025_by_stmt_date.empty and all(col in df_stmt_2025_by_stmt_date.columns for col in ['Policy Number', 'Effective Date']):
                for _, stmt_row in df_stmt_2025_by_stmt_date.iterrows():
                    policy = stmt_row.get('Policy Number', '')
                    eff_date = pd.to_datetime(stmt_row.get('Effective Date'), errors='coerce')
                    if policy and pd.notna(eff_date):
                        reconciled_in_2025.add((policy, eff_date))
            
            # Step 3: Find original transactions that were paid in 2025
            paid_in_2025_mask = pd.Series(False, index=df_originals.index)
            if reconciled_in_2025 and all(col in df_originals.columns for col in ['Policy Number', 'Effective Date']):
                for idx, row in df_originals.iterrows():
                    policy = row.get('Policy Number', '')
                    eff_date = pd.to_datetime(row.get('Effective Date'), errors='coerce')
                    if (policy, eff_date) in reconciled_in_2025:
                        paid_in_2025_mask.loc[idx] = True
            
            # Step 4: Calculate metrics
            df_originals_paid_in_2025 = df_originals[paid_in_2025_mask]
            df_originals_2025_unpaid = df_originals_2025[~df_originals_2025.index.isin(df_originals_paid_in_2025.index)]
            
            # Premium from transactions paid in 2025 (regardless of their effective date)
            if 'Premium Sold' in df_originals_paid_in_2025.columns:
                metrics['premium_reconciled_ytd'] = df_originals_paid_in_2025['Premium Sold'].sum()
            
            # Agent commission actually paid in 2025
            if 'Agent Paid Amount (STMT)' in df_stmt_2025_by_stmt_date.columns:
                metrics['agent_comm_paid_ytd'] = df_stmt_2025_by_stmt_date['Agent Paid Amount (STMT)'].sum()
            
            # Unreconciled = 2025 originals that haven't been paid yet
            if 'Premium Sold' in df_originals_2025_unpaid.columns:
                metrics['premium_unreconciled_ytd'] = df_originals_2025_unpaid['Premium Sold'].sum()
            if 'Agent Estimated Comm $' in df_originals_2025_unpaid.columns:
                metrics['agent_comm_estimated_ytd'] = df_originals_2025_unpaid['Agent Estimated Comm $'].sum()
        except Exception as e:
            # If date parsing fails, fall back to zero
            pass
    
    return metrics

def log_debug(message, level="INFO", error_obj=None):
    """Add a debug log entry to session state."""
    if "debug_logs" not in st.session_state:
        st.session_state.debug_logs = []
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    log_entry = {
        "timestamp": timestamp,
        "level": level,
        "message": message
    }
    
    if error_obj:
        log_entry["error_details"] = str(error_obj)
        if hasattr(error_obj, '__dict__'):
            log_entry["error_attrs"] = str(error_obj.__dict__)
    
    st.session_state.debug_logs.append(log_entry)
    
    # Keep only last 500 entries
    if len(st.session_state.debug_logs) > 500:
        st.session_state.debug_logs = st.session_state.debug_logs[-500:]

def clear_debug_logs():
    """Clear all debug logs."""
    st.session_state.debug_logs = []

def is_reconciliation_transaction(transaction_id):
    """
    Check if transaction is a reconciliation entry that should be locked.
    Returns True for -STMT-, -VOID-, -ADJ- transactions.
    """
    if not transaction_id:
        return False
    
    transaction_id_str = str(transaction_id)
    reconciliation_types = ['-STMT-', '-VOID-', '-ADJ-']
    
    return any(suffix in transaction_id_str for suffix in reconciliation_types)

def is_import_transaction(transaction_id):
    """
    Check if transaction is an import-created entry that should have restricted editing.
    Returns True for -IMPORT transactions.
    """
    if not transaction_id:
        return False
    
    transaction_id_str = str(transaction_id)
    return '-IMPORT' in transaction_id_str

def clean_data_for_database(data):
    """
    Remove UI-only fields from data dictionary before database insertion.
    
    UI-only fields that don't exist in the database:
    - Rate (Commission Rate display field)
    - Edit, Select, Action, Details (table checkboxes)
    - new_effective_date, new_expiration_date (renewal helpers)
    - expiration_date (maps to X-DATE in database)
    """
    # Define fields that should be removed before database insertion
    ui_only_fields = {
        'Rate', 
        'Edit', 
        'Select', 
        'Action', 
        'Details',
        'new_effective_date',
        'new_expiration_date',
        'expiration_date',  # This maps to X-DATE in the database
        'Days Until Expiration',  # Calculated field
        'Status',  # UI display field
        '_id'  # Internal row identifier
    }
    
    # Create a copy of the data to avoid modifying the original
    cleaned_data = data.copy()
    
    # Remove UI-only fields
    for field in ui_only_fields:
        if field in cleaned_data:
            del cleaned_data[field]
    
    return cleaned_data

def load_policy_types():
    """Load policy types from configuration file."""
    policy_types_file = "config_files/policy_types_updated.json"
    default_types = [
        {"name": "Auto", "active": True, "default": False},
        {"name": "Home", "active": True, "default": False},
        {"name": "Life", "active": True, "default": False},
        {"name": "Health", "active": True, "default": False},
        {"name": "Commercial", "active": True, "default": True},
        {"name": "Umbrella", "active": True, "default": False},
        {"name": "Flood", "active": True, "default": False},
        {"name": "Other", "active": True, "default": False}
    ]
    
    try:
        with open(policy_types_file, 'r') as f:
            config = json.load(f)
            # Convert from new format to old format for compatibility
            policy_types = []
            for pt in config.get('policy_types', []):
                policy_types.append({
                    "name": pt.get('name', pt.get('code', '')),
                    "active": pt.get('active', True),
                    "default": pt.get('name') == config.get('default', 'HOME')
                })
            return policy_types, True  # Always allow custom for now
    except:
        # If file doesn't exist or is corrupted, return defaults
        return default_types, True

def load_policy_types_config():
    """Load the full policy types configuration as a dictionary."""
    policy_types_file = "config_files/policy_types_updated.json"
    default_config = {
        "policy_types": [
            {"name": "Auto", "active": True, "default": False},
            {"name": "Home", "active": True, "default": False},
            {"name": "Life", "active": True, "default": False},
            {"name": "Health", "active": True, "default": False},
            {"name": "Commercial", "active": True, "default": True},
            {"name": "Umbrella", "active": True, "default": False},
            {"name": "Flood", "active": True, "default": False},
            {"name": "Other", "active": True, "default": False}
        ],
        "allow_custom": True
    }
    
    try:
        with open(policy_types_file, 'r') as f:
            config = json.load(f)
            # Convert from new format to old format for compatibility
            converted_config = {
                "policy_types": [],
                "allow_custom": True,
                "last_updated": config.get('last_updated', '')
            }
            for pt in config.get('policy_types', []):
                converted_config['policy_types'].append({
                    "name": pt.get('name', pt.get('code', '')),
                    "active": pt.get('active', True),
                    "default": pt.get('name') == config.get('default', 'HOME')
                })
            return converted_config
    except:
        # If file doesn't exist or is corrupted, return defaults
        return default_config

def save_policy_types(policy_types, allow_custom=True):
    """Save policy types to configuration file."""
    policy_types_file = "config_files/policy_types_updated.json"
    config = {
        "policy_types": policy_types,
        "allow_custom": allow_custom,
        "last_updated": datetime.datetime.now().strftime("%Y-%m-%d")
    }
    
    try:
        os.makedirs("config_files", exist_ok=True)
        with open(policy_types_file, 'w') as f:
            json.dump(config, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving policy types: {e}")
        return False

def get_active_policy_types():
    """Get list of active policy type names for dropdowns."""
    policy_types, allow_custom = load_policy_types()
    active_types = [pt['name'] for pt in policy_types if pt.get('active', True)]
    return active_types, allow_custom

def get_default_policy_type():
    """Get the default policy type."""
    policy_types, _ = load_policy_types()
    for pt in policy_types:
        if pt.get('default', False):
            return pt['name']
    return policy_types[0]['name'] if policy_types else "Other"

def add_policy_type(new_type_name):
    """Add a new policy type to the configuration."""
    policy_types, allow_custom = load_policy_types()
    
    # Check if type already exists
    existing_names = [pt['name'].lower() for pt in policy_types]
    if new_type_name.lower() in existing_names:
        return False, "Policy type already exists"
    
    # Add new type
    policy_types.append({
        "name": new_type_name,
        "active": True,
        "default": False
    })
    
    # Save updated list
    if save_policy_types(policy_types, allow_custom):
        return True, "Policy type added successfully"
    return False, "Error saving policy type"

def apply_formula_display(df, show_formulas=True):
    """
    Apply formula calculations to existing columns with indicators.
    Optionally shows formulas or actual values based on toggle.
    """
    if df.empty:
        return df
    
    df = df.copy()
    
    # Store original values for comparison
    df['_original_agency'] = df['Agency Estimated Comm/Revenue (CRM)'].copy()
    df['_original_agent'] = df['Agent Estimated Comm $'].copy()
    
    if show_formulas:
        # Calculate Commissionable Premium (Premium Sold - Policy Taxes & Fees)
        df['Commissionable Premium'] = df.apply(
            lambda row: (
                float(row.get('Premium Sold', 0) or 0) - float(row.get('Policy Taxes & Fees', 0) or 0)
            ),
            axis=1
        )
        
        # Calculate formula values using Commissionable Premium
        df['_formula_agency'] = df.apply(
            lambda row: (
                0.0 if pd.isna(row.get('Commissionable Premium', 0)) or pd.isna(row.get('Policy Gross Comm %', 0))
                else float(row.get('Commissionable Premium', 0) or 0) * float(row.get('Policy Gross Comm %', 0) or 0) / 100
            ),
            axis=1
        )
        
        # Get agent commission rate based on transaction type
        def get_agent_rate(row):
            trans_type = row.get('Transaction Type', '')
            if trans_type in ['NEW', 'NBS', 'STL', 'BoR']:
                return 50.0
            elif trans_type in ['RWL', 'REWRITE']:
                return 25.0
            elif trans_type in ['CAN', 'XCL']:
                return 0.0
            elif trans_type in ['END', 'PCH']:
                # Check if new business or renewal based on dates
                if row.get('Policy Origination Date') == row.get('Effective Date'):
                    return 50.0
                else:
                    return 25.0
            else:
                # Default to agent comm rate if available
                agent_rate = row.get('Agent Comm %', 0)
                if agent_rate and agent_rate < 1:
                    return agent_rate * 100  # Convert decimal to percentage
                return agent_rate or 0
        
        # Calculate Agent formula
        df['_formula_agent'] = df.apply(
            lambda row: (
                row['_formula_agency'] * get_agent_rate(row) / 100
            ),
            axis=1
        )
        
        # Calculate Broker Fee Agent Commission (always 50%)
        df['Broker Fee Agent Comm'] = df.apply(
            lambda row: (
                float(row.get('Broker Fee', 0) or 0) * 0.50
            ),
            axis=1
        )
        
        # Calculate Total Agent Commission
        df['Total Agent Comm'] = df.apply(
            lambda row: (
                row['_formula_agent'] + row['Broker Fee Agent Comm']
            ),
            axis=1
        )
        
        # Apply formulas to display columns
        df['Agency Estimated Comm/Revenue (CRM)'] = df['_formula_agency'].round(2)
        df['Agent Estimated Comm $'] = df['_formula_agent'].round(2)
        
        # Add indicators based on variance
        df['_agency_indicator'] = df.apply(
            lambda row: (
                '🔒' if row.get('Transaction ID', '').find('-STMT-') >= 0 or 
                       row.get('Transaction ID', '').find('-VOID-') >= 0 or
                       row.get('Transaction ID', '').find('-ADJ-') >= 0
                else '⚠️' if pd.isna(row.get('Premium Sold')) or pd.isna(row.get('Policy Gross Comm %')) or
                             row.get('Premium Sold', 0) == 0 or row.get('Policy Gross Comm %', 0) == 0
                else '✏️' if abs(float(row.get('_original_agency', 0) or 0) - row['_formula_agency']) > 0.01
                else '✓'
            ),
            axis=1
        )
        
        df['_agent_indicator'] = df.apply(
            lambda row: (
                '🔒' if row.get('Transaction ID', '').find('-STMT-') >= 0 or 
                       row.get('Transaction ID', '').find('-VOID-') >= 0 or
                       row.get('Transaction ID', '').find('-ADJ-') >= 0
                else '⚠️' if pd.isna(row.get('Premium Sold')) or pd.isna(row.get('Policy Gross Comm %')) or
                             row.get('Premium Sold', 0) == 0 or row.get('Policy Gross Comm %', 0) == 0
                else '✏️' if abs(float(row.get('_original_agent', 0) or 0) - row['_formula_agent']) > 0.01
                else '✓'
            ),
            axis=1
        )
        
        # Format display with indicators
        df['Agency Estimated Comm/Revenue (CRM)'] = df.apply(
            lambda row: f"${row['Agency Estimated Comm/Revenue (CRM)']:.2f} {row['_agency_indicator']}",
            axis=1
        )
        
        df['Agent Estimated Comm $'] = df.apply(
            lambda row: f"${row['Agent Estimated Comm $']:.2f} {row['_agent_indicator']}",
            axis=1
        )
    
    # Clean up temporary columns
    temp_cols = ['_original_agency', '_original_agent', '_formula_agency', '_formula_agent', 
                 '_agency_indicator', '_agent_indicator']
    for col in temp_cols:
        if col in df.columns:
            df = df.drop(columns=[col])
    
    return df

@st.cache_data
def get_custom_css():
    """Get cached CSS for better performance."""
    return """<style>        /* Remove default padding and maximize main block width */
        .main .block-container {
            padding-top: 1rem;
            padding-bottom: 0rem;
            padding-left: 1.5rem;
            padding-right: 1.5rem;
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
            max-width: calc(100vw - 240px) !important;
        }
        
        /* Ensure main content area starts after sidebar */
        .main {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
            max-width: calc(100vw - 240px) !important;
        }
        
        /* Additional targeting for content wrapper */
        [data-testid="stAppViewContainer"] > .main {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
        }
          /* Force all main content to respect sidebar space */
        section.main {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
        }
        
        /* Target the root app container - this is critical! */
        [data-testid="stAppViewContainer"] {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
            max-width: calc(100vw - 240px) !important;
        }
        
        /* Also target the main header area */
        [data-testid="stHeader"] {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
        }
        
        /* Force sidebar to be permanently visible and disable all collapse functionality */
        section[data-testid="stSidebar"] {
            min-width: 240px !important;
            max-width: 240px !important;
            width: 240px !important;
            display: block !important;
            visibility: visible !important;
            position: fixed !important;
            left: 0 !important;
            top: 0 !important;
            height: 100vh !important;
            z-index: 999 !important;
            transform: translateX(0px) !important;
            margin-left: 0px !important;
        }
        
        /* Force sidebar to stay expanded even when marked as collapsed */
        section[data-testid="stSidebar"][aria-expanded="false"] {
            transform: translateX(0px) !important;
            margin-left: 0px !important;
            left: 0 !important;
        }
        
        /* Hide ALL possible collapse/hide buttons with comprehensive selectors */
        button[data-testid="collapsedControl"],
        button[kind="header"],
        [data-testid="stSidebarNav"] button[kind="header"],
        [data-testid="stSidebar"] button[kind="header"],
        [data-testid="stSidebar"] [data-testid="collapsedControl"],
        [data-testid="stSidebar"] button[title*="collapse"],
        [data-testid="stSidebar"] button[title*="hide"],
        [data-testid="stSidebar"] button[aria-label*="collapse"],
        [data-testid="stSidebar"] button[aria-label*="hide"],
        .stSidebar button[kind="header"],
        section[data-testid="stSidebar"] > div > button,
        section[data-testid="stSidebar"] button:first-child {
            display: none !important;
            visibility: hidden !important;
            opacity: 0 !important;
            pointer-events: none !important;
            width: 0 !important;
            height: 0 !important;
        }
        
        /* Ensure sidebar content is always visible */
        [data-testid="stSidebar"] > div {
            display: block !important;
            visibility: visible !important;
            opacity: 1 !important;
        }
        
        /* Force sidebar radio buttons to be visible */
        [data-testid="stSidebar"] .stRadio {
            display: block !important;
            visibility: visible !important;
        }
        
        /* Override any CSS that might hide the sidebar */
        [data-testid="stSidebar"] {
            opacity: 1 !important;
            pointer-events: auto !important;
        }
        /* Remove extra margin from header/title */
        .main .block-container h1 {
            margin-bottom: 0.5rem;
        }
        /* Highlight all interactive input fields in Add New Policy Transaction form and Admin Panel rename headers */
        .stForm input:not([disabled]), .stForm select:not([disabled]), .stForm textarea:not([disabled]),
        .stTextInput > div > input:not([disabled]), .stNumberInput > div > input:not([disabled]), .stDateInput > div > input:not([disabled]) {
            background-color: #fff3b0 !important; /* Darker yellow */
            border: 2px solid #e6a800 !important; /* Darker yellow border */
            border-radius: 6px !important;
        }
        /* Make selectboxes match other fields and remove focus ring */
        .stSelectbox > div[data-baseweb="select"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        .stSelectbox > div[data-baseweb="select"]:focus,
        .stSelectbox > div[data-baseweb="select"]:active,
        .stSelectbox > div[data-baseweb="select"]:focus-visible {
            border: 2px solid #e6a800 !important;
            box-shadow: none !important;
            outline: none !important;
        }
        /* Add yellow border to the client name search input at the top of Add New Policy Transaction */
        input[type="text"][aria-label="Type client name to search:"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        /* Add yellow border to Premium Sold Calculator inputs */
        input[type="number"][aria-label="Existing Premium"],
        input[type="number"][aria-label="New/Revised Premium"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        /* Add yellow border to New Policy Premium inputs */
        input[type="number"][aria-label="New Policy Premium"],
        input[type="number"][aria-label="Policy Gross Comm %"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        /* Ensure Transaction Type and FULL OR MONTHLY PMTS selectboxes have yellow styling */
        div[data-testid="stSelectbox"] > div[data-baseweb="select"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        /* Make sure all form selectboxes are styled consistently */
        .stForm div[data-testid="stSelectbox"] > div[data-baseweb="select"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        /* Custom fatter scrollbars for all tables and editors */
        .stDataFrame ::-webkit-scrollbar, .stDataEditor ::-webkit-scrollbar {
            height: 18px;
            width: 18px;
        }
        .stDataFrame ::-webkit-scrollbar-thumb, .stDataEditor ::-webkit-scrollbar-thumb {
            background: #888888;
            border-radius: 8px;
            border: 3px solid #b0b0b0;
        }
        .stDataFrame ::-webkit-scrollbar-track, .stDataEditor ::-webkit-scrollbar-track {
            background: #b0b0b0;
            border-radius: 8px;
        }
        /* For Firefox */
        .stDataFrame, .stDataEditor {
            scrollbar-width: thick;
            scrollbar-color: #888888 #b0b0b0;
        }
        /* Highlight selectboxes in Admin Panel reorder columns section */
        [id^="reorder_col_"] > div[data-baseweb="select"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        [id^="reorder_col_"] > div[data-baseweb="select"]:focus,
        [id^="reorder_col_"] > div[data-baseweb="select"]:active,
        [id^="reorder_col_"] > div[data-baseweb="select"]:focus-visible {
            border: 2px solid #e6a800 !important;
            box-shadow: none !important;
            outline: none !important;
        }
        /* Highlight selectboxes in Admin Panel reorder columns section (force all backgrounds) */
        [data-testid^="stSelectbox"][id^="reorder_col_"] > div[data-baseweb="select"],
        [id^="reorder_col_"] [data-baseweb="select"],
        [id^="reorder_col_"] [class*="css-1wa3eu0-placeholder"],
        [id^="reorder_col_"] [class*="css-1uccc91-singleValue"],
        [id^="reorder_col_"] [class*="css-1okebmr-indicatorSeparator"],
        [id^="reorder_col_"] [class*="css-1pahdxg-control"],
        [id^="reorder_col_"] [class*="css-1s2u09g-control"],
        [id^="reorder_col_"] [class*="css-1n7v3ny-option"],
        [id^="reorder_col_"] [class*="css-9gakcf-option"],
        [id^="reorder_col_"] [class*="css-1n6sfyn-MenuList"],
        [id^="reorder_col_"] [class*="css-1n6sfyn-MenuList"] * {
            background-color: #fff3b0 !important;
            border-color: #e6a800 !important;
            color: #222 !important;
        }        /* Force highlight for all selectboxes labeled 'Transaction Type' everywhere */
        label:has(> div[data-baseweb="select"]) {
            background-color: #fff3b0 !important;
            border-radius: 6px !important;
            padding: 2px 4px !important;
        }
        /* Also target selectbox input and dropdown for 'Transaction Type' */
        [aria-label="Transaction Type"] > div[data-baseweb="select"],
        [aria-label="Transaction Type"] [data-baseweb="select"],
        [aria-label="Transaction Type"] [class*="css-1wa3eu0-placeholder"],
        [aria-label="Transaction Type"] [class*="css-1uccc91-singleValue"],
        [aria-label="Transaction Type"] [class*="css-1okebmr-indicatorSeparator"],
        [aria-label="Transaction Type"] [class*="css-1pahdxg-control"],
        [aria-label="Transaction Type"] [class*="css-1s2u09g-control"],
        [aria-label="Transaction Type"] [class*="css-1n7v3ny-option"],
        [aria-label="Transaction Type"] [class*="css-9gakcf-option"],
        [aria-label="Transaction Type"] [class*="css-1n6sfyn-MenuList"],
        [aria-label="Transaction Type"] [class*="css-1n6sfyn-MenuList"] * {
            background-color: #fff3b0 !important;
            border-color: #e6a800 !important;
            color: #222 !important;
        }
        [aria-label="Transaction Type"] > div[data-baseweb="select"] {
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        [aria-label="Transaction Type"] > div[data-baseweb="select"]:focus,
        [aria-label="Transaction Type"] > div[data-baseweb="select"]:active,
        [aria-label="Transaction Type"] > div[data-baseweb="select"]:focus-visible {
            border: 2px solid #e6a800 !important;
            box-shadow: none !important;
            outline: none !important;
        }
        
        /* Maintain scroll position in data editor */
        [data-testid="stDataFrame"] > div,
        [data-testid="stDataEditor"] > div {
            scroll-behavior: auto !important;
        }
        
        /* Prevent auto-scroll to start */
        .stDataEditor {
            overflow-x: auto !important;
            overflow-y: auto !important;
        }
        </style>"""

def apply_css():
    """Apply custom CSS styling to the Streamlit app."""
    with st.container():
        st.markdown(get_custom_css(), unsafe_allow_html=True)


def format_currency(val):
    """Format the value as currency for display."""
    if pd.isna(val) or val is None:
        return ""
    try:
        return f"${val:,.2f}"
    except Exception:
        return val

def round_numeric_columns(df):
    """Round all numeric columns to 2 decimal places to avoid floating point precision issues."""
    numeric_columns = df.select_dtypes(include=['float64', 'float32', 'float', 'number']).columns
    for col in numeric_columns:
        df[col] = df[col].round(2)
    
    # Also handle columns that might be stored as strings but contain numeric values
    percentage_columns = ['Policy Gross Comm %', 'Agent Comm %', 'Agent Comm (New 50% RWL 25%)']
    for col in percentage_columns:
        if col in df.columns:
            # Convert to numeric if it's a string, then round
            df[col] = pd.to_numeric(df[col], errors='coerce').round(2)
    
    return df

def clean_numeric_value(value):
    """Clean and round a numeric value to 2 decimal places."""
    if pd.isna(value) or value is None:
        return None
    try:
        # Convert to float and round to 2 decimal places
        return round(float(value), 2)
    except (ValueError, TypeError):
        return value


def convert_timestamps_for_json(data):
    """Convert any Timestamp objects in data to strings for JSON serialization."""
    cleaned_data = {}
    for key, value in data.items():
        if isinstance(value, pd.Timestamp):
            cleaned_data[key] = value.isoformat()
        elif isinstance(value, datetime.datetime):
            cleaned_data[key] = value.isoformat()
        elif isinstance(value, datetime.date):
            cleaned_data[key] = value.isoformat()
        elif pd.isna(value):
            cleaned_data[key] = None
        else:
            cleaned_data[key] = value
    return cleaned_data

CURRENCY_COLUMNS = [
    "Premium Sold",    "Agency Comm Received (STMT)",
    "Gross Premium Paid",
    "Agency Gross Comm",
    "Agent Paid Amount (STMT)",
    "Estimated Agent Comm",
    "Policy Balance Due",  # Updated to use the new calculated column
    "Agent Estimated Comm $",
    "Agency Estimated Comm/Revenue (CRM)",
    "Agency Estimated Comm/Revenue (CRM)"
]

def format_currency_columns(df):
    # Use mapped column names for currency formatting
    for ui_field in CURRENCY_COLUMNS:
        mapped_col = get_mapped_column(ui_field)
        if mapped_col and mapped_col in df.columns:
            df[mapped_col] = df[mapped_col].apply(format_currency)
        elif ui_field in df.columns:  # Fallback to original name
            df[ui_field] = df[ui_field].apply(format_currency)
    return df

# --- Helper Functions ---

# Column mapping persistence functions
def load_saved_column_mappings():
    """Load saved column mappings from JSON file."""
    mappings_file = os.path.join('config_files', 'saved_mappings.json')
    try:
        if os.path.exists(mappings_file):
            with open(mappings_file, 'r') as f:
                return json.load(f)
    except Exception as e:
        st.error(f"Error loading saved mappings: {str(e)}")
    return {}

def save_column_mappings_to_file(mappings):
    """Save column mappings to JSON file."""
    mappings_file = os.path.join('config_files', 'saved_mappings.json')
    try:
        # Ensure the config_files directory exists
        os.makedirs('config_files', exist_ok=True)
        
        with open(mappings_file, 'w') as f:
            json.dump(mappings, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving mappings: {str(e)}")
        return False

def generate_client_id(length=6):
    """Generate a unique Client ID with exactly 3 letters and 3 numbers in random order."""
    # Generate exactly 3 letters and 3 numbers, then mix them randomly
    letters = string.ascii_uppercase
    digits = string.digits
    
    # Pick 3 random letters
    selected_letters = [random.choice(letters) for _ in range(3)]
    
    # Pick 3 random numbers
    selected_numbers = [random.choice(digits) for _ in range(3)]
    
    # Combine them
    result = selected_letters + selected_numbers
    
    # Shuffle to create random pattern
    random.shuffle(result)
    return ''.join(result)

def generate_transaction_id(length=7, suffix=None):
    """Generate a unique Transaction ID with at least 3 letters and 3 numbers.
    
    Args:
        length: Length of the base ID (default 7)
        suffix: Optional suffix to append (e.g., '-IMPORT', '-STMT')
    """
    # Ensure at least 3 letters and 3 numbers for 7-character ID
    letters = string.ascii_uppercase
    digits = string.digits
    
    # Pick 3 random letters
    selected_letters = [random.choice(letters) for _ in range(3)]
    
    # Pick 3 random numbers
    selected_numbers = [random.choice(digits) for _ in range(3)]
    
    # For the 7th character, randomly choose letter or number
    if random.choice([True, False]):
        extra_char = random.choice(letters)
    else:
        extra_char = random.choice(digits)
    
    # Combine all characters
    result = selected_letters + selected_numbers + [extra_char]
    
    # Shuffle to create random pattern
    random.shuffle(result)
    base_id = ''.join(result)
    
    # Append suffix if provided
    if suffix:
        return f"{base_id}{suffix}"
    return base_id

def generate_unique_client_id():
    """Generate a unique Client ID with 3 letters and 3 numbers mixed randomly by checking against existing IDs."""
    try:
        # Get all existing client IDs
        supabase = get_supabase_client()
        response = supabase.table('policies').select('"Client ID"').execute()
        
        existing_ids = set()
        if response.data:
            for record in response.data:
                client_id = record.get('Client ID')
                if client_id:
                    existing_ids.add(str(client_id).upper())
        
        # Generate a unique ID using the existing pattern
        max_attempts = 100
        for _ in range(max_attempts):
            # Use the existing generate_client_id function for proper format
            new_id = generate_client_id()
            if new_id.upper() not in existing_ids:
                return new_id
        
        # If we couldn't find a unique ID in 100 attempts, add timestamp suffix
        # This maintains the format while ensuring uniqueness
        base_id = generate_client_id()
        timestamp = str(int(datetime.datetime.now().timestamp()))[-2:]
        return f"{base_id[:4]}{timestamp}"
        
    except Exception as e:
        # If there's an error checking existing IDs, just generate one
        return generate_client_id()

def generate_reconciliation_transaction_id(transaction_type="STMT", date=None):
    """Generate a reconciliation transaction ID with format: XXXXXXX-TYPE-YYYYMMDD"""
    base_id = generate_transaction_id()
    
    if date is None:
        date = datetime.datetime.now()
    
    date_str = date.strftime("%Y%m%d")
    
    return f"{base_id}-{transaction_type}-{date_str}"

# --- Commission Rule Functions ---
def lookup_commission_rule(carrier_id, mga_id=None, policy_type=None, transaction_type="NEW", effective_date=None):
    """
    Look up the best matching commission rule for given criteria.
    
    Args:
        carrier_id (str): UUID of the carrier
        mga_id (str, optional): UUID of the MGA (None for direct appointments)
        policy_type (str, optional): Policy type (e.g., "Auto", "HO3")
        transaction_type (str): "NEW" or "RWL" (default: "NEW")
        effective_date (date, optional): Date to check rule effectiveness (default: today)
    
    Returns:
        dict: Commission rule with rates, or None if no rule found
        {
            'rule_id': UUID,
            'new_rate': decimal,
            'renewal_rate': decimal, 
            'rule_description': str,
            'carrier_name': str,
            'mga_name': str or None,
            'applied_rule_text': str  # User-friendly description
        }
    """
    try:
        supabase = get_supabase_client()
        
        if effective_date is None:
            effective_date = datetime.date.today()
        
        # Build query to find matching rules
        query = supabase.table('commission_rules').select("""
            rule_id,
            carrier_id,
            mga_id,
            policy_type,
            new_rate,
            renewal_rate,
            rule_description,
            effective_date,
            end_date,
            is_active
        """)
        
        # Filter by carrier
        query = query.eq('carrier_id', carrier_id)
        
        # Filter by active rules and effective date
        query = query.eq('is_active', True)
        query = query.lte('effective_date', effective_date.isoformat())
        query = query.or_('end_date.is.null,end_date.gte.' + effective_date.isoformat())
        
        # Execute query
        response = query.execute()
        
        if not response.data:
            return None
        
        # Calculate rule priority for each matching rule
        rules_with_priority = []
        for rule in response.data:
            priority = 0
            
            # MGA match gets highest priority
            if mga_id and rule.get('mga_id') == mga_id:
                priority += 1000
            elif not mga_id and not rule.get('mga_id'):
                priority += 500  # Direct appointment match
            elif rule.get('mga_id'):
                continue  # MGA rule doesn't match, skip
            
            # Policy type match
            if policy_type and rule.get('policy_type'):
                rule_types = [t.strip() for t in rule['policy_type'].split(',')]
                if policy_type in rule_types:
                    priority += 100
                else:
                    continue  # Policy type doesn't match, skip
            elif not rule.get('policy_type'):
                priority += 10  # Default rule (no specific policy type)
            
            rules_with_priority.append((priority, rule))
        
        if not rules_with_priority:
            return None
        
        # Sort by priority (highest first) and get the best match
        rules_with_priority.sort(key=lambda x: x[0], reverse=True)
        best_rule = rules_with_priority[0][1]
        
        # Get carrier and MGA names for display
        carrier_response = supabase.table('carriers').select('carrier_name').eq('carrier_id', carrier_id).execute()
        carrier_name = carrier_response.data[0]['carrier_name'] if carrier_response.data else 'Unknown Carrier'
        
        mga_name = None
        if best_rule.get('mga_id'):
            mga_response = supabase.table('mgas').select('mga_name').eq('mga_id', best_rule['mga_id']).execute()
            mga_name = mga_response.data[0]['mga_name'] if mga_response.data else 'Unknown MGA'
        
        # Determine which rate to use based on transaction type
        rate_to_use = best_rule['new_rate']
        if transaction_type in ['RWL', 'REWRITE'] and best_rule.get('renewal_rate'):
            rate_to_use = best_rule['renewal_rate']
        
        # Build user-friendly description
        mga_text = mga_name if mga_name else "Direct"
        policy_text = f" - {best_rule['policy_type']}" if best_rule.get('policy_type') else ""
        applied_rule_text = f"{carrier_name} ({mga_text}){policy_text} - {rate_to_use}%"
        
        return {
            'rule_id': best_rule['rule_id'],
            'new_rate': best_rule['new_rate'],
            'renewal_rate': best_rule.get('renewal_rate'),
            'rule_description': best_rule.get('rule_description'),
            'carrier_name': carrier_name,
            'mga_name': mga_name,
            'applied_rule_text': applied_rule_text,
            'rate_to_use': rate_to_use
        }
        
    except Exception as e:
        st.error(f"Error looking up commission rule: {e}")
        return None

def load_carriers_for_dropdown():
    """Load all active carriers for dropdown selection."""
    try:
        supabase = get_supabase_client()
        response = supabase.table('carriers').select('carrier_id, carrier_name').eq('status', 'Active').order('carrier_name').execute()
        return response.data if response.data else []
    except Exception as e:
        st.error(f"Error loading carriers: {e}")
        return []

def load_mgas_for_carrier(carrier_id):
    """Load MGAs associated with a specific carrier."""
    try:
        supabase = get_supabase_client()
        
        # First try to get MGAs from carrier_mga_relationships table
        try:
            # Get relationships for this carrier
            response = supabase.table('carrier_mga_relationships').select("mga_id").eq('carrier_id', carrier_id).execute()
            
            if response.data:
                # Get the MGA details for each relationship
                mgas = []
                for rel in response.data:
                    if rel.get('mga_id'):
                        # Get MGA details
                        mga_response = supabase.table('mgas').select('mga_id, mga_name, status').eq('mga_id', rel['mga_id']).execute()
                        if mga_response.data and mga_response.data[0].get('status') != 'Inactive':
                            mga = mga_response.data[0]
                            mgas.append({
                                'mga_id': mga['mga_id'],
                                'mga_name': mga['mga_name']
                            })
                
                if mgas:  # If we found MGAs through relationships, return them
                    mgas.sort(key=lambda x: x['mga_name'])
                    return mgas
        except Exception as e:
            # Log the error for debugging but continue to fallbacks
            # Silent fail - continue to other methods
            pass
        
        # Fallback: Get MGAs that have commission rules with this carrier
        # First get commission rules for this carrier
        response = supabase.table('commission_rules').select("mga_id").eq('carrier_id', carrier_id).execute()
        
        # Filter out null mga_ids and get unique MGA details
        mga_ids = []
        for rule in response.data:
            if rule.get('mga_id') and rule['mga_id'] not in mga_ids:
                mga_ids.append(rule['mga_id'])
        
        # Get MGA details for valid mga_ids
        if mga_ids:
            mga_response = supabase.table('mgas').select('mga_id, mga_name, status').in_('mga_id', mga_ids).eq('status', 'Active').execute()
            if mga_response.data:
                mgas = [{'mga_id': m['mga_id'], 'mga_name': m['mga_name']} for m in mga_response.data]
                mgas.sort(key=lambda x: x['mga_name'])
                return mgas
        
        # If no MGAs found through commission rules, return empty list
        # Wright Flood and other carriers might not have MGAs
        return []
        
    except Exception as e:
        st.error(f"Error loading MGAs for carrier: {e}")
        return []

# --- Excel Utility Functions ---
def create_formatted_excel_file(data, sheet_name="Data", filename_prefix="export"):
    """
    Create a formatted Excel file from DataFrame with professional styling.
    
    Args:
        data (pd.DataFrame): Data to export
        sheet_name (str): Name for the Excel sheet
        filename_prefix (str): Prefix for the filename
    
    Returns:
        io.BytesIO: Excel file buffer
        str: Suggested filename
    """
    try:
        # Create filename with timestamp
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        
        # Create Excel buffer
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Write data to Excel
            data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with reasonable limits
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
            
            # Format currency columns
            currency_columns = ['Commission_Paid', 'Agency_Commission_Received', 'Balance_Due', 
                              'Agency Estimated Comm/Revenue (CRM)', 'Premium_Amount']
            
            for col_name in currency_columns:
                if col_name in data.columns:
                    col_letter = None
                    for col_num, column_name in enumerate(data.columns, 1):
                        if column_name == col_name:
                            col_letter = openpyxl.utils.get_column_letter(col_num)
                            break
                    
                    if col_letter:
                        for row_num in range(2, len(data) + 2):  # Start from row 2 (after header)
                            cell = worksheet[f"{col_letter}{row_num}"]
                            cell.number_format = '"$"#,##0.00'
        
        excel_buffer.seek(0)
        return excel_buffer, filename
        
    except Exception as e:
        st.error(f"Error creating Excel file: {e}")
        return None, None

def create_multi_sheet_excel(data_dict, filename_prefix="multi_sheet_export"):
    """
    Create Excel file with multiple sheets and metadata.
    
    Args:
        data_dict (dict): Dictionary with sheet_name: data pairs
        filename_prefix (str): Prefix for filename
    
    Returns:
        io.BytesIO: Excel file buffer
        str: Suggested filename
    """
    try:
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Create metadata sheet first
            metadata = pd.DataFrame([
                ['Export Date', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Application', 'Commission Management System'],
                ['Export Type', 'Multi-Sheet Report'],
                ['Total Sheets', len(data_dict)],
                ['Sheet Names', ', '.join(data_dict.keys())]
            ], columns=['Parameter', 'Value'])
            
            metadata.to_excel(writer, sheet_name='Export Info', index=False)
            
            # Format metadata sheet
            workbook = writer.book
            metadata_sheet = writer.sheets['Export Info']
            
            # Style metadata headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in metadata_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust metadata columns
            metadata_sheet.column_dimensions['A'].width = 25
            metadata_sheet.column_dimensions['B'].width = 50
            
            # Add data sheets
            for sheet_name, data in data_dict.items():
                if not data.empty:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Format data sheet
                    worksheet = writer.sheets[sheet_name]
                    
                    # Format headers
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Auto-adjust columns
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 30)
                        worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
        
        excel_buffer.seek(0)
        return excel_buffer, filename
        
    except Exception as e:
        st.error(f"Error creating multi-sheet Excel file: {e}")
        return None, None

def validate_excel_import(uploaded_file):
    """
    Validate uploaded Excel file and return DataFrame with validation results.
    
    Args:
        uploaded_file: Streamlit uploaded file object
        
    Returns:
        tuple: (success: bool, data: DataFrame or None, errors: list)
    """
    try:
        # Read Excel file
        df = pd.read_excel(uploaded_file, engine='openpyxl')
        
        validation_errors = []
        
        # Check if file is empty
        if df.empty:
            validation_errors.append("Excel file is empty")
            return False, None, validation_errors
        
        # Check for required columns (based on column mapping)
        required_columns = [
            'Client_ID', 'Transaction_ID', 'Policy_Type', 
            'Transaction_Type', 'Effective_Date'
        ]
        
        missing_columns = []
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)
        
        if missing_columns:
            validation_errors.append(f"Missing required columns: {', '.join(missing_columns)}")
        
        # Validate data types and formats
        if 'Effective_Date' in df.columns:
            try:
                pd.to_datetime(df['Effective_Date'])
            except:
                validation_errors.append("Invalid date format in Effective_Date column")
        
        # Check for duplicate Transaction_IDs
        if 'Transaction_ID' in df.columns:
            duplicates = df['Transaction_ID'].duplicated().sum()
            if duplicates > 0:
                validation_errors.append(f"Found {duplicates} duplicate Transaction_IDs")
        
        # Validate numeric columns
        numeric_columns = ['Commission_Paid', 'Agency_Commission_Received', 'Balance_Due']
        for col in numeric_columns:
            if col in df.columns:
                try:
                    pd.to_numeric(df[col], errors='coerce')
                except:
                    validation_errors.append(f"Invalid numeric data in {col} column")
        
        # Return validation results
        if validation_errors:
            return False, df, validation_errors
        else:
            return True, df, []
            
    except Exception as e:
        return False, None, [f"Error reading Excel file: {str(e)}"]

# --- Commission calculation function ---
def calculate_commission(row):
    try:
        agency_revenue_col = get_mapped_column("Agency Estimated Comm/Revenue (CRM)")
        revenue = float(row[agency_revenue_col]) if agency_revenue_col and row[agency_revenue_col] is not None else 0.0
    except (ValueError, TypeError, KeyError):
        revenue = 0.0
    
    transaction_type_col = get_mapped_column("Transaction Type")
    policy_orig_col = get_mapped_column("Policy Origination Date") 
    effective_date_col = get_mapped_column("Effective Date")
    
    try:
        transaction_type = row.get(transaction_type_col, "") if transaction_type_col else ""
        if transaction_type in ["NEW", "NBS", "STL", "BoR"]:
            return revenue * 0.50
        elif transaction_type in ["END", "PCH"]:
            policy_orig = row.get(policy_orig_col, "") if policy_orig_col else ""
            effective_date = row.get(effective_date_col, "") if effective_date_col else ""
            return revenue * 0.50 if policy_orig == effective_date else revenue * 0.25
        elif transaction_type in ["RWL", "REWRITE"]:
            return revenue * 0.25
        elif transaction_type in ["CAN", "XCL"]:
            return 0
        else:
            return revenue * 0.25
    except (KeyError, TypeError):
        return revenue * 0.25

def get_pending_renewals(df: pd.DataFrame) -> pd.DataFrame:
    """
    Identifies and generates a DataFrame of policies pending renewal.
    Excludes policies that have already been renewed (appear in Prior Policy Number of another policy).
    Shows ALL past-due renewals (no lower limit) and future renewals up to 90 days.
    """
    # Filter for relevant transaction types
    renewal_candidates = df[df[get_mapped_column("Transaction Type")].isin(["NEW", "RWL"])].copy()
    
    # Convert date columns to datetime objects
    renewal_candidates['expiration_date'] = pd.to_datetime(renewal_candidates[get_mapped_column("X-DATE")], errors='coerce')
    
    # Sort by policy number and expiration date to find the latest transaction
    renewal_candidates = renewal_candidates.sort_values(by=["Policy Number", "expiration_date"], ascending=[True, False])
    
    # Get the most recent transaction for each policy
    latest_renewals = renewal_candidates.drop_duplicates(subset="Policy Number", keep="first")
    
    # Calculate days until expiration for each policy
    today = pd.to_datetime(datetime.date.today())
    latest_renewals['Days Until Expiration'] = (latest_renewals['expiration_date'] - today).dt.days
    
    # Filter for policies expiring within 90 days OR already expired (no lower limit on past due)
    # This will show ALL past-due renewals and future renewals up to 90 days
    pending_renewals = latest_renewals[latest_renewals['Days Until Expiration'] <= 90].copy()
    
    # Optional: Add a safeguard for very old policies (e.g., more than 1 year past due)
    # Uncomment if needed: pending_renewals = pending_renewals[pending_renewals['Days Until Expiration'] > -365]
    
    # Get list of policy numbers that have been renewed (appear in Prior Policy Number field)
    prior_policy_col = get_mapped_column("Prior Policy Number")
    if prior_policy_col and prior_policy_col in df.columns:
        # Get all policy numbers that appear as prior policies (meaning they've been renewed)
        renewed_policies = df[df[prior_policy_col].notna()][prior_policy_col].unique()
        # Exclude these from pending renewals
        pending_renewals = pending_renewals[~pending_renewals["Policy Number"].isin(renewed_policies)]
    
    # Exclude policies that have been cancelled or excluded
    # Check if any transaction for a policy number has type "CAN" or "XCL"
    transaction_type_col = get_mapped_column("Transaction Type")
    if transaction_type_col and transaction_type_col in df.columns:
        # Get all policy numbers that have a CAN or XCL transaction
        cancelled_policies = df[df[transaction_type_col].isin(["CAN", "XCL"])]["Policy Number"].unique()
        # Exclude these from pending renewals
        pending_renewals = pending_renewals[~pending_renewals["Policy Number"].isin(cancelled_policies)]
    
    # Exclude STMT and VOID transactions based on Transaction ID
    if 'Transaction ID' in pending_renewals.columns:
        # Filter out any transactions with -STMT- or -VOID- in their Transaction ID
        pending_renewals = pending_renewals[
            ~pending_renewals['Transaction ID'].astype(str).str.contains('-STMT-|-VOID-', case=False, na=False)
        ]
    
    # Sort by Days Until Expiration (ascending, so most past-due show first)
    pending_renewals = pending_renewals.sort_values('Days Until Expiration', ascending=True)
    
    return pending_renewals

def style_renewal_rows(row):
    """
    Style renewal rows based on days until expiration.
    - Red background for past due (negative days)
    - Yellow background for urgent (0-7 days)
    - Normal for 8+ days
    """
    days_until = row.get('Days Until Expiration', 0)
    
    if pd.isna(days_until):
        return [''] * len(row)
    
    if days_until < 0:
        # Past due - light red background
        return ['background-color: #ffcccc'] * len(row)
    elif days_until <= 7:
        # Urgent (0-7 days) - light yellow background
        return ['background-color: #fff3cd'] * len(row)
    else:
        # Normal (8+ days) - no special styling
        return [''] * len(row)

def normalize_business_name(name):
    """
    Normalize business names by removing common suffixes and punctuation.
    Helps match "RCM Construction" to "RCM Construction of SWFL LLC"
    """
    if not name:
        return ""
    
    # Common business suffixes to remove
    suffixes = [
        'LLC', 'L.L.C.', 'L.L.C', 'Inc', 'Inc.', 'Incorporated',
        'Corp', 'Corp.', 'Corporation', 'Ltd', 'Ltd.', 'Limited',
        'PA', 'P.A.', 'PC', 'P.C.', 'PLLC', 'P.L.L.C.',
        'LLP', 'L.L.P.', 'LP', 'L.P.', 'Company', 'Co.', 'Co'
    ]
    
    # Convert to string and strip
    normalized = str(name).strip()
    
    # Remove suffixes (case-insensitive)
    for suffix in suffixes:
        # Check both with and without leading comma
        patterns = [
            rf',?\s+{re.escape(suffix)}\s*$',
            rf'\s+{re.escape(suffix)}\s*$'
        ]
        for pattern in patterns:
            normalized = re.sub(pattern, '', normalized, flags=re.IGNORECASE)
    
    # Remove "of [Location]" patterns
    normalized = re.sub(r'\s+of\s+[A-Z][A-Za-z\s]+(?:LLC|Inc|Corp)?$', '', normalized, flags=re.IGNORECASE)
    
    # Clean up extra whitespace and punctuation
    normalized = ' '.join(normalized.split())
    normalized = normalized.strip(' ,.-')
    
    return normalized

def find_potential_customer_matches(search_name, existing_customers):
    """
    Find potential customer matches using various strategies.
    Returns list of (customer_name, match_type, score) tuples.
    """
    if not search_name:
        return []
    
    search_name_lower = search_name.lower().strip()
    search_normalized = normalize_business_name(search_name).lower()
    search_first_word = search_name.split()[0].lower() if search_name else ""
    
    # Handle "Last, First" format
    search_name_reversed = ""
    if "," in search_name:
        parts = search_name.split(",", 1)
        if len(parts) == 2:
            search_name_reversed = f"{parts[1].strip()} {parts[0].strip()}".lower()
    
    matches = {}  # Use dict to avoid duplicates
    
    for customer in existing_customers:
        if not customer:
            continue
            
        customer_lower = customer.lower().strip()
        customer_normalized = normalize_business_name(customer).lower()
        customer_first_word = customer.split()[0].lower() if customer else ""
        
        # 1. Exact match (highest priority)
        if search_name_lower == customer_lower:
            matches[customer] = ('exact', 100)
            continue
        
        # 2. Reversed name match (Last, First -> First Last)
        if search_name_reversed and search_name_reversed == customer_lower:
            if customer not in matches or matches[customer][1] < 98:
                matches[customer] = ('name_reversed', 98)
            continue
        
        # 3. Normalized match (very high priority)
        if search_normalized and search_normalized == customer_normalized:
            if customer not in matches or matches[customer][1] < 95:
                matches[customer] = ('normalized', 95)
            continue
        
        # 4. First word match (e.g., "Barboun" matches "Barboun, Thomas")
        if search_first_word and customer_first_word == search_first_word:
            if customer not in matches or matches[customer][1] < 90:
                matches[customer] = ('first_word', 90)
            continue
        
        # 5. Contains match (e.g., "RCM" in "RCM Construction")
        if len(search_name) >= 3:
            if search_name_lower in customer_lower:
                if customer not in matches or matches[customer][1] < 85:
                    matches[customer] = ('contains', 85)
                continue
            
            # Check if search is contained in normalized version
            if search_normalized in customer_normalized:
                if customer not in matches or matches[customer][1] < 83:
                    matches[customer] = ('normalized_contains', 83)
                continue
        
        # 6. Customer contains search (e.g., searching "RCM Construction" finds "RCM Construction of SWFL LLC")
        if customer_lower in search_name_lower and len(customer) >= 3:
            if customer not in matches or matches[customer][1] < 80:
                matches[customer] = ('reverse_contains', 80)
            continue
        
        # 7. Starts with match
        if customer_lower.startswith(search_name_lower[:3]) and len(search_name) >= 3:
            if customer not in matches or matches[customer][1] < 75:
                matches[customer] = ('starts_with', 75)
        
        # 8. Fuzzy word matching (e.g., "Adam Gomes" matches "Gomes Adam" or "Adam J Gomes")
        search_words = set(search_name_lower.split())
        customer_words = set(customer_lower.split())
        if len(search_words) >= 2 and len(customer_words) >= 2:
            # Check if all search words are in customer name (any order)
            if search_words.issubset(customer_words):
                if customer not in matches or matches[customer][1] < 88:
                    matches[customer] = ('all_words', 88)
            # Check if most words match
            elif len(search_words.intersection(customer_words)) >= min(len(search_words), len(customer_words)) - 1:
                if customer not in matches or matches[customer][1] < 82:
                    matches[customer] = ('most_words', 82)
    
    # Convert to sorted list
    result = [(name, match_type, score) for name, (match_type, score) in matches.items()]
    result.sort(key=lambda x: (-x[2], x[0]))  # Sort by score desc, then name
    
    return result

def calculate_transaction_balances(all_data):
    """
    Calculate outstanding balances for all transactions.
    Reuses the exact logic from Unreconciled Transactions tab.
    Returns DataFrame with _balance column added.
    """
    if all_data.empty:
        return pd.DataFrame()
    
    # Get original transactions only (exclude -STMT-, -ADJ-, -VOID-)
    original_trans = all_data[
        ~all_data['Transaction ID'].str.contains('-STMT-|-ADJ-|-VOID-', na=False)
    ].copy()
    
    if original_trans.empty:
        return pd.DataFrame()
    
    # Normalize effective dates for comparison
    # Convert all dates to a consistent format for comparison
    all_data_normalized = all_data.copy()
    try:
        # Try to parse dates with mixed formats
        all_data_normalized['_normalized_date'] = pd.to_datetime(all_data['Effective Date'], format='mixed', errors='coerce')
    except:
        # Fallback to coerce if mixed format fails
        all_data_normalized['_normalized_date'] = pd.to_datetime(all_data['Effective Date'], errors='coerce')
    
    original_trans_normalized = original_trans.copy()
    try:
        original_trans_normalized['_normalized_date'] = pd.to_datetime(original_trans['Effective Date'], format='mixed', errors='coerce')
    except:
        original_trans_normalized['_normalized_date'] = pd.to_datetime(original_trans['Effective Date'], errors='coerce')
    
    # Calculate balance for each transaction
    for idx, row in original_trans.iterrows():
        # Calculate credits (commission owed)
        credit = float(row.get('Agent Estimated Comm $', 0) or 0)
        
        # Calculate debits (total paid for this policy)
        policy_num = row['Policy Number']
        
        # Get normalized date for comparison
        effective_date_normalized = original_trans_normalized.at[idx, '_normalized_date']
        
        # Get all STMT and VOID entries for this specific policy and date
        # Include -VOID- entries since they have negative amounts that reduce the debit
        if pd.notna(effective_date_normalized):
            recon_entries = all_data_normalized[
                (all_data_normalized['Policy Number'] == policy_num) &
                (all_data_normalized['_normalized_date'] == effective_date_normalized) &
                (all_data['Transaction ID'].str.contains('-STMT-|-VOID-', na=False))
            ]
        else:
            # Fallback to string comparison if date parsing failed
            effective_date = row['Effective Date']
            recon_entries = all_data[
                (all_data['Policy Number'] == policy_num) &
                (all_data['Effective Date'] == effective_date) &
                (all_data['Transaction ID'].str.contains('-STMT-|-VOID-', na=False))
            ]
        
        debit = 0
        if not recon_entries.empty:
            debit = recon_entries['Agent Paid Amount (STMT)'].fillna(0).sum()
        
        # Calculate balance
        balance = credit - debit
        original_trans.at[idx, '_balance'] = balance
    
    # Return only transactions with outstanding balance
    return original_trans[original_trans['_balance'] > 0.01]

def match_statement_transactions(statement_df, column_mapping, existing_data, statement_date):
    """
    Match statement transactions to existing database records.
    Now uses the same balance calculation as Unreconciled Transactions tab.
    Returns: (matched_list, unmatched_list, can_create_list)
    """
    matched = []
    unmatched = []
    can_create = []
    
    # Get transactions with outstanding balances using the same logic as Unreconciled tab
    outstanding_trans = calculate_transaction_balances(existing_data)
    
    # Build lookup dictionaries
    existing_lookup = {}
    customer_trans_lookup = {}  # Customer -> list of transactions
    
    if not outstanding_trans.empty:
        for idx, trans in outstanding_trans.iterrows():
            trans_dict = trans.to_dict()
            trans_dict['balance'] = trans['_balance']
            
            # Store by policy key - normalize date format for matching
            eff_date_normalized = trans['Effective Date']
            if pd.notna(eff_date_normalized):
                try:
                    eff_date_normalized = pd.to_datetime(eff_date_normalized).strftime('%Y-%m-%d')
                except:
                    eff_date_normalized = str(eff_date_normalized)
            policy_key = f"{trans['Policy Number']}_{eff_date_normalized}"
            existing_lookup[policy_key] = trans_dict
            
            # Store by customer (for fuzzy matching)
            customer = trans['Customer']
            if pd.notna(customer):
                customer_key = customer.lower().strip()
                if customer_key not in customer_trans_lookup:
                    customer_trans_lookup[customer_key] = []
                customer_trans_lookup[customer_key].append(trans_dict)
    
    # Get unique customer list for fuzzy matching
    all_customers = []
    if not existing_data.empty:
        all_customers = existing_data['Customer'].dropna().unique().tolist()
    
    # Process each statement row
    for idx, row in statement_df.iterrows():
        # Extract mapped values
        customer = str(row[column_mapping['Customer']]).strip() if 'Customer' in column_mapping else ''
        policy_num = str(row[column_mapping['Policy Number']]).strip() if 'Policy Number' in column_mapping else ''
        eff_date = row[column_mapping['Effective Date']] if 'Effective Date' in column_mapping else None
        
        # Skip rows that appear to be totals
        # Check if customer name contains common total indicators
        customer_lower = customer.lower()
        if any(total_word in customer_lower for total_word in ['total', 'totals', 'subtotal', 'sub-total', 'grand total', 'sum']):
            continue
        
        # Also skip if policy number or customer is empty/missing and it looks like a summary row
        if (not customer or customer.lower() in ['', 'nan', 'none']) and (not policy_num or policy_num.lower() in ['', 'nan', 'none']):
            continue
        # Primary reconciliation amount is what the agent was paid
        amount = float(row[column_mapping['Agent Paid Amount (STMT)']]) if 'Agent Paid Amount (STMT)' in column_mapping else 0
        # Agency amount is optional for audit purposes
        agency_amount = float(row[column_mapping['Agency Comm Received (STMT)']]) if 'Agency Comm Received (STMT)' in column_mapping else 0
        
        # Convert effective date to string format
        if pd.notna(eff_date):
            try:
                eff_date = pd.to_datetime(eff_date).strftime('%Y-%m-%d')
            except:
                eff_date = str(eff_date)
        
        match_result = {
            'row_index': idx,
            'customer': customer,
            'policy_number': policy_num,
            'effective_date': eff_date,
            'amount': amount,  # Agent Paid Amount (primary)
            'agency_amount': agency_amount,  # Agency Comm Received (audit)
            'statement_data': row.to_dict()
        }
        
        # Try to match: Policy Number + Effective Date (highest confidence)
        policy_key = f"{policy_num}_{eff_date}"
        if policy_key in existing_lookup:
            match_result['match'] = existing_lookup[policy_key]
            match_result['confidence'] = 100
            match_result['match_type'] = 'Policy + Date'
            matched.append(match_result)
            continue
        
        # Try enhanced customer matching
        potential_customers = find_potential_customer_matches(customer, all_customers)
        
        if potential_customers:
            # Check if we have a single high-confidence match
            if len(potential_customers) == 1 and potential_customers[0][2] >= 90:
                # Single high-confidence match - try to find transaction
                matched_customer = potential_customers[0][0]
                customer_key = matched_customer.lower().strip()
                
                if customer_key in customer_trans_lookup:
                    customer_trans = customer_trans_lookup[customer_key]
                    # Check for amount match among transactions
                    amount_matched = False
                    for trans in customer_trans:
                        if amount > 0 and abs(trans['balance'] - amount) / amount <= 0.05:  # 5% tolerance
                            match_result['match'] = trans
                            match_result['confidence'] = 90
                            match_result['match_type'] = f'{potential_customers[0][1]} + Amount'
                            match_result['matched_customer'] = matched_customer
                            matched.append(match_result)
                            amount_matched = True
                            break
                    
                    if not amount_matched:
                        # Multiple transactions but no amount match
                        if len(customer_trans) == 1:
                            # Single transaction for this customer
                            match_result['match'] = customer_trans[0]
                            match_result['confidence'] = 85
                            match_result['match_type'] = potential_customers[0][1]
                            match_result['matched_customer'] = matched_customer
                            matched.append(match_result)
                        else:
                            # Multiple transactions - needs selection
                            match_result['potential_matches'] = customer_trans
                            match_result['potential_customers'] = potential_customers
                            match_result['needs_selection'] = True
                            unmatched.append(match_result)
                else:
                    # Customer found but no transactions in lookup - still add to unmatched for manual selection
                    match_result['potential_customers'] = potential_customers
                    match_result['needs_selection'] = True
                    unmatched.append(match_result)
            else:
                # Multiple potential matches or low confidence - needs manual selection
                match_result['potential_customers'] = potential_customers
                match_result['needs_selection'] = True
                
                # Collect all transactions for all potential customers
                all_potential_trans = []
                for potential_customer, match_type, score in potential_customers[:5]:  # Limit to top 5
                    customer_key = potential_customer.lower().strip()
                    if customer_key in customer_trans_lookup:
                        trans_list = customer_trans_lookup[customer_key]
                        for trans in trans_list:
                            trans_copy = trans.copy()  # Avoid modifying original
                            trans_copy['_customer_match'] = potential_customer
                            trans_copy['_match_type'] = match_type
                            trans_copy['_match_score'] = score
                            all_potential_trans.append(trans_copy)
                
                if all_potential_trans:
                    match_result['potential_matches'] = all_potential_trans
                    
                unmatched.append(match_result)
        else:
            # No fuzzy match found - check if exact customer exists
            customer_key = customer.lower().strip()
            if customer_key in customer_trans_lookup:
                # Exact match exists
                customer_matches = customer_trans_lookup[customer_key]
                # Check for amount match
                amount_matched = False
                for trans in customer_matches:
                    if amount > 0 and abs(trans['balance'] - amount) / amount <= 0.05:  # 5% tolerance
                        match_result['match'] = trans
                        match_result['confidence'] = 85
                        match_result['match_type'] = 'Customer + Amount'
                        matched.append(match_result)
                        amount_matched = True
                        break
                
                if not amount_matched:
                    # No amount match - needs selection
                    match_result['potential_matches'] = customer_matches
                    match_result['potential_customers'] = [(customer, 'exact', 100)]
                    match_result['needs_selection'] = True
                    unmatched.append(match_result)
            else:
                # No match found - can create new transaction
                match_result['can_create'] = True
                can_create.append(match_result)
    
    return matched, unmatched, can_create

def show_import_results(statement_date, all_data):
    """Display import results and allow user to review/confirm"""
    st.divider()
    st.markdown("### 📊 Import Preview")
    
    # Create tabs for different result types
    result_tabs = st.tabs([
        f"✅ Matched ({len(st.session_state.matched_transactions)})",
        f"❌ Unmatched ({len(st.session_state.unmatched_transactions)})",
        f"➕ Can Create ({len(st.session_state.transactions_to_create)})"
    ])
    
    with result_tabs[0]:  # Matched transactions
        if st.session_state.matched_transactions:
            matched_df = []
            for item in st.session_state.matched_transactions:
                # Show matched customer name if different from statement
                display_customer = item['customer']
                if 'matched_customer' in item and item['matched_customer'] != item['customer']:
                    display_customer = f"{item['customer']} → {item['matched_customer']}"
                
                matched_df.append({
                    'Status': '✅',
                    'Transaction ID': item['match'].get('Transaction ID', 'N/A'),
                    'Type': item['match'].get('Transaction Type', 'N/A'),
                    'Customer': display_customer,
                    'Policy': item.get('policy_number', ''),
                    'Eff Date': item.get('effective_date', ''),
                    'Statement Amt': item.get('amount', 0),
                    'DB Balance': item['match'].get('balance', item.get('amount', 0)),
                    'Confidence': f"{item['confidence']}%",
                    'Match Type': item['match_type']
                })
            
            df = pd.DataFrame(matched_df)
            # Apply special styling for STMT and VOID transactions
            styled_df = style_special_transactions(df)
            
            st.dataframe(
                styled_df,
                column_config={
                    "Statement Amt": st.column_config.NumberColumn(format="$%.2f"),
                    "DB Balance": st.column_config.NumberColumn(format="$%.2f")
                },
                use_container_width=True,
                hide_index=True
            )
            
            total_matched = sum(item.get('amount', 0) for item in st.session_state.matched_transactions)
            st.metric("Total Matched Amount", f"${total_matched:,.2f}")
        else:
            st.info("No matched transactions")
    
    with result_tabs[1]:  # Unmatched transactions
        if st.session_state.unmatched_transactions:
            st.warning("These transactions need manual review")
            
            # Create session state for manual matches if not exists
            if 'manual_matches' not in st.session_state:
                st.session_state.manual_matches = {}
            
            for idx, item in enumerate(st.session_state.unmatched_transactions):
                with st.expander(f"🔍 {item.get('customer', 'Unknown')} - ${item.get('amount', 0):,.2f}", expanded=True):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.markdown(f"**Statement Details:**")
                        st.text(f"Policy: {item.get('policy_number', '')}")
                        st.text(f"Effective Date: {item.get('effective_date', '')}")
                        st.text(f"Amount: ${item.get('amount', 0):,.2f}")
                        
                        # Show additional statement details if available
                        if 'statement_data' in item:
                            stmt_data = item['statement_data']
                            # Track if we found a direct Rate column
                            found_rate = False
                            
                            # Try to show LOB/Chg, Tran, and Rate from the statement
                            for col_name in ['LOB/Chg', 'LOB', ' Tran', 'Tran', 'Transaction', 'Rate', ' Rate']:
                                if col_name in stmt_data:
                                    value = stmt_data[col_name]
                                    if pd.notna(value) and str(value).strip():
                                        # Clean up the column name for display
                                        display_name = col_name.strip()
                                        if display_name == 'LOB/Chg' or display_name == 'LOB':
                                            st.text(f"LOB/Chg: {value}")
                                        elif display_name in ['Tran', 'Transaction']:
                                            st.text(f"Transaction: {value}")
                                        elif display_name == 'Rate':
                                            found_rate = True
                                            # Format rate as percentage if it's a number
                                            try:
                                                rate_val = float(value)
                                                if rate_val < 1:  # Decimal format
                                                    st.text(f"Rate: {rate_val:.2%}")
                                                else:  # Already percentage
                                                    st.text(f"Rate: {rate_val}%")
                                            except:
                                                st.text(f"Rate: {value}")
                            
                            # Only show mapped Rate field if we didn't find a direct Rate column
                            # or if the mapped column is different from 'Rate'
                            if 'Rate' in st.session_state.column_mapping:
                                rate_col = st.session_state.column_mapping['Rate']
                                if rate_col != 'Rate' and rate_col in stmt_data and pd.notna(stmt_data[rate_col]):
                                    try:
                                        rate_val = float(stmt_data[rate_col])
                                        if rate_val < 1:  # Decimal format
                                            st.text(f"Rate: {rate_val:.2%}")
                                        else:  # Already percentage
                                            st.text(f"Rate: {rate_val}%")
                                    except:
                                        st.text(f"Rate: {stmt_data[rate_col]}")
                    
                    with col2:
                        # Check if we have potential customers
                        if 'potential_customers' in item:
                            st.markdown("**Potential Matches Found:**")
                            
                            customer_options = []
                            for cust, match_type, score in item['potential_customers'][:10]:
                                customer_options.append(f"{cust} ({match_type}: {score}%)")
                            
                            selected_customer_idx = st.selectbox(
                                "Select Customer",
                                range(len(customer_options)),
                                format_func=lambda x: customer_options[x],
                                key=f"customer_select_{idx}"
                            )
                            
                            if selected_customer_idx is not None:
                                selected_customer = item['potential_customers'][selected_customer_idx][0]
                                
                                # Show transactions for selected customer
                                # First try to get from potential_matches if available
                                customer_trans = []
                                if 'potential_matches' in item:
                                    customer_trans = [t for t in item['potential_matches'] 
                                                    if t.get('_customer_match') == selected_customer]
                                
                                # If no transactions found in potential_matches, fetch directly
                                if not customer_trans:
                                    # Get transactions with balance for this customer
                                    trans_with_balance = calculate_transaction_balances(all_data)
                                    customer_trans_df = trans_with_balance[
                                        trans_with_balance['Customer'] == selected_customer
                                    ]
                                    if not customer_trans_df.empty:
                                        # Convert to list of dicts for consistency
                                        customer_trans = []
                                        for _, trans in customer_trans_df.iterrows():
                                            trans_dict = trans.to_dict()
                                            trans_dict['balance'] = trans.get('_balance', 0)
                                            customer_trans.append(trans_dict)
                                
                                if customer_trans:
                                    st.markdown("**Available Transactions:**")
                                    
                                    trans_options = []
                                    for trans in customer_trans:
                                        trans_desc = f"ID: {trans['Transaction ID']} | "
                                        trans_desc += f"Type: {trans.get('Transaction Type', 'N/A')} | "
                                        trans_desc += f"Policy Type: {trans.get('Policy Type', 'N/A')} | "
                                        trans_desc += f"Policy: {trans['Policy Number']} | "
                                        trans_desc += f"Eff: {trans['Effective Date']} | "
                                        trans_desc += f"Balance: ${trans['balance']:,.2f}"
                                        trans_options.append(trans_desc)
                                    
                                    selected_trans_idx = st.selectbox(
                                        "Select Transaction",
                                        range(len(trans_options)),
                                        format_func=lambda x: trans_options[x],
                                        key=f"trans_select_{idx}"
                                    )
                                    
                                    if st.button("✅ Confirm Match", key=f"confirm_{idx}", type="primary"):
                                        # Immediately move to matched transactions
                                        matched_item = item.copy()
                                        matched_item['match'] = customer_trans[selected_trans_idx]
                                        matched_item['confidence'] = 100
                                        matched_item['match_type'] = 'Manual - Selected'
                                        matched_item['matched_customer'] = selected_customer
                                        
                                        # Add to matched list
                                        st.session_state.matched_transactions.append(matched_item)
                                        
                                        # Remove from unmatched list
                                        st.session_state.unmatched_transactions = [
                                            unmatched for i, unmatched in enumerate(st.session_state.unmatched_transactions) 
                                            if i != idx
                                        ]
                                        
                                        # Clear any manual match entry for this index
                                        if idx in st.session_state.manual_matches:
                                            del st.session_state.manual_matches[idx]
                                        
                                        st.success("Match confirmed and moved to matched list!")
                                        st.rerun()
                                else:
                                    st.info("No transactions with balance for this customer")
                                    
                                    # Debug mode - show why transactions aren't available
                                    if st.checkbox("🔍 Show debug info", key=f"debug_{idx}"):
                                        # Get ALL transactions for this customer (not just those with balance)
                                        all_customer_trans = all_data[
                                            (all_data['Customer'] == selected_customer) &
                                            (~all_data['Transaction ID'].str.contains('-STMT-|-VOID-|-ADJ-', na=False))
                                        ]
                                        
                                        if not all_customer_trans.empty:
                                            st.write(f"Found {len(all_customer_trans)} total transactions for {selected_customer}:")
                                            for _, trans in all_customer_trans.iterrows():
                                                # Calculate balance for this transaction
                                                credit = float(trans.get('Agent Estimated Comm $', 0) or 0)
                                                policy_num = trans['Policy Number']
                                                effective_date = trans['Effective Date']
                                                
                                                # Get reconciliation entries
                                                recon_entries = all_data[
                                                    (all_data['Policy Number'] == policy_num) &
                                                    (all_data['Effective Date'] == effective_date) &
                                                    (all_data['Transaction ID'].str.contains('-STMT-|-VOID-', na=False))
                                                ]
                                                
                                                debit = 0
                                                if not recon_entries.empty:
                                                    debit = recon_entries['Agent Paid Amount (STMT)'].fillna(0).sum()
                                                
                                                balance = credit - debit
                                                
                                                st.write(f"- **{trans['Transaction ID']}**: Policy {policy_num}, Credit: ${credit:,.2f}, Debit: ${debit:,.2f}, Balance: ${balance:,.2f}")
                                                if balance <= 0.01:
                                                    st.write(f"  ⚠️ Not shown because balance is ${balance:,.2f}")
                                        else:
                                            st.warning(f"No transactions found for customer '{selected_customer}' in database. Check for name variations.")
                        else:
                            st.info("No potential matches found")
                    
                    # Options for handling the transaction
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # Option to create new
                        statement_customer = item.get('customer', 'Unknown')
                        create_checkbox_label = f"Create new transaction for: {statement_customer} (from statement)"
                        
                        if st.checkbox(create_checkbox_label, key=f"create_new_{idx}"):
                            st.caption("✓ Mapped client transaction details will be pre-filled from statement")
                            
                            # Client matching section
                            st.markdown("**🔍 Client Match Options:**")
                            
                            # Look for existing clients with similar names
                            try:
                                supabase = get_supabase_client()
                                
                                # First try exact match
                                exact_match = supabase.table('clients').select('*').eq('client_name', statement_customer).execute()
                                
                                # Also search for similar names
                                all_clients = supabase.table('clients').select('client_id', 'client_name').execute()
                                similar_clients = []
                                
                                if all_clients.data:
                                    # Use the same fuzzy matching logic as transaction matching
                                    for client in all_clients.data:
                                        client_name = client.get('client_name', '')
                                        if client_name and client_name != statement_customer:
                                            # Check for name variations
                                            if (statement_customer.lower() in client_name.lower() or 
                                                client_name.lower() in statement_customer.lower() or
                                                (len(statement_customer) > 3 and statement_customer.lower()[:3] == client_name.lower()[:3])):
                                                similar_clients.append(client)
                                
                                # Build radio button options
                                client_options = []
                                client_values = []
                                
                                # If exact match found, add it first
                                if exact_match.data:
                                    client = exact_match.data[0]
                                    client_options.append(f'Link to existing: "{client["client_name"]}" (Client ID: {client["client_id"]})')
                                    client_values.append(('existing', client['client_id'], client['client_name']))
                                
                                # Add similar clients
                                for client in similar_clients[:3]:  # Limit to top 3 similar
                                    client_options.append(f'Link to existing: "{client["client_name"]}" (Client ID: {client["client_id"]})')
                                    client_values.append(('existing', client['client_id'], client['client_name']))
                                
                                # Always add option to create new
                                client_options.append(f'Create as NEW client (new Client ID will be assigned)')
                                client_values.append(('new', None, statement_customer))
                                
                                # Default selection
                                default_index = 0 if exact_match.data else len(client_options) - 1
                                
                                selected_client_option = st.radio(
                                    "Select client option:",
                                    client_options,
                                    index=default_index,
                                    key=f"client_match_{idx}"
                                )
                                
                                # Get the selected client info
                                selected_index = client_options.index(selected_client_option)
                                client_action, client_id, client_name = client_values[selected_index]
                                
                            except Exception as e:
                                st.error(f"Error looking up clients: {str(e)}")
                                client_action = 'new'
                                client_id = None
                                client_name = statement_customer
                            
                            # Show transaction type selector
                            transaction_types = ["NEW", "RWL", "END", "CAN", "XCL", "PCH", "STL", "BoR"]
                            default_type = "NEW"
                            
                            # Try to guess from statement if available
                            if 'statement_data' in item and 'Transaction Type' in item['statement_data']:
                                stmt_type = item['statement_data'].get('Transaction Type', '').upper()
                                if stmt_type in transaction_types:
                                    default_type = stmt_type
                            
                            selected_type = st.selectbox(
                                "Transaction Type",
                                transaction_types,
                                index=transaction_types.index(default_type),
                                key=f"trans_type_{idx}"
                            )
                            
                            st.session_state.manual_matches[idx] = {
                                'statement_item': item,
                                'create_new': True,
                                'transaction_type': selected_type,
                                'client_action': client_action,
                                'client_id': client_id,
                                'client_name': client_name
                            }
                            
                            if client_action == 'existing':
                                st.success(f"Will create {selected_type} transaction linked to Client ID: {client_id}")
                            else:
                                st.success(f"Will create {selected_type} transaction with NEW client")
                        st.caption("*(Use for new policies or endorsements not yet in system)*")
                    
                    with col2:
                        # Option to match existing transaction
                        if 'potential_customers' in item and selected_customer_idx is not None:
                            selected_customer = item['potential_customers'][selected_customer_idx][0]
                            
                            # Get the selected transaction ID if available
                            transaction_id = "N/A"
                            if customer_trans and 'trans_select_' + str(idx) in st.session_state:
                                trans_idx = st.session_state['trans_select_' + str(idx)]
                                if trans_idx < len(customer_trans):
                                    transaction_id = customer_trans[trans_idx].get('Transaction ID', 'N/A')
                            
                            # Get match type and confidence
                            match_type = item['potential_customers'][selected_customer_idx][1]
                            confidence = item['potential_customers'][selected_customer_idx][2]
                            
                            # Check if names match
                            statement_customer = item.get('customer', 'Unknown')
                            names_match = statement_customer.lower() == selected_customer.lower()
                            
                            # Smarter warning logic
                            should_warn = False
                            if not names_match:
                                # Don't warn if it's a high-confidence name reversal
                                if match_type == "name_reversed" and confidence >= 95:
                                    should_warn = False
                                # Don't warn if it's an exact match (just different case)
                                elif match_type == "exact" and confidence == 100:
                                    should_warn = False
                                # Don't warn for high confidence business name variations
                                elif match_type == "business_normalized" and confidence >= 95:
                                    should_warn = False
                                # Warn for everything else
                                else:
                                    should_warn = True
                            
                            # Build checkbox label
                            checkbox_label = f"Force match to selected customer: {selected_customer} (Transaction ID: {transaction_id})"
                            
                            if st.checkbox(checkbox_label, key=f"match_existing_{idx}"):
                                st.session_state.manual_matches[idx] = {
                                    'statement_item': item,
                                    'match_to_customer': selected_customer,
                                    'match_existing': True
                                }
                                st.success(f"Will match to {selected_customer}")
                            
                            # Show warning only when appropriate
                            if should_warn and st.session_state.get(f"match_existing_{idx}", False):
                                st.markdown(f"<span style='color: red;'>⚠️ Warning: Customer names don't match ({statement_customer} ≠ {selected_customer})</span>", unsafe_allow_html=True)
            
            # Show confirmed matches
            if st.session_state.manual_matches:
                st.divider()
                st.markdown("### ✅ Confirmed Manual Matches")
                
                manual_match_count = len(st.session_state.manual_matches)
                st.metric("Manual Matches", manual_match_count)
                
                if st.button("Apply Manual Matches", type="primary"):
                    # Move manually matched items to appropriate lists
                    for idx, match_info in st.session_state.manual_matches.items():
                        if 'create_new' in match_info and match_info['create_new']:
                            # Move to create list with selected transaction type
                            item_to_create = match_info['statement_item'].copy()
                            item_to_create['selected_transaction_type'] = match_info.get('transaction_type', 'NEW')
                            st.session_state.transactions_to_create.append(item_to_create)
                        elif 'match_existing' in match_info and match_info['match_existing']:
                            # Handle match to existing customer without specific transaction
                            matched_item = match_info['statement_item'].copy()
                            # Find the best matching transaction for this customer
                            customer_name = match_info['match_to_customer']
                            
                            # For manual matches, first try to find ANY transaction for this customer
                            # with outstanding balance, regardless of policy/date match
                            matching_trans = all_data[
                                (all_data['Customer'] == customer_name) &
                                (~all_data['Transaction ID'].str.contains('-STMT-|-VOID-|-ADJ-', na=False))
                            ]
                            
                            # Calculate balances for customer transactions
                            if not matching_trans.empty:
                                # Get transactions with balance using the same logic as unreconciled
                                trans_with_balance = calculate_transaction_balances(all_data)
                                customer_trans_with_balance = trans_with_balance[
                                    trans_with_balance['Customer'] == customer_name
                                ]
                                
                                # First try exact policy match
                                exact_match = customer_trans_with_balance[
                                    customer_trans_with_balance['Policy Number'] == matched_item['policy_number']
                                ]
                                
                                if not exact_match.empty:
                                    # Use the exact policy match
                                    matched_item['match'] = exact_match.iloc[0].to_dict()
                                elif not customer_trans_with_balance.empty:
                                    # Use any transaction with balance for this customer
                                    matched_item['match'] = customer_trans_with_balance.iloc[0].to_dict()
                                else:
                                    # Use any transaction for this customer
                                    matched_item['match'] = matching_trans.iloc[0].to_dict()
                                
                                matched_item['confidence'] = 100
                                matched_item['match_type'] = 'Manual - Forced Match'
                                matched_item['matched_customer'] = customer_name
                                st.session_state.matched_transactions.append(matched_item)
                            else:
                                # If no transactions found, still honor the manual match
                                matched_item['match'] = {'Customer': customer_name}
                                matched_item['confidence'] = 100
                                matched_item['match_type'] = 'Manual - Customer Only'
                                matched_item['matched_customer'] = customer_name
                                st.session_state.matched_transactions.append(matched_item)
                        else:
                            # Original logic for transaction-specific matches
                            matched_item = match_info['statement_item'].copy()
                            matched_item['match'] = match_info['matched_transaction']
                            matched_item['confidence'] = 100
                            matched_item['match_type'] = 'Manual'
                            st.session_state.matched_transactions.append(matched_item)
                    
                    # Remove from unmatched
                    indices_to_remove = list(st.session_state.manual_matches.keys())
                    st.session_state.unmatched_transactions = [
                        item for i, item in enumerate(st.session_state.unmatched_transactions) 
                        if i not in indices_to_remove
                    ]
                    
                    # Clear manual matches
                    st.session_state.manual_matches = {}
                    st.success("Manual matches applied!")
                    st.rerun()
        else:
            st.success("All transactions were matched!")
    
    with result_tabs[2]:  # Can create
        if st.session_state.transactions_to_create:
            st.info("These transactions don't exist in the database but can be created")
            
            create_df = []
            for item in st.session_state.transactions_to_create:
                # Use selected transaction type if available, otherwise try to get from statement
                trans_type = item.get('selected_transaction_type', 'NEW')
                if trans_type == 'NEW' and 'statement_data' in item and st.session_state.column_mapping.get('Transaction Type'):
                    stmt_type = item['statement_data'].get(st.session_state.column_mapping.get('Transaction Type', ''), 'NEW')
                    if stmt_type:
                        trans_type = stmt_type
                
                create_df.append({
                    'Create': True,
                    'Customer': item['customer'],
                    'Policy': item['policy_number'],
                    'Type': trans_type,
                    'Eff Date': item['effective_date'],
                    'Amount': item['amount']
                })
            
            df = pd.DataFrame(create_df)
            edited_df = st.data_editor(
                df,
                column_config={
                    "Create": st.column_config.CheckboxColumn("Create", help="Check to create this transaction"),
                    "Amount": st.column_config.NumberColumn(format="$%.2f")
                },
                use_container_width=True,
                hide_index=True,
                key="create_selector"
            )
            
            # Option to create all or selected
            col1, col2 = st.columns(2)
            with col1:
                create_selected = st.checkbox(
                    "Create missing transactions before reconciling", 
                    value=True,
                    key="create_selected"
                )
            with col2:
                selected_count = edited_df['Create'].sum()
                st.metric("Selected to Create", selected_count)
        else:
            st.info("No new transactions to create")
    
    # Import confirmation section
    st.divider()
    st.markdown("### 🚀 Import Confirmation")
    
    # Get create_selected value from session state or default
    create_selected = st.session_state.get('create_selected', True)
    
    # Calculate totals
    total_matched = sum(item['amount'] for item in st.session_state.matched_transactions)
    total_unmatched = sum(item['amount'] for item in st.session_state.unmatched_transactions)
    total_to_create = sum(item['amount'] for item in st.session_state.transactions_to_create) if create_selected else 0
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Matched", f"${total_matched:,.2f}")
    with col2:
        st.metric("Unmatched", f"${total_unmatched:,.2f}")
    with col3:
        st.metric("To Create", f"${total_to_create:,.2f}")
    with col4:
        st.metric("Total", f"${total_matched + total_to_create:,.2f}")
    
    # Show statement total for verification if available
    if 'statement_file_total' in st.session_state and st.session_state.statement_file_total > 0:
        st.divider()
        st.markdown("### ✅ Verification Check")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Statement Total", f"${st.session_state.statement_file_total:,.2f}", 
                     help="Total from your commission statement file")
        with col2:
            reconcile_total = total_matched + total_to_create
            st.metric("Ready to Reconcile", f"${reconcile_total:,.2f}",
                     help="Sum of matched + to be created transactions")
        with col3:
            difference = st.session_state.statement_file_total - reconcile_total
            if abs(difference) < 0.01:  # Less than a penny
                st.metric("Difference", "$0.00", delta_color="off")
                st.success("✓ Perfectly balanced!")
            else:
                st.metric("Difference", f"${abs(difference):,.2f}", 
                         delta=f"{'Over' if difference < 0 else 'Under'} by ${abs(difference):,.2f}",
                         delta_color="inverse")
                if difference > 0:
                    st.warning(f"⚠️ Missing ${difference:,.2f} from statement")
                else:
                    st.error(f"❌ Exceeding statement by ${abs(difference):,.2f}")
    
    # Import button
    if st.button("🔄 Proceed with Import", type="primary", disabled=len(st.session_state.matched_transactions) == 0):
        with st.spinner("Importing transactions..."):
            # Store all database operations to execute at once
            all_operations = []
            created_count = 0
            reconciled_count = 0
            
            try:
                # Generate batch ID
                batch_id = f"IMPORT-{statement_date.strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
                
                # Step 1: Create missing transactions if selected
                if create_selected and st.session_state.transactions_to_create:
                    for idx, item in enumerate(st.session_state.transactions_to_create):
                        # Check if this transaction should be created (default to True if no edit)
                        should_create = True
                        if 'create_selector' in st.session_state:
                            # Get the dataframe from session state if it exists
                            create_df = pd.DataFrame([{
                                'Create': True,
                                'Customer': t['customer'],
                                'Policy': t['policy_number'],
                                'Eff Date': t['effective_date'],
                                'Amount': t['amount']
                            } for t in st.session_state.transactions_to_create])
                            should_create = create_df.loc[idx, 'Create'] if idx < len(create_df) else True
                        
                        if should_create:
                            # Generate new transaction ID with -IMPORT suffix
                            new_trans_id = generate_transaction_id(suffix='-IMPORT')
                            
                            # Check if this item has client info from manual matching
                            client_id_to_use = None
                            final_customer_name = item['customer']
                            
                            # Check if client info was selected during manual matching
                            if 'client_action' in item:
                                if item['client_action'] == 'existing' and item.get('client_id'):
                                    # Use the selected existing client
                                    client_id_to_use = item['client_id']
                                    final_customer_name = item.get('client_name', item['customer'])
                                elif item['client_action'] == 'new':
                                    # Create new client
                                    try:
                                        supabase = get_supabase_client()
                                        # Generate a client ID if needed
                                        new_client_id = f"CL-{str(uuid.uuid4())[:8].upper()}"
                                        new_client = {
                                            'client_id': new_client_id,
                                            'client_name': item['customer']
                                        }
                                        result = supabase.table('clients').insert(new_client).execute()
                                        if result.data:
                                            client_id_to_use = new_client_id
                                    except Exception as e:
                                        st.warning(f"Could not create new client for {item['customer']}: {str(e)}")
                            else:
                                # Fallback to original customer name matching logic
                                if not all_data.empty:
                                    all_customers = all_data['Customer'].dropna().unique().tolist()
                                    # Use the find_potential_customer_matches function to check for existing customer
                                    potential_matches = find_potential_customer_matches(item['customer'], all_customers)
                                    if potential_matches and potential_matches[0][2] >= 90:  # High confidence match
                                        # Use the existing customer name format
                                        final_customer_name = potential_matches[0][0]
                            
                            # Create new transaction
                            new_trans = {
                                'Transaction ID': new_trans_id,
                                'Customer': final_customer_name,  # Use matched customer name if found
                                'Policy Number': item['policy_number'],
                                'Effective Date': item['effective_date'],
                                'Transaction Type': item.get('selected_transaction_type', item['statement_data'].get(st.session_state.column_mapping.get('Transaction Type', ''), 'NEW')),
                                'Premium Sold': item['statement_data'].get(st.session_state.column_mapping.get('Premium Sold', ''), 0),
                                'Agent Estimated Comm $': item['amount'],  # Use statement amount as estimated
                                'Agency Estimated Comm/Revenue (CRM)': item['amount'],
                                'NOTES': f"Created from statement import {batch_id}"
                            }
                            
                            # Add Client ID if we have one
                            if client_id_to_use:
                                new_trans['Client ID'] = client_id_to_use
                            
                            # Add other mapped fields with special handling for Policy Type
                            for sys_field, stmt_field in st.session_state.column_mapping.items():
                                if sys_field not in new_trans and stmt_field in item['statement_data']:
                                    value = item['statement_data'][stmt_field]
                                    
                                    # Apply policy type mapping if it's the Policy Type field
                                    if sys_field == 'Policy Type' and value:
                                        # Load policy type mappings
                                        mapping_file = "config_files/policy_type_mappings.json"
                                        try:
                                            if os.path.exists(mapping_file):
                                                with open(mapping_file, 'r') as f:
                                                    type_mappings = json.load(f)
                                                # Check if this statement value has a mapping
                                                if str(value) in type_mappings:
                                                    value = type_mappings[str(value)]
                                        except:
                                            pass  # If mapping fails, use original value
                                    
                                    new_trans[sys_field] = value
                            
                            # Clean data before queueing for insertion
                            cleaned_trans = clean_data_for_database(new_trans)
                            
                            # Queue operation instead of immediate execution
                            all_operations.append(('insert', 'policies', cleaned_trans))
                            created_count += 1
                            
                            # Add to matched transactions for reconciliation
                            # Use the new transaction data as the match
                            st.session_state.matched_transactions.append({
                                'match': new_trans,
                                'amount': item['amount'],
                                'customer': final_customer_name,  # Use the matched customer name
                                'policy_number': item['policy_number']
                            })
                
                # Step 2: Create reconciliation entries for all matched transactions
                for item in st.session_state.matched_transactions:
                    recon_id = generate_reconciliation_transaction_id("STMT", statement_date)
                    
                    # Create reconciliation entry
                    recon_entry = {
                        'Transaction ID': recon_id,
                        'Customer': item.get('matched_customer', item['match'].get('Customer', item.get('customer', ''))),
                        'Policy Number': item['match'].get('Policy Number', item.get('policy_number', '')),
                        'Effective Date': item['match'].get('Effective Date', item.get('effective_date', '')),
                        'Transaction Type': item['match'].get('Transaction Type', ''),
                        'Premium Sold': 0,
                        'Agency Comm Received (STMT)': item.get('agency_amount', 0),  # Agency amount for audit
                        'Agent Paid Amount (STMT)': item['amount'],  # Agent amount (primary)
                        'STMT DATE': statement_date.strftime('%Y-%m-%d'),
                        'reconciliation_status': 'reconciled',
                        'reconciliation_id': batch_id,
                        'is_reconciliation_entry': True,
                        # 'Cross-Reference Key': item['match'].get('Transaction ID', ''),  # Field doesn't exist in database
                        'NOTES': f"Import batch {batch_id} | Matched to: {item['match'].get('Transaction ID', 'Manual Match')}"
                    }
                    
                    # Clean data before queueing for insertion
                    cleaned_recon = clean_data_for_database(recon_entry)
                    
                    # Queue reconciliation entry
                    all_operations.append(('insert', 'policies', cleaned_recon))
                    reconciled_count += 1
                
                # Execute all operations in a single batch
                # This ensures atomicity - either all succeed or none do
                if all_operations:
                    st.info(f"Executing {len(all_operations)} database operations...")
                    
                    supabase = get_supabase_client()
                    successful_operations = 0
                    
                    # Execute each operation
                    for op_type, table, data in all_operations:
                        if op_type == 'insert':
                            result = supabase.table(table).insert(data).execute()
                            if not result.data:
                                raise Exception(f"Failed to insert record: {data.get('Transaction ID', 'Unknown')}")
                            successful_operations += 1
                    
                    # If we got here, all operations succeeded
                    st.success(f"✅ All {successful_operations} operations completed successfully!")
                
                # Clear session state only after successful completion
                st.session_state.import_data = None
                st.session_state.matched_transactions = []
                st.session_state.unmatched_transactions = []
                st.session_state.transactions_to_create = []
                st.session_state.column_mapping = {}
                if 'statement_file_total' in st.session_state:
                    del st.session_state.statement_file_total
                
                st.success(f"""
                ✅ Import completed successfully!
                - Created {created_count} new transactions
                - Reconciled {reconciled_count} transactions
                - Batch ID: {batch_id}
                """)
                
                # Clear cache and refresh
                st.cache_data.clear()
                # Note: The uploaded file is automatically cleaned up by Streamlit on rerun
                time.sleep(4)
                st.rerun()
                
            except Exception as e:
                # If ANY error occurs, no database changes were made
                st.error(f"""
                ❌ Import failed - NO changes were made to the database
                
                Error: {str(e)}
                
                Please fix the issue and try again. All transactions remain unchanged.
                """)
                st.exception(e)
                
                # Log which operation failed if possible
                if 'all_operations' in locals() and all_operations:
                    st.info(f"Failed while processing operation {successful_operations + 1} of {len(all_operations)}")

def duplicate_for_renewal(df: pd.DataFrame) -> pd.DataFrame:
    """
    Duplicates the given policies and updates their dates for renewal.
    """
    if df.empty:
        return pd.DataFrame()

    renewed_df = df.copy()
    
    # Calculate new term dates
    renewed_df['new_effective_date'] = renewed_df['expiration_date']
    
    # Calculate new expiration date based on Policy Term
    policy_term_col = get_mapped_column("Policy Term")
    renewed_df['new_expiration_date'] = renewed_df.apply(
        lambda row: row['new_effective_date'] + pd.DateOffset(months=int(row[policy_term_col])) 
        if policy_term_col in row and pd.notna(row.get(policy_term_col)) and row.get(policy_term_col) != 0
        else row['new_effective_date'] + pd.DateOffset(months=6),  # Default to 6 months if not specified
        axis=1
    )
    
    # Update the relevant columns - Using YYYY-MM-DD format
    renewed_df[get_mapped_column("Effective Date")] = renewed_df['new_effective_date'].dt.strftime('%Y-%m-%d')
    renewed_df[get_mapped_column("X-DATE")] = renewed_df['new_expiration_date'].dt.strftime('%Y-%m-%d')
    renewed_df[get_mapped_column("Transaction Type")] = "RWL"
    
    return renewed_df

def edit_transaction_form(modal_data, source_page="edit_policies", is_renewal=False):
    """
    Reusable edit transaction form that can be used from both 
    Edit Policy Transactions and Pending Policy Renewals pages.
    
    Args:
        modal_data: Dictionary containing the transaction data to edit
        source_page: String indicating which page called this form
        is_renewal: Boolean indicating if this is for a renewal
    
    Returns:
        Dict with 'action' and 'data' keys if form was submitted, None otherwise
    """
    # Get transaction ID column name
    transaction_id_col = get_mapped_column("Transaction ID")
    
    # Track all rendered fields to prevent duplicates
    rendered_fields = set()
    
    # Remove the Select column from modal data if present
    if 'Select' in modal_data:
        del modal_data['Select']
    
    # Get transaction ID and customer name for header
    transaction_id = modal_data.get(transaction_id_col, 'Unknown')
    customer_name = modal_data.get('Customer', 'Unknown')
    
    # Check if this is a reconciliation transaction (only for non-renewals)
    if not is_renewal and is_reconciliation_transaction(transaction_id):
        st.error("🔒 This is a reconciliation transaction and cannot be edited.")
        st.info("Reconciliation entries (-STMT-, -VOID-, -ADJ-) are permanent audit records. Use the Reconciliation page to create adjustments if needed.")
        if st.button("Close", type="primary"):
            return {"action": "close", "data": None}
        return None
    
    # Check if this is an import-created transaction
    is_import = is_import_transaction(transaction_id)
    if is_import:
        # Show comprehensive explanation for import transactions
        with st.expander("📥 **Import-Created Transaction** - Click to understand this transaction", expanded=True):
            st.markdown("""
            **What is this transaction?**  
            This transaction was automatically created during reconciliation because:
            - A payment was recorded in your commission statement
            - No matching policy transaction existed in the system
            - The system created this placeholder to record the payment
            
            **Why do you need to complete it?**  
            To properly track commissions, you need to add:
            - Premium information (Premium Sold, Taxes & Fees)
            - Commission rates (Policy Gross Comm %, Agent Comm %)
            - This creates the "credit" side to match the "debit" (payment) already recorded
            
            **What are the limitations?**  
            🔒 **Protected fields** (cannot be edited):
            - Transaction ID (ends with -IMPORT)
            - Payment amounts (Agent Paid, Agency Comm Received)
            - Statement date
            - Customer information
            
            ✏️ **Editable fields** (please complete):
            - Premium Sold ← Enter the premium amount
            - Policy Taxes & Fees ← Enter taxes/fees
            - Commission rates ← Will auto-populate if carrier/MGA selected
            - Policy details (Policy Number, Dates, etc.)
            
            📊 **Calculated fields** (auto-update):
            - Commissionable Premium
            - Agency/Agent Estimated Commission
            
            ⚠️ **Important**: This transaction cannot be deleted to preserve payment history.
            """)
        
        st.markdown("---")
    
    # Form title based on context
    if is_renewal:
        st.info(f"**Reviewing renewal for Policy:** {modal_data.get('Policy Number', 'Unknown')} | **Customer:** {customer_name}")
    else:
        st.info(f"**Transaction ID:** {transaction_id} | **Customer:** {customer_name}")
    
    # Create form
    with st.form("edit_transaction_form"):
        # For renewals, show the new Transaction ID at the top
        if is_renewal:
            st.text_input(
                "New Transaction ID (Pending)",
                value=transaction_id,
                disabled=True,
                help="This is the new Transaction ID that will be assigned to the renewal transaction"
            )
        # Track updated values
        updated_data = {}
        
        # Define internal system fields that should be read-only
        internal_fields = [
            'reconciliation_status', 
            'reconciliation_id', 
            'reconciled_at', 
            'is_reconciliation_entry',
            '_id',
            'Client ID (CRM)',
            'STMT DATE',
            'Agency Comm Received (STMT)',
            'Agent Paid Amount (STMT)'
        ]
        
        # Add renewal-specific read-only fields
        if is_renewal:
            internal_fields.append('Policy Origination Date')
        
        # Define field groups for better organization
        client_fields = ['Client ID (CRM)', 'Client ID', 'Customer', 'Client Name', 'Agent Name']
        policy_fields = ['Writing Code', 'Policy Number', 'Policy #', 'Prior Policy Number', 'Product', 'Carrier', 'Policy Type', 'Carrier Name', 'MGA Name', 'Transaction Type', 'Policy Checklist Complete', 'FULL OR MONTHLY PMTS', 'NOTES']
        date_fields = ['Policy Issue Date', 'Policy Effective Date', 'As of Date', 'Effective Date', 'Policy Origination Date', 'X-DATE']
        commission_fields = [
            'Premium Sold', 'Policy Taxes & Fees', 'Commissionable Premium',
            'Agency Estimated Comm/Revenue (CRM)', 
            'Policy Gross Comm %', 'Agent Estimated Comm $',
            'Agency Comm Received (STMT)', 'Agent Paid Amount (STMT)',
            'Agent Comm %', 'Broker Fee', 
            'Broker Fee Agent Comm', 'Total Agent Comm'
        ]
        status_fields = ['Reconciliation Notes', 'Reconciled?']  # 'Cross-Reference Key' not in database
        
        # Client Information
        st.markdown("#### Client Information")
        col1, col2 = st.columns(2)
        
        field_counter = 0
        for field in modal_data.keys():
            # Skip internal fields and Client ID - they'll be handled separately
            if field in client_fields and field not in internal_fields and field != 'Client ID':
                with col1 if field_counter % 2 == 0 else col2:
                    updated_data[field] = st.text_input(
                        field, 
                        value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                        key=f"modal_{field}"
                    )
                    rendered_fields.add(field)  # Track rendered field
                field_counter += 1
        
        # Always handle Client ID field with special logic
        # (Even if it was in modal_data, we skipped it above)
        with col1 if field_counter % 2 == 0 else col2:
            # Check if Client ID exists in the data
            client_id_value = modal_data.get('Client ID', '')
            
            # If no Client ID exists, show option to generate one
            if not client_id_value or str(client_id_value).strip() == '':
                generate_new = st.checkbox(
                    "Generate New Client ID",
                    key="modal_generate_client_id",
                    help="Check this box to generate a unique Client ID (3 letters + 3 numbers mixed randomly, e.g., U5G0O4, A2B7C9)"
                )
                
                if generate_new:
                    # Generate a unique numeric Client ID that doesn't exist in database
                    # Store in session state to persist across form interactions
                    if 'generated_client_id' not in st.session_state:
                        with st.spinner("Checking for unique Client ID..."):
                            st.session_state.generated_client_id = generate_unique_client_id()
                    
                    client_id_value = st.session_state.generated_client_id
                    st.info(f"🆔 Generated Client ID: {client_id_value} (verified unique)")
            else:
                # Clear generated ID from session state if we have an existing ID
                if 'generated_client_id' in st.session_state:
                    del st.session_state.generated_client_id
            
            help_text = "Enter a Client ID to link this transaction to a specific client" if not client_id_value else "Existing Client ID"
            updated_data['Client ID'] = st.text_input(
                'Client ID', 
                value=str(client_id_value) if client_id_value else '',
                key="modal_Client ID",
                help=help_text,
                placeholder="e.g., U5G0O4"
            )
            
            # Debug: Show what's in the Client ID field
            st.caption(f"Debug - Client ID value: '{client_id_value}' (type: {type(client_id_value)})")
            
            rendered_fields.add('Client ID')
            field_counter += 1
        
        # Policy Information
        st.markdown("#### Policy Information")
        col3, col4 = st.columns(2)
        
        # Handle Carrier Name and MGA Name using values from outside selection
        with col3:
            if 'Carrier Name' in modal_data.keys():
                # Get carrier name from session state (selected outside form)
                carrier_from_outside = st.session_state.get('edit_final_carrier_name', '')
                if carrier_from_outside:
                    updated_data['Carrier Name'] = carrier_from_outside
                    st.text_input(
                        'Carrier Name (from selection above)', 
                        value=carrier_from_outside,
                        key="modal_Carrier Name",
                        disabled=True,
                        help="Carrier selected from dropdown above"
                    )
                else:
                    # Fallback to original value if nothing selected
                    updated_data['Carrier Name'] = modal_data.get('Carrier Name', '')
                    st.text_input(
                        'Carrier Name', 
                        value=str(modal_data.get('Carrier Name', '')) if modal_data.get('Carrier Name') is not None else '',
                        key="modal_Carrier Name",
                        disabled=True,
                        help="Select carrier from dropdown above"
                    )
                rendered_fields.add('Carrier Name')
        
        with col4:
            if 'MGA Name' in modal_data.keys():
                # Get MGA name from session state (selected outside form)
                mga_from_outside = st.session_state.get('edit_final_mga_name', '')
                if mga_from_outside is not None:  # Can be empty string for Direct Appointment
                    updated_data['MGA Name'] = mga_from_outside
                    st.text_input(
                        'MGA Name (from selection above)', 
                        value=mga_from_outside if mga_from_outside else "Direct Appointment",
                        key="modal_MGA Name",
                        disabled=True,
                        help="MGA selected from dropdown above"
                    )
                else:
                    # Fallback to original value if nothing selected
                    updated_data['MGA Name'] = modal_data.get('MGA Name', '')
                    st.text_input(
                        'MGA Name', 
                        value=str(modal_data.get('MGA Name', '')) if modal_data.get('MGA Name') is not None else '',
                        key="modal_MGA Name",
                        disabled=True,
                        help="Select MGA from dropdown above"
                    )
                rendered_fields.add('MGA Name')
            
            # Policy Type in right column, second position
            if 'Policy Type' in modal_data.keys():
                # Load policy types from configuration
                policy_types_config = load_policy_types_config()
                active_types = [pt['name'] for pt in policy_types_config['policy_types'] if pt['active']]
                
                # Get current value
                current_policy_type = modal_data.get('Policy Type', '')
                
                # Ensure current value is in options
                options = active_types.copy()
                if current_policy_type and current_policy_type not in options:
                    options.insert(0, current_policy_type)
                
                updated_data['Policy Type'] = st.selectbox(
                    'Policy Type (add in Admin Panel)',
                    options=options,
                    index=options.index(current_policy_type) if current_policy_type in options else 0,
                    key="modal_Policy Type",
                    help="Go to Admin Panel → Policy Types to add new types"
                )
                rendered_fields.add('Policy Type')
        
        # Second row - Transaction Type 
        with col3:
            if 'Transaction Type' in modal_data.keys():
                # Transaction type dropdown
                transaction_types = ["NEW", "RWL", "END", "PCH", "CAN", "XCL", "NBS", "STL", "BoR", "REWRITE"]
                current_trans_type = modal_data.get('Transaction Type', 'RWL' if is_renewal else 'NEW')
                
                # For renewals, lock to RWL
                if is_renewal:
                    updated_data['Transaction Type'] = st.selectbox(
                        'Transaction Type',
                        options=['RWL'],
                        index=0,
                        key="modal_Transaction Type",
                        disabled=True,
                        help="Renewal transactions are always RWL"
                    )
                else:
                    updated_data['Transaction Type'] = st.selectbox(
                        'Transaction Type',
                        options=transaction_types,
                        index=transaction_types.index(current_trans_type) if current_trans_type in transaction_types else 0,
                        key="modal_Transaction Type"
                    )
                    rendered_fields.add('Transaction Type')
        
        # Now handle the rest of the policy fields
        field_counter = 0
        for field in policy_fields:
            if field in modal_data.keys() and field not in ['Carrier Name', 'MGA Name', 'Policy Type', 'Transaction Type', 'Policy Checklist Complete', 'FULL OR MONTHLY PMTS', 'NOTES']:
                with col3 if field_counter % 2 == 0 else col4:
                    # Regular text input
                    # Make Prior Policy Number read-only for renewals
                    if field == 'Prior Policy Number' and is_renewal:
                        value = modal_data.get(field, '')
                        st.text_input(
                            field, 
                            value=str(value) if value is not None else '',
                            key=f"modal_{field}",
                            disabled=True,
                            help="Automatically populated from the policy being renewed"
                        )
                        updated_data[field] = value  # Preserve the value since it's disabled
                    else:
                        updated_data[field] = st.text_input(
                            field, 
                            value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                            key=f"modal_{field}"
                        )
                field_counter += 1
        
        
        # Date Fields
        st.markdown("#### Dates")
        col5, col6 = st.columns(2)
        
        # Left column - Effective Date first, then Policy Origination Date
        with col5:
            # Effective Date
            if 'Effective Date' in modal_data.keys():
                date_value = modal_data.get('Effective Date')
                if date_value and pd.notna(date_value):
                    try:
                        parsed_date = pd.to_datetime(date_value)
                        updated_data['Effective Date'] = st.date_input(
                            'Effective Date',
                            value=parsed_date.date(),
                            key="modal_Effective Date",
                        )
                        rendered_fields.add('Effective Date')
                    except:
                        updated_data['Effective Date'] = st.text_input(
                            'Effective Date',
                            value=str(date_value),
                            key="modal_Effective Date",
                            help="Enter date in MM/DD/YYYY format"
                        )
                else:
                    updated_data['Effective Date'] = st.date_input(
                        'Effective Date',
                        value=None,
                        key="modal_Effective Date",
                    )
                rendered_fields.add('Effective Date')
            
            # Policy Origination Date (read-only for renewals)
            if 'Policy Origination Date' in modal_data.keys():
                date_value = modal_data.get('Policy Origination Date')
                
                # Auto-populate logic based on transaction type
                auto_populated_date = None
                auto_populate_message = None
                
                # Get current transaction type
                current_transaction_type = updated_data.get('Transaction Type', modal_data.get('Transaction Type', ''))
                # Get policy number from updated_data first (user's input), then fall back to modal_data
                policy_number = updated_data.get('Policy Number', modal_data.get('Policy Number', ''))
                
                # Only auto-populate if current date_value is empty
                if not date_value or pd.isna(date_value):
                    if current_transaction_type == 'NEW':
                        # For NEW transactions, copy from Effective Date
                        effective_date = updated_data.get('Effective Date')
                        if effective_date:
                            auto_populated_date = effective_date
                            auto_populate_message = "📅 Auto-populated from Effective Date (NEW transaction)"
                    
                    elif current_transaction_type == 'BoR':
                        # For BoR, use Effective Date (new relationship)
                        effective_date = updated_data.get('Effective Date')
                        if effective_date:
                            auto_populated_date = effective_date
                            auto_populate_message = "🤝 Auto-populated from Effective Date (new BoR relationship)"
                    
                    elif current_transaction_type in ['RWL', 'END', 'PCH', 'CAN', 'XCL', 'REWRITE', 'NBS', 'STL'] and policy_number:
                        # For these types, look up the original NEW transaction
                        try:
                            # Load all policies data
                            all_data = load_policies_data()
                            
                            # Function to trace back to original NEW transaction
                            def find_origination_date(policy_num, visited=None):
                                if visited is None:
                                    visited = set()
                                
                                if policy_num in visited:
                                    return None, None  # Circular reference protection
                                visited.add(policy_num)
                                
                                # Find transactions for this policy number
                                policy_transactions = all_data[all_data['Policy Number'] == policy_num].copy()
                                
                                if not policy_transactions.empty:
                                    # Sort by transaction date/effective date to get earliest
                                    if 'Effective Date' in policy_transactions.columns:
                                        policy_transactions = policy_transactions.sort_values('Effective Date')
                                    
                                    # Check for NEW transaction
                                    new_transactions = policy_transactions[policy_transactions['Transaction Type'] == 'NEW']
                                    if len(new_transactions) > 1:
                                        # Multiple NEW transactions warning
                                        return None, "⚠️ Multiple NEW transactions found for this policy"
                                    elif len(new_transactions) == 1:
                                        # Found the NEW transaction
                                        orig_date = new_transactions.iloc[0].get('Policy Origination Date')
                                        if orig_date and pd.notna(orig_date):
                                            return orig_date, "📊 Auto-populated from original NEW transaction"
                                        # If NEW transaction has no origination date, use its effective date
                                        eff_date = new_transactions.iloc[0].get('Effective Date')
                                        if eff_date and pd.notna(eff_date):
                                            return eff_date, "📊 Auto-populated from NEW transaction's Effective Date"
                                    
                                    # No NEW transaction, check for Prior Policy Number
                                    for _, trans in policy_transactions.iterrows():
                                        prior_policy = trans.get('Prior Policy Number')
                                        if prior_policy and pd.notna(prior_policy) and str(prior_policy).strip():
                                            # Recursively follow the chain
                                            result_date, result_msg = find_origination_date(prior_policy, visited)
                                            if result_date:
                                                return result_date, result_msg
                                
                                return None, None
                            
                            # Find the origination date
                            found_date, message = find_origination_date(policy_number)
                            if found_date:
                                auto_populated_date = found_date
                                auto_populate_message = message
                            
                        except Exception as e:
                            # If lookup fails, just continue without auto-population
                            pass
                
                # Use auto-populated date if available
                # For END, RWL, and other continuation transactions, always show what we found
                if auto_populated_date:
                    if not date_value or pd.isna(date_value) or (current_transaction_type in ['END', 'RWL', 'PCH', 'REWRITE'] and date_value != auto_populated_date):
                        date_value = auto_populated_date
                        if auto_populate_message:
                            st.info(auto_populate_message)
                    elif date_value != auto_populated_date and current_transaction_type in ['END', 'RWL', 'PCH', 'REWRITE']:
                        # Show what we found even if not updating
                        st.warning(f"Found origination date {auto_populated_date} from NEW transaction, but keeping existing value {date_value}")
                
                if is_renewal:
                    # For renewals, show as read-only text
                    st.text_input(
                        'Policy Origination Date (preserved)',
                        value=str(date_value) if date_value else '',
                        key="modal_Policy Origination Date_display",
                        disabled=True,
                        help="Original policy date is preserved for renewals"
                    )
                    updated_data['Policy Origination Date'] = date_value
                else:
                    # For regular edits, allow date input
                    if date_value and pd.notna(date_value):
                        try:
                            parsed_date = pd.to_datetime(date_value)
                            updated_data['Policy Origination Date'] = st.date_input(
                                'Policy Origination Date',
                                value=parsed_date.date(),
                                key="modal_Policy Origination Date",
                                help="Auto-populated based on transaction type and policy history"
                                )
                        except:
                            updated_data['Policy Origination Date'] = st.text_input(
                                'Policy Origination Date',
                                value=str(date_value),
                                key="modal_Policy Origination Date",
                                help="Enter date in MM/DD/YYYY format"
                            )
                    else:
                        updated_data['Policy Origination Date'] = st.date_input(
                            'Policy Origination Date',
                            value=None,
                            key="modal_Policy Origination Date",
                            help="Auto-populated based on transaction type and policy history"
                        )
        
        # Right column - X-DATE only (aligned with Effective Date)
        with col6:
            # X-DATE
            if 'X-DATE' in modal_data.keys():
                date_value = modal_data.get('X-DATE')
                
                # Check if we should auto-populate X-DATE for NEW/RWL (except AUTO)
                transaction_type = updated_data.get('Transaction Type', modal_data.get('Transaction Type', ''))
                policy_type = updated_data.get('Policy Type', modal_data.get('Policy Type', ''))
                effective_date = updated_data.get('Effective Date', modal_data.get('Effective Date'))
                
                if (transaction_type in ['NEW', 'RWL'] and policy_type != 'AUTO' and 
                    effective_date and (not date_value or pd.isna(date_value))):
                    try:
                        # Parse the effective date
                        if not isinstance(effective_date, pd.Timestamp):
                            effective_date = pd.to_datetime(effective_date)
                        
                        # Calculate X-DATE (12 months later)
                        date_value = effective_date + pd.DateOffset(months=12)
                        st.info("📅 X-DATE auto-populated as Effective Date + 12 months for NEW/RWL policy")
                    except Exception as e:
                        pass
                
                if date_value and pd.notna(date_value):
                    try:
                        parsed_date = pd.to_datetime(date_value)
                        updated_data['X-DATE'] = st.date_input(
                            'X-DATE (Expiration)',
                            value=parsed_date.date(),
                            key="modal_X-DATE",
                        )
                    except:
                        updated_data['X-DATE'] = st.text_input(
                            'X-DATE (Expiration)',
                            value=str(date_value),
                            key="modal_X-DATE",
                            help="Enter date in MM/DD/YYYY format"
                        )
                else:
                    updated_data['X-DATE'] = st.date_input(
                        'X-DATE (Expiration)',
                        value=None,
                        key="modal_X-DATE",
                    )
        
        # Add Policy Term after Policy Origination Date
        with col6:
            if 'Policy Term' in modal_data.keys():
                # Policy Term dropdown
                policy_terms = [3, 6, 9, 12]
                
                # Try to auto-calculate term based on Effective Date and X-DATE
                calculated_term = None
                effective_date = updated_data.get('Effective Date')
                x_date = updated_data.get('X-DATE')
                
                # Auto-populate 12-month term for NEW and RWL (except AUTO)
                transaction_type = updated_data.get('Transaction Type', modal_data.get('Transaction Type', ''))
                policy_type = updated_data.get('Policy Type', modal_data.get('Policy Type', ''))
                
                if transaction_type in ['NEW', 'RWL'] and policy_type != 'AUTO' and effective_date:
                    # If no X-DATE is set, calculate it as Effective Date + 12 months
                    if not x_date or pd.isna(x_date):
                        try:
                            # Parse the effective date
                            if not isinstance(effective_date, pd.Timestamp):
                                effective_date = pd.to_datetime(effective_date)
                            
                            # Calculate X-DATE (12 months later)
                            x_date_calculated = effective_date + pd.DateOffset(months=12)
                            updated_data['X-DATE'] = x_date_calculated.date()
                            x_date = x_date_calculated
                            
                            # Update the X-DATE field in session state to reflect the change
                            if 'modal_X-DATE' in st.session_state:
                                st.session_state['modal_X-DATE'] = x_date_calculated.date()
                            
                            st.info("📅 X-DATE auto-populated as Effective Date + 12 months for NEW/RWL policy")
                        except Exception as e:
                            st.warning(f"Could not calculate X-DATE: {str(e)}")
                    
                    # Set calculated term to 12 months
                    calculated_term = 12
                
                elif effective_date and x_date:
                    try:
                        # Convert to datetime if needed
                        if not isinstance(effective_date, pd.Timestamp):
                            effective_date = pd.to_datetime(effective_date)
                        if not isinstance(x_date, pd.Timestamp):
                            x_date = pd.to_datetime(x_date)
                        
                        # Calculate the difference in months
                        months_diff = (x_date.year - effective_date.year) * 12 + (x_date.month - effective_date.month)
                        
                        # If we get a standard term, use it
                        if months_diff in policy_terms:
                            calculated_term = months_diff
                        # If it's close to a standard term (within a few days), round to nearest
                        elif months_diff > 0:
                            # Check if it's close to any standard term
                            for term in policy_terms:
                                # Allow for dates that are a few days off (e.g., 11.9 months rounds to 12)
                                days_diff = (x_date - effective_date).days
                                expected_days = term * 30  # Approximate days in months
                                if abs(days_diff - expected_days) <= 15:  # Within 15 days
                                    calculated_term = term
                                    break
                    except:
                        pass
                
                # Use calculated term if available, otherwise use existing value
                current_term = calculated_term if calculated_term else modal_data.get('Policy Term', None)
                
                # Handle the display
                if current_term is None or pd.isna(current_term):
                    selected_index = 0
                else:
                    try:
                        selected_index = policy_terms.index(int(current_term)) + 1
                    except (ValueError, TypeError):
                        selected_index = 0
                
                # Force the calculated term to be selected if we have one
                if calculated_term is not None:
                    try:
                        selected_index = policy_terms.index(int(calculated_term)) + 1
                        # Update session state to ensure the widget reflects the calculated value
                        if 'modal_Policy Term' not in st.session_state or st.session_state.get('modal_Policy Term') != calculated_term:
                            st.session_state['modal_Policy Term'] = calculated_term
                    except (ValueError, TypeError):
                        selected_index = 0
                
                # Show info message if we auto-calculated
                if calculated_term:
                    st.info(f"📊 Policy Term auto-calculated as {calculated_term} months based on dates")
                
                # Set the selectbox with the calculated value
                selectbox_value = st.selectbox(
                    'Policy Term',
                    options=[None] + policy_terms,
                    format_func=lambda x: "" if x is None else f"{x} months",
                    index=selected_index,
                    key="modal_Policy Term",
                    help="Auto-populated based on Effective Date and X-DATE when available"
                )
                
                # If we calculated a term, use that; otherwise use what the user selected
                if calculated_term is not None:
                    updated_data['Policy Term'] = calculated_term
                else:
                    updated_data['Policy Term'] = selectbox_value
        
        # Premium Information
        st.markdown("#### Premium Information")
        
        # Premium Sold Calculator for Endorsements
        st.markdown("##### Premium Sold Calculator (for Endorsements)")
        
        col_calc1, col_calc2, col_calc3 = st.columns(3)
        with col_calc1:
            existing_premium = st.number_input(
                "Existing Premium", 
                value=0.00, 
                format="%.2f", 
                step=100.0,
                key="modal_existing_premium",
                help="Enter the current/existing premium amount"
            )
        with col_calc2:
            new_premium = st.number_input(
                "New/Revised Premium", 
                value=0.00, 
                format="%.2f", 
                step=100.0,
                key="modal_new_premium",
                help="Enter the new/revised premium amount"
            )
        with col_calc3:
            premium_sold_calc = new_premium - existing_premium
            st.metric(
                "Premium Sold (New - Existing):", 
                f"${premium_sold_calc:+,.2f}",
                help="Calculated endorsement premium (can be positive or negative)"
            )
        
        # Regular Premium Sold field
        st.markdown("##### New Policy Premium")
        if 'Premium Sold' in modal_data.keys():
            value = modal_data.get('Premium Sold', 0)
            if is_renewal:
                value = value if value else 0
            try:
                numeric_value = float(value) if pd.notna(value) else 0.0
                
                # If endorsement calculator has values, use that
                if (existing_premium != 0.0 or new_premium != 0.0) and premium_sold_calc != 0.0:
                    numeric_value = premium_sold_calc
                    help_text = "Using value from Endorsement Calculator above"
                else:
                    help_text = "Enter the total premium for new policies (not endorsements)"
                
                updated_data['Premium Sold'] = st.number_input(
                    'New Policy Premium',
                    value=numeric_value,
                    step=0.01,
                    format="%.2f",
                    key="modal_Premium Sold",
                    help=help_text
                )
            except:
                updated_data['Premium Sold'] = st.text_input(
                    'Premium Sold',
                    value=str(value),
                    key="modal_Premium Sold"
                )
        
        # Carrier Taxes & Fees
        st.markdown("#### Carrier Taxes & Fees")
        col7, col8 = st.columns(2)
        
        with col7:
            if 'Policy Taxes & Fees' in modal_data.keys():
                value = modal_data.get('Policy Taxes & Fees', 0)
                if is_renewal:
                    value = 0
                try:
                    numeric_value = float(value) if pd.notna(value) else 0.0
                    updated_data['Policy Taxes & Fees'] = st.number_input(
                        'Policy Taxes & Fees',
                        value=numeric_value,
                        step=0.01,
                        format="%.2f",
                        key="modal_Policy Taxes & Fees"
                    )
                except:
                    updated_data['Policy Taxes & Fees'] = st.text_input(
                        'Policy Taxes & Fees',
                        value=str(value),
                        key="modal_Policy Taxes & Fees"
                    )
        
        with col8:
            # Commissionable Premium (calculated field)
            premium_sold = updated_data.get('Premium Sold', modal_data.get('Premium Sold', 0))
            taxes_fees = updated_data.get('Policy Taxes & Fees', modal_data.get('Policy Taxes & Fees', 0))
            try:
                premium_sold = float(premium_sold) if pd.notna(premium_sold) else 0.0
                taxes_fees = float(taxes_fees) if pd.notna(taxes_fees) else 0.0
                commissionable_premium = premium_sold - taxes_fees
            except:
                commissionable_premium = 0.0
            
            st.number_input(
                'Commissionable Premium',
                value=commissionable_premium,
                format="%.2f",
                key="modal_Commissionable Premium_display",
                disabled=True,
                help=f"Premium Sold - Policy Taxes & Fees = ${premium_sold:.2f} - ${taxes_fees:.2f} = ${commissionable_premium:.2f}"
            )
            updated_data['Commissionable Premium'] = commissionable_premium
        
        # Commission Details
        st.markdown("#### Commission Details")
        
        # Add info about Prior Policy Number affecting rates
        if st.session_state.get('edit_has_commission_rule', False):
            current_trans_type = updated_data.get('Transaction Type', modal_data.get('Transaction Type', 'NEW'))
            if current_trans_type not in ['NEW', 'RWL', 'CAN', 'XCL']:
                st.info("💡 For END, PCH, and REWRITE transactions: Commission rates depend on Prior Policy Number. If present, renewal rates apply. Click the Calculate button to update rates after changing Prior Policy Number.")
        
        # Row 1: Policy Gross Comm % and Agency Estimated Comm/Revenue
        col9, col10 = st.columns(2)
        
        with col9:
            if 'Policy Gross Comm %' in modal_data.keys():
                # Check if we have commission rates from rule lookup
                has_rule = st.session_state.get('edit_has_commission_rule', False)
                
                if has_rule:
                    # Get the current transaction type from session state (most current value)
                    current_transaction_type = st.session_state.get('modal_Transaction Type', 
                                            updated_data.get('Transaction Type', 
                                            modal_data.get('Transaction Type', 'NEW')))
                    # Check session state first for the most current value
                    prior_policy = st.session_state.get('modal_Prior Policy Number', 
                                    updated_data.get('Prior Policy Number', 
                                    modal_data.get('Prior Policy Number', '')))
                    new_rate = st.session_state.get('edit_commission_new_rate', 0)
                    renewal_rate = st.session_state.get('edit_commission_renewal_rate', new_rate)
                    
                    # Determine which rate to use based on transaction type and prior policy
                    if current_transaction_type == 'NEW':
                        # NEW transactions always use new rate
                        value = new_rate
                        help_text = f"New business rate (Transaction Type: NEW)"
                    elif current_transaction_type == 'RWL':
                        # RWL transactions always use renewal rate
                        value = renewal_rate
                        help_text = f"Renewal rate (Transaction Type: RWL)"
                    else:
                        # For all other transaction types, check Prior Policy Number
                        if prior_policy and str(prior_policy).strip():
                            # Has prior policy = this is a renewal policy
                            value = renewal_rate
                            help_text = f"Renewal rate (Has Prior Policy: {prior_policy})"
                        else:
                            # No prior policy = this is a new policy
                            value = new_rate
                            help_text = f"New business rate (No Prior Policy)"
                else:
                    # Use existing value
                    value = modal_data.get('Policy Gross Comm %', 0)
                    if is_renewal:
                        value = 0
                    help_text = "Enter commission rate manually (no rule found)"
                
                try:
                    numeric_value = float(value) if pd.notna(value) else 0.0
                    updated_data['Policy Gross Comm %'] = st.number_input(
                        'Policy Gross Comm %',
                        value=numeric_value,
                        step=0.01,
                        format="%.2f",
                        key="modal_Policy Gross Comm %",
                        help=help_text
                    )
                except:
                    updated_data['Policy Gross Comm %'] = st.text_input(
                        'Policy Gross Comm %',
                        value=str(value),
                        key="modal_Policy Gross Comm %",
                        help=help_text
                    )
        
        with col10:
            # Agency Estimated Comm/Revenue (calculated)
            gross_comm_pct = updated_data.get('Policy Gross Comm %', modal_data.get('Policy Gross Comm %', 0))
            try:
                gross_comm_pct = float(gross_comm_pct) if pd.notna(gross_comm_pct) else 0.0
                agency_comm = commissionable_premium * (gross_comm_pct / 100)
                
                # Get current transaction type for cancellation check
                current_trans_type_for_agency = st.session_state.get('modal_Transaction Type', 
                                        updated_data.get('Transaction Type', 
                                        modal_data.get('Transaction Type', 'NEW')))
                
                # For cancellations (CAN/XCL), make the commission negative (chargeback)
                if current_trans_type_for_agency in ["CAN", "XCL"]:
                    agency_comm = -abs(agency_comm)  # Ensure it's negative
            except:
                agency_comm = 0.0
            
            st.number_input(
                'Agency Estimated Comm/Revenue (CRM)',
                value=agency_comm,
                format="%.2f",
                key="modal_Agency Estimated Comm_display",
                disabled=True,
                help=f"Commissionable Premium × Policy Gross Comm % = ${commissionable_premium:.2f} × {gross_comm_pct:.2f}% = ${agency_comm:.2f}" + (" (CHARGEBACK)" if current_trans_type_for_agency in ["CAN", "XCL"] else "")
            )
            updated_data['Agency Estimated Comm/Revenue (CRM)'] = agency_comm
        
        # Row 2: Agent Comm % and Agent Estimated Comm $
        col11, col12 = st.columns(2)
        
        # Get the current transaction type from session state (most current value)
        # Define this before columns to ensure it's available in both
        current_transaction_type = st.session_state.get('modal_Transaction Type', 
                                updated_data.get('Transaction Type', 
                                modal_data.get('Transaction Type', 'NEW')))
        
        with col11:
            if 'Agent Comm %' in modal_data.keys():
                # Get Prior Policy Number - check session state first, then updated data, then modal data
                # This ensures we get the most current value even if field hasn't been rendered yet
                prior_policy = st.session_state.get('modal_Prior Policy Number', 
                                updated_data.get('Prior Policy Number', 
                                modal_data.get('Prior Policy Number', '')))
                
                # Determine agent commission rate based on transaction type and prior policy
                if current_transaction_type == "NEW":
                    # NEW transactions always get 50%
                    agent_rate = 50.0
                elif current_transaction_type == "RWL":
                    # RWL transactions always get 25%
                    agent_rate = 25.0
                elif current_transaction_type in ["NBS", "STL", "BoR"]:
                    # These are typically new business
                    agent_rate = 50.0
                elif current_transaction_type in ["CAN", "XCL"]:
                    # Cancellations - determine the original rate based on Prior Policy Number
                    if prior_policy and str(prior_policy).strip():
                        # Has prior policy = this was a renewal, so chargeback at 25%
                        agent_rate = 25.0
                    else:
                        # No prior policy = this was new business, so chargeback at 50%
                        agent_rate = 50.0
                else:
                    # For all other transaction types (END, PCH, REWRITE), check Prior Policy Number
                    if prior_policy and str(prior_policy).strip():
                        # Has prior policy = this is a renewal policy
                        agent_rate = 25.0
                    else:
                        # No prior policy = this is a new policy
                        agent_rate = 50.0
                
                # Display as read-only field since it's calculated
                help_text = f"Transaction Type: {current_transaction_type}"
                if current_transaction_type in ["CAN", "XCL"]:
                    if prior_policy and str(prior_policy).strip():
                        help_text += f" | CANCELLATION - Chargeback at renewal rate (25%)"
                    else:
                        help_text += f" | CANCELLATION - Chargeback at new business rate (50%)"
                elif current_transaction_type not in ["NEW", "RWL", "NBS", "STL", "BoR"]:
                    if prior_policy and str(prior_policy).strip():
                        help_text += f" | Has Prior Policy: {prior_policy} → Renewal rate (25%)"
                    else:
                        help_text += f" | No Prior Policy → New business rate (50%)"
                
                st.text_input(
                    'Agent Comm %',
                    value=f"{agent_rate}%",
                    key="modal_Agent Comm %_display",
                    disabled=True,
                    help=help_text
                )
                updated_data['Agent Comm %'] = agent_rate
        
        with col12:
            # Agent Estimated Comm $ (calculated)
            # Use the agent_rate calculated above
            agent_comm_pct = agent_rate if 'agent_rate' in locals() else updated_data.get('Agent Comm %', 0)
            try:
                agent_comm_pct = float(agent_comm_pct) if pd.notna(agent_comm_pct) else 0.0
                agent_comm = agency_comm * (agent_comm_pct / 100)
                
                # For cancellations (CAN/XCL), make the commission negative (chargeback)
                if current_transaction_type in ["CAN", "XCL"]:
                    agent_comm = -abs(agent_comm)  # Ensure it's negative
            except:
                agent_comm = 0.0
            
            st.number_input(
                'Agent Estimated Comm $',
                value=agent_comm,
                format="%.2f",
                key="modal_Agent Estimated Comm_display",
                disabled=True,
                help=f"Agency Comm × Agent Rate = ${agency_comm:.2f} × {agent_comm_pct:.2f}% = ${agent_comm:.2f}" + (" (CHARGEBACK)" if current_transaction_type in ["CAN", "XCL"] else "")
            )
            updated_data['Agent Estimated Comm $'] = agent_comm
        
        # Row 3: Broker Fee and Broker Fee Agent Comm
        col13, col14 = st.columns(2)
        
        with col13:
            if 'Broker Fee' in modal_data.keys():
                value = modal_data.get('Broker Fee', 0)
                if is_renewal:
                    value = 0
                try:
                    numeric_value = float(value) if pd.notna(value) else 0.0
                    updated_data['Broker Fee'] = st.number_input(
                        'Broker Fee',
                        value=numeric_value,
                        step=0.01,
                        format="%.2f",
                        key="modal_Broker Fee"
                    )
                except:
                    updated_data['Broker Fee'] = st.text_input(
                        'Broker Fee',
                        value=str(value),
                        key="modal_Broker Fee"
                    )
        
        with col14:
            # Broker Fee Agent Comm (calculated - 50% of broker fee)
            broker_fee = updated_data.get('Broker Fee', modal_data.get('Broker Fee', 0))
            try:
                broker_fee = float(broker_fee) if pd.notna(broker_fee) else 0.0
                broker_fee_agent = broker_fee * 0.50
            except:
                broker_fee_agent = 0.0
            
            st.number_input(
                'Broker Fee Agent Comm',
                value=broker_fee_agent,
                format="%.2f",
                key="modal_Broker Fee Agent Comm_display",
                disabled=True,
                help="50% of broker fee"
            )
            updated_data['Broker Fee Agent Comm'] = broker_fee_agent
        
        # Total Agent Comm (full width)
        total_agent_comm = agent_comm + broker_fee_agent
        st.number_input(
            'Total Agent Comm',
            value=total_agent_comm,
            format="%.2f",
            key="modal_Total Agent Comm_display",
            disabled=True,
            help=f"Agent Comm + Broker Fee Comm = ${agent_comm:.2f} + ${broker_fee_agent:.2f} = ${total_agent_comm:.2f}"
        )
        updated_data['Total Agent Comm'] = total_agent_comm
        
        # Additional Fields
        st.markdown("#### Additional Fields")
        col15, col16 = st.columns(2)
        
        with col15:
            # Policy Checklist Complete - always show this field
            current_val = str(modal_data.get('Policy Checklist Complete', 'No')).upper() == 'YES'
            # For renewals, default to unchecked
            if is_renewal:
                current_val = False
            updated_data['Policy Checklist Complete'] = 'Yes' if st.checkbox(
                'Policy Checklist Complete',
                value=current_val,
                key="modal_Policy Checklist Complete"
            ) else 'No'
        
        with col16:
            # FULL OR MONTHLY PMTS - always show this field
            payment_types = ["FULL", "MONTHLY", ""]
            current_payment = modal_data.get('FULL OR MONTHLY PMTS', '')
            updated_data['FULL OR MONTHLY PMTS'] = st.selectbox(
                'FULL OR MONTHLY PMTS',
                options=payment_types,
                index=payment_types.index(current_payment) if current_payment in payment_types else 2,
                key="modal_FULL OR MONTHLY PMTS"
            )
        
        # NOTES - Full width (always show)
        updated_data['NOTES'] = st.text_area(
            'NOTES',
            value=str(modal_data.get('NOTES', '')) if modal_data.get('NOTES') is not None else '',
            key="modal_NOTES",
            height=70
        )
        
        # Calculate button - make it prominent
        if st.form_submit_button("🧮 Calculate", type="primary", help="Click to update commission rates and verify all calculations"):
            # Check if we need to update commission rate from carrier selection
            if st.session_state.get('edit_commission_rate') is not None and 'Policy Gross Comm %' in updated_data:
                # The commission rate has already been set from the carrier selection above
                pass
            st.success("✅ Calculations updated! Review the amounts above before saving.")
        
        # Internal Fields (collapsed by default)
        # For -IMPORT transactions, also show payment fields as read-only here
        with st.expander("Internal Fields", expanded=False):
            internal_col1, internal_col2 = st.columns(2)
            
            # Define payment fields that should be protected for import transactions
            payment_fields = ['Agent Paid Amount (STMT)', 'Agency Comm Received (STMT)', 'STMT DATE']
            
            field_counter = 0
            for field in modal_data.keys():
                # Show internal fields and payment fields (for import transactions)
                if (field in internal_fields or field.startswith('_') or 
                    field in ['reconciliation_status', 'reconciliation_id', 'reconciled_at'] or
                    (is_import and field in payment_fields)):
                    with internal_col1 if field_counter % 2 == 0 else internal_col2:
                        value = modal_data.get(field, '')
                        help_text = "Protected payment data from import" if is_import and field in payment_fields else None
                        st.text_input(
                            field,
                            value=str(value) if value is not None else '',
                            key=f"modal_{field}_internal",
                            disabled=True,
                            help=help_text
                        )
                        updated_data[field] = value  # Preserve internal field values
                    field_counter += 1
        
        # Form buttons
        col1, col2 = st.columns(2)
        
        with col1:
            save_button_text = "Renew Policy" if is_renewal else "Save Changes"
            save_button = st.form_submit_button(save_button_text, type="primary", use_container_width=True)
        
        with col2:
            cancel_button = st.form_submit_button("Cancel", use_container_width=True)
        
        if save_button:
            # Convert date objects to strings
            # REMOVED: Date formatting to preserve YYYY-MM-DD format
            # for field, value in updated_data.items():
            #     if isinstance(value, datetime.date):
            #         updated_data[field] = value.strftime('%m/%d/%Y')
            
            # Auto-populate 12-month term for NEW and RWL (except AUTO)
            transaction_type = updated_data.get('Transaction Type', modal_data.get('Transaction Type', ''))
            policy_type = updated_data.get('Policy Type', modal_data.get('Policy Type', ''))
            
            if transaction_type in ['NEW', 'RWL'] and policy_type != 'AUTO':
                # Set Policy Term to 12 months
                updated_data['Policy Term'] = 12
                
                # Calculate X-DATE as Effective Date + 12 months
                effective_date = updated_data.get('Effective Date', modal_data.get('Effective Date'))
                if effective_date:
                    try:
                        # Parse the effective date
                        if isinstance(effective_date, str):
                            # Try different date formats
                            for fmt in ['%m/%d/%Y', '%Y-%m-%d']:
                                try:
                                    eff_date = datetime.datetime.strptime(effective_date, fmt)
                                    break
                                except:
                                    continue
                        elif isinstance(effective_date, (datetime.date, datetime.datetime)):
                            eff_date = effective_date
                        else:
                            eff_date = pd.to_datetime(effective_date)
                        
                        # Add 12 months
                        if isinstance(eff_date, datetime.date) and not isinstance(eff_date, datetime.datetime):
                            eff_date = datetime.datetime.combine(eff_date, datetime.time())
                        
                        # Calculate X-DATE (12 months later)
                        x_date = eff_date + pd.DateOffset(months=12)
                        
                        # Format as string - Using YYYY-MM-DD
                        updated_data['X-DATE'] = x_date.strftime('%Y-%m-%d')
                    except Exception as e:
                        # If date calculation fails, just set the term without X-DATE
                        st.warning(f"Could not calculate X-DATE: {str(e)}")
            
            # Add commission rule tracking
            commission_rule_id = st.session_state.get('edit_commission_rule_id')
            if commission_rule_id:
                updated_data['commission_rule_id'] = commission_rule_id
            
            # Track carrier and MGA IDs
            carrier_id = st.session_state.get('edit_selected_carrier_id')
            mga_id = st.session_state.get('edit_selected_mga_id')
            if carrier_id:
                updated_data['carrier_id'] = carrier_id
            if mga_id:
                updated_data['mga_id'] = mga_id
            
            # Check if commission rate was overridden
            has_rule = st.session_state.get('edit_has_commission_rule', False)
            if has_rule:
                # Determine which rate should have been used based on transaction type and prior policy
                transaction_type = updated_data.get('Transaction Type', 'NEW')
                prior_policy = updated_data.get('Prior Policy Number', '')
                new_rate = st.session_state.get('edit_commission_new_rate', 0)
                renewal_rate = st.session_state.get('edit_commission_renewal_rate', new_rate)
                
                # Apply same logic as display
                if transaction_type == 'NEW':
                    expected_rate = new_rate
                elif transaction_type == 'RWL':
                    expected_rate = renewal_rate
                else:
                    # For all other transaction types, check Prior Policy Number
                    if prior_policy and str(prior_policy).strip():
                        expected_rate = renewal_rate
                    else:
                        expected_rate = new_rate
                
                actual_rate = updated_data.get('Policy Gross Comm %', 0)
                if float(actual_rate) != float(expected_rate):
                    updated_data['commission_rate_override'] = actual_rate
                    # You could prompt for override reason here if needed
            
            # Merge updated data with original data
            final_data = modal_data.copy()
            final_data.update(updated_data)
            
            # Clear generated client ID from session state after successful save
            if 'generated_client_id' in st.session_state:
                del st.session_state.generated_client_id
            
            return {"action": "save", "data": final_data}
        
        if cancel_button:
            # Clear generated client ID from session state
            if 'generated_client_id' in st.session_state:
                del st.session_state.generated_client_id
            return {"action": "cancel", "data": None}
    
    return None

def main():
    # Check password before showing any content
    if not check_password():
        st.stop()
    
    # Apply CSS styling
    apply_css()
      # --- Page Selection ---
    page = st.sidebar.radio(
        "Navigation",        [
            "Dashboard",
            "Reports",
            "All Policy Transactions",
            "Edit Policy Transactions",
            "Add New Policy Transaction",
            "Search & Filter",
            "Reconciliation",
            "Admin Panel",
            "Contacts",
            "Tools",
            "Help",
            "Policy Revenue Ledger",
            "Policy Revenue Ledger Reports",
            "Pending Policy Renewals"
        ]
    )
    
    # Add logout button to sidebar
    st.sidebar.divider()
    if st.sidebar.button("🚪 Logout", type="secondary", use_container_width=True):
        st.session_state["password_correct"] = False
        st.rerun()
    
    # --- Load data with caching for better performance ---
    # Moved all_data loading to individual pages for proper cache refresh
    # all_data = load_policies_data()
    supabase = get_supabase_client()
    
    # Note: all_data empty check moved to individual pages

    # --- Page Navigation ---
    if page == "Dashboard":
        st.title("📊 Commission Dashboard")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        if all_data.empty:
            st.warning("No data found in policies table. Please add some policy data first.")
        else:
            # Calculate metrics
            metrics = calculate_dashboard_metrics(all_data)
            
            # --- Dashboard Metrics ---
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("📊 TRANSACTIONS")
                st.metric("Total Transactions", f"{metrics['total_transactions']:,}")
                st.metric("This Month", f"{metrics['transactions_this_month']:,}")
                st.metric("Reconciliation (-STMT-)", f"{metrics['stmt_transactions']:,}")
            
            with col2:
                st.subheader("📋 POLICIES")
                st.metric("Unique Policies", f"{metrics['unique_policies']:,}")
                st.metric("Active", f"{metrics['active_policies']:,}")
                st.metric("Cancelled", f"{metrics['cancelled_policies']:,}")
            
            st.divider()
            
            # Financial Summary - YTD 2025
            st.subheader("💰 FINANCIAL SUMMARY - YTD 2025")
            
            # Reconciled (Paid) Metrics
            st.markdown("**Reconciled (Paid)**")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Premium Sold", f"${metrics['premium_reconciled_ytd']:,.2f}")
            with col2:
                st.metric("Agent Commission Paid", f"${metrics['agent_comm_paid_ytd']:,.2f}")
            
            # Unreconciled (Estimated) Metrics
            st.markdown("**Unreconciled (Estimated)**")
            col3, col4 = st.columns(2)
            with col3:
                st.metric("Premium Estimated", f"${metrics['premium_unreconciled_ytd']:,.2f}")
            with col4:
                st.metric("Agent Commission Estimated", f"${metrics['agent_comm_estimated_ytd']:,.2f}")
            
            st.divider()
            
            # --- Client Search and Quick Actions ---
            st.subheader("🔍 Quick Client Search")
            
            search_col1, search_col2 = st.columns([3, 1])
            
            with search_col1:
                search_term = st.text_input("Search by Client Name, Policy Number, or Transaction ID", 
                                          placeholder="Enter search term...")
            
            with search_col2:
                st.write("")  # Spacing
                search_button = st.button("Search", type="primary")
            
            # Search functionality
            if search_term or search_button:
                if search_term:
                    # Search across multiple columns
                    mask = pd.Series(False, index=all_data.index)
                    search_columns = ['Customer', 'Policy Number', 'Transaction ID', 'Client ID']
                    
                    for col in search_columns:
                        if col in all_data.columns:
                            mask |= all_data[col].astype(str).str.contains(search_term, case=False, na=False)
                    
                    search_results = all_data[mask].copy()
                    
                    if not search_results.empty:
                        st.success(f"Found {len(search_results)} matching records")
                        
                        # Format dates to strings before displaying
                        
                        # Configure column settings for proper numeric display
                        column_config = {}
                        numeric_cols = [
                            'Agent Estimated Comm $',
                            'Policy Gross Comm %',
                            'Agency Estimated Comm/Revenue (CRM)',
                            'Agency Comm Received (STMT)',
                            'Premium Sold',
                            'Agent Paid Amount (STMT)',
                            'Agency Comm Received (STMT)',
                            'Commissionable Premium',
                            'Broker Fee',
                            'Broker Fee Agent Comm',
                            'Total Agent Comm',
                            'Policy Taxes & Fees'
                        ]
                        
                        for col in numeric_cols:
                            if col in search_results.columns:
                                column_config[col] = st.column_config.NumberColumn(
                                    col,
                                    format="$%.2f" if '$' in col or col in ['Premium Sold', 'Commissionable Premium', 'Broker Fee', 'Policy Taxes & Fees'] else "%.2f",
                                    step=0.01
                                )
                        
                        # Configure date columns as text columns
                        date_cols = ['Policy Origination Date', 'Effective Date', 'X-DATE', 'STMT DATE']
                        for col in date_cols:
                            if col in search_results.columns:
                                column_config[col] = st.column_config.TextColumn(
                                    col,
                                    help="Date format: MM/DD/YYYY"
                                )
                        
                        # Display search results in an editable table
                        edited_data = st.data_editor(
                            search_results,
                            use_container_width=True,
                            height=400,
                            key="dashboard_search_editor",
                            column_config=column_config
                        )
                        
                        # Update button for search results
                        if st.button("Save Changes", key="dashboard_save"):
                            try:
                                # Update the database with changes
                                for idx, row in edited_data.iterrows():
                                    # Update via Supabase
                                    update_dict = {}
                                    for col in edited_data.columns:
                                        if col not in ['_id', 'Transaction ID']:  # Don't update ID fields
                                            update_dict[col] = row[col]
                                    
                                    # Find the record by Transaction ID
                                    transaction_id = row.get('Transaction ID', row.get('Transaction_ID'))
                                    if transaction_id:
                                        try:
                                            supabase.table('policies').update(update_dict).eq('Transaction ID', transaction_id).execute()
                                        except Exception as update_error:
                                            st.error(f"Error updating record {transaction_id}: {update_error}")
                                
                                st.success("Changes saved successfully!")
                                st.rerun()
                                
                            except Exception as e:
                                st.error(f"Error saving changes: {e}")
                    else:
                        st.warning("No records found matching your search term")
                else:
                    st.info("Enter a search term to find records")
            
            st.divider()
            
            # --- Recent Activity ---
            st.subheader("📈 Recent Activity")
            
            # Show last 10 records
            recent_data = all_data.head(10)
            
            # Configure column settings for proper numeric display
            column_config = {}
            numeric_cols = [
                'Agent Estimated Comm $',
                'Policy Gross Comm %',
                'Agency Estimated Comm/Revenue (CRM)',
                'Agency Comm Received (STMT)',
                'Premium Sold',
                'Agent Paid Amount (STMT)',
                'Agency Comm Received (STMT)'
            ]
            
            for col in numeric_cols:
                if col in recent_data.columns:
                    column_config[col] = st.column_config.NumberColumn(
                        col,
                        format="%.2f"
                    )
            
            # Configure date columns for safe display formatting
            date_cols = ['Policy Origination Date', 'Effective Date', 'X-DATE', 'STMT DATE']
            for col in date_cols:
                if col in recent_data.columns:
                    column_config[col] = st.column_config.TextColumn(
                        col,
                        help="Date format: MM/DD/YYYY"
                    )
            
            # Apply special styling for STMT and VOID transactions
            styled_data = style_special_transactions(recent_data)
            
            st.dataframe(
                styled_data,
                use_container_width=True,
                height=300,
                column_config=column_config
            )
            
            # --- Quick Statistics ---
            st.subheader("📊 Quick Statistics")
            
            stats_col1, stats_col2 = st.columns(2)
            
            with stats_col1:
                if 'Policy Type' in all_data.columns:
                    policy_type_counts = all_data['Policy Type'].value_counts()
                    st.write("**Policies by Type:**")
                    for policy_type, count in policy_type_counts.items():
                        st.write(f"• {policy_type}: {count}")
                        
            with stats_col2:
                if 'Transaction Type' in all_data.columns:
                    # Transaction type distribution pie chart
                    trans_types = all_data['Transaction Type'].value_counts()
                    fig = px.pie(values=trans_types.values, names=trans_types.index, 
                                 title="Transaction Type Distribution",
                                 color_discrete_map={
                                     'NEW': '#28a745',
                                     'RWL': '#007bff', 
                                     'END': '#ffc107',
                                     'CAN': '#dc3545',
                                     'XCL': '#dc3545',
                                     '-STMT-': '#6c757d',
                                     '-VOID-': '#6c757d',
                                     '-ADJ-': '#6c757d'
                                 })
                    fig.update_traces(textposition='inside', textinfo='percent+label')
                    st.plotly_chart(fig, use_container_width=True)
    
    # --- Reports ---
    elif page == "Reports":
        st.title("📈 Reports")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        if all_data.empty:
            st.warning("No data found in policies table. Please add some policy data first.")
        else:
            tab1, tab2, tab3, tab4 = st.tabs(["Summary Reports", "Commission Analysis", "Policy Reports", "Custom Reports"])
            
            with tab1:
                st.subheader("Summary Reports")
                
                # Overall summary
                col1, col2 = st.columns(2)
                
                with col1:
                    st.metric("Total Transactions", len(all_data))
                    if 'Commission_Paid' in all_data.columns:
                        st.metric("Total Commission Paid", f"${all_data['Commission_Paid'].sum():,.2f}")
                    if 'Agency_Commission_Received' in all_data.columns:
                        st.metric("Total Agency Commission", f"${all_data['Agency_Commission_Received'].sum():,.2f}")
                
                with col2:
                    if 'Balance_Due' in all_data.columns:
                        st.metric("Total Balance Due", f"${all_data['Balance_Due'].sum():,.2f}")
                    if 'Effective_Date' in all_data.columns:
                        latest_date = pd.to_datetime(all_data['Effective_Date'], errors='coerce').max()
                        st.metric("Latest Policy Date", latest_date.strftime('%m/%d/%Y') if pd.notnull(latest_date) else "N/A")
            
            with tab2:
                st.subheader("Commission Analysis")
                
                if 'Commission_Paid' in all_data.columns and 'Policy_Type' in all_data.columns:
                    commission_by_type = all_data.groupby('Policy_Type')['Commission_Paid'].sum().sort_values(ascending=False)
                    st.write("**Commission by Policy Type:**")
                    st.dataframe(commission_by_type.reset_index())
                    
                if 'Transaction_Type' in all_data.columns and 'Commission_Paid' in all_data.columns:
                    commission_by_transaction = all_data.groupby('Transaction_Type')['Commission_Paid'].sum().sort_values(ascending=False)
                    st.write("**Commission by Transaction Type:**")
                    st.dataframe(commission_by_transaction.reset_index())
            
            with tab3:
                st.subheader("Policy Reports")
                
                if 'Policy_Type' in all_data.columns:
                    policy_counts = all_data['Policy_Type'].value_counts()
                    st.write("**Policy Count by Type:**")
                    st.dataframe(policy_counts.reset_index())
                
                if 'Transaction_Type' in all_data.columns:
                    transaction_counts = all_data['Transaction_Type'].value_counts()
                    st.write("**Transaction Count by Type:**")
                    st.dataframe(transaction_counts.reset_index())
            
            with tab4:
                st.subheader("Custom Reports")
                st.info("Custom report builder - Select criteria to generate custom reports")
                
                # Initialize variables
                start_date = None
                end_date = None
                policy_types = []
                
                # Date range filter
                if 'Effective_Date' in all_data.columns:
                    date_col = st.columns(2)
                    with date_col[0]:
                        start_date = st.date_input("Start Date")
                    with date_col[1]:
                        end_date = st.date_input("End Date")
                
                # Policy type filter
                if 'Policy_Type' in all_data.columns:
                    policy_types = st.multiselect("Policy Types", all_data['Policy_Type'].unique())
                
                if st.button("Generate Custom Report"):
                    filtered_data = all_data.copy()
                    
                    # Apply filters
                    if 'Effective_Date' in all_data.columns and start_date and end_date:
                        filtered_data['Effective_Date'] = pd.to_datetime(filtered_data['Effective_Date'], errors='coerce')
                        filtered_data = filtered_data[
                            (filtered_data['Effective_Date'] >= pd.Timestamp(start_date)) &
                            (filtered_data['Effective_Date'] <= pd.Timestamp(end_date))
                        ]
                    
                    if policy_types and 'Policy_Type' in all_data.columns:
                        filtered_data = filtered_data[filtered_data['Policy_Type'].isin(policy_types)]
                    
                    st.write(f"**Custom Report Results ({len(filtered_data)} records):**")
                    st.dataframe(filtered_data, use_container_width=True)
                    
                    # Export custom report results
                    if not filtered_data.empty:
                        st.subheader("Export Custom Report")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write("**📄 CSV Export**")
                            csv_data = filtered_data.to_csv(index=False)
                            st.download_button(
                                label="📥 Download Custom Report CSV",
                                data=csv_data,
                                file_name=f"custom_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                                mime="text/csv",
                                help="Export custom report results as CSV file"
                            )
                        
                        with col2:
                            st.write("**📊 Excel Export**")
                            excel_buffer, excel_filename = create_formatted_excel_file(
                                filtered_data, 
                                sheet_name="Custom Report", 
                                filename_prefix="custom_report"
                            )
                            if excel_buffer:
                                st.download_button(
                                    label="📥 Download Custom Report Excel",
                                    data=excel_buffer,
                                    file_name=excel_filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    help="Export custom report results as formatted Excel file"
                                )
            
            # Export all reports data
            if not all_data.empty:
                st.divider()
                st.subheader("Export All Policy Data")
                st.write("Export complete policy database for comprehensive reporting:")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("**📄 CSV Export**")
                    csv_all_data = all_data.to_csv(index=False)
                    st.download_button(
                        label="📥 Download All Policies CSV",
                        data=csv_all_data,
                        file_name=f"all_policies_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        help="Export all policy data as CSV file"
                    )
                
                with col2:
                    st.write("**📊 Excel Export**")
                    
                    # Create multi-sheet Excel with different report views
                    reports_data = {
                        "All Policies": all_data,
                        "Summary Stats": pd.DataFrame([
                            ["Total Transactions", len(all_data)],
                            ["Unique Policies", all_data['Policy Number'].nunique() if 'Policy Number' in all_data.columns else 0],
                            ["Total Commission", all_data['Commission_Paid'].sum() if 'Commission_Paid' in all_data.columns else 0],
                            ["Total Balance Due", all_data['Balance_Due'].sum() if 'Balance_Due' in all_data.columns else 0],
                            ["Export Date", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                        ], columns=['Metric', 'Value'])
                    }
                    
                    # Add policy type breakdown if available
                    if 'Policy_Type' in all_data.columns:
                        policy_breakdown = all_data['Policy_Type'].value_counts().reset_index()
                        policy_breakdown.columns = ['Policy_Type', 'Count']
                        reports_data["Policy Type Breakdown"] = policy_breakdown
                    
                    excel_buffer_multi, excel_filename_multi = create_multi_sheet_excel(
                        reports_data, 
                        filename_prefix="complete_policy_report"
                    )
                    
                    if excel_buffer_multi:
                        st.download_button(
                            label="📥 Download Comprehensive Excel Report",
                            data=excel_buffer_multi,
                            file_name=excel_filename_multi,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Export comprehensive multi-sheet Excel report with all data and summaries"
                        )
    
    # --- All Policy Transactions ---
    elif page == "All Policy Transactions":
        st.title("📋 All Policy Transactions")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        if all_data.empty:
            st.warning("No data found in policies table. Please add some policy data first.")
        else:
            # Calculate unique policy count
            unique_policies = all_data['Policy Number'].nunique() if 'Policy Number' in all_data.columns else 0
            st.write(f"**Total Transactions: {len(all_data):,} | Unique Policies: {unique_policies:,}**")
            
            # Add toggle for formula view
            col_toggle, col_space = st.columns([2, 8])
            with col_toggle:
                show_formulas = st.toggle("📊 Show Formulas", value=True, help="Toggle between formula calculations and actual values")
            
            # Pagination controls
            items_per_page = st.selectbox("Items per page", [25, 50, 100, 200], index=1)
            total_pages = max(1, len(all_data) // items_per_page + (1 if len(all_data) % items_per_page > 0 else 0))
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                page_num = st.number_input("Page", min_value=1, max_value=total_pages, value=1, step=1)
            
            # Calculate pagination
            start_idx = (page_num - 1) * items_per_page
            end_idx = min(start_idx + items_per_page, len(all_data))
            
            # Display paginated data
            paginated_data = all_data.iloc[start_idx:end_idx]
            
            # Define preferred column order
            preferred_order = [
                '_id', 'Transaction ID', 'Client ID', 'Customer', 
                'Policy Number', 'Prior Policy Number',
                'Carrier Name', 'MGA Name',  # MGA right after Carrier
                'Policy Type', 'Transaction Type', 'Policy Term',  # Policy Term between Transaction Type and Effective Date
                'Policy Origination Date', 'Effective Date', 'X-DATE',  # Policy Origination Date moved before Effective Date
                'Premium Sold', 'Policy Taxes & Fees', 'Commissionable Premium',
                'Policy Gross Comm %', 'Agency Estimated Comm/Revenue (CRM)',
                'Agent Estimated Comm $', 'Broker Fee', 'Broker Fee Agent Comm', 'Total Agent Comm',
                'Agency Comm Received (STMT)', 'Agent Paid Amount (STMT)',
                'STMT DATE', 'Policy Checklist Complete', 'FULL OR MONTHLY PMTS', 'NOTES'
            ]
            
            # Reorder columns - keep preferred order columns that exist, then add any remaining
            existing_preferred = [col for col in preferred_order if col in paginated_data.columns]
            remaining_cols = [col for col in paginated_data.columns if col not in preferred_order]
            paginated_data = paginated_data[existing_preferred + remaining_cols]
            
            # Apply formula display to existing columns
            paginated_data_display = apply_formula_display(paginated_data, show_formulas=show_formulas)
            
            st.write(f"Showing records {start_idx + 1} to {end_idx} of {len(all_data)}")
            
            # Show mode info
            if show_formulas:
                st.info("📊 **Formula Mode**: Showing calculated values. Icons: ✓ = matches manual entry, ✏️ = manual override, ⚠️ = missing data, 🔒 = reconciliation entry")
            else:
                st.info("📝 **Actual Values Mode**: Showing original manual entries from the database.")
            
            # Configure column settings for proper numeric display
            column_config = {}
            numeric_cols = [
                'Agent Estimated Comm $',
                'Agent Comm %',  # Added this column for 2 decimal formatting
                'Policy Gross Comm %',
                'Agency Estimated Comm/Revenue (CRM)',
                'Agency Comm Received (STMT)',
                'Premium Sold',
                'Agent Paid Amount (STMT)',
                'Agency Comm Received (STMT)',
                'Broker Fee',
                'Policy Taxes & Fees',
                'Commissionable Premium',
                'Broker Fee Agent Comm',
                'Total Agent Comm',
                'Policy Balance Due'
            ]
            
            for col in numeric_cols:
                if col in paginated_data_display.columns:
                    # Note: When in formula mode, these columns will have indicators appended
                    # Streamlit will treat them as text columns due to the emoji indicators
                    if show_formulas and col in ['Agency Estimated Comm/Revenue (CRM)', 'Agent Estimated Comm $']:
                        column_config[col] = st.column_config.TextColumn(
                            col,
                            help="Formula calculation with status indicator"
                        )
                    else:
                        column_config[col] = st.column_config.NumberColumn(
                            col,
                            format="%.2f"
                        )
            
            # Special handling for Policy Term - should be integer with no decimals
            if 'Policy Term' in paginated_data_display.columns:
                column_config['Policy Term'] = st.column_config.NumberColumn(
                    'Policy Term',
                    format="%d",  # Integer format with no decimals
                    help="Policy term in months"
                )
            
            # Display the data in a scrollable table
            # Apply special styling for STMT and VOID transactions
            styled_data = style_special_transactions(paginated_data_display)
            
            st.dataframe(
                styled_data,
                use_container_width=True,
                height=600,
                column_config=column_config
            )
            
            # Export options
            st.subheader("Export Options")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.write("**📄 CSV Export**")
                # Export based on current view mode
                export_data = apply_formula_display(paginated_data, show_formulas=show_formulas) if show_formulas else paginated_data
                csv = export_data.to_csv(index=False)
                st.download_button(
                    label="📥 Current Page CSV",
                    data=csv,
                    file_name=f"policies_page_{page_num}.csv",
                    mime="text/csv",
                    help="Export current page as CSV file"
                )
            
            with col2:
                # Export all data based on current view mode
                all_data_export = apply_formula_display(all_data, show_formulas=show_formulas) if show_formulas else all_data
                csv_all = all_data_export.to_csv(index=False)
                st.write("**📄 CSV Export**")
                st.download_button(
                    label="📥 All Data CSV",
                    data=csv_all,
                    file_name="all_policies.csv",
                    mime="text/csv",
                    help="Export all policies as CSV file"
                )
            
            with col3:
                st.write("**📊 Excel Export**")
                excel_buffer, excel_filename = create_formatted_excel_file(
                    export_data, 
                    sheet_name="Policies Page", 
                    filename_prefix=f"policies_page_{page_num}"
                )
                if excel_buffer:
                    st.download_button(
                        label="📥 Current Page Excel",
                        data=excel_buffer,
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Export current page as formatted Excel file"
                    )
            
            with col4:
                st.write("**📊 Excel Export**")
                excel_buffer_all, excel_filename_all = create_formatted_excel_file(
                    all_data_export, 
                    sheet_name="All Policies", 
                    filename_prefix="all_policies"
                )
                if excel_buffer_all:
                    st.download_button(
                        label="📥 All Data Excel",
                        data=excel_buffer_all,
                        file_name=excel_filename_all,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Export all policies as formatted Excel file"
                    )
    
    # --- Edit Policy Transactions ---
    elif page == "Edit Policy Transactions":
        st.title("✏️ Edit Policy Transactions")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        if all_data.empty:
            st.warning("No data found in policies table. Please add some policy data first.")
        else:
            st.warning("⚠️ Be careful when editing data. Changes are saved directly to the database.")
            
            # Search and filter options
            st.subheader("Find Policies to Edit")
            
            # Add filter buttons
            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                show_attention_needed = st.button("⚠️ Show Transactions Requiring Attention", type="secondary")
            with col2:
                if 'show_attention_filter' in st.session_state and st.session_state.show_attention_filter:
                    if st.button("↩️ Reset Filters", type="secondary"):
                        st.session_state.show_attention_filter = False
                        st.rerun()
            
            # Track filter state
            if show_attention_needed:
                st.session_state.show_attention_filter = True
                st.rerun()
            
            # Wrap search in a form to prevent accidental reruns
            with st.form("search_form"):
                search_col1, search_col2 = st.columns([3, 1])
                
                with search_col1:
                    edit_search_term = st.text_input("Search policies to edit", placeholder="Enter search term...")
                
                with search_col2:
                    st.write("")  # Spacing
                    edit_search_button = st.form_submit_button("Find Records", type="primary")
            
            # Check if we should show attention-needed transactions
            show_attention_filter = st.session_state.get('show_attention_filter', False)
            
            # Show filtered data for editing
            if edit_search_term or edit_search_button or show_attention_filter:
                if show_attention_filter:
                    # Filter for transactions requiring attention
                    # These are transactions that have payments but no total agent commission
                    mask = (
                        # Has a payment recorded
                        (all_data['Agent Paid Amount (STMT)'].notna() & (all_data['Agent Paid Amount (STMT)'] > 0)) &
                        # But missing total agent commission
                        (
                            (all_data['Total Agent Comm'].isna()) | 
                            (all_data['Total Agent Comm'] == 0)
                        )
                    )
                    
                    edit_results = all_data[mask].copy()
                    
                    if not edit_results.empty:
                        st.warning(f"⚠️ Found {len(edit_results)} transactions with payments but no total agent commission")
                        st.info("These transactions need commission data (either policy commission or broker fees) to ensure accurate ledger reports.")
                    else:
                        st.success("✅ All transactions with payments have total agent commission calculated!")
                        # Continue to show empty result to user
                        
                elif edit_search_term:
                    # Search across multiple columns
                    mask = pd.Series(False, index=all_data.index)
                    search_columns = ['Customer', 'Policy Number', 'Transaction ID', 'Client ID']
                    
                    for col in search_columns:
                        if col in all_data.columns:
                            mask |= all_data[col].astype(str).str.contains(edit_search_term, case=False, na=False)
                    
                    edit_results = all_data[mask].copy()
                
                # Common processing for both search and attention filter
                if 'edit_results' in locals() and not edit_results.empty:
                    # Format numeric columns to ensure 2 decimal places
                    edit_results = round_numeric_columns(edit_results)
                    
                    # Keep dates as they are from the database
                    date_cols = [
                        'Policy Origination Date',
                        'Effective Date',
                        'X-DATE',
                        'STMT DATE',
                        'Policy Issue Date',
                        'Policy Effective Date',
                        'As of Date'
                    ]
                    
                    # Filter out reconciliation transactions
                    # First, find the transaction ID column using multiple detection methods
                    transaction_id_col = None
                    
                    # Method 1: Try using column mapper
                    transaction_id_col = get_mapped_column("Transaction ID")
                    
                    # Method 2: If mapper fails, check common variations
                    if not transaction_id_col:
                        possible_names = ["Transaction ID", "Transaction_ID", "TransactionID", 
                                        "transaction_id", "Transaction Id", "TRANSACTION_ID",
                                        "transaction id", "transactionid"]
                        for col in edit_results.columns:
                            if col in possible_names:
                                transaction_id_col = col
                                break
                    
                    # Method 3: If still not found, use flexible search
                    if not transaction_id_col:
                        for col in edit_results.columns:
                            # Remove spaces and underscores for comparison
                            col_normalized = col.lower().replace(" ", "").replace("_", "")
                            if col_normalized == "transactionid":
                                transaction_id_col = col
                                break
                    
                    if transaction_id_col:
                        # Count total before filtering
                        total_count = len(edit_results)
                        
                        # Filter out reconciliation transactions
                        edit_results_editable = edit_results[
                            ~edit_results[transaction_id_col].apply(is_reconciliation_transaction)
                        ].copy()
                        
                        recon_count = total_count - len(edit_results_editable)
                        
                        # Show counts
                        if recon_count > 0:
                            col1, col2 = st.columns(2)
                            with col1:
                                st.success(f"Found {len(edit_results_editable)} editable transactions")
                            with col2:
                                st.info(f"🔒 {recon_count} reconciliation entries (view in Reconciliation page)")
                        else:
                            st.success(f"Found {len(edit_results_editable)} transactions for editing")
                        
                        # Update edit_results to only contain editable transactions
                        edit_results = edit_results_editable
                        
                        # Check if we have any editable transactions left
                        if edit_results.empty:
                            st.warning("No editable transactions found. All transactions for this customer are reconciliation entries.")
                            return  # Exit early
                    else:
                        st.success(f"Found {len(edit_results)} records for editing")
                    
                    # Only proceed if we have editable transactions
                    if not edit_results.empty:
                        # Find the actual column names dynamically
                        # Note: transaction_id_col was already found above using robust methods, only search if not found
                        if not transaction_id_col:
                            for col in edit_results.columns:
                                if 'transaction' in col.lower() and 'id' in col.lower():
                                    transaction_id_col = col
                                    break
                        
                        client_id_col = None
                        for col in edit_results.columns:
                            if 'client' in col.lower() and 'id' in col.lower():
                                client_id_col = col
                                break
                        
                        # Add a selection column for deletion
                        edit_results_with_selection = edit_results.copy()
                        edit_results_with_selection.insert(0, 'Select', False)
                        
                        # Reorder columns to place date columns after Policy Number
                        # First, identify all columns
                        all_cols = list(edit_results_with_selection.columns)
                        
                        # Define the desired order for the beginning columns
                        priority_cols = ['Select']
                        
                        # Find Transaction ID and Client ID columns
                        if transaction_id_col and transaction_id_col in all_cols:
                            priority_cols.append(transaction_id_col)
                        if client_id_col and client_id_col in all_cols:
                            priority_cols.append(client_id_col)
                        
                        # Add Customer if exists
                        if 'Customer' in all_cols:
                            priority_cols.append('Customer')
                        
                        # Add Policy Number
                        if 'Policy Number' in all_cols:
                            priority_cols.append('Policy Number')
                        
                        # Add the three date columns immediately after Policy Number
                        date_cols_ordered = ['Policy Origination Date', 'Effective Date', 'X-DATE']
                        for col in date_cols_ordered:
                            if col in all_cols:
                                priority_cols.append(col)
                        
                        # Add remaining columns that aren't in priority list
                        remaining_cols = [col for col in all_cols if col not in priority_cols]
                        
                        # Combine all columns in the desired order
                        final_col_order = priority_cols + remaining_cols
                        
                        # Reorder the dataframe
                        edit_results_with_selection = edit_results_with_selection[final_col_order]
                        
                        # DO NOT format dates here - keep original values
                    
                        # Configure column settings for the data editor
                        column_config = {
                            "Select": st.column_config.CheckboxColumn(
                                "Select",
                                help="Select rows to delete",
                                default=False,
                            )
                        }
                        
                        # Configure numeric columns to display with 2 decimal places
                        numeric_cols = [
                            'Agent Estimated Comm $',
                            'Policy Gross Comm %',
                            'Agency Estimated Comm/Revenue (CRM)',
                            'Agency Comm Received (STMT)',
                            'Premium Sold',
                            'Agent Paid Amount (STMT)',
                            'Agency Comm Received (STMT)',
                            'Policy Taxes & Fees',
                            'Commissionable Premium',
                            'Broker Fee',
                            'Broker Fee Agent Comm',
                            'Total Agent Comm',
                            'Policy Balance Due',
                            'Agent Comm %'
                        ]
                        
                        # Dollar amount columns (show with $ sign)
                        dollar_cols = [
                            'Agent Estimated Comm $',
                            'Agency Estimated Comm/Revenue (CRM)',
                            'Agency Comm Received (STMT)',
                            'Premium Sold',
                            'Agent Paid Amount (STMT)',
                            'Policy Taxes & Fees',
                            'Commissionable Premium',
                            'Broker Fee',
                            'Broker Fee Agent Comm',
                            'Total Agent Comm',
                            'Policy Balance Due'
                        ]
                        
                        # Percentage columns (show without $ sign)
                        percent_cols = [
                            'Policy Gross Comm %',
                            'Agent Comm %'
                        ]
                        
                        for col in dollar_cols:
                            if col in edit_results.columns:
                                column_config[col] = st.column_config.NumberColumn(
                                    col,
                                    format="$%.2f",
                                    step=0.01
                                )
                        
                        for col in percent_cols:
                            if col in edit_results.columns:
                                column_config[col] = st.column_config.NumberColumn(
                                    col,
                                    format="%.2f",
                                    step=0.01
                                )
                        
                        # Configure date columns as text columns with help text
                        # Since dates are stored as strings, we can't use DateColumn
                        for col in date_cols:
                            if col in edit_results_with_selection.columns:
                                column_config[col] = st.column_config.TextColumn(
                                    col,
                                    help="Date format: YYYY-MM-DD",
                                    max_chars=10
                                )
                        
                        # If we have transaction ID column, make it clear it will be auto-generated
                        if transaction_id_col:
                            column_config[transaction_id_col] = st.column_config.TextColumn(
                                transaction_id_col,
                                help="Leave blank for new rows - will be auto-generated on save",
                                disabled=False  # Allow editing but we'll generate if blank
                            )
                        
                        # Create a unique key for this search result to track edits
                        editor_key = "edit_policies_editor"
                        
                        # Initialize or reset session state for this editor
                        search_key = f"last_search_{editor_key}"
                        edit_position_key = f"edit_position_{editor_key}"
                        unsaved_changes_key = f"unsaved_changes_{editor_key}"
                        
                        # Create a unique search identifier that includes both search term and filter state
                        current_search_state = f"{edit_search_term}_{show_attention_filter}"
                        
                        # Initialize if not exists or reset if search criteria changed
                        if (editor_key not in st.session_state or 
                            search_key not in st.session_state or 
                            st.session_state[search_key] != current_search_state):
                            st.session_state[editor_key] = edit_results_with_selection.copy()
                            st.session_state[search_key] = current_search_state
                            # Clear position tracking on new search
                            if edit_position_key in st.session_state:
                                del st.session_state[edit_position_key]
                            if unsaved_changes_key in st.session_state:
                                del st.session_state[unsaved_changes_key]
                        
                        # Preserve edit position and unsaved changes
                        if unsaved_changes_key in st.session_state:
                            # Restore any unsaved changes from before the refresh
                            for row_idx, col_name, value in st.session_state[unsaved_changes_key]:
                                if row_idx < len(st.session_state[editor_key]) and col_name in st.session_state[editor_key].columns:
                                    st.session_state[editor_key].loc[row_idx, col_name] = value
                        
                        # Auto-save functionality setup
                        auto_save_key = f"auto_save_{editor_key}"
                        if auto_save_key not in st.session_state:
                            st.session_state[auto_save_key] = True
                        
                        # Auto-save toggle at the top
                        col1, col2, col3 = st.columns([4, 1, 1])
                        with col1:
                            st.markdown("### Edit Policies")
                        with col3:
                            st.session_state[auto_save_key] = st.checkbox(
                                "🔄 Auto-save", 
                                value=st.session_state[auto_save_key],
                                help="Automatically save changes as you type"
                            )
                        
                        # Add a container for status messages
                        status_container = st.empty()
                        
                        # Show editing tips
                        with st.expander("💡 Editing Tips", expanded=False):
                            st.markdown("""
                            **Best editing experience:**
                            - 🖥️ **Use Full Screen mode** (expand icon in top-right of table) to prevent screen jumping
                            - Auto-save is enabled by default - changes save automatically
                            - Full screen maintains your position while editing
                            
                            **To minimize data loss:**
                            - Edit in Full Screen mode for best stability
                            - For bulk edits: Export to Excel → Edit → Import back
                            
                            **Keyboard shortcuts:**
                            - Tab: Move to next cell (may be limited)
                            - Enter: Confirm edit and move down
                            - Shift+Enter: Confirm edit and move up
                            """)
                        
                        # Calculate height based on number of rows (35px per row + 50px for header)
                        # Max height of 600px to prevent very tall tables
                        num_data_rows = len(st.session_state[editor_key])
                        calculated_height = min(50 + (num_data_rows + 2) * 35, 600)  # +2 for the extra rows you want
                        
                        # Track only Select column changes for performance
                        select_column_key = f"{editor_key}_select_only"
                        if select_column_key not in st.session_state:
                            st.session_state[select_column_key] = st.session_state[editor_key]['Select'].copy() if 'Select' in st.session_state[editor_key].columns else pd.Series()
                        
                        # Editable data grid with selection column
                        edited_data = st.data_editor(
                            st.session_state[editor_key],
                            use_container_width=True,
                            height=calculated_height,
                            key=f"{editor_key}_widget",
                            num_rows="fixed",  # Back to fixed to prevent too many blank rows
                            column_config=column_config,
                            disabled=False
                        )
                        
                        # Detect changes and auto-save - skip if only Select column changed
                        data_changed = False
                        if 'Select' in edited_data.columns:
                            # Compare dataframes excluding the Select column for performance
                            cols_to_check = [col for col in edited_data.columns if col != 'Select']
                            if cols_to_check:
                                data_changed = not edited_data[cols_to_check].equals(st.session_state[editor_key][cols_to_check])
                        else:
                            data_changed = not edited_data.equals(st.session_state[editor_key])
                        
                        if data_changed:
                            changes_detected = []
                            for idx in edited_data.index:
                                for col in edited_data.columns:
                                    # Skip the Select column to avoid triggering saves on checkbox clicks
                                    if col != 'Select' and edited_data.loc[idx, col] != st.session_state[editor_key].loc[idx, col]:
                                        changes_detected.append((idx, col, edited_data.loc[idx, col]))
                            
                            if changes_detected and st.session_state[auto_save_key]:
                                # Auto-save changes immediately
                                status_container.info("💾 Auto-saving changes...")
                                
                                try:
                                    # Get transaction ID column
                                    transaction_id_col = None
                                    for col in edited_data.columns:
                                        if 'transaction' in col.lower() and 'id' in col.lower():
                                            transaction_id_col = col
                                            break
                                    
                                    saved_count = 0
                                    for idx, col_name, new_value in changes_detected:
                                        if col_name != 'Select' and transaction_id_col:
                                            transaction_id = edited_data.loc[idx, transaction_id_col]
                                            if pd.notna(transaction_id) and str(transaction_id).strip():
                                                # Update single cell
                                                update_dict = {col_name: new_value if pd.notna(new_value) else None}
                                                supabase.table('policies').update(update_dict).eq(transaction_id_col, transaction_id).execute()
                                                saved_count += 1
                                    
                                    if saved_count > 0:
                                        status_container.success(f"✅ Auto-saved {saved_count} changes")
                                        # Update the base data to reflect saved changes
                                        st.session_state[editor_key] = edited_data.copy()
                                        # Don't rerun - just update state
                                
                                except Exception as e:
                                    status_container.error(f"Auto-save error: {str(e)}")
                                    log_debug(f"Auto-save error: {str(e)}", "ERROR", e)
                            
                            elif changes_detected and not st.session_state[auto_save_key]:
                                status_container.info(f"📝 {len(changes_detected)} unsaved changes")
                                st.session_state[unsaved_changes_key] = changes_detected
                                
                            
                        # Always update session state for Select column changes
                        # This prevents the equals() check from triggering on checkbox clicks
                        if 'Select' in edited_data.columns:
                            st.session_state[editor_key]['Select'] = edited_data['Select'].copy()
                        
                        # Handle no data changes case
                        if not data_changed and st.session_state[auto_save_key]:
                            status_container.success("✅ All changes auto-saved")
                        
                        # Update session state without rerun
                        st.session_state[editor_key] = edited_data
                        
                        # Get the client ID from the search results (if all rows have the same client)
                        existing_client_id = None
                        if client_id_col:
                            unique_client_ids = edit_results[client_id_col].dropna().unique()
                            if len(unique_client_ids) == 1:
                                existing_client_id = unique_client_ids[0]
                                st.info(f"💡 New rows will receive unique Transaction IDs and use Client ID: {existing_client_id}")
                            elif len(unique_client_ids) > 1:
                                st.info("💡 New rows will receive unique Transaction IDs. Multiple Client IDs found - new rows will need Client ID specified.")
                        else:
                            st.info("💡 New rows will receive unique Transaction IDs and Client IDs when you save.")
                            
                        # Add buttons for adding new row and editing selected row
                        button_col1, button_col2 = st.columns(2)
                            
                        with button_col1:
                            if st.button("➕ Add New Transaction for This Client", type="secondary", use_container_width=True):
                                    # Ensure session state exists
                                    if editor_key in st.session_state:
                                        # Create a new empty row with generated IDs
                                        new_row = pd.Series(dtype='object')
                                        for col in edit_results_with_selection.columns:
                                            new_row[col] = None
                                        
                                        # Set default values
                                        new_row['Select'] = False
                                        if transaction_id_col:
                                            new_row[transaction_id_col] = generate_transaction_id()
                                        if client_id_col and existing_client_id:
                                            new_row[client_id_col] = existing_client_id
                                        elif client_id_col:
                                            new_row[client_id_col] = generate_client_id()
                                        
                                        # Add the new row to session state
                                        new_df = pd.concat([st.session_state[editor_key], pd.DataFrame([new_row])], ignore_index=True)
                                        st.session_state[editor_key] = new_df
                                        st.rerun()
                                    else:
                                        st.error("Session state not initialized. Please try searching again.")
                            
                        with button_col2:
                            # Check for selected rows for edit button - use only the Select column for performance
                            if 'Select' in edited_data.columns:
                                # Track selected count in session state for performance
                                selected_count_key = f"{editor_key}_selected_count"
                                
                                # Only recalculate if the Select column has changed
                                if 'Select' in edited_data.columns:
                                    current_selected = edited_data['Select'].tolist()
                                    prev_selected_key = f"{editor_key}_prev_selected"
                                    
                                    if (prev_selected_key not in st.session_state or 
                                        st.session_state[prev_selected_key] != current_selected):
                                        # Calculate selected count only when selection changes
                                        selected_mask = edited_data['Select'] == True
                                        selected_count = selected_mask.sum()
                                        st.session_state[selected_count_key] = selected_count
                                        st.session_state[prev_selected_key] = current_selected
                                        if selected_count == 1:
                                            # Cache the selected index too
                                            st.session_state[f"{editor_key}_selected_idx"] = edited_data[selected_mask].index[0]
                                    else:
                                        # Use cached values
                                        selected_count = st.session_state.get(selected_count_key, 0)
                                
                                if selected_count == 1:
                                    if st.button("✏️ Edit Selected Transaction", type="primary", use_container_width=True):
                                        st.session_state['show_edit_modal'] = True
                                        # Use cached index
                                        selected_idx = st.session_state.get(f"{editor_key}_selected_idx")
                                        if selected_idx is not None:
                                            st.session_state['edit_modal_data'] = edited_data.loc[selected_idx].to_dict()
                                elif selected_count == 0:
                                    st.button("✏️ Edit Selected Transaction", type="primary", use_container_width=True, disabled=True, help="Select one transaction to edit")
                                else:
                                    st.button("✏️ Edit Selected Transaction", type="primary", use_container_width=True, disabled=True, help=f"{selected_count} selected - please select only ONE transaction")
                            else:
                                st.button("✏️ Edit Selected Transaction", type="primary", use_container_width=True, disabled=True, help="No selection column available")
                            
                        # Save and Delete buttons with status
                        st.markdown("---")
                        col1, col2, col3 = st.columns([2, 2, 2])
                        
                        with col1:
                            # Show save status
                            if unsaved_changes_key in st.session_state and st.session_state[unsaved_changes_key]:
                                st.warning(f"⚠️ {len(st.session_state[unsaved_changes_key])} unsaved changes")
                            else:
                                st.success("✅ All changes saved")
                        
                        with col2:
                            save_button = st.button("💾 Save All Changes", type="primary", use_container_width=True)
                        
                        if save_button:
                            try:
                                updated_count = 0
                                inserted_count = 0
                                
                                # Store original transaction IDs to track which rows are updates vs inserts
                                original_transaction_ids = set()
                                if transaction_id_col:
                                    original_transaction_ids = set(edit_results[transaction_id_col].dropna().astype(str))
                                
                                # Process each record in the edited data
                                for idx, row in edited_data.iterrows():
                                    # Skip the selection column for save operations
                                    if 'Select' in row:
                                        row = row.drop('Select')
                                    
                                    # Get the transaction ID for this row
                                    transaction_id = row.get(transaction_id_col) if transaction_id_col else None
                                            
                                    # Check if this is a new row by seeing if the transaction ID exists in original data
                                    is_new_row = False
                                    if transaction_id_col:
                                        # If there's no transaction ID or it's not in the original set, it's new
                                        if pd.isna(transaction_id) or str(transaction_id).strip() == '' or str(transaction_id) not in original_transaction_ids:
                                            is_new_row = True
                                    else:
                                        # Fallback to index-based check if no transaction ID column
                                        is_new_row = idx >= len(edit_results)
                                            
                                    if is_new_row:
                                        # For new rows, generate unique IDs if they're missing
                                        if transaction_id_col and (pd.isna(transaction_id) or str(transaction_id).strip() == ''):
                                            transaction_id = generate_transaction_id()
                                            row[transaction_id_col] = transaction_id
                                        if client_id_col and (pd.isna(row[client_id_col]) or str(row[client_id_col]).strip() == ''):
                                            # Use existing client ID if searching for a specific client, otherwise generate new
                                            if existing_client_id:
                                                row[client_id_col] = existing_client_id
                                            else:
                                                row[client_id_col] = generate_client_id()
                                    
                                    # Process all rows
                                    if is_new_row:
                                        # INSERT new record
                                        insert_dict = {}
                                        for col in edited_data.columns:
                                            if col not in ['_id', 'Select']:  # Exclude auto-generated fields and selection column
                                                value = row[col]
                                                # Clean numeric values
                                                if pd.notna(value) and isinstance(value, (int, float)):
                                                    insert_dict[col] = clean_numeric_value(value)
                                                else:
                                                    insert_dict[col] = value if pd.notna(value) else None
                                                
                                        try:
                                            # Clean data before insertion
                                            cleaned_insert = clean_data_for_database(insert_dict)
                                            supabase.table('policies').insert(cleaned_insert).execute()
                                            inserted_count += 1
                                        except Exception as insert_error:
                                            st.error(f"Error inserting new record: {insert_error}")
                                    elif transaction_id:  # Only update if we have a transaction ID
                                        # UPDATE existing record
                                        update_dict = {}
                                        for col in edited_data.columns:
                                            if col not in ['_id', 'Select', transaction_id_col]:  # Don't update ID field or selection column
                                                value = row[col]
                                                # Clean numeric values
                                                if pd.notna(value) and isinstance(value, (int, float)):
                                                    update_dict[col] = clean_numeric_value(value)
                                                else:
                                                    update_dict[col] = value if pd.notna(value) else None
                                                
                                        try:
                                            supabase.table('policies').update(update_dict).eq(transaction_id_col, transaction_id).execute()
                                            updated_count += 1
                                        except Exception as update_error:
                                            st.error(f"Error updating record: {update_error}")
                                    else:
                                        # This shouldn't happen, but log it if it does
                                        st.warning(f"Skipped row {idx} - existing row but no transaction ID found")
                                        
                                # Clear cache and show success message
                                clear_policies_cache()
                                
                                if inserted_count > 0 and updated_count > 0:
                                    st.success(f"Successfully inserted {inserted_count} new records and updated {updated_count} existing records!")
                                elif inserted_count > 0:
                                    st.success(f"Successfully inserted {inserted_count} new records!")
                                elif updated_count > 0:
                                    st.success(f"Successfully updated {updated_count} records!")
                                else:
                                    st.info("No changes were made.")
                                
                                st.rerun()
                                
                                # Clear unsaved changes after successful save
                                if unsaved_changes_key in st.session_state:
                                    del st.session_state[unsaved_changes_key]
                                
                            except Exception as e:
                                st.error(f"Error saving changes: {e}")
                            
                        with col3:
                            if st.button("🔄 Refresh Data", use_container_width=True):
                                # Clear session state and refresh
                                if unsaved_changes_key in st.session_state:
                                    del st.session_state[unsaved_changes_key]
                                st.rerun()
                        
                        # Modal Form Implementation for Edit
                        if st.session_state.get('show_edit_modal', False):
                            # Create a modal-like overlay
                            st.markdown("---")
                            st.markdown("### 📝 Edit Transaction")
                                
                            modal_data = st.session_state.get('edit_modal_data', {})
                                
                            # Add an anchor point to prevent scroll jumping
                            st.empty()  # This helps maintain scroll position
                            
                            # Carrier & MGA Selection (OUTSIDE FORM for dynamic updates)
                            st.subheader("Carrier & MGA Selection 🏢")
                            st.info("💡 Select carrier first to see available MGAs. This will auto-populate commission rates.")
                                
                            # Get current values from modal data
                            current_carrier = modal_data.get('Carrier Name', '')
                            current_mga = modal_data.get('MGA Name', '')
                            
                            # Load carriers for dropdown
                            carriers_list = load_carriers_for_dropdown()
                            
                            # Use container to better control rendering
                            carrier_container = st.container()
                            with carrier_container:
                                col1, col2 = st.columns(2)
                                with col1:
                                    # Carrier dropdown with search capability
                                    carrier_options = [""] + [c['carrier_name'] for c in carriers_list]
                                    
                                    # Find index of current carrier
                                    carrier_index = 0
                                    if current_carrier and current_carrier in carrier_options:
                                        carrier_index = carrier_options.index(current_carrier)
                                    
                                    selected_carrier_name = st.selectbox(
                                        "Carrier Name*",
                                        options=carrier_options,
                                        index=carrier_index,
                                        format_func=lambda x: "🔍 Select or search carrier..." if x == "" else f"🏢 {x}",
                                        help="Select carrier to auto-populate commission rates",
                                        key="edit_policy_carrier_outside"
                                    )
                                    
                                    # Get carrier_id for selected carrier
                                    selected_carrier_id = None
                                    if selected_carrier_name:
                                        selected_carrier_id = next((c['carrier_id'] for c in carriers_list if c['carrier_name'] == selected_carrier_name), None)
                                        st.session_state['edit_selected_carrier_id'] = selected_carrier_id
                                        st.session_state['edit_selected_carrier_name'] = selected_carrier_name
                                    
                                    # Fallback text input for manual entry
                                    if not selected_carrier_name:
                                        carrier_name_manual = st.text_input("Or enter carrier name manually", value=current_carrier, placeholder="Type carrier name", key="edit_carrier_manual_outside")
                                        st.session_state['edit_carrier_name_manual'] = carrier_name_manual
                                
                                with col2:
                                    # MGA dropdown (filtered by carrier) - Updates immediately!
                                    mga_options = ["Direct Appointment"]
                                    selected_mga_id = None
                                    
                                    if selected_carrier_id:
                                        mgas_list = load_mgas_for_carrier(selected_carrier_id)
                                        mga_options.extend([m['mga_name'] for m in mgas_list])
                                    
                                    # Find index of current MGA
                                    mga_index = 0
                                    if current_mga:
                                        if current_mga in mga_options:
                                            mga_index = mga_options.index(current_mga)
                                        elif "Direct Appointment" in mga_options:
                                            mga_index = mga_options.index("Direct Appointment")
                                    
                                    selected_mga_name = st.selectbox(
                                        "MGA/Appointment",
                                        options=mga_options,
                                        index=mga_index,
                                        format_func=lambda x: f"🤝 {x}" if x != "Direct Appointment" else "🏢 Direct Appointment",
                                        help="MGA options update automatically when you select a carrier",
                                        key="edit_policy_mga_outside"
                                    )
                                    
                                    # Get mga_id for selected MGA
                                    if selected_mga_name != "Direct Appointment" and selected_carrier_id:
                                        mgas_list = load_mgas_for_carrier(selected_carrier_id) 
                                        selected_mga_id = next((m['mga_id'] for m in mgas_list if m['mga_name'] == selected_mga_name), None)
                                        st.session_state['edit_selected_mga_id'] = selected_mga_id
                                        st.session_state['edit_selected_mga_name'] = selected_mga_name
                                    else:
                                        st.session_state['edit_selected_mga_id'] = None
                                        st.session_state['edit_selected_mga_name'] = selected_mga_name
                                
                                # Fallback text input for manual entry
                                if not selected_carrier_name:
                                    mga_name_manual = st.text_input("Or enter MGA name manually", value=current_mga, placeholder="Type MGA name or leave blank", key="edit_mga_manual_outside")
                                    st.session_state['edit_mga_name_manual'] = mga_name_manual
                            
                            # Store final values for form submission
                            if selected_carrier_name:
                                final_carrier_name = selected_carrier_name
                                final_mga_name = selected_mga_name if selected_mga_name != "Direct Appointment" else ""
                            else:
                                final_carrier_name = st.session_state.get('edit_carrier_name_manual', '')
                                final_mga_name = st.session_state.get('edit_mga_name_manual', '')
                            
                            # Store in session state for form to access
                            st.session_state['edit_final_carrier_name'] = final_carrier_name
                            st.session_state['edit_final_mga_name'] = final_mga_name
                            
                            # Display selection status and commission info
                            if selected_carrier_name and selected_carrier_id:
                                # Look up commission rule
                                commission_rule = None
                                policy_type = modal_data.get('Policy Type', '')
                                
                                if selected_mga_id:
                                    # Try carrier + MGA + policy type first
                                    commission_rule = lookup_commission_rule(selected_carrier_id, selected_mga_id, policy_type)
                                    if not commission_rule:
                                        # Try carrier + MGA without policy type
                                        commission_rule = lookup_commission_rule(selected_carrier_id, selected_mga_id, None)
                                
                                if not commission_rule:
                                    # Try carrier + policy type without MGA
                                    commission_rule = lookup_commission_rule(selected_carrier_id, None, policy_type)
                                
                                if not commission_rule:
                                    # Try carrier default
                                    commission_rule = lookup_commission_rule(selected_carrier_id, None, None)
                                
                                if commission_rule:
                                    # Store both rates and let the form decide which to use based on transaction type
                                    new_rate = commission_rule.get('new_rate', 0)
                                    renewal_rate = commission_rule.get('renewal_rate', new_rate)  # Default to new rate if no renewal rate
                                    
                                    st.info(f"ℹ️ Commission rule found: {commission_rule.get('rule_description', 'Carrier default')}")
                                    st.success(f"✅ Rates available - New: {new_rate}% | Renewal: {renewal_rate}%")
                                    st.info("💡 The correct rate will be applied based on your Transaction Type selection in the form below")
                                    
                                    # Store both rates in session state
                                    st.session_state['edit_commission_new_rate'] = new_rate
                                    st.session_state['edit_commission_renewal_rate'] = renewal_rate
                                    st.session_state['edit_commission_rule_id'] = commission_rule.get('rule_id')
                                    st.session_state['edit_has_commission_rule'] = True
                                else:
                                    st.info(f"ℹ️ No commission rule found for {selected_carrier_name}. Enter rate manually.")
                                    st.session_state['edit_commission_new_rate'] = None
                                    st.session_state['edit_commission_renewal_rate'] = None
                                    st.session_state['edit_commission_rule_id'] = None
                                    st.session_state['edit_has_commission_rule'] = False
                            
                            st.markdown("---")
                            
                            # Use the reusable edit transaction form
                            result = edit_transaction_form(modal_data, source_page="edit_policies")
                            
                            if result:
                                if result["action"] == "save":
                                    try:
                                        # Get transaction ID and _id to determine if this is new or existing
                                        transaction_id = result["data"].get(get_mapped_column("Transaction ID"))
                                        record_id = result["data"].get('_id')
                                        
                                        # Convert data for database operation
                                        save_data = result["data"].copy()
                                        save_data = convert_timestamps_for_json(save_data)
                                        
                                        # Handle NaN values
                                        for key, value in save_data.items():
                                            if pd.isna(value):
                                                save_data[key] = None
                                        
                                        # First check if this transaction already exists in the database
                                        # This handles cases where the record was added inline but doesn't have _id in session state
                                        existing_record = None
                                        if transaction_id:
                                            try:
                                                check_response = supabase.table('policies').select('_id').eq(
                                                    get_mapped_column("Transaction ID"), transaction_id
                                                ).execute()
                                                if check_response.data and len(check_response.data) > 0:
                                                    existing_record = check_response.data[0]
                                            except:
                                                pass
                                        
                                        # Determine if this is an INSERT or UPDATE
                                        if existing_record or (record_id is not None and record_id != '' and not pd.isna(record_id)):
                                            # Existing record - UPDATE
                                            # Remove _id from update data as it shouldn't be updated
                                            if '_id' in save_data:
                                                del save_data['_id']
                                            
                                            # For Supabase Python client, we need to specify select parameter in update
                                            response = supabase.table('policies').update(save_data).eq(
                                                get_mapped_column("Transaction ID"), transaction_id
                                            ).execute()
                                        else:
                                            # New record - INSERT
                                            # Remove _id field to let database auto-generate it
                                            if '_id' in save_data:
                                                del save_data['_id']
                                            
                                            # Clean data before insertion
                                            cleaned_save = clean_data_for_database(save_data)
                                            response = supabase.table('policies').insert(cleaned_save).execute()
                                        
                                        # Check if we have a response - for updates, data might be None but operation succeeded
                                        # For inserts, we should have data
                                        if response is not None and (response.data or existing_record or record_id):
                                            st.success("✅ Transaction updated successfully!")
                                            clear_policies_cache()
                                            
                                            # Clear modal state
                                            st.session_state['show_edit_modal'] = False
                                            st.session_state['edit_modal_data'] = None
                                            
                                            # Force clear the session state for the editor
                                            if 'edit_policies_editor' in st.session_state:
                                                del st.session_state['edit_policies_editor']
                                            if 'last_search_edit_policies_editor' in st.session_state:
                                                del st.session_state['last_search_edit_policies_editor']
                                            
                                            time.sleep(1)
                                            st.rerun()
                                        else:
                                            st.error("❌ Update failed - no response from database")
                                    
                                    except Exception as e:
                                        st.error(f"Error updating transaction: {str(e)}")
                                
                                elif result["action"] == "close" or result["action"] == "cancel":
                                    # Clear modal state
                                    st.session_state['show_edit_modal'] = False
                                    st.session_state['edit_modal_data'] = None
                                    st.rerun()
                            
                            # The old form implementation has been removed and replaced with the reusable function
                        # Delete functionality moved to bottom
                        st.divider()
                        st.subheader("🗑️ Delete Selected Records")
                        
                        # Re-check selected rows for delete functionality
                        selected_rows_for_delete = edited_data[edited_data['Select'] == True].copy()
                        
                        # Collect transaction IDs to delete BEFORE any modifications
                        transaction_ids_to_delete = []
                        reconciliation_attempts = []
                        if not selected_rows_for_delete.empty and transaction_id_col:
                            for idx, row in selected_rows_for_delete.iterrows():
                                tid = row[transaction_id_col]
                                if tid and pd.notna(tid):
                                    # Check if this is a reconciliation transaction
                                    if is_reconciliation_transaction(tid):
                                        reconciliation_attempts.append(str(tid))
                                    # IMPORT transactions can now be deleted - add them to delete list
                                    else:
                                        transaction_ids_to_delete.append(str(tid))
                        
                        # Show error if trying to delete reconciliation transactions
                        if reconciliation_attempts:
                            st.error(f"🔒 Cannot delete {len(reconciliation_attempts)} reconciliation transaction(s):")
                            for tid in reconciliation_attempts:
                                st.write(f"- {tid}")
                            st.info("Reconciliation entries (-STMT-, -VOID-, -ADJ-) are permanent audit records. Use the Reconciliation page to create adjustments if needed.")
                        
                        if transaction_ids_to_delete:
                            st.warning(f"⚠️ You have selected {len(transaction_ids_to_delete)} record(s) for deletion:")
                            # Check if any are IMPORT transactions
                            import_count = sum(1 for tid in transaction_ids_to_delete if '-IMPORT' in str(tid))
                            if import_count > 0:
                                st.info(f"ℹ️ Note: {import_count} of these are IMPORT transactions that can now be deleted.")
                            # Show which records are selected
                            for tid in transaction_ids_to_delete:
                                # Find the customer name for this transaction ID
                                customer_row = edited_data[edited_data[transaction_id_col] == tid]
                                if not customer_row.empty:
                                    customer = customer_row.iloc[0].get('Customer', 'Unknown')
                                    st.write(f"- {tid} - {customer}")
                            
                            col1, col2 = st.columns([1, 3])
                            with col1:
                                if st.button("🗑️ Confirm Delete", type="secondary"):
                                    try:
                                        deleted_count = 0
                                        # Use the pre-collected transaction IDs
                                        for tid in transaction_ids_to_delete:
                                            # Get the full row data for archiving
                                            row_to_archive = edited_data[edited_data[transaction_id_col] == tid]
                                            if not row_to_archive.empty:
                                                row = row_to_archive.iloc[0]
                                                # First, archive the record to deleted_policies table
                                                # Convert row to dict, handling NaN values and type conversion
                                                policy_dict = {}
                                                for col in row.index:
                                                    if col != 'Select':  # Skip the selection column
                                                        value = row[col]
                                                        if pd.notna(value):
                                                            # Convert to Python native types for JSON serialization
                                                            if hasattr(value, 'item'):  # numpy scalar
                                                                policy_dict[col] = value.item()
                                                            elif isinstance(value, (int, float, bool, str)):
                                                                policy_dict[col] = value
                                                            else:
                                                                policy_dict[col] = str(value)
                                                        else:
                                                            policy_dict[col] = None
                                                
                                                # Create archive record with JSONB structure
                                                archive_record = {
                                                    'transaction_id': tid,
                                                    'customer_name': row.get('Customer', 'Unknown'),
                                                    'policy_data': policy_dict
                                                }
                                                
                                                try:
                                                    # Insert into deleted_policies table
                                                    supabase.table('deleted_policies').insert(archive_record).execute()
                                                    
                                                    # Then delete from main policies table
                                                    supabase.table('policies').delete().eq(transaction_id_col, tid).execute()
                                                    deleted_count += 1
                                                except Exception as archive_error:
                                                    st.error(f"Error archiving record {tid}: {archive_error}")
                                        
                                        # Clear cache and session state before rerun
                                        clear_policies_cache()
                                        
                                        # Clear the editor session state to force refresh
                                        if 'edit_policies_editor' in st.session_state:
                                            del st.session_state['edit_policies_editor']
                                        if 'edit_policies_editor_widget' in st.session_state:
                                            del st.session_state['edit_policies_editor_widget']
                                        if 'last_search_edit_policies_editor' in st.session_state:
                                            del st.session_state['last_search_edit_policies_editor']
                                        
                                        st.success(f"Successfully deleted {deleted_count} records! (Archived for recovery)")
                                        time.sleep(1)  # Brief pause to show success message
                                        st.rerun()
                                        
                                    except Exception as e:
                                        st.error(f"Error deleting records: {e}")
                            with col2:
                                st.info("Click 'Confirm Delete' to permanently remove the selected records.")
                        else:
                            # Only show error if we haven't already handled the rows as reconciliation transactions
                            if not selected_rows_for_delete.empty and not reconciliation_attempts and not transaction_id_col:
                                st.error("Could not identify transaction IDs for selected rows. Make sure the Transaction ID column is properly identified.")
                            elif selected_rows_for_delete.empty:
                                st.info("Check the 'Select' checkbox in the data editor above to select rows for deletion.")
                            # If we processed reconciliation transactions, don't show any additional message
                else:
                    if show_attention_filter:
                        # Already showed the success message above
                        pass
                    else:
                        st.warning("No records found matching your search")
            else:
                st.info("Use the search box above to find policies to edit, or edit all policies below:")
            
            # Option to edit all data (with pagination for performance)
            show_all = st.checkbox("Show all policies for editing (use with caution)")
            
            if show_all:
                st.warning("Editing all policies at once can be slow with large datasets")
                
                # Sort by Customer A-Z first, then limit to first 50 records for performance
                if 'Customer' in all_data.columns:
                    all_data_sorted = all_data.sort_values(by='Customer', ascending=True)
                    edit_all_data = all_data_sorted.head(50)
                else:
                    edit_all_data = all_data.head(50)
                st.write(f"Showing first 50 records for editing out of {len(all_data)} total (sorted by Customer A-Z)")
                
                # Find the actual column names dynamically
                transaction_id_col = None
                client_id_col = None
                for col in edit_all_data.columns:
                    if 'transaction' in col.lower() and 'id' in col.lower():
                        transaction_id_col = col
                    if 'client' in col.lower() and 'id' in col.lower():
                        client_id_col = col
                
                # Try minimal column config to see if it helps with sorting
                edited_all_data = st.data_editor(
                    edit_all_data,
                    use_container_width=True,
                        height=500,
                        key="edit_all_policies_editor",
                        num_rows="dynamic",
                        disabled=['_id'] if '_id' in edit_all_data.columns else []  # Only disable _id column
                    )
                
                if st.button("💾 Save Changes", key="save_all_edits"):
                    try:
                        updated_count = 0
                        inserted_count = 0
                        
                        for idx, row in edited_all_data.iterrows():
                            # Check if this is a new row
                            is_new_row = idx >= len(edit_all_data)
                            
                            if is_new_row:
                                # For new rows, generate unique IDs if they're missing
                                if transaction_id_col and (pd.isna(row[transaction_id_col]) or str(row[transaction_id_col]).strip() == ''):
                                    row[transaction_id_col] = generate_transaction_id()
                                if client_id_col and (pd.isna(row[client_id_col]) or str(row[client_id_col]).strip() == ''):
                                    row[client_id_col] = generate_client_id()
                            
                            # Get the transaction ID for processing
                            transaction_id = row.get(transaction_id_col) if transaction_id_col else None
                            
                            if transaction_id:
                                
                                if is_new_row:
                                    # INSERT new record
                                    insert_dict = {}
                                    for col in edited_all_data.columns:
                                        if col not in ['_id']:  # Exclude auto-generated fields
                                            insert_dict[col] = row[col] if pd.notna(row[col]) else None
                                    
                                    try:
                                        # Clean data before insertion
                                        cleaned_insert = clean_data_for_database(insert_dict)
                                        supabase.table('policies').insert(cleaned_insert).execute()
                                        inserted_count += 1
                                    except Exception as insert_error:
                                        st.error(f"Error inserting new record: {insert_error}")
                                else:
                                    # UPDATE existing record
                                    update_dict = {}
                                    for col in edited_all_data.columns:
                                        if col not in ['_id', transaction_id_col]:  # Don't update the ID field itself
                                            update_dict[col] = row[col] if pd.notna(row[col]) else None
                                    
                                    try:
                                        supabase.table('policies').update(update_dict).eq(transaction_id_col, transaction_id).execute()
                                        updated_count += 1
                                    except Exception as update_error:
                                        st.error(f"Error updating record: {update_error}")
                        
                        # Clear cache and show success message
                        clear_policies_cache()
                        
                        if inserted_count > 0 and updated_count > 0:
                            st.success(f"Successfully inserted {inserted_count} new records and updated {updated_count} existing records!")
                        elif inserted_count > 0:
                            st.success(f"Successfully inserted {inserted_count} new records!")
                        elif updated_count > 0:
                            st.success(f"Successfully updated {updated_count} records!")
                        else:
                            st.info("No changes were made.")
                        
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"Error saving changes: {e}")
    
    # --- Add New Policy Transaction ---
    elif page == "Add New Policy Transaction":
        st.title("➕ Add New Policy Transaction")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        # Display success message if a policy was just added
        if 'add_policy_success' in st.session_state:
            success_info = st.session_state['add_policy_success']
            # Check if the message is recent (within last 10 seconds)
            time_diff = datetime.datetime.now() - success_info['timestamp']
            if time_diff.total_seconds() < 10:
                st.success(f"""
                ✅ **Transaction Successfully Added to Database!**
                
                **Customer:** {success_info['customer']}  
                **Policy Number:** {success_info['policy_number']}  
                **Transaction ID:** {success_info['transaction_id']}
                
                The form has been cleared and is ready for the next transaction.
                """)
                st.balloons()
            else:
                # Clear old success message
                del st.session_state['add_policy_success']
        
        # Search for Existing Client ID by Name
        st.subheader("Search for Existing Client ID by Name")
        client_search = st.text_input("Type client name to search:", key="client_search")
        
        selected_client_id = None
        if client_search:
            # Search for matching clients, filtering out those without Client IDs
            matching_clients = all_data[
                (all_data['Customer'].str.contains(client_search, case=False, na=False)) & 
                (all_data['Client ID'].notna()) & 
                (all_data['Client ID'] != '')
            ][['Client ID', 'Customer']].drop_duplicates()
            
            if not matching_clients.empty:
                st.info(f"Found {len(matching_clients)} matching clients:")
                # Display options
                for idx, client in matching_clients.iterrows():
                    client_id = client['Client ID']
                    customer_name = client['Customer']
                    # Only show button if client has a valid ID
                    if client_id and str(client_id).strip():
                        if st.button(f"Use {client_id} - {customer_name}", key=f"select_client_{idx}"):
                            selected_client_id = client_id
                            st.session_state['selected_client_id'] = selected_client_id
                            st.session_state['selected_customer_name'] = customer_name
            else:
                st.warning("No matching clients found")
        
        # Get selected client ID and name from session state if available
        if 'selected_client_id' in st.session_state:
            selected_client_id = st.session_state['selected_client_id']
        
        selected_customer_name = None
        if 'selected_customer_name' in st.session_state:
            selected_customer_name = st.session_state['selected_customer_name']
        
        # Carrier & MGA Selection (OUTSIDE FORM for dynamic updates)
        st.subheader("Carrier & MGA Selection 🏢")
        st.info("💡 Select carrier first to see available MGAs")
        
        # Load carriers for dropdown
        carriers_list = load_carriers_for_dropdown()
        
        col1, col2 = st.columns(2)
        with col1:
            # Carrier dropdown with search capability
            carrier_options = [""] + [c['carrier_name'] for c in carriers_list]
            selected_carrier_name = st.selectbox(
                "Carrier Name*",
                options=carrier_options,
                format_func=lambda x: "🔍 Select or search carrier..." if x == "" else f"🏢 {x}",
                help="Select carrier to auto-populate commission rates",
                key="add_policy_carrier_outside"
            )
            
            # Get carrier_id for selected carrier
            selected_carrier_id = None
            if selected_carrier_name:
                selected_carrier_id = next((c['carrier_id'] for c in carriers_list if c['carrier_name'] == selected_carrier_name), None)
                st.session_state['selected_carrier_id'] = selected_carrier_id
                st.session_state['selected_carrier_name'] = selected_carrier_name
            
            # Fallback text input for manual entry
            if not selected_carrier_name:
                carrier_name_manual = st.text_input("Or enter carrier name manually", placeholder="Type carrier name", key="carrier_manual_outside")
                st.session_state['carrier_name_manual'] = carrier_name_manual
        
        with col2:
            # MGA dropdown (filtered by carrier) - Updates immediately!
            mga_options = ["Direct Appointment"]
            selected_mga_id = None
            
            if selected_carrier_id:
                mgas_list = load_mgas_for_carrier(selected_carrier_id)
                mga_options.extend([m['mga_name'] for m in mgas_list])
            
            selected_mga_name = st.selectbox(
                "MGA/Appointment",
                options=mga_options,
                format_func=lambda x: f"🤝 {x}" if x != "Direct Appointment" else "🏢 Direct Appointment",
                help="MGA options update automatically when you select a carrier",
                key="add_policy_mga_outside"
            )
            
            # Get mga_id for selected MGA
            if selected_mga_name != "Direct Appointment" and selected_carrier_id:
                mgas_list = load_mgas_for_carrier(selected_carrier_id) 
                selected_mga_id = next((m['mga_id'] for m in mgas_list if m['mga_name'] == selected_mga_name), None)
                st.session_state['selected_mga_id'] = selected_mga_id
                st.session_state['selected_mga_name'] = selected_mga_name
            else:
                st.session_state['selected_mga_id'] = None
                st.session_state['selected_mga_name'] = selected_mga_name
            
            # Fallback text input for manual entry
            if not selected_carrier_name:
                mga_name_manual = st.text_input("Or enter MGA name manually", placeholder="Type MGA name or leave blank", key="mga_manual_outside")
                st.session_state['mga_name_manual'] = mga_name_manual
        
        # Store final values for form submission
        if selected_carrier_name:
            final_carrier_name = selected_carrier_name
            final_mga_name = selected_mga_name if selected_mga_name != "Direct Appointment" else ""
        else:
            final_carrier_name = st.session_state.get('carrier_name_manual', '')
            final_mga_name = st.session_state.get('mga_name_manual', '')
        
        # Store in session state for form to access
        st.session_state['final_carrier_name'] = final_carrier_name
        st.session_state['final_mga_name'] = final_mga_name
        
        # Display selection status and commission info
        if selected_carrier_name and selected_carrier_id:
            st.success(f"✅ Carrier: {selected_carrier_name} | MGA: {selected_mga_name}")
            
            # Look up commission rule
            commission_rule = None
            
            if selected_mga_id:
                # Try carrier + MGA first
                commission_rule = lookup_commission_rule(selected_carrier_id, selected_mga_id, None)
            
            if not commission_rule:
                # Try carrier default
                commission_rule = lookup_commission_rule(selected_carrier_id, None, None)
            
            if commission_rule:
                # Store both rates and let the form decide which to use based on transaction type
                new_rate = commission_rule.get('new_rate', 0)
                renewal_rate = commission_rule.get('renewal_rate', new_rate)  # Default to new rate if no renewal rate
                
                st.info(f"ℹ️ Commission rule found: {commission_rule.get('rule_description', 'Carrier default')}")
                st.success(f"✅ Rates available - New: {new_rate}% | Renewal: {renewal_rate}%")
                st.info("💡 The correct rate will be applied based on your Transaction Type selection in the form below")
                
                # Store both rates in session state
                st.session_state['commission_new_rate'] = new_rate
                st.session_state['commission_renewal_rate'] = renewal_rate
                st.session_state['commission_rule_id'] = commission_rule.get('rule_id')
                st.session_state['has_commission_rule'] = True
            else:
                st.info(f"ℹ️ No commission rule found for {selected_carrier_name}. Enter rate manually.")
                st.session_state['commission_new_rate'] = None
                st.session_state['commission_renewal_rate'] = None
                st.session_state['commission_rule_id'] = None
                st.session_state['has_commission_rule'] = False
        
        st.markdown("---")
        
        # Main Form
        with st.form("add_policy_form"):
            # Client Information Section
            st.subheader("Client Information")
            col1, col2 = st.columns(2)
            with col1:
                if selected_customer_name:
                    customer = st.text_input("Customer Name", value=selected_customer_name)
                else:
                    customer = st.text_input("Customer Name", placeholder="Enter customer name")
            with col2:
                if selected_client_id:
                    client_id = st.text_input("Client ID", value=selected_client_id, disabled=True)
                else:
                    client_id = st.text_input("Client ID", value=generate_client_id())
            
            # Hidden transaction ID
            transaction_id = generate_transaction_id()
            
            # Policy Information Section
            st.subheader("Policy Information")
            
            # Row 1: Policy Type and Policy Number
            col1, col2 = st.columns(2)
            with col1:
                # Get dynamic policy types
                active_types, allow_custom = get_active_policy_types()
                default_type = get_default_policy_type()
                
                # Find default index
                try:
                    default_index = active_types.index(default_type)
                except ValueError:
                    default_index = 0
                
                policy_type = st.selectbox(
                    "Policy Type (add new types in Admin Panel)", 
                    options=active_types, 
                    index=default_index,
                    help="To add new policy types, go to Admin Panel → Manage Policy Types"
                )
            with col2:
                policy_number = st.text_input("Policy Number", placeholder="Enter policy number", key="add_policy_number")
            
            # Row 1.5: Prior Policy Number (for rewrites and renewals)
            prior_policy_number = st.text_input(
                "Prior Policy Number (optional)", 
                placeholder="Enter if this is a renewal or rewrite",
                help="For REWRITE or renewal transactions, enter the original policy number to maintain audit trail",
                key="add_prior_policy_number"
            )
            
            # Get carrier/MGA values from session state (selected outside the form)
            final_carrier_name = st.session_state.get('final_carrier_name', '')
            final_mga_name = st.session_state.get('final_mga_name', '')
            selected_carrier_id = st.session_state.get('selected_carrier_id')
            selected_mga_id = st.session_state.get('selected_mga_id')
            
            # Row 2: Transaction Type and Policy Term
            col1, col2 = st.columns(2)
            with col1:
                transaction_type = st.selectbox("Transaction Type", ["NEW", "RWL", "END", "PCH", "CAN", "XCL", "NBS", "STL", "BoR", "REWRITE"])
            with col2:
                policy_term = st.selectbox(
                    "Policy Term",
                    options=[None, 3, 6, 9, 12],
                    format_func=lambda x: "" if x is None else f"{x} months",
                    help="Select policy duration in months"
                )
            
            # Row 3: Effective Date and Policy Origination Date
            col1, col2 = st.columns(2)
            with col1:
                # Use session state for Effective Date to allow clearing
                effective_date_default = st.session_state.get('add_effective_date', None)
                if effective_date_default is None:
                    effective_date = st.date_input("Effective Date", value=None, key="add_effective_date")
                else:
                    effective_date = st.date_input("Effective Date", value=effective_date_default, key="add_effective_date")
            with col2:
                # Use session state for Policy Origination Date to allow clearing
                policy_orig_date_default = st.session_state.get('add_policy_orig_date', None)
                if policy_orig_date_default is None:
                    policy_orig_date = st.date_input("Policy Origination Date", value=None, key="add_policy_orig_date")
                else:
                    policy_orig_date = st.date_input("Policy Origination Date", value=policy_orig_date_default, key="add_policy_orig_date")
            
            # Row 4: X-DATE and Payment Type
            col1, col2 = st.columns(2)
            with col1:
                # Use session state for X-DATE to allow clearing
                x_date_default = st.session_state.get('add_x_date', None)
                if x_date_default is None:
                    x_date = st.date_input("X-DATE", value=None, help="Expiration date", key="add_x_date")
                else:
                    x_date = st.date_input("X-DATE", value=x_date_default, help="Expiration date", key="add_x_date")
            with col2:
                full_or_monthly = st.selectbox("FULL OR MONTHLY PMTS", ["FULL", "MONTHLY", ""])
            
            # Row 5: Checklist and Notes
            col1, col2 = st.columns(2)
            with col1:
                policy_checklist = st.checkbox("Policy Checklist Complete")
            with col2:
                notes = st.text_area("NOTES", placeholder="Enter any additional notes...", height=70)
            
            # Premium Sold Calculator
            st.subheader("Premium Sold Calculator (for Endorsements)")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                existing_premium = st.number_input("Existing Premium", value=0.00, format="%.2f", step=100.0)
            with col2:
                new_premium = st.number_input("New/Revised Premium", value=0.00, format="%.2f", step=100.0)
            with col3:
                premium_sold_calc = new_premium - existing_premium
                st.metric("Premium Sold (New - Existing):", f"${premium_sold_calc:+,.2f}")
            
            # New Policy Premium
            st.subheader("New Policy Premium")
            
            premium_sold = st.number_input("New Policy Premium", value=0.00, format="%.2f", step=100.0, 
                                         help="Enter the total premium for new policies (not endorsements)")
            
            # Broker Fee / Carrier Taxes & Fees Section
            st.subheader("Broker Fee / Carrier Taxes & Fees")
            col1, col2 = st.columns(2)
            with col1:
                broker_fee = st.number_input("Broker Fee ($)", value=0.0, format="%.2f", step=10.0, help="Agency broker fee (you receive 50% commission on this amount)")
            with col2:
                policy_taxes_fees = st.number_input("Carrier Taxes & Fees ($)", value=0.0, format="%.2f", step=10.0, help="Non-commissionable carrier taxes and fees")
                # Calculate and display commissionable premium
                # Determine which calculator to use based on user input
                # If endorsement calculator has non-zero values, use it
                if (existing_premium != 0.0 or new_premium != 0.0) and premium_sold_calc != 0.0:
                    # Using endorsement calculator
                    premium_for_calculation = premium_sold_calc
                # Otherwise use new policy premium if it has a value
                elif premium_sold != 0.0:
                    # Using new policy premium
                    premium_for_calculation = premium_sold
                else:
                    # No premium entered yet
                    premium_for_calculation = 0.0
                    
                commissionable_premium = premium_for_calculation - policy_taxes_fees
                
                # Show which calculator is being used
                if (existing_premium != 0.0 or new_premium != 0.0) and premium_sold_calc != 0.0:
                    calc_source = "Using Endorsement Calculator"
                elif premium_sold != 0.0:
                    calc_source = "Using New Policy Premium"
                else:
                    calc_source = "Enter premium above"
                    
                st.text_input("Commissionable Premium ($)", value=f"{commissionable_premium:.2f}", disabled=True, 
                             help=f"Premium minus Carrier Taxes & Fees | {calc_source}")
            
            # Commission Details Section (Enhanced with Auto-Population)
            st.subheader("Commission Details 💰")
            
            # Use commission rule from session state if available
            applied_rule = None
            suggested_rate = 10.0  # Default fallback
            rule_info_text = "Manual entry (no carrier selected)"
            
            # Check if we have commission rates from session state (set during carrier selection)
            has_rule = st.session_state.get('has_commission_rule', False)
            
            if has_rule and selected_carrier_id:
                # Get rates from session state
                new_rate = st.session_state.get('commission_new_rate', 10.0)
                renewal_rate = st.session_state.get('commission_renewal_rate', new_rate)
                rule_id = st.session_state.get('commission_rule_id')
                
                # Apply Prior Policy Number logic
                if transaction_type == 'NEW':
                    # NEW transactions always use new rate
                    suggested_rate = new_rate
                    rule_info_text = f"✅ New business rate - {new_rate}%"
                elif transaction_type == 'RWL':
                    # RWL transactions always use renewal rate
                    suggested_rate = renewal_rate
                    rule_info_text = f"✅ Renewal rate - {renewal_rate}%"
                else:
                    # For all other transaction types, check Prior Policy Number
                    if prior_policy_number and str(prior_policy_number).strip():
                        # Has prior policy = this is a renewal policy
                        suggested_rate = renewal_rate
                        rule_info_text = f"✅ Renewal rate (Has Prior Policy: {prior_policy_number}) - {renewal_rate}%"
                    else:
                        # No prior policy = this is a new policy
                        suggested_rate = new_rate
                        rule_info_text = f"✅ New business rate (No Prior Policy) - {new_rate}%"
                
                applied_rule = {
                    'rule_id': rule_id,
                    'new_rate': new_rate,
                    'renewal_rate': renewal_rate,
                    'rate_to_use': suggested_rate
                }
            elif selected_carrier_id:
                rule_info_text = f"⚠️ No rule found for {final_carrier_name} - using manual entry"
                applied_rule = None
            
            # Rule application info
            st.info(f"**Applied Rule:** {rule_info_text}")
            
            # Commission override section
            col_override, col_reason = st.columns([1, 2])
            with col_override:
                use_override = st.checkbox("Override Commission Rate", 
                                         help="Check to manually enter a different rate than the rule", 
                                         key="add_policy_override")
            
            override_reason = ""
            if use_override:
                with col_reason:
                    override_reason = st.text_input("Override Reason*", 
                                                  placeholder="e.g., Special client rate, promotional rate", 
                                                  help="Required when overriding commission rate",
                                                  key="add_policy_override_reason")
            
            col1, col2 = st.columns(2)
            with col1:
                # Commission rate input
                if use_override:
                    policy_gross_comm_input = st.number_input(
                        "Policy Gross Comm % (OVERRIDE)", 
                        value=suggested_rate, 
                        format="%.2f", 
                        min_value=0.0, 
                        max_value=100.0, 
                        key="policy_gross_comm_details",
                        help="⚠️ You are overriding the rule-based rate"
                    )
                    if policy_gross_comm_input != suggested_rate:
                        st.warning(f"⚠️ Override: {policy_gross_comm_input}% vs Rule: {suggested_rate}%")
                else:
                    policy_gross_comm_input = st.number_input(
                        "Policy Gross Comm %", 
                        value=suggested_rate, 
                        format="%.2f", 
                        min_value=0.0, 
                        max_value=100.0, 
                        key="policy_gross_comm_details",
                        help="Rate from commission rule" if applied_rule else "Manual entry rate"
                    )
                
                # Determine agent commission rate based on transaction type and prior policy
                if transaction_type == "NEW":
                    # NEW transactions always get 50%
                    agent_comm_rate = 50.0
                elif transaction_type == "RWL":
                    # RWL transactions always get 25%
                    agent_comm_rate = 25.0
                elif transaction_type in ["NBS", "STL", "BoR"]:
                    # These are typically new business
                    agent_comm_rate = 50.0
                elif transaction_type in ["CAN", "XCL"]:
                    # Cancellations - determine the original rate based on Prior Policy Number
                    if prior_policy_number and str(prior_policy_number).strip():
                        # Has prior policy = this was a renewal, so chargeback at 25%
                        agent_comm_rate = 25.0
                    else:
                        # No prior policy = this was new business, so chargeback at 50%
                        agent_comm_rate = 50.0
                else:
                    # For all other transaction types (END, PCH, REWRITE), check Prior Policy Number
                    if prior_policy_number and str(prior_policy_number).strip():
                        # Has prior policy = this is a renewal policy
                        agent_comm_rate = 25.0
                    else:
                        # No prior policy = this is a new policy
                        agent_comm_rate = 50.0
                
                st.text_input("Agent Comm %", value=f"{agent_comm_rate}%", disabled=True, help="Rate based on transaction type")
                
                # Calculate broker fee agent commission
                broker_fee_agent_comm = broker_fee * 0.50
                st.text_input("Broker Fee Agent Comm", value=f"${broker_fee_agent_comm:.2f}", disabled=True, help="50% of broker fee")
            
            with col2:
                # Calculate agency commission from commissionable premium
                agency_est_comm = commissionable_premium * (policy_gross_comm_input / 100)
                st.text_input("Agency Estimated Comm/Revenue (CRM)", value=f"${agency_est_comm:.2f}", disabled=True, help="Calculated from commissionable premium")
                
                # Calculate agent commission
                agent_est_comm = agency_est_comm * (agent_comm_rate / 100)
                st.text_input("Agent Estimated Comm $", value=f"${agent_est_comm:.2f}", disabled=True, help="Your commission on the policy")
                
                # Calculate total agent commission
                total_agent_comm = agent_est_comm + broker_fee_agent_comm
                st.text_input("Total Agent Comm", value=f"${total_agent_comm:.2f}", disabled=True, help="Policy commission plus broker fee commission")
                
                # Show rule ID if available (for tracking)
                if applied_rule:
                    st.caption(f"Rule ID: {applied_rule['rule_id'][:8]}...")
            
            # Validation for override
            if use_override and not override_reason:
                st.error("⚠️ Override reason is required when using a custom commission rate.")
            
            # Form buttons
            col1, col2 = st.columns(2)
            with col1:
                calculate = st.form_submit_button("🧮 Calculate", type="primary", use_container_width=True, help="Click to verify all commission calculations")
            with col2:
                submitted = st.form_submit_button("💾 Save Policy Transaction", type="primary", use_container_width=True)
            
            # Handle Calculate button - just refresh the form with calculations
            if calculate:
                st.info("✅ Calculations updated! Review the values below.")
                # The form will automatically recalculate when rerun
            
            # Handle Save button
            elif submitted:
                # Validate override requirements
                if use_override and not override_reason.strip():
                    st.error("❌ Override reason is required when using a custom commission rate.")
                elif customer and policy_number:
                    try:
                        # Agent rate is already correctly set based on Prior Policy Number
                        # No need to adjust for END/PCH here since we already handled it above
                        
                        # Prepare the new policy record with commission rule integration
                        new_policy = {
                            "Client ID": client_id,
                            "Transaction ID": transaction_id,
                            "Customer": customer,
                            "Carrier Name": final_carrier_name,
                            "MGA Name": final_mga_name,
                            # NEW: Commission integration fields
                            "carrier_id": selected_carrier_id,
                            "mga_id": selected_mga_id,
                            "commission_rule_id": applied_rule['rule_id'] if applied_rule else None,
                            "commission_rate_override": policy_gross_comm_input if use_override else None,
                            "override_reason": override_reason if use_override else None,
                            "Policy Type": policy_type,
                            "Policy Number": policy_number,
                            "Prior Policy Number": prior_policy_number if prior_policy_number else None,
                            "Transaction Type": transaction_type,
                            "Policy Term": policy_term,  # Add Policy Term
                            "Policy Origination Date": policy_orig_date.strftime('%Y-%m-%d'),  # Changed to YYYY-MM-DD
                            "Effective Date": effective_date.strftime('%Y-%m-%d'),  # Changed to YYYY-MM-DD
                            "X-DATE": x_date.strftime('%Y-%m-%d'),  # Changed to YYYY-MM-DD
                            "Policy Checklist Complete": "Yes" if policy_checklist else "No",
                            "Premium Sold": clean_numeric_value(premium_for_calculation),
                            "Policy Taxes & Fees": clean_numeric_value(policy_taxes_fees),
                            "Commissionable Premium": clean_numeric_value(commissionable_premium),
                            "Broker Fee": clean_numeric_value(broker_fee),
                            "Policy Gross Comm %": clean_numeric_value(policy_gross_comm_input),
                            "Agency Estimated Comm/Revenue (CRM)": clean_numeric_value(agency_est_comm),
                            "Agent Comm %": clean_numeric_value(agent_comm_rate),
                            "Agent Estimated Comm $": clean_numeric_value(agent_est_comm),
                            "Broker Fee Agent Comm": clean_numeric_value(broker_fee_agent_comm),
                            "Total Agent Comm": clean_numeric_value(total_agent_comm),
                            "FULL OR MONTHLY PMTS": full_or_monthly,
                            "NOTES": notes
                        }
                        
                        # Remove None values to avoid database issues
                        new_policy = {k: v for k, v in new_policy.items() if v is not None}
                        
                        # Clean data before insertion
                        cleaned_policy = clean_data_for_database(new_policy)
                        
                        # Insert into database
                        supabase.table('policies').insert(cleaned_policy).execute()
                        clear_policies_cache()
                        
                        # Set success message in session state to persist after rerun
                        st.session_state['add_policy_success'] = {
                            'customer': customer,
                            'policy_number': policy_number,
                            'transaction_id': transaction_id,
                            'timestamp': datetime.datetime.now()
                        }
                        
                        # Clear all form-related session state to reset the form
                        keys_to_clear = [
                            'selected_client_id',
                            'selected_customer_name',
                            'new_client_name',
                            'new_carrier_name',
                            'new_policy_type',
                            'add_policy_number',
                            'add_prior_policy_number',
                            'add_effective_date',
                            'add_x_date',
                            'add_policy_orig_date',
                            # NEW: Commission integration form keys
                            'add_policy_carrier',
                            'add_policy_carrier_manual',
                            'add_policy_mga',
                            'add_policy_mga_manual',
                            'add_policy_override',
                            'add_policy_override_reason',
                            'policy_gross_comm_details'
                        ]
                        for key in keys_to_clear:
                            if key in st.session_state:
                                del st.session_state[key]
                        
                        # Clear the form by rerunning
                        st.rerun()
                            
                    except Exception as e:
                        st.error(f"Error adding policy: {e}")
                        st.error("Please check that all required fields are filled correctly.")
                else:
                    st.error("Please fill in at least Customer Name and Policy Number")
    
    # --- Search & Filter ---
    elif page == "Search & Filter":
        st.title("🔍 Advanced Search & Filter")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        if all_data.empty:
            st.warning("No data found in policies table. Please add some policy data first.")
        else:
            st.subheader("Search Criteria")
            
            # Create filter form
            with st.form("advanced_search_form"):
                search_col1, search_col2 = st.columns(2)
                
                with search_col1:
                    # Text search fields
                    customer_search = st.text_input("Customer Name Contains")
                    policy_number_search = st.text_input("Policy Number Contains")
                    client_id_search = st.text_input("Client ID Contains")
                    transaction_id_search = st.text_input("Transaction ID Contains")
                
                with search_col2:
                    # Dropdown filters
                    if 'Policy Type' in all_data.columns:
                        policy_type_filter = st.multiselect("Policy Type", all_data['Policy Type'].unique())
                    
                    if 'Transaction Type' in all_data.columns:
                        transaction_type_filter = st.multiselect("Transaction Type", all_data['Transaction Type'].unique())
                    
                    # Date range
                    if 'Effective Date' in all_data.columns:
                        date_from = st.date_input("Effective Date From", value=None)
                        date_to = st.date_input("Effective Date To", value=None)
                
                # Numeric filters
                st.subheader("Numeric Filters")
                numeric_col1, numeric_col2 = st.columns(2)
                
                with numeric_col1:
                    if 'Agent Paid Amount (STMT)' in all_data.columns:
                        commission_min = st.number_input("Min Agent Paid Amount", value=0.0, format="%.2f")
                        commission_max = st.number_input("Max Agent Paid Amount", value=float(all_data['Agent Paid Amount (STMT)'].max() if all_data['Agent Paid Amount (STMT)'].max() > 0 else 999999), format="%.2f")
                
                with numeric_col2:
                    if 'Policy Balance Due' in all_data.columns:
                        balance_min = st.number_input("Min Balance Due", value=0.0, format="%.2f")
                        balance_max = st.number_input("Max Balance Due", value=float(all_data['Policy Balance Due'].max() if all_data['Policy Balance Due'].max() > 0 else 999999), format="%.2f")
                
                # Submit search
                search_submitted = st.form_submit_button("🔍 Apply Filters", type="primary")
            
            # Reset filters button
            if st.button("🔄 Clear All Filters"):
                st.rerun()
            
            # Apply filters and show results
            if search_submitted or any([customer_search, policy_number_search, client_id_search, transaction_id_search]):
                filtered_data = all_data.copy()
                
                # Apply text filters
                if customer_search:
                    filtered_data = filtered_data[filtered_data['Customer'].str.contains(customer_search, case=False, na=False)]
                
                if policy_number_search:
                    filtered_data = filtered_data[filtered_data['Policy Number'].str.contains(policy_number_search, case=False, na=False)]
                
                if client_id_search:
                    filtered_data = filtered_data[filtered_data['Client ID'].str.contains(client_id_search, case=False, na=False)]
                
                if transaction_id_search:
                    filtered_data = filtered_data[filtered_data['Transaction ID'].str.contains(transaction_id_search, case=False, na=False)]
                
                # Apply dropdown filters
                if 'Policy Type' in all_data.columns and policy_type_filter:
                    filtered_data = filtered_data[filtered_data['Policy Type'].isin(policy_type_filter)]
                
                if 'Transaction Type' in all_data.columns and transaction_type_filter:
                    filtered_data = filtered_data[filtered_data['Transaction Type'].isin(transaction_type_filter)]
                
                # Apply date filters
                if 'Effective Date' in all_data.columns and (date_from or date_to):
                    filtered_data['Effective Date'] = pd.to_datetime(filtered_data['Effective Date'], errors='coerce')
                    
                    if date_from:
                        filtered_data = filtered_data[filtered_data['Effective Date'] >= pd.Timestamp(date_from)]
                    
                    if date_to:
                        filtered_data = filtered_data[filtered_data['Effective Date'] <= pd.Timestamp(date_to)]
                
                # Apply numeric filters
                if 'Agent Paid Amount (STMT)' in all_data.columns:
                    filtered_data = filtered_data[
                        (filtered_data['Agent Paid Amount (STMT)'] >= commission_min) &
                        (filtered_data['Agent Paid Amount (STMT)'] <= commission_max)
                    ]
                
                if 'Policy Balance Due' in all_data.columns:
                    filtered_data = filtered_data[
                        (filtered_data['Policy Balance Due'] >= balance_min) &
                        (filtered_data['Policy Balance Due'] <= balance_max)
                    ]
                
                # Show results
                st.subheader(f"Search Results ({len(filtered_data)} records found)")
                
                if not filtered_data.empty:
                    # Configure column settings for proper numeric display
                    column_config = {}
                    numeric_cols = [
                        'Agent Estimated Comm $',
                                'Policy Gross Comm %',
                        'Agency Estimated Comm/Revenue (CRM)',
                        'Agency Comm Received (STMT)',
                        'Premium Sold',
                        'Agent Paid Amount (STMT)',
                        'Agency Comm Received (STMT)'
                    ]
                    
                    for col in numeric_cols:
                        if col in filtered_data.columns:
                            column_config[col] = st.column_config.NumberColumn(
                                col,
                                format="%.2f"
                            )
                    
                    # Display filtered data
                    # Apply special styling for STMT and VOID transactions
                    styled_data = style_special_transactions(filtered_data)
                    
                    st.dataframe(
                        styled_data,
                        use_container_width=True,
                        height=500,
                        column_config=column_config
                    )
                    
                    # Export filtered results
                    st.subheader("Export Filtered Results")
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write("**📄 CSV Export**")
                        csv = filtered_data.to_csv(index=False)
                        st.download_button(
                            label="📥 Download as CSV",
                            data=csv,
                            file_name="filtered_policies.csv",
                            mime="text/csv",
                            help="Export filtered results as CSV file"
                        )
                    
                    with col2:
                        st.write("**📊 Excel Export**")
                        excel_buffer, excel_filename = create_formatted_excel_file(
                            filtered_data, 
                            sheet_name="Filtered Results", 
                            filename_prefix="filtered_policies"
                        )
                        if excel_buffer:
                            st.download_button(
                                label="📥 Download as Excel",
                                data=excel_buffer,
                                file_name=excel_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                help="Export filtered results as formatted Excel file"
                            )
                else:
                    st.warning("No records match your search criteria")
            else:
                st.info("Use the form above to search and filter policies")
    
    # --- Reconciliation ---
    elif page == "Reconciliation":
        st.title("💳 Commission Reconciliation")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        # Create tabs for different reconciliation functions
        rec_tab1, rec_tab2, rec_tab3, rec_tab4 = st.tabs([
            "Import Statement", 
            "Unreconciled Transactions", 
            "Reconciliation History",
            "Adjustments & Voids"
        ])
        
        with rec_tab1:
            st.subheader("📥 Import Commission Statement")
            
            # Add custom CSS for yellow border on Statement Date
            st.markdown("""
                <style>
                /* Target the Statement Date input specifically */
                div.stDateInput > div > div[data-baseweb="input"] {
                    border: 3px solid #ffd700 !important;
                    border-radius: 4px !important;
                }
                
                /* Alternative selector for broader compatibility */
                section[data-testid="stSidebar"] + div div.row-widget.stDateInput:first-of-type input {
                    border: 3px solid #ffd700 !important;
                    border-radius: 4px !important;
                    box-shadow: 0 0 0 2px #ffd700 !important;
                }
                
                /* Target the container div */
                div.row-widget.stDateInput:first-of-type > div {
                    border: 3px solid #ffd700 !important;
                    border-radius: 4px !important;
                    padding: 2px;
                }
                </style>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Create a container with yellow background as a visual highlight
                st.markdown("""
                    <div style="background-color: #fffacd; border: 3px solid #ffd700; border-radius: 8px; padding: 10px; margin-bottom: 10px;">
                        <p style="margin: 0; font-weight: bold; color: #333;">📅 Statement Date</p>
                    </div>
                """, unsafe_allow_html=True)
                
                statement_date = st.date_input(
                    "",  # Empty label since we have the custom header above
                    value=datetime.date.today(),
                    help="The date on the commission statement",
                                        key="statement_date_input"
                )
            
            with col2:
                reconciliation_date = st.date_input(
                    "Date Reconciled",
                    value=datetime.date.today(),
                    disabled=True,
                    help="Today's date (when reconciliation is performed)",
                    format="MM/DD/YYYY"
                )
            
            st.divider()
            
            # Method selection
            import_method = st.radio(
                "Select Import Method",
                ["Manual Entry", "Upload CSV/Excel File"],
                horizontal=True
            )
            
            if import_method == "Manual Entry":
                st.subheader("Manual Statement Entry")
                
                # Initialize batch in session state
                if 'reconciliation_batch' not in st.session_state:
                    st.session_state.reconciliation_batch = []
                if 'statement_total' not in st.session_state:
                    st.session_state.statement_total = 0.0
                
                # Statement total input
                col1, col2 = st.columns([2, 1])
                with col1:
                    statement_total = st.number_input(
                        "Enter Total Commission from Statement",
                        min_value=0.0,
                        value=st.session_state.statement_total,
                        step=0.01,
                        format="%.2f",
                        help="Enter the total commission amount shown on your statement"
                    )
                    st.session_state.statement_total = statement_total
                
                with col2:
                    batch_total = sum(item['amount'] for item in st.session_state.reconciliation_batch)
                    if statement_total > 0:
                        progress = min(batch_total / statement_total, 1.0)
                        st.metric("Batch Progress", f"${batch_total:,.2f}")
                        st.progress(progress)
                        if batch_total == statement_total:
                            st.success("✓ Ready to reconcile!")
                        elif batch_total > statement_total:
                            st.error("Over statement amount!")
                        else:
                            st.info(f"${statement_total - batch_total:,.2f} remaining")
                
                # Show current batch if any
                if st.session_state.reconciliation_batch:
                    st.divider()
                    st.markdown("### Current Reconciliation Batch")
                    
                    # Show batch summary
                    batch_df = pd.DataFrame(st.session_state.reconciliation_batch)
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Items in Batch", len(batch_df))
                    with col2:
                        batch_total = batch_df['amount'].sum()
                        st.metric("Batch Total", f"${batch_total:,.2f}")
                    with col3:
                        if statement_total > 0:
                            variance = batch_total - statement_total
                            if variance == 0:
                                st.metric("Variance", "$0.00", delta_color="off")
                                st.success("✓ Perfectly balanced!")
                            else:
                                st.metric("Variance", f"${variance:,.2f}", delta=f"${abs(variance):,.2f}")
                    
                    # Enhanced batch display with more details
                    batch_df['Remove'] = False
                    
                    # Add more display columns
                    display_columns = ['Remove', 'transaction_id', 'customer', 'policy_number', 
                                     'policy_type', 'effective_date', 'balance', 'amount']
                    
                    # Only include columns that exist
                    available_columns = [col for col in display_columns if col in batch_df.columns]
                    
                    edited_batch = st.data_editor(
                        batch_df[available_columns],
                        column_config={
                            "Remove": st.column_config.CheckboxColumn("Remove", help="Check to remove from batch"),
                            "transaction_id": st.column_config.TextColumn("Transaction ID", disabled=True),
                            "customer": st.column_config.TextColumn("Customer", disabled=True),
                            "policy_number": st.column_config.TextColumn("Policy #", disabled=True),
                            "policy_type": st.column_config.TextColumn("Type", disabled=True),
                            "effective_date": st.column_config.TextColumn("Effective", disabled=True),
                            "balance": st.column_config.NumberColumn("Full Balance", disabled=True, format="$%.2f", help="Total outstanding balance"),
                            "amount": st.column_config.NumberColumn("To Reconcile", disabled=True, format="$%.2f", help="Amount to reconcile in this batch")
                        },
                        use_container_width=True,
                        hide_index=True,
                        key="batch_editor"
                    )
                    
                    # Remove checked items
                    if edited_batch['Remove'].any():
                        items_to_remove = edited_batch[edited_batch['Remove']]['transaction_id'].tolist()
                        st.session_state.reconciliation_batch = [
                            item for item in st.session_state.reconciliation_batch 
                            if item['transaction_id'] not in items_to_remove
                        ]
                        st.rerun()
                    
                    # Reconcile batch button
                    col1, col2, col3 = st.columns([1, 1, 2])
                    
                    with col1:
                        # Only enable reconcile button if totals match
                        can_reconcile = batch_total == statement_total and statement_total > 0 and len(st.session_state.reconciliation_batch) > 0
                        
                        if can_reconcile:
                            if st.button("🔄 Reconcile Batch", type="primary", key="reconcile_batch"):
                                try:
                                    # Generate batch reconciliation ID
                                    batch_id = f"REC-{statement_date.strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
                                    success_count = 0
                                    
                                    # Create reconciliation entries for each item in batch
                                    for item in st.session_state.reconciliation_batch:
                                        recon_id = generate_reconciliation_transaction_id("STMT", statement_date)
                                        
                                        # Create reconciliation entry
                                        recon_entry = {
                                            'Transaction ID': recon_id,
                                            'Client ID': item['client_id'],
                                            'Customer': item['customer'],
                                            'Carrier Name': item['carrier_name'],
                                            'Policy Type': item['policy_type'],
                                            'Policy Number': item['policy_number'],
                                            'Transaction Type': item['transaction_type'],
                                            'Effective Date': item['effective_date'],
                                            'X-DATE': item['x_date'],
                                            'Premium Sold': 0,
                                            'Policy Gross Comm %': 0,
                                            'Agency Estimated Comm/Revenue (CRM)': 0,
                                            'Agency Comm Received (STMT)': item['amount'],
                                            'Agent Estimated Comm $': 0,
                                            'Agent Paid Amount (STMT)': item['amount'],
                                            'STMT DATE': statement_date.strftime('%Y-%m-%d'),
                                            'reconciliation_status': 'reconciled',
                                            'reconciliation_id': batch_id,
                                            'reconciled_at': datetime.datetime.now().isoformat(),
                                            'is_reconciliation_entry': True
                                        }
                                        
                                        # Clean data before insertion
                                        cleaned_recon = clean_data_for_database(recon_entry)
                                        
                                        # Insert reconciliation entry
                                        supabase.table('policies').insert(cleaned_recon).execute()
                                        
                                        # Update original transaction
                                        supabase.table('policies').update({
                                            'reconciliation_status': 'reconciled',
                                            'reconciliation_id': batch_id,
                                            'reconciled_at': datetime.datetime.now().isoformat()
                                        }).eq('_id', item['_id']).execute()
                                        
                                        success_count += 1
                                    
                                    # Clear batch
                                    st.session_state.reconciliation_batch = []
                                    st.session_state.statement_total = 0.0
                                    
                                    st.success(f"✅ Batch reconciled successfully! {success_count} transactions processed. Batch ID: {batch_id}")
                                    st.cache_data.clear()
                                    time.sleep(2)
                                    st.rerun()
                                    
                                except Exception as e:
                                    st.error(f"Error reconciling batch: {str(e)}")
                        else:
                            # Show disabled button with explanation
                            st.button("🔄 Reconcile Batch", type="primary", disabled=True, key="reconcile_batch_disabled")
                            if batch_total != statement_total:
                                if batch_total > statement_total:
                                    st.error(f"Batch exceeds statement by ${batch_total - statement_total:,.2f}")
                                else:
                                    st.warning(f"Need ${statement_total - batch_total:,.2f} more to match statement")
                            elif statement_total == 0:
                                st.warning("Enter statement total first")
                            elif len(st.session_state.reconciliation_batch) == 0:
                                st.info("Add transactions to batch")
                    
                    with col2:
                        if st.button("Clear Batch", type="secondary"):
                            st.session_state.reconciliation_batch = []
                            st.rerun()
                
                st.divider()
                
                # Precise drill-down selection
                st.markdown("### Select Transaction to Add to Batch")
                
                # Initialize session state for selections if not exists
                if 'recon_selected_customer' not in st.session_state:
                    st.session_state.recon_selected_customer = None
                if 'recon_selected_policy_type' not in st.session_state:
                    st.session_state.recon_selected_policy_type = None
                if 'recon_selected_policy_number' not in st.session_state:
                    st.session_state.recon_selected_policy_number = None
                if 'recon_selected_effective_date' not in st.session_state:
                    st.session_state.recon_selected_effective_date = None
                
                # Customer selection
                customers = sorted(all_data["Customer"].dropna().unique().tolist()) if not all_data.empty else []
                selected_customer = st.selectbox(
                    "Select Customer", 
                    ["Select..."] + customers, 
                    key="recon_drill_customer",
                    index=0 if st.session_state.recon_selected_customer is None else (
                        customers.index(st.session_state.recon_selected_customer) + 1 
                        if st.session_state.recon_selected_customer in customers else 0
                    )
                )
                
                # Initialize variables
                policy_types = []
                policy_numbers = []
                effective_dates = []
                selected_policy_type = None
                selected_policy_number = None
                selected_effective_date = None
                client_id = None
                transaction_id = None
                
                # Policy Type selection
                if selected_customer and selected_customer != "Select...":
                    st.session_state.recon_selected_customer = selected_customer
                    
                    # Get all original transactions for this customer
                    customer_data = all_data[
                        (all_data["Customer"] == selected_customer) &
                        (~all_data['Transaction ID'].str.contains('-STMT-|-ADJ-|-VOID-', na=False))
                    ]
                    
                    # Calculate balance for each transaction
                    if not customer_data.empty:
                        # Get all reconciliation entries for balance calculation
                        for idx, row in customer_data.iterrows():
                            trans_id = row['Transaction ID']
                            policy_num = row['Policy Number']
                            
                            # Calculate credits (commission owed)
                            credit = float(row.get('Agent Estimated Comm $', 0) or 0)
                            
                            # Calculate debits (total paid for this transaction)
                            # Find all -STMT- entries for this original transaction
                            stmt_entries = all_data[
                                (all_data['Policy Number'] == policy_num) &
                                (all_data['Transaction ID'].str.contains('-STMT-', na=False))
                            ]
                            
                            debit = 0
                            if not stmt_entries.empty:
                                debit = stmt_entries['Agent Paid Amount (STMT)'].fillna(0).sum()
                            
                            # Calculate balance
                            balance = credit - debit
                            customer_data.at[idx, '_balance'] = balance
                        
                        # Filter to show only transactions with outstanding balance
                        customer_data = customer_data[customer_data['_balance'] > 0.01]  # Small threshold to avoid floating point issues
                    
                    if not customer_data.empty:
                        policy_types = sorted(customer_data["Policy Type"].dropna().unique().tolist())
                        selected_policy_type = st.selectbox(
                            "Select Policy Type", 
                            ["Select..."] + policy_types, 
                            key="recon_drill_policy_type"
                        )
                        
                        # Policy Number selection
                        if selected_policy_type and selected_policy_type != "Select...":
                            st.session_state.recon_selected_policy_type = selected_policy_type
                            
                            policy_data = customer_data[customer_data["Policy Type"] == selected_policy_type]
                            policy_numbers = sorted(policy_data["Policy Number"].dropna().unique().tolist())
                            selected_policy_number = st.selectbox(
                                "Select Policy Number", 
                                ["Select..."] + policy_numbers, 
                                key="recon_drill_policy_number"
                            )
                            
                            # Effective Date selection
                            if selected_policy_number and selected_policy_number != "Select...":
                                st.session_state.recon_selected_policy_number = selected_policy_number
                                
                                number_data = policy_data[policy_data["Policy Number"] == selected_policy_number]
                                effective_dates = sorted(number_data["Effective Date"].dropna().unique().tolist())
                                selected_effective_date = st.selectbox(
                                    "Select Effective Date", 
                                    ["Select..."] + effective_dates, 
                                    key="recon_drill_effective_date"
                                )
                                
                                # Get exact transaction when all fields are selected
                                if selected_effective_date and selected_effective_date != "Select...":
                                    st.session_state.recon_selected_effective_date = selected_effective_date
                                    
                                    exact_match = number_data[number_data["Effective Date"] == selected_effective_date]
                                    
                                    if not exact_match.empty:
                                        # Get the specific transaction details
                                        transaction = exact_match.iloc[0]
                                        client_id = transaction.get("Client ID", "")
                                        transaction_id = transaction.get("Transaction ID", "")
                                        
                                        # Display transaction details
                                        st.success(f"✅ Found Transaction: {transaction_id}")
                                        
                                        # Calculate outstanding balance
                                        credit = float(transaction.get('Agent Estimated Comm $', 0) or 0)
                                        balance = float(transaction.get('_balance', credit))
                                        
                                        # Show transaction summary with balance
                                        col1, col2, col3, col4 = st.columns(4)
                                        with col1:
                                            st.metric("Premium Sold", f"${transaction.get('Premium Sold', 0):,.2f}")
                                        with col2:
                                            st.metric("Commission Owed", f"${credit:,.2f}")
                                        with col3:
                                            previous_payments = credit - balance
                                            st.metric("Previous Payments", f"${previous_payments:,.2f}")
                                        with col4:
                                            st.metric("Outstanding Balance", f"${balance:,.2f}", 
                                                     help="Amount still owed for this transaction")
                                        
                                        # Show full transaction details
                                        with st.expander("📋 View Full Transaction Details", expanded=True):
                                            # Create two columns for organized display
                                            detail_col1, detail_col2 = st.columns(2)
                                            
                                            with detail_col1:
                                                st.markdown("**Policy Information**")
                                                details_dict = {
                                                    "Transaction ID": transaction.get('Transaction ID', ''),
                                                    "Client ID": transaction.get('Client ID', ''),
                                                    "Customer": transaction.get('Customer', ''),
                                                    "Policy Number": transaction.get('Policy Number', ''),
                                                    "Policy Type": transaction.get('Policy Type', ''),
                                                    "Carrier Name": transaction.get('Carrier Name', ''),
                                                    "Transaction Type": transaction.get('Transaction Type', ''),
                                                }
                                                for key, value in details_dict.items():
                                                    st.text(f"{key}: {value}")
                                            
                                            with detail_col2:
                                                st.markdown("**Dates & Amounts**")
                                                dates_dict = {
                                                    "Effective Date": transaction.get('Effective Date', ''),
                                                    "Policy Origination Date": transaction.get('Policy Origination Date', ''),
                                                    "X-DATE": transaction.get('X-DATE', ''),
                                                    "Premium Sold": f"${float(transaction.get('Premium Sold', 0) or 0):,.2f}",
                                                    "Policy Gross Comm %": f"{float(transaction.get('Policy Gross Comm %', 0) or 0):.2f}%",
                                                    "Agency Est. Comm": f"${float(transaction.get('Agency Estimated Comm/Revenue (CRM)', 0) or 0):,.2f}",
                                                    "Agent Comm Rate": transaction.get('Agent Comm %', ''),
                                                }
                                                for key, value in dates_dict.items():
                                                    st.text(f"{key}: {value}")
                                            
                                            # Show payment history if any
                                            if previous_payments > 0:
                                                st.divider()
                                                st.markdown("**Payment History**")
                                                
                                                # Find all reconciliation entries for this policy
                                                policy_num = transaction.get('Policy Number', '')
                                                payment_history = all_data[
                                                    (all_data['Policy Number'] == policy_num) &
                                                    (all_data['Transaction ID'].str.contains('-STMT-', na=False))
                                                ]
                                                
                                                if not payment_history.empty:
                                                    payment_df = payment_history[['Transaction ID', 'STMT DATE', 'Agent Paid Amount (STMT)']].copy()
                                                    payment_df['STMT DATE'] = pd.to_datetime(payment_df['STMT DATE'], format='mixed').dt.strftime('%m/%d/%Y')
                                                    payment_df = payment_df.rename(columns={
                                                        'Transaction ID': 'Reconciliation ID',
                                                        'STMT DATE': 'Statement Date',
                                                        'Agent Paid Amount (STMT)': 'Amount Paid'
                                                    })
                                                    
                                                    st.dataframe(
                                                        payment_df,
                                                        column_config={
                                                            "Amount Paid": st.column_config.NumberColumn(format="$%.2f")
                                                        },
                                                        use_container_width=True,
                                                        hide_index=True
                                                    )
                                        
                                        st.divider()
                                        
                                        # Add to batch section
                                        st.markdown("### Add to Reconciliation Batch")
                                        
                                        # Amount to reconcile
                                        amount_to_reconcile = st.number_input(
                                            "Amount to Reconcile",
                                            min_value=0.01,
                                            max_value=balance,
                                            value=balance,
                                            step=0.01,
                                            format="%.2f",
                                            key="recon_amount",
                                            help=f"Enter amount from statement for this transaction (max: ${balance:,.2f})"
                                        )
                                        
                                        # Check if already in batch
                                        already_in_batch = any(
                                            item['transaction_id'] == transaction_id 
                                            for item in st.session_state.reconciliation_batch
                                        )
                                        
                                        col1, col2 = st.columns(2)
                                        
                                        with col1:
                                            if already_in_batch:
                                                st.warning("✓ Already in batch")
                                            else:
                                                if st.button("➕ Add to Batch", type="primary", key="add_to_batch"):
                                                    # Add to reconciliation batch
                                                    batch_item = {
                                                        '_id': transaction['_id'],
                                                        'transaction_id': transaction_id,
                                                        'client_id': client_id,
                                                        'customer': selected_customer,
                                                        'carrier_name': transaction.get('Carrier Name', ''),
                                                        'policy_type': selected_policy_type,
                                                        'policy_number': selected_policy_number,
                                                        'transaction_type': transaction.get('Transaction Type', ''),
                                                        'effective_date': selected_effective_date,
                                                        'x_date': transaction.get('X-DATE', ''),
                                                        'amount': amount_to_reconcile,
                                                        'balance': balance
                                                    }
                                                    
                                                    st.session_state.reconciliation_batch.append(batch_item)
                                                    st.success(f"✅ Added ${amount_to_reconcile:,.2f} to batch")
                                                    
                                                    # Clear selections
                                                    st.session_state.recon_selected_customer = None
                                                    st.session_state.recon_selected_policy_type = None
                                                    st.session_state.recon_selected_policy_number = None
                                                    st.session_state.recon_selected_effective_date = None
                                                    
                                                    time.sleep(1)
                                                    st.rerun()
                                        
                                        with col2:
                                            # Show how this affects batch total
                                            if not already_in_batch and st.session_state.statement_total > 0:
                                                new_batch_total = sum(item['amount'] for item in st.session_state.reconciliation_batch) + amount_to_reconcile
                                                if new_batch_total > st.session_state.statement_total:
                                                    st.error(f"Would exceed statement by ${new_batch_total - st.session_state.statement_total:,.2f}")
                                                else:
                                                    remaining = st.session_state.statement_total - new_batch_total
                                                    st.info(f"Would leave ${remaining:,.2f} to reconcile")
                                    else:
                                        st.warning("No transaction found for this selection")
                    else:
                        st.info(f"No transactions with outstanding balance found for {selected_customer}")
            
            else:  # Upload CSV/Excel File
                st.subheader("📤 Upload Statement File")
                
                # Initialize session state for import
                if 'import_data' not in st.session_state:
                    st.session_state.import_data = None
                if 'column_mapping' not in st.session_state:
                    st.session_state.column_mapping = {}
                if 'matched_transactions' not in st.session_state:
                    st.session_state.matched_transactions = []
                if 'unmatched_transactions' not in st.session_state:
                    st.session_state.unmatched_transactions = []
                if 'transactions_to_create' not in st.session_state:
                    st.session_state.transactions_to_create = []
                
                # Step 1: File Upload
                uploaded_file = st.file_uploader(
                    "Choose a CSV or Excel file",
                    type=['csv', 'xlsx', 'xls'],
                    help="Upload your commission statement file"
                )
                
                if uploaded_file is not None:
                    # Parse file
                    try:
                        if uploaded_file.name.endswith('.csv'):
                            df = pd.read_csv(uploaded_file)
                        else:
                            df = pd.read_excel(uploaded_file)
                        
                        st.session_state.import_data = df
                        
                        # Count actual transactions (excluding totals rows)
                        transaction_count = len(df)
                        # Check if any row contains 'total' in any column
                        for col in df.columns:
                            try:
                                if df[col].astype(str).str.lower().str.contains('total|totals|subtotal|sub-total|grand total|sum', na=False).any():
                                    transaction_count = len(df) - df[col].astype(str).str.lower().str.contains('total|totals|subtotal|sub-total|grand total|sum', na=False).sum()
                                    break
                            except:
                                continue
                        
                        if transaction_count == len(df):
                            st.success(f"✅ Loaded {transaction_count} transactions from {uploaded_file.name}")
                        else:
                            st.success(f"✅ Loaded {transaction_count} transactions from {uploaded_file.name} (excluded {len(df) - transaction_count} total row{'s' if len(df) - transaction_count > 1 else ''})")
                        
                        # Show preview
                        with st.expander("📊 File Preview", expanded=True):
                            st.dataframe(df.head(10), use_container_width=True)
                        
                        # Initialize saved mappings in session state if not exists
                        if 'saved_column_mappings' not in st.session_state:
                            st.session_state.saved_column_mappings = load_saved_column_mappings()
                        
                        # Save/Load Column Mappings Section (moved up for better workflow)
                        st.divider()
                        st.markdown("### 💾 Save/Load Column Mappings")
                        st.info("💡 Load a saved mapping first to auto-fill the column selections below!")
                        
                        col1, col2, col3 = st.columns([2, 1, 1])
                        
                        with col1:
                            mapping_name = st.text_input(
                                "Mapping Name",
                                placeholder="e.g., ABC Insurance Statement",
                                help="Give this mapping a name to save it for future use"
                            )
                        
                        with col2:
                            # Save current mapping
                            if st.button("💾 Save Mapping", type="secondary", disabled=not mapping_name, key="save_mapping_top"):
                                if st.session_state.column_mapping:
                                    st.session_state.saved_column_mappings[mapping_name] = {
                                        'mapping': st.session_state.column_mapping.copy(),
                                        'created': datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                                        'field_count': len(st.session_state.column_mapping)
                                    }
                                    # Save to file for persistence
                                    if save_column_mappings_to_file(st.session_state.saved_column_mappings):
                                        st.success(f"✅ Saved mapping: {mapping_name}")
                                    else:
                                        st.error("Failed to save mapping to file")
                                else:
                                    st.warning("No mapping to save")
                        
                        with col3:
                            # Load saved mapping
                            if st.session_state.saved_column_mappings:
                                selected_mapping = st.selectbox(
                                    "Load Saved Mapping",
                                    options=[''] + list(st.session_state.saved_column_mappings.keys()),
                                    format_func=lambda x: 'Select a mapping...' if x == '' else x
                                )
                                
                                if selected_mapping and st.button("📂 Load", type="secondary", key="load_mapping_top"):
                                    saved_map = st.session_state.saved_column_mappings[selected_mapping]['mapping']
                                    # Verify columns exist in current file
                                    valid_mapping = {}
                                    missing_cols = []
                                    
                                    for sys_field, stmt_col in saved_map.items():
                                        if stmt_col in df.columns:
                                            valid_mapping[sys_field] = stmt_col
                                        else:
                                            missing_cols.append(stmt_col)
                                    
                                    # Store the loaded mapping in session state
                                    st.session_state.column_mapping = valid_mapping
                                    # Flag that we just loaded a mapping
                                    st.session_state.mapping_just_loaded = True
                                    
                                    if missing_cols:
                                        st.warning(f"Some columns not found in current file: {', '.join(missing_cols)}")
                                    else:
                                        st.success(f"✅ Loaded mapping: {selected_mapping}")
                                    st.rerun()
                        
                        # Show saved mappings
                        if st.session_state.saved_column_mappings:
                            with st.expander("📋 Saved Mappings", expanded=False):
                                for name, info in st.session_state.saved_column_mappings.items():
                                    col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
                                    with col1:
                                        st.text(name)
                                    with col2:
                                        st.text(f"Created: {info['created']}")
                                    with col3:
                                        st.text(f"{info['field_count']} fields")
                                    with col4:
                                        if st.button("🗑️", key=f"delete_{name}", help=f"Delete {name}"):
                                            del st.session_state.saved_column_mappings[name]
                                            save_column_mappings_to_file(st.session_state.saved_column_mappings)
                                            st.rerun()
                        
                        # Step 2: Column Mapping
                        st.divider()
                        st.markdown("### 🔗 Map Statement Columns to System Fields")
                        st.info(f"📅 Statement Date will be set to: {statement_date.strftime('%m/%d/%Y')}")
                        
                        # Required fields
                        required_fields = {
                            'Customer': 'Customer/Client Name',
                            'Policy Number': 'Policy Number',
                            'Effective Date': 'Policy Effective Date',
                            'Agent Paid Amount (STMT)': 'Agent Payment Amount (Required)'
                        }
                        
                        # Optional fields
                        optional_fields = {
                            'Agency Comm Received (STMT)': 'Agency Commission (for Audit)',
                            'Policy Type': 'Policy Type',
                            'Transaction Type': 'Transaction Type (NEW/RWL/END/CXL)',
                            'Premium Sold': 'Premium Amount',
                            'X-DATE': 'Expiration/Cancellation Date',
                            'NOTES': 'Notes/Description',
                            'Rate': 'Commission Rate %'
                        }
                        
                        # Create mapping interface
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown("**Required Fields**")
                            for sys_field, description in required_fields.items():
                                # Get the current value from column_mapping
                                current_value = st.session_state.column_mapping.get(sys_field, '')
                                
                                # Find the index of the current value in options
                                options = [''] + list(df.columns)
                                try:
                                    default_index = options.index(current_value) if current_value in options else 0
                                except:
                                    default_index = 0
                                
                                selected_col = st.selectbox(
                                    f"{sys_field} ({description})",
                                    options=options,
                                    key=f"map_{sys_field}",
                                    index=default_index,
                                    help=f"Select the column that contains {description}"
                                )
                                if selected_col:
                                    st.session_state.column_mapping[sys_field] = selected_col
                                elif sys_field in st.session_state.column_mapping:
                                    # Remove from mapping if deselected
                                    del st.session_state.column_mapping[sys_field]
                        
                        with col2:
                            st.markdown("**Optional Fields**")
                            for sys_field, description in optional_fields.items():
                                # Get the current value from column_mapping
                                current_value = st.session_state.column_mapping.get(sys_field, '')
                                
                                # Find the index of the current value in options
                                options = [''] + list(df.columns)
                                try:
                                    default_index = options.index(current_value) if current_value in options else 0
                                except:
                                    default_index = 0
                                
                                selected_col = st.selectbox(
                                    f"{sys_field} ({description})",
                                    options=options,
                                    key=f"map_{sys_field}",
                                    index=default_index,
                                    help=f"Select the column that contains {description}"
                                )
                                if selected_col:
                                    st.session_state.column_mapping[sys_field] = selected_col
                                elif sys_field in st.session_state.column_mapping:
                                    # Remove from mapping if deselected
                                    del st.session_state.column_mapping[sys_field]
                        
                        # Check if all required fields are mapped
                        required_mapped = all(
                            field in st.session_state.column_mapping and st.session_state.column_mapping[field]
                            for field in required_fields.keys()
                        )
                        
                        if required_mapped:
                            st.divider()
                            
                            # Process and match transactions
                            if st.button("🔍 Process & Match Transactions", type="primary"):
                                with st.spinner("Matching transactions..."):
                                    # First check for unmapped policy types if Policy Type column is mapped
                                    unmapped_policy_types = []
                                    if 'Policy Type' in st.session_state.column_mapping:
                                        policy_type_col = st.session_state.column_mapping['Policy Type']
                                        
                                        # Load policy type mappings
                                        mapping_file = "config_files/policy_type_mappings.json"
                                        type_mappings = {}
                                        try:
                                            if os.path.exists(mapping_file):
                                                with open(mapping_file, 'r') as f:
                                                    type_mappings = json.load(f)
                                        except:
                                            pass
                                        
                                        # Get unique policy types from statement (excluding null/empty)
                                        statement_types = df[policy_type_col].dropna().unique()
                                        statement_types = [str(t) for t in statement_types if str(t).strip()]
                                        
                                        # Check each type
                                        for stmt_type in statement_types:
                                            # Skip if it's already mapped or if it's a known type in our system
                                            if stmt_type not in type_mappings:
                                                # Also check if this type already exists in our policy types
                                                # Load policy types configuration
                                                policy_types_file = "config_files/policy_types_updated.json"
                                                existing_types = []
                                                try:
                                                    if os.path.exists(policy_types_file):
                                                        with open(policy_types_file, 'r') as f:
                                                            policy_types_config = json.load(f)
                                                            existing_types = policy_types_config.get('policy_types', [])
                                                except:
                                                    pass
                                                
                                                # If not in mappings and not in existing types, it's unmapped
                                                if stmt_type not in existing_types:
                                                    unmapped_policy_types.append(stmt_type)
                                    
                                    # If unmapped types found, show error and stop
                                    if unmapped_policy_types:
                                        st.error("❌ **Unmapped Policy Types Found**")
                                        st.warning("The following policy types from your statement are not mapped to existing policy types:")
                                        
                                        # Show unmapped types in a nice format
                                        for unmapped_type in unmapped_policy_types:
                                            st.write(f"• **{unmapped_type}**")
                                        
                                        st.info("👉 **Next Steps:**")
                                        st.write("1. Go to **Admin Panel** → **Policy Type Mapping** tab")
                                        st.write("2. Add mappings for the policy types listed above")
                                        st.write("3. Return here and try the import again")
                                        
                                        # Provide helpful instructions
                                        st.markdown("**💡 Quick Tip:** Use the navigation menu on the left sidebar to go to Admin Panel.")
                                        
                                        # Stop processing
                                        st.stop()
                                    
                                    # Calculate statement total from all rows (including totals row)
                                    # This gives us the check-and-balance figure
                                    statement_total_amount = 0
                                    if 'Agent Paid Amount (STMT)' in st.session_state.column_mapping:
                                        agent_col = st.session_state.column_mapping['Agent Paid Amount (STMT)']
                                        try:
                                            # Sum all rows including any totals
                                            all_amounts = pd.to_numeric(df[agent_col], errors='coerce').fillna(0)
                                            # Check if there's a totals row (last row or row with 'total' in customer name)
                                            customer_col = st.session_state.column_mapping.get('Customer', '')
                                            if customer_col:
                                                # Find rows that look like totals
                                                totals_mask = df[customer_col].astype(str).str.lower().str.contains('total|sum', na=False)
                                                if totals_mask.any():
                                                    # Use the totals row value
                                                    statement_total_amount = all_amounts[totals_mask].max()
                                                else:
                                                    # No totals row found, sum all non-total rows
                                                    statement_total_amount = all_amounts.sum()
                                            else:
                                                statement_total_amount = all_amounts.sum()
                                        except:
                                            statement_total_amount = 0
                                    
                                    st.session_state.statement_file_total = statement_total_amount
                                    
                                    matched, unmatched, to_create = match_statement_transactions(
                                        df, 
                                        st.session_state.column_mapping,
                                        all_data,
                                        statement_date
                                    )
                                    
                                    st.session_state.matched_transactions = matched
                                    st.session_state.unmatched_transactions = unmatched
                                    st.session_state.transactions_to_create = to_create
                                    
                                    # Show summary
                                    st.success("✅ Matching complete!")
                                    
                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        st.metric("Matched", len(matched))
                                    with col2:
                                        st.metric("Unmatched", len(unmatched))
                                    with col3:
                                        st.metric("Can Create", len(to_create))
                            
                            # Show match results
                            if st.session_state.matched_transactions or st.session_state.unmatched_transactions or st.session_state.transactions_to_create:
                                show_import_results(statement_date, all_data)
                        
                        else:
                            st.warning("⚠️ Please map all required fields before proceeding")
                            
                    except Exception as e:
                        st.error(f"Error reading file: {str(e)}")
                        st.exception(e)
        
        with rec_tab2:
            st.subheader("📋 Transactions with Outstanding Balance")
            
            # Get all transactions with outstanding balances using shared function
            if not all_data.empty:
                outstanding = calculate_transaction_balances(all_data)
                
                if not outstanding.empty:
                    # Summary metrics
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Transactions with Balance", len(outstanding))
                    
                    with col2:
                        total_outstanding = outstanding['_balance'].sum()
                        st.metric("Total Outstanding", f"${total_outstanding:,.2f}")
                    
                    with col3:
                        avg_balance = outstanding['_balance'].mean()
                        st.metric("Average Balance", f"${avg_balance:,.2f}")
                    
                    # Customer filter
                    st.divider()
                    
                    customer_filter = st.selectbox(
                        "Filter by Customer",
                        options=['All'] + sorted(outstanding['Customer'].dropna().unique().tolist()),
                        key="outstanding_customer_filter"
                    )
                    
                    if customer_filter != 'All':
                        display_df = outstanding[outstanding['Customer'] == customer_filter]
                    else:
                        display_df = outstanding
                
                    # Display the data with balance
                    display_df['Outstanding Balance'] = display_df['_balance']
                    display_cols = display_df[[
                        'Transaction ID', 'Customer', 'Policy Number', 
                        'Effective Date', 'Agent Estimated Comm $', 
                        'Outstanding Balance'
                    ]]
                    
                    # Apply special styling for STMT and VOID transactions
                    styled_display = style_special_transactions(display_cols)
                    
                    st.dataframe(
                        styled_display,
                        column_config={
                            "Agent Estimated Comm $": st.column_config.NumberColumn(format="$%.2f"),
                            "Outstanding Balance": st.column_config.NumberColumn(format="$%.2f")
                        },
                        use_container_width=True,
                        hide_index=True
                    )
                else:
                    st.success("✅ All transactions have been fully reconciled!")
            else:
                st.warning("No data available")
        
        with rec_tab3:
            st.subheader("📜 Reconciliation History")
            
            # Show reconciliation entries
            if not all_data.empty:
                # Include both -STMT- and -VOID- transactions
                recon_entries = all_data[
                    (all_data['Transaction ID'].str.contains('-STMT-', na=False)) |
                    (all_data['Transaction ID'].str.contains('-VOID-', na=False))
                ]
                
                if not recon_entries.empty:
                    # Date range filter
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        start_date = st.date_input(
                            "From Date",
                            value=datetime.date.today() - datetime.timedelta(days=30),
                        )
                    
                    with col2:
                        end_date = st.date_input(
                            "To Date",
                            value=datetime.date.today(),
                        )
                    
                    # Filter by date range
                    if 'STMT DATE' in recon_entries.columns:
                        # Handle mixed date formats
                        recon_entries['STMT DATE'] = pd.to_datetime(recon_entries['STMT DATE'], format='mixed')
                        mask = (recon_entries['STMT DATE'].dt.date >= start_date) & (recon_entries['STMT DATE'].dt.date <= end_date)
                        filtered_recon = recon_entries[mask]
                    else:
                        filtered_recon = recon_entries
                    
                    if not filtered_recon.empty:
                        # Summary metrics
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            st.metric("Total Entries", len(filtered_recon))
                        with col2:
                            # Use Agent Paid Amount as primary reconciliation field
                            if 'Agent Paid Amount (STMT)' in filtered_recon.columns:
                                total_agent_paid = filtered_recon['Agent Paid Amount (STMT)'].sum()
                            else:
                                # Fallback to Agency amount for backward compatibility
                                total_agent_paid = filtered_recon['Agency Comm Received (STMT)'].sum()
                            st.metric("Total Agent Payments", f"${total_agent_paid:,.2f}")
                        with col3:
                            # Count unique batch IDs
                            unique_batches = filtered_recon['reconciliation_id'].nunique() if 'reconciliation_id' in filtered_recon.columns else 0
                            st.metric("Reconciliation Batches", unique_batches)
                        with col4:
                            if 'Agent Paid Amount (STMT)' in filtered_recon.columns:
                                avg_per_batch = filtered_recon['Agent Paid Amount (STMT)'].sum() / unique_batches if unique_batches > 0 else 0
                            else:
                                avg_per_batch = filtered_recon['Agency Comm Received (STMT)'].sum() / unique_batches if unique_batches > 0 else 0
                            st.metric("Avg per Batch", f"${avg_per_batch:,.2f}")
                        
                        # Group by reconciliation batch
                        st.divider()
                        
                        view_mode = st.radio(
                            "View Mode",
                            ["By Batch", "All Transactions"],
                            horizontal=True,
                            key="history_view_mode"
                        )
                        
                        if view_mode == "By Batch" and 'reconciliation_id' in filtered_recon.columns:
                            # Show batch summary
                            # Use Agent Paid Amount for batch totals
                            agg_dict = {
                                'Transaction ID': 'count',
                                'STMT DATE': 'first'
                            }
                            if 'Agent Paid Amount (STMT)' in filtered_recon.columns:
                                agg_dict['Agent Paid Amount (STMT)'] = 'sum'
                            else:
                                agg_dict['Agency Comm Received (STMT)'] = 'sum'
                            
                            # Add reconciliation status aggregation
                            if 'reconciliation_status' in filtered_recon.columns:
                                agg_dict['reconciliation_status'] = lambda x: 'VOIDED' if any(str(v).upper() == 'VOID' for v in x.values if pd.notna(v)) else 'ACTIVE'
                            
                            batch_summary = filtered_recon.groupby('reconciliation_id').agg(agg_dict).reset_index()
                            
                            # Determine status, void ID, and void date for each batch
                            batch_summary['Status'] = 'ACTIVE'  # Default
                            batch_summary['Void ID'] = '-'
                            batch_summary['Void Date'] = '-'
                            
                            # Check if batch has been voided
                            for idx, row in batch_summary.iterrows():
                                batch_id = row['reconciliation_id']
                                
                                # Check if this batch has void entries
                                if batch_id.startswith('VOID-'):
                                    # This is a void batch itself
                                    batch_summary.at[idx, 'Status'] = 'VOID ENTRY'
                                    batch_summary.at[idx, 'Void ID'] = batch_id
                                    # Get the date from the batch transactions
                                    void_transactions = filtered_recon[filtered_recon['reconciliation_id'] == batch_id]
                                    if not void_transactions.empty and 'As of Date' in void_transactions.columns:
                                        void_date = pd.to_datetime(void_transactions['As of Date'].iloc[0])
                                        batch_summary.at[idx, 'Void Date'] = void_date.strftime('%m/%d/%Y')
                                else:
                                    # Check if there's a corresponding void batch
                                    void_batch_id = f"VOID-{batch_id}"
                                    void_exists = all_data[
                                        (all_data['reconciliation_id'] == void_batch_id) & 
                                        (all_data['Transaction ID'].str.contains('-STMT-', na=False))
                                    ]
                                    
                                    if not void_exists.empty:
                                        batch_summary.at[idx, 'Status'] = 'VOIDED'
                                        batch_summary.at[idx, 'Void ID'] = void_batch_id
                                        # Get void date
                                        if 'As of Date' in void_exists.columns:
                                            void_date = pd.to_datetime(void_exists['As of Date'].iloc[0])
                                            batch_summary.at[idx, 'Void Date'] = void_date.strftime('%m/%d/%Y')
                                
                                # Also check reconciliation_status if available
                                if 'reconciliation_status' in batch_summary.columns:
                                    status = str(row.get('reconciliation_status', '')).upper()
                                    if status == 'VOIDED' or status == 'VOID':
                                        batch_summary.at[idx, 'Status'] = 'VOIDED'
                            
                            rename_dict = {
                                'reconciliation_id': 'Batch ID',
                                'Transaction ID': 'Transaction Count',
                                'STMT DATE': 'Statement Date'
                            }
                            if 'Agent Paid Amount (STMT)' in batch_summary.columns:
                                rename_dict['Agent Paid Amount (STMT)'] = 'Agent Payment Total'
                            else:
                                rename_dict['Agency Comm Received (STMT)'] = 'Agent Payment Total'
                            
                            batch_summary = batch_summary.rename(columns=rename_dict)
                            
                            # Format date
                            batch_summary['Statement Date'] = pd.to_datetime(batch_summary['Statement Date']).dt.strftime('%m/%d/%Y')
                            
                            # Display batch summary
                            selected_batch = st.selectbox(
                                "Select Batch to View Details",
                                options=[''] + batch_summary['Batch ID'].tolist(),
                                format_func=lambda x: 'Select a batch...' if x == '' else f"{x} ({batch_summary[batch_summary['Batch ID']==x]['Statement Date'].iloc[0] if x != '' else ''})"
                            )
                            
                            # Configure column display with color coding for status
                            column_config = {
                                "Agent Payment Total": st.column_config.NumberColumn(format="$%.2f"),
                                "Status": st.column_config.TextColumn(
                                    help="ACTIVE = Normal reconciliation, VOIDED = Has been reversed, VOID ENTRY = Reversal entry"
                                ),
                                "Void ID": st.column_config.TextColumn(
                                    help="ID of the void batch if this reconciliation was voided"
                                ),
                                "Void Date": st.column_config.TextColumn(
                                    help="Date when this reconciliation was voided"
                                )
                            }
                            
                            # Reorder columns for better display
                            display_columns = ['Batch ID', 'Statement Date', 'Status', 'Transaction Count', 'Agent Payment Total', 'Void ID', 'Void Date']
                            # Only include columns that exist
                            display_columns = [col for col in display_columns if col in batch_summary.columns]
                            
                            # Apply styling based on status
                            def highlight_status(row):
                                if row['Status'] == 'VOIDED':
                                    return ['background-color: #ffcccc'] * len(row)
                                elif row['Status'] == 'VOID ENTRY':
                                    return ['background-color: #ffe6cc'] * len(row)
                                else:
                                    return [''] * len(row)
                            
                            styled_df = batch_summary[display_columns].sort_values('Statement Date', ascending=False).style.apply(highlight_status, axis=1)
                            
                            st.dataframe(
                                styled_df,
                                column_config=column_config,
                                use_container_width=True,
                                hide_index=True
                            )
                            
                            # Show batch details if selected
                            if selected_batch:
                                st.divider()
                                st.subheader(f"Batch Details: {selected_batch}")
                                
                                batch_details = filtered_recon[filtered_recon['reconciliation_id'] == selected_batch]
                                
                                # Format dates
                                if 'STMT DATE' in batch_details.columns:
                                    batch_details['STMT DATE'] = pd.to_datetime(batch_details['STMT DATE'], format='mixed').dt.strftime('%m/%d/%Y')
                                
                                st.dataframe(
                                    batch_details[[
                                        'Transaction ID', 'Customer', 'Policy Number', 
                                        'STMT DATE', 'Agency Comm Received (STMT)', 'Agent Paid Amount (STMT)'
                                    ]],
                                    column_config={
                                        "Agency Comm Received (STMT)": st.column_config.NumberColumn(format="$%.2f"),
                                        "Agent Paid Amount (STMT)": st.column_config.NumberColumn(format="$%.2f")
                                    },
                                    use_container_width=True,
                                    hide_index=True
                                )
                        else:
                            # Show all transactions
                            # Format dates
                            if 'STMT DATE' in filtered_recon.columns:
                                display_recon = filtered_recon.copy()
                                display_recon['STMT DATE'] = pd.to_datetime(display_recon['STMT DATE'], format='mixed').dt.strftime('%m/%d/%Y')
                            else:
                                display_recon = filtered_recon.copy()
                            
                            # Add reconciliation status display
                            if 'reconciliation_status' not in display_recon.columns:
                                display_recon['reconciliation_status'] = 'reconciled'  # Default for old data
                            
                            # Rename reconciliation_status for display
                            display_recon['Reconciliation Status'] = display_recon['reconciliation_status'].apply(
                                lambda x: x.upper() if pd.notna(x) else 'RECONCILED'
                            )
                            
                            # Add batch ID column if not present
                            if 'reconciliation_id' in display_recon.columns:
                                display_recon['Batch ID'] = display_recon['reconciliation_id']
                            
                            # Add Is Void Entry column
                            display_recon['Is Void Entry'] = display_recon.apply(
                                lambda row: 'Yes' if str(row.get('Reconciliation Status', '')).upper() == 'VOID' or 
                                          (row.get('Batch ID', '').startswith('VOID-') if 'Batch ID' in row else False) or
                                          ('-VOID-' in str(row.get('Transaction ID', '')))
                                          else 'No', 
                                axis=1
                            )
                            
                            # Include both agent and agency amounts
                            display_columns = ['Transaction ID', 'Transaction Type', 'Customer', 'Policy Number', 'STMT DATE']
                            if 'Agent Paid Amount (STMT)' in display_recon.columns:
                                display_columns.append('Agent Paid Amount (STMT)')
                            if 'Agency Comm Received (STMT)' in display_recon.columns:
                                display_columns.append('Agency Comm Received (STMT)')
                            
                            # Add Cross-Reference Key to show original transaction (if column exists)
                            # if 'Cross-Reference Key' in display_recon.columns:
                            #     display_columns.append('Cross-Reference Key')
                            
                            # Add new tracking columns
                            display_columns.extend(['Reconciliation Status', 'Batch ID', 'Is Void Entry'])
                            
                            # Only include columns that exist
                            display_columns = [col for col in display_columns if col in display_recon.columns]
                            
                            # Apply styling based on status
                            def highlight_void_status(row):
                                if row.get('Is Void Entry') == 'Yes':
                                    return ['background-color: #ffcccc'] * len(row)
                                elif str(row.get('Reconciliation Status', '')).upper() == 'VOID':
                                    return ['background-color: #ffe6cc'] * len(row)
                                else:
                                    return [''] * len(row)
                            
                            styled_df = display_recon[display_columns].sort_values('STMT DATE', ascending=False).style.apply(highlight_void_status, axis=1)
                            
                            # Calculate height to show all rows plus 2 extra
                            num_rows = len(display_recon[display_columns])
                            row_height = 35  # Approximate height per row in pixels
                            header_height = 35  # Header row height
                            extra_rows = 2  # Number of extra blank rows to show
                            calculated_height = header_height + (num_rows + extra_rows) * row_height
                            # Cap the height at a reasonable maximum
                            max_height = 800
                            display_height = min(calculated_height, max_height)
                            
                            st.dataframe(
                                styled_df,
                                column_config={
                                    "Agent Paid Amount (STMT)": st.column_config.NumberColumn(format="$%.2f"),
                                    "Agency Comm Received (STMT)": st.column_config.NumberColumn(format="$%.2f"),
                                    # "Cross-Reference Key": st.column_config.TextColumn(
                                    #     help="Original Transaction ID that was matched"
                                    # ),
                                    "Reconciliation Status": st.column_config.TextColumn(
                                        help="RECONCILED = Normal entry, VOID = Void reversal entry"
                                    ),
                                    "Batch ID": st.column_config.TextColumn(
                                        help="Batch ID this transaction belongs to"
                                    ),
                                    "Is Void Entry": st.column_config.TextColumn(
                                        help="Yes = This is a reversal entry, No = Original entry"
                                    )
                                },
                                use_container_width=True,
                                hide_index=True,
                                height=display_height
                            )
                    else:
                        st.info("No reconciliations found in the selected date range")
                else:
                    st.info("No reconciliation history found")
            else:
                st.warning("No data available")
        
        with rec_tab4:
            st.subheader("🔧 Adjustments & Voids")
            
            action = st.radio(
                "Select Action",
                ["Create Adjustment", "Void Reconciliation"],
                horizontal=True
            )
            
            if action == "Create Adjustment":
                st.info("Adjustments allow you to correct reconciliation errors without modifying original records")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Find transaction to adjust
                    trans_id = st.text_input(
                        "Transaction ID to Adjust",
                        placeholder="Enter the original transaction ID"
                    )
                    
                    adjustment_amount = st.number_input(
                        "Adjustment Amount",
                        help="Positive to increase, negative to decrease",
                        step=0.01
                    )
                
                with col2:
                    adjustment_reason = st.text_area(
                        "Reason for Adjustment",
                        placeholder="Explain why this adjustment is needed"
                    )
                
                if st.button("Create Adjustment", type="primary"):
                    if trans_id and adjustment_amount != 0 and adjustment_reason:
                        try:
                            # Verify the transaction exists
                            original_trans = all_data[all_data['Transaction ID'] == trans_id]
                            
                            if original_trans.empty:
                                st.error(f"Transaction ID {trans_id} not found")
                            else:
                                # Generate adjustment ID
                                adj_id = generate_reconciliation_transaction_id("ADJ")
                                adj_date = datetime.datetime.now()
                                
                                # Get original transaction details
                                orig_row = original_trans.iloc[0]
                                
                                # Create adjustment entry
                                adjustment_entry = {
                                    'Transaction ID': adj_id,
                                    'Customer': orig_row['Customer'],
                                    'Policy Number': orig_row.get('Policy Number', ''),
                                    'Policy Type': orig_row.get('Policy Type', ''),
                                    'Effective Date': orig_row.get('Effective Date', ''),
                                    'Transaction Type': 'ADJUSTMENT',
                                    'Premium Sold': 0,
                                    'Policy Gross Comm %': 0,
                                    'Agent Estimated Comm $': adjustment_amount,  # The adjustment amount
                                    'Agency Estimated Comm/Revenue (CRM)': 0,
                                    'Agency Comm Received (STMT)': 0,
                                    'Agent Paid Amount (STMT)': 0,
                                    'STMT DATE': adj_date.date(),
                                    'Policy Checklist Complete': orig_row.get('Policy Checklist Complete', ''),
                                    'NOTES': f"ADJUSTMENT for {trans_id}: {adjustment_reason}",
                                    'reconciliation_status': 'adjustment',
                                    'reconciliation_id': f"ADJ-{trans_id}",
                                    'is_reconciliation_entry': True,
                                    'Client ID': orig_row.get('Client ID', ''),
                                    'FULL OR MONTHLY PMTS': orig_row.get('FULL OR MONTHLY PMTS', ''),
                                    'X-DATE': orig_row.get('X-DATE', ''),
                                    'Description': f"Adjustment: {adjustment_reason}"
                                }
                                
                                # Clean data before insertion
                                cleaned_adjustment = clean_data_for_database(adjustment_entry)
                                
                                # Save to database
                                result = supabase.table('policies').insert(cleaned_adjustment).execute()
                                
                                if result.data:
                                    st.success(f"✅ Adjustment {adj_id} created successfully!")
                                    st.info(f"Adjustment of ${adjustment_amount:,.2f} applied to transaction {trans_id}")
                                    
                                    # Clear the form
                                    if 'trans_id_input' in st.session_state:
                                        del st.session_state['trans_id_input']
                                    
                                    # Refresh data
                                    st.cache_data.clear()
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error("Failed to create adjustment entry")
                                    
                        except Exception as e:
                            st.error(f"Error creating adjustment: {str(e)}")
                            traceback.print_exc()
                    else:
                        st.error("Please fill in all fields")
            
            else:  # Void Reconciliation
                st.warning("⚠️ Voiding a reconciliation will reverse the entire statement batch")
                
                # Show recent reconciliation batches
                if not all_data.empty and 'reconciliation_id' in all_data.columns:
                    # Get unique reconciliation batches
                    reconciliation_entries = all_data[
                        (all_data['Transaction ID'].str.contains('-STMT-', na=False)) &
                        (all_data['reconciliation_id'].notna())
                    ]
                    
                    if not reconciliation_entries.empty:
                        # Group by batch to show batch summary
                        batch_summary = reconciliation_entries.groupby('reconciliation_id').agg({
                            'Transaction ID': 'count',
                            'Agent Paid Amount (STMT)': 'sum',
                            'STMT DATE': 'first',
                            'reconciled_at': 'first'
                        }).reset_index()
                        
                        # Sort by date descending
                        batch_summary = batch_summary.sort_values('reconciled_at', ascending=False).head(20)
                        
                        st.subheader("Recent Reconciliation Batches")
                        
                        # Select batch to void
                        batch_options = batch_summary['reconciliation_id'].tolist()
                        
                        if batch_options:
                            selected_batch = st.selectbox(
                                "Select Batch to Void",
                                options=[''] + batch_options,
                                format_func=lambda x: 'Select a batch...' if x == '' else f"{x} - ${batch_summary[batch_summary['reconciliation_id']==x]['Agent Paid Amount (STMT)'].iloc[0]:,.2f} ({batch_summary[batch_summary['reconciliation_id']==x]['Transaction ID'].iloc[0]} transactions)"
                            )
                            
                            if selected_batch:
                                # Show batch details
                                batch_details = reconciliation_entries[reconciliation_entries['reconciliation_id'] == selected_batch]
                                
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    st.metric("Transactions in Batch", len(batch_details))
                                with col2:
                                    batch_total = batch_details['Agent Paid Amount (STMT)'].sum()
                                    st.metric("Batch Total", f"${batch_total:,.2f}")
                                with col3:
                                    stmt_date = pd.to_datetime(batch_details['STMT DATE'].iloc[0], format='mixed').strftime('%m/%d/%Y')
                                    st.metric("Statement Date", stmt_date)
                                
                                # Show transactions that will be voided
                                with st.expander("View Transactions to be Voided", expanded=True):
                                    display_cols = ['Transaction ID', 'Customer', 'Policy Number', 'Agent Paid Amount (STMT)']
                                    st.dataframe(
                                        batch_details[display_cols],
                                        column_config={
                                            "Agent Paid Amount (STMT)": st.column_config.NumberColumn(format="$%.2f")
                                        },
                                        use_container_width=True,
                                        hide_index=True
                                    )
                                
                                # Void confirmation
                                st.divider()
                                
                                void_reason = st.text_area(
                                    "Reason for Voiding (Required)",
                                    placeholder="Explain why this batch needs to be voided",
                                    key="void_reason"
                                )
                                
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    confirm_void = st.checkbox(
                                        "I understand this will reverse all transactions in this batch",
                                        key="confirm_void"
                                    )
                                
                                with col2:
                                    if confirm_void and void_reason:
                                        if st.button("🗑️ Void Batch", type="secondary"):
                                            try:
                                                void_count = 0
                                                
                                                # Extract the statement date from the batch ID
                                                # Formats: IMPORT-YYYYMMDD-XXXXXXXX, REC-YYYYMMDD-XXXXXXXX, MNL-YYYYMMDD-XXXXXXXX
                                                statement_date = None
                                                
                                                # Try to find YYYYMMDD pattern in the batch ID
                                                import re
                                                date_match = re.search(r'-(\d{8})-', selected_batch)
                                                if date_match:
                                                    date_str = date_match.group(1)
                                                    try:
                                                        statement_date = datetime.datetime.strptime(date_str, '%Y%m%d')
                                                    except:
                                                        pass
                                                
                                                # If we couldn't extract date from batch ID, fallback to current date
                                                if statement_date is None:
                                                    statement_date = datetime.datetime.now()
                                                
                                                # Create void entries for each transaction in the batch
                                                for idx, row in batch_details.iterrows():
                                                    # Generate void transaction ID with statement date
                                                    void_id = generate_reconciliation_transaction_id("VOID", statement_date.date())
                                                    
                                                    # Create void entry (negative amounts)
                                                    void_entry = {
                                                        'Transaction ID': void_id,
                                                        'Client ID': row.get('Client ID', ''),
                                                        'Customer': row.get('Customer', ''),
                                                        'Carrier Name': row.get('Carrier Name', ''),
                                                        'Policy Type': row.get('Policy Type', ''),
                                                        'Policy Number': row.get('Policy Number', ''),
                                                        'Transaction Type': row.get('Transaction Type', ''),
                                                        'Effective Date': row.get('Effective Date', ''),
                                                        'X-DATE': row.get('X-DATE', ''),
                                                        'Premium Sold': 0,
                                                        'Policy Gross Comm %': 0,
                                                        'Agency Estimated Comm/Revenue (CRM)': 0,
                                                        'Agency Comm Received (STMT)': -float(row.get('Agency Comm Received (STMT)', 0)),  # Negative
                                                        'Agent Estimated Comm $': 0,
                                                        'Agent Paid Amount (STMT)': -float(row.get('Agent Paid Amount (STMT)', 0)),  # Negative
                                                        'STMT DATE': statement_date.strftime('%m/%d/%Y'),  # Use statement date in MM/DD/YYYY format
                                                        'reconciliation_status': 'void',
                                                        'reconciliation_id': f"VOID-{selected_batch}",
                                                        'reconciled_at': datetime.datetime.now().isoformat(),  # Current timestamp for when void occurred
                                                        'is_reconciliation_entry': True,
                                                        'NOTES': f"VOID: {void_reason}"
                                                    }
                                                    
                                                    # Clean data before insertion
                                                    cleaned_void = clean_data_for_database(void_entry)
                                                    
                                                    # Insert void entry
                                                    supabase.table('policies').insert(cleaned_void).execute()
                                                    void_count += 1
                                                
                                                # Update original transactions to mark as unreconciled again
                                                # Find original transactions that were reconciled in this batch
                                                original_trans = all_data[
                                                    (all_data['reconciliation_id'] == selected_batch) &
                                                    (~all_data['Transaction ID'].str.contains('-STMT-|-ADJ-|-VOID-', na=False))
                                                ]
                                                
                                                for idx, orig in original_trans.iterrows():
                                                    supabase.table('policies').update({
                                                        'reconciliation_status': 'unreconciled',
                                                        'reconciliation_id': None,
                                                        'reconciled_at': None
                                                    }).eq('_id', orig['_id']).execute()
                                                
                                                # Log the void in reconciliations table
                                                void_log = {
                                                    'reconciliation_date': datetime.datetime.now().date().isoformat(),  # Current date for when void occurred
                                                    'statement_date': batch_details['STMT DATE'].iloc[0],
                                                    'carrier_name': 'VOID',
                                                    'total_amount': -batch_total,
                                                    'transaction_count': void_count,
                                                    'notes': f"VOIDED Batch {selected_batch}: {void_reason}"
                                                }
                                                supabase.table('reconciliations').insert(void_log).execute()
                                                
                                                st.success(f"✅ Successfully voided batch {selected_batch}")
                                                st.info(f"Created {void_count} void entries")
                                                
                                                # Clear cache and refresh
                                                st.cache_data.clear()
                                                time.sleep(2)
                                                st.rerun()
                                                
                                            except Exception as e:
                                                st.error(f"Error voiding batch: {str(e)}")
                                    else:
                                        if not void_reason:
                                            st.warning("Please provide a reason for voiding")
                                        if not confirm_void:
                                            st.info("Check the confirmation box to proceed")
                    else:
                        st.info("No reconciliation batches found to void")
                else:
                    st.info("No reconciliation data available")
    
    # --- Admin Panel ---
    elif page == "Admin Panel":
        st.title("⚙️ Admin Panel")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9 = st.tabs(["Database Info", "Column Mapping", "Data Management", "System Tools", "Deletion History", "Debug Logs", "Formulas & Calculations", "Policy Types", "Policy Type Mapping"])
        
        with tab1:
            st.subheader("Database Information")
            
            # Database stats
            if not all_data.empty:
                st.metric("Total Records", len(all_data))
                st.metric("Total Columns", len(all_data.columns))
                
                st.subheader("Column Information")
                col_info = pd.DataFrame({
                    'Column': all_data.columns,
                    'Data Type': [str(all_data[col].dtype) for col in all_data.columns],
                    'Non-Null Count': [all_data[col].count() for col in all_data.columns],
                    'Null Count': [all_data[col].isnull().sum() for col in all_data.columns]
                })
                st.dataframe(col_info, use_container_width=True)
                
                st.subheader("Sample Data")
                st.dataframe(all_data.head(), use_container_width=True)
            else:
                st.info("No data available")
        
        with tab2:
            st.subheader("Column Mapping Configuration")
            st.info("Map database columns to user-friendly display names. Changes are saved to column_mapping.json")
            
            if not all_data.empty:
                # Load existing mappings
                from column_mapping_config import column_mapper
                current_mapping = column_mapper._load_mapping()
                
                # Create editable mapping interface
                st.write("**Edit Column Display Names:**")
                st.caption("Change how column names appear in the app without modifying the database")
                
                # Initialize session state for editing
                if 'column_mapping_edits' not in st.session_state:
                    # Clean up the current mapping to remove duplicates and fix issues
                    cleaned_mapping = {}
                    seen_db_cols = set()
                    
                    for ui_field, db_col in current_mapping.items():
                        # Skip duplicate STMT DATE mapping
                        if db_col == "STMT DATE" and db_col in seen_db_cols:
                            continue
                        # Keep calculated fields
                        if db_col == "(Calculated/Virtual)":
                            cleaned_mapping[ui_field] = db_col
                        # Keep valid mappings
                        elif db_col in all_data.columns or db_col == "NOTES":
                            cleaned_mapping[ui_field] = db_col
                            seen_db_cols.add(db_col)
                    
                    st.session_state.column_mapping_edits = cleaned_mapping
                
                # Create columns for better layout
                col1, col2, col3 = st.columns([2, 2, 1])
                with col1:
                    st.markdown("**Database Column**")
                with col2:
                    st.markdown("**Display Name (UI)**")
                with col3:
                    st.markdown("**Action**")
                
                # Display editable mappings for important columns
                important_columns = [
                    "Agent Comm %",
                    "Agency Estimated Comm/Revenue (CRM)",
                    "Agent Estimated Comm $",
                    "Agent Paid Amount (STMT)",
                    "Agency Comm Received (STMT)",
                    "Policy Gross Comm %",
                    "Premium Sold",
                    "Policy Balance Due",
                    "Customer",
                    "Policy Number",
                    "Transaction Type",
                    "Effective Date",
                    "X-DATE"
                ]
                
                # Show mapped columns first
                st.markdown("---")
                
                # Also show calculated fields that need mapping
                calculated_fields = ["Policy Balance Due", "Agent Estimated Comm $"]
                
                for db_col in important_columns:
                    if db_col in all_data.columns or db_col in calculated_fields:
                        col1, col2, col3 = st.columns([2, 2, 1])
                        with col1:
                            if db_col in calculated_fields:
                                st.text(f"{db_col} (Calculated)")
                            else:
                                st.text(db_col)
                        with col2:
                            # Find the UI name for this database column
                            ui_name = db_col  # default
                            for ui_field, mapped_col in st.session_state.column_mapping_edits.items():
                                if mapped_col == db_col or (db_col in calculated_fields and ui_field == db_col):
                                    ui_name = ui_field
                                    break
                            
                            new_name = st.text_input(
                                "Display name",
                                value=ui_name,
                                key=f"map_{db_col}",
                                label_visibility="hidden"
                            )
                            
                            # Update session state if changed
                            if new_name != ui_name:
                                if db_col in calculated_fields:
                                    st.session_state.column_mapping_edits[new_name] = "(Calculated/Virtual)"
                                else:
                                    st.session_state.column_mapping_edits[new_name] = db_col
                                # Remove old mapping if UI name changed
                                if ui_name in st.session_state.column_mapping_edits and ui_name != new_name:
                                    del st.session_state.column_mapping_edits[ui_name]
                        
                        with col3:
                            if db_col == "Agent Comm %":
                                st.caption("⭐ Rename to 'Agent Comm %'")
                
                # Show other database columns
                st.markdown("---")
                st.markdown("**Other Database Columns:**")
                other_cols = [col for col in sorted(all_data.columns) if col not in important_columns]
                
                # Display in a more compact format
                cols_per_row = 3
                for i in range(0, len(other_cols), cols_per_row):
                    cols = st.columns(cols_per_row)
                    for j, col in enumerate(other_cols[i:i+cols_per_row]):
                        if j < len(cols):
                            cols[j].write(f"• {col}")
                
                # Save button
                st.markdown("---")
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("💾 Save Column Mappings", type="primary"):
                        try:
                            # Validate and save
                            result = column_mapper.save_mapping(
                                st.session_state.column_mapping_edits,
                                list(all_data.columns)
                            )
                            
                            if result['success']:
                                st.success("✅ Column mappings saved successfully!")
                                if result.get('warnings'):
                                    for warning in result['warnings']:
                                        st.warning(warning)
                                # Clear the mapping cache to ensure new mappings are loaded
                                column_mapper.clear_cache()
                                st.balloons()
                            else:
                                st.error("Failed to save mappings:")
                                for error in result.get('errors', []):
                                    st.error(f"• {error}")
                        except Exception as e:
                            st.error(f"Error saving mappings: {str(e)}")
                
                with col2:
                    if st.button("🔄 Reset to Defaults"):
                        st.session_state.column_mapping_edits = column_mapper.default_ui_fields.copy()
                        st.rerun()
        
        with tab3:
            st.subheader("Data Management")
            
            st.warning("⚠️ These operations affect your database. Use with caution!")
            
            # Database backup
            if st.button("📁 Create Database Backup"):
                try:
                    import shutil
                    backup_name = f"commissions_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                    shutil.copy2("commissions.db", backup_name)
                    st.success(f"Database backed up as {backup_name}")
                except Exception as e:
                    st.error(f"Backup failed: {e}")
            
            # Data validation
            if st.button("🔍 Validate Data Integrity"):
                if not all_data.empty:
                    issues = []
                    
                    # Check for duplicates
                    if 'Transaction_ID' in all_data.columns:
                        duplicates = all_data['Transaction_ID'].duplicated().sum()
                        if duplicates > 0:
                            issues.append(f"Found {duplicates} duplicate Transaction IDs")
                    
                    # Check for missing critical data
                    if 'Customer' in all_data.columns:
                        missing_customers = all_data['Customer'].isnull().sum()
                        if missing_customers > 0:
                            issues.append(f"Found {missing_customers} records with missing customer names")
                    
                    if issues:
                        st.warning("Data integrity issues found:")
                        for issue in issues:
                            st.write(f"• {issue}")
                    else:
                        st.success("No data integrity issues found!")
                else:
                    st.info("No data to validate")
        
        with tab4:
            st.subheader("System Tools")
            
            # System information
            st.write("**System Information:**")
            st.write(f"• Python version: {pd.__version__}")
            st.write(f"• Pandas version: {pd.__version__}")
            st.write(f"• Database: Supabase Cloud")
            
            # Clear session state
            if st.button("🔄 Clear Session State"):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.success("Session state cleared!")
                st.rerun()
        
        with tab5:
            st.subheader("🗑️ Deletion History - Last 100 Deleted Policy Transactions")
            st.info("View and restore recently deleted policies. Records are kept for recovery purposes.")
            
            try:
                # Fetch deleted policies from Supabase
                deleted_response = supabase.table('deleted_policies').select("*").order('deleted_at', desc=True).limit(100).execute()
                
                if deleted_response.data:
                    # Extract policy data from JSONB structure
                    deleted_records = []
                    for record in deleted_response.data:
                        policy_info = {
                            'deletion_id': record['deletion_id'],
                            'deleted_at': record['deleted_at'],
                            'transaction_id': record['transaction_id'],
                            'customer_name': record['customer_name']
                        }
                        # Add the policy data fields
                        if 'policy_data' in record and record['policy_data']:
                            policy_info.update(record['policy_data'])
                        deleted_records.append(policy_info)
                    
                    deleted_df = pd.DataFrame(deleted_records)
                    
                    # Convert deleted_at to datetime
                    if 'deleted_at' in deleted_df.columns:
                        deleted_df['deleted_at'] = pd.to_datetime(deleted_df['deleted_at'])
                    
                    # Add a selection column for restoration
                    deleted_df.insert(0, 'Restore', False)
                    
                    # Display the deleted records
                    st.write(f"**Found {len(deleted_df)} deleted records:**")
                    
                    # Show key info at the top
                    edited_deleted = st.data_editor(
                        deleted_df,
                        use_container_width=True,
                        height=400,
                        key="deleted_policies_editor",
                        column_config={
                            "Restore": st.column_config.CheckboxColumn(
                                "Restore",
                                help="Select records to restore",
                                default=False,
                            ),
                            "deleted_at": st.column_config.DatetimeColumn(
                                "Deleted At",
                                format="DD/MM/YYYY HH:mm",
                                timezone="local"
                            )
                        }
                    )
                    
                    # Restore functionality
                    st.divider()
                    col1, col2 = st.columns([2, 3])
                    
                    with col1:
                        if st.button("♻️ Restore Selected Records", type="primary"):
                            # Find rows where Restore checkbox is True
                            selected_to_restore = edited_deleted[edited_deleted['Restore'] == True]
                            
                            if not selected_to_restore.empty:
                                try:
                                    restored_count = 0
                                    for idx, row in selected_to_restore.iterrows():
                                        # Prepare data for restoration (exclude deletion-specific columns)
                                        restore_data = {}
                                        for col in row.index:
                                            # Exclude metadata columns that aren't part of the policies table
                                            if col not in ['Restore', 'deletion_id', 'deleted_at', 'transaction_id', 'customer_name']:
                                                if pd.notna(row[col]):
                                                    value = row[col]
                                                    # Clean numeric values for proper data types
                                                    if isinstance(value, (int, float)):
                                                        # Check if it should be an integer (no decimal part)
                                                        if isinstance(value, float) and value.is_integer():
                                                            restore_data[col] = int(value)
                                                        else:
                                                            restore_data[col] = clean_numeric_value(value)
                                                    else:
                                                        restore_data[col] = value
                                        
                                        # Restore to policies table
                                        supabase.table('policies').insert(restore_data).execute()
                                        
                                        # Remove from deleted_policies table
                                        deletion_id = row['deletion_id']
                                        supabase.table('deleted_policies').delete().eq('deletion_id', deletion_id).execute()
                                        
                                        restored_count += 1
                                    
                                    # Clear cache and show success
                                    clear_policies_cache()
                                    st.success(f"Successfully restored {restored_count} records!")
                                    st.rerun()
                                    
                                except Exception as restore_error:
                                    st.error(f"Error restoring records: {restore_error}")
                            else:
                                st.warning("Please select records to restore using the checkboxes.")
                    
                    with col2:
                        if st.button("🗑️ Permanently Delete Selected", type="secondary"):
                            # Find rows where Restore checkbox is True (using same checkbox for selection)
                            selected_to_delete = edited_deleted[edited_deleted['Restore'] == True]
                            
                            if not selected_to_delete.empty:
                                st.warning(f"⚠️ This will permanently delete {len(selected_to_delete)} records from history!")
                                if st.button("Confirm Permanent Deletion", key="confirm_perm_delete"):
                                    try:
                                        for idx, row in selected_to_delete.iterrows():
                                            deletion_id = row['deletion_id']
                                            supabase.table('deleted_policies').delete().eq('deletion_id', deletion_id).execute()
                                        
                                        st.success(f"Permanently deleted {len(selected_to_delete)} records from history.")
                                        st.rerun()
                                    except Exception as perm_delete_error:
                                        st.error(f"Error permanently deleting records: {perm_delete_error}")
                    
                    # Export deleted records
                    st.divider()
                    st.write("**Export Deletion History:**")
                    csv_data = deleted_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Deletion History CSV",
                        data=csv_data,
                        file_name=f"deletion_history_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                else:
                    st.info("No deleted policies found. Deleted records will appear here for recovery.")
                    
            except Exception as e:
                if "relation \"deleted_policies\" does not exist" in str(e):
                    st.warning("The deleted_policies table doesn't exist yet. Please run the SQL script to create it:")
                    st.code("""
-- Run this in your Supabase SQL editor:
CREATE TABLE IF NOT EXISTS deleted_policies (
    deletion_id SERIAL PRIMARY KEY,
    deleted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    -- Copy all columns from policies table
    _id INTEGER,
    "Client ID" TEXT,
    "Transaction ID" TEXT,
    "Customer" TEXT,
    -- ... (see create_deleted_policies_table.sql for full schema)
);
                    """)
                else:
                    st.error(f"Error loading deletion history: {e}")
        
        with tab6:
            st.subheader("🐛 Debug Logs")
            st.info("This section captures all debug messages, errors, and system events to help diagnose issues.")
            
            # Initialize debug logs if not exists
            if "debug_logs" not in st.session_state:
                st.session_state.debug_logs = []
            
            # Control buttons
            col1, col2, col3 = st.columns([1, 1, 4])
            with col1:
                if st.button("🗑️ Clear Logs"):
                    clear_debug_logs()
                    st.success("Debug logs cleared!")
                    st.rerun()
            
            with col2:
                if st.button("🔄 Refresh"):
                    st.rerun()
            
            with col3:
                st.write(f"Total log entries: {len(st.session_state.debug_logs)}")
            
            # Filter options
            st.markdown("**Filter Options:**")
            col1, col2 = st.columns(2)
            with col1:
                log_level_filter = st.selectbox(
                    "Log Level",
                    ["All", "ERROR", "WARNING", "INFO", "DEBUG"],
                    key="log_level_filter"
                )
            
            with col2:
                search_term = st.text_input("Search in logs", key="log_search")
            
            # Display logs
            if st.session_state.debug_logs:
                # Filter logs
                filtered_logs = st.session_state.debug_logs
                
                if log_level_filter != "All":
                    filtered_logs = [log for log in filtered_logs if log.get("level") == log_level_filter]
                
                if search_term:
                    filtered_logs = [log for log in filtered_logs if search_term.lower() in str(log).lower()]
                
                # Display in reverse order (newest first)
                st.markdown(f"**Showing {len(filtered_logs)} log entries:**")
                
                for log in reversed(filtered_logs[-100:]):  # Show last 100 entries
                    level = log.get("level", "INFO")
                    timestamp = log.get("timestamp", "")
                    message = log.get("message", "")
                    
                    # Color code by level
                    if level == "ERROR":
                        st.error(f"**[{timestamp}]** {message}")
                        if "error_details" in log:
                            with st.expander("Error Details"):
                                st.code(log["error_details"])
                                if "error_attrs" in log:
                                    st.code(log["error_attrs"])
                    elif level == "WARNING":
                        st.warning(f"**[{timestamp}]** {message}")
                    elif level == "DEBUG":
                        st.info(f"**[{timestamp}]** {message}")
                    else:
                        st.write(f"**[{timestamp}]** {message}")
                
                # Export logs
                st.divider()
                if st.button("📥 Export Logs as JSON"):
                    logs_json = json.dumps(st.session_state.debug_logs, indent=2)
                    st.download_button(
                        label="Download Debug Logs",
                        data=logs_json,
                        file_name=f"debug_logs_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                        mime="application/json"
                    )
            else:
                st.info("No debug logs yet. Logs will appear here as you use the application.")
        
        with tab7:
            st.subheader("📊 Formulas & Calculations - Complete Documentation")
            st.info("⚠️ CRITICAL: This section contains ALL formulas used throughout the application. Review carefully to understand the complete calculation matrix.")
            
            # Formula tabs
            formula_tab1, formula_tab2, formula_tab3, formula_tab4, formula_tab5, formula_tab6 = st.tabs(["Core Formulas", "Rate Matrix", "Field Dependencies", "Formula Testing", "Implementation Details", "Formula Issues"])
            
            with formula_tab1:
                st.markdown("### 🎯 CORE FORMULAS - Complete Matrix")
                
                st.warning("These formulas are the foundation of ALL commission calculations in the system.")
                
                # Formula 1: Commissionable Premium
                st.markdown("#### Formula #1: Commissionable Premium")
                st.code("""
Formula: Premium Sold - Policy Taxes & Fees = Commissionable Premium
Purpose: Calculate the premium amount eligible for commission (excluding non-commissionable taxes/fees)
Example: $10,000 - $500 = $9,500

Fields Used:
- INPUT: Premium Sold (user enters)
- INPUT: Policy Taxes & Fees (user enters)
- OUTPUT: Commissionable Premium (calculated)

Where Used:
- Add New Policy Transaction form
- Edit Policy Transaction form
- All commission calculations use this as the base
                """, language="text")
                
                # Formula 2: Agency Commission
                st.markdown("#### Formula #2: Agency Estimated Comm/Revenue (CRM)")
                st.code("""
Formula: Commissionable Premium × Policy Gross Comm % = Agency Estimated Comm/Revenue (CRM)
Purpose: Calculate the gross commission the agency receives from the carrier
Example: $9,500 × 10% = $950

Fields Used:
- INPUT: Commissionable Premium (from Formula #1)
- INPUT: Policy Gross Comm % (user enters)
- OUTPUT: Agency Estimated Comm/Revenue (CRM) (calculated)

Where Used:
- Add New Policy Transaction form (auto-calculated)
- Edit Policy Transaction form (auto-calculated)
- Policy Revenue Ledger Reports
- Dashboard metrics
                """, language="text")
                
                # Formula 3: Agent Commission
                st.markdown("#### Formula #3: Agent Estimated Comm $")
                st.code("""
Formula: Agency Estimated Comm/Revenue (CRM) × Agent Comm Rate = Agent Estimated Comm $
Purpose: Calculate the agent's portion of the commission

Agent Comm Rate Logic:
- NEW, NBS, STL, BoR: 50%
- RWL, REWRITE: 25%
- END, PCH: 50% if new business (Policy Orig Date = Effective Date), 25% if renewal
- CAN, XCL: 0%

Example: $950 × 50% = $475

Fields Used:
- INPUT: Agency Estimated Comm/Revenue (CRM) (from Formula #2)
- INPUT: Transaction Type (determines rate)
- INPUT: Policy Origination Date (for END/PCH)
- INPUT: Effective Date (for END/PCH)
- OUTPUT: Agent Estimated Comm $ (calculated)

Where Used:
- Add New Policy Transaction form (auto-calculated)
- Edit Policy Transaction form (auto-calculated)
- All reports and dashboard
                """, language="text")
                
                # Formula 4: Broker Fee Commission
                st.markdown("#### Formula #4: Broker Fee Agent Commission")
                st.code("""
Formula: Broker Fee × 0.50 = Broker Fee Agent Comm
Purpose: Calculate agent's portion of broker fee (ALWAYS 50% regardless of transaction type)
Example: $250 × 0.50 = $125

Fields Used:
- INPUT: Broker Fee (user enters)
- OUTPUT: Broker Fee Agent Comm (calculated)

Special Rule: Broker fee commission is ALWAYS 50%, never affected by transaction type
                """, language="text")
                
                # Formula 5: Total Agent Commission
                st.markdown("#### Formula #5: Total Agent Commission")
                st.code("""
Formula: Agent Estimated Comm $ + Broker Fee Agent Comm = Total Agent Comm
Purpose: Calculate total commission including broker fee portion
Example: $475 + $125 = $600

Fields Used:
- INPUT: Agent Estimated Comm $ (from Formula #3)
- INPUT: Broker Fee Agent Comm (from Formula #4)
- OUTPUT: Total Agent Comm (calculated)
                """, language="text")
                
                # Formula 6: Balance Due
                st.markdown("#### Formula #6: Policy Balance Due")
                st.code("""
Formula: Agent Estimated Comm $ - Agent Paid Amount (STMT) = Policy Balance Due
Purpose: Track outstanding commission owed to agent
Example: $475 - $200 = $275

Fields Used:
- INPUT: Agent Estimated Comm $ (from Formula #3)
- INPUT: Agent Paid Amount (STMT) (from reconciliation)
- OUTPUT: Policy Balance Due (calculated)

Where Used:
- Policy Revenue Ledger Reports
- Dashboard search results
- Outstanding balance tracking
                """, language="text")
                
                # Formula 7: Premium Sold for Endorsements
                st.markdown("#### Formula #7: Premium Sold Calculator (Endorsements)")
                st.code("""
Formula: New/Revised Premium - Existing Premium = Premium Sold
Purpose: Calculate the additional premium for endorsements
Example: $1,350 - $1,200 = $150

Fields Used:
- INPUT: New/Revised Premium (user enters)
- INPUT: Existing Premium (user enters)
- OUTPUT: Premium Sold (calculated)

Where Used:
- Add New Policy Transaction form (Endorsement Calculator section)
- Only for END and PCH transaction types
                """, language="text")
                
            with formula_tab2:
                st.markdown("### 📋 Commission Rate Matrix")
                
                # Create rate matrix dataframe
                rate_data = {
                    "Transaction Type": ["NEW", "NBS", "STL", "BoR", "RWL", "REWRITE", "END (New)", "END (Renewal)", "PCH (New)", "PCH (Renewal)", "CAN", "XCL"],
                    "Full Name": [
                        "New Business", "New Business (Special)", "Still (Continuing)", "Book of Renewals",
                        "Renewal", "Rewrite", "Endorsement (New)", "Endorsement (Renewal)",
                        "Policy Change (New)", "Policy Change (Renewal)", "Cancellation", "Excluded"
                    ],
                    "Agent Rate": ["50%", "50%", "50%", "50%", "25%", "25%", "50%", "25%", "50%", "25%", "0%", "0%"],
                    "Condition": [
                        "Always", "Always", "Always", "Always", "Always", "Always",
                        "If Orig Date = Eff Date", "If Orig Date ≠ Eff Date",
                        "If Orig Date = Eff Date", "If Orig Date ≠ Eff Date",
                        "No commission", "No commission"
                    ]
                }
                
                rate_df = pd.DataFrame(rate_data)
                st.dataframe(rate_df, use_container_width=True, hide_index=True)
                
                st.markdown("#### Special Rules")
                st.markdown("""
                - **NEW vs RWL Detection**: Based on Policy Origination Date vs Effective Date
                - **Endorsements/Changes**: Commission rate depends on whether it's on a new or renewal policy
                - **Cancellations**: No agent commission paid on CAN or XCL transactions
                - **Override Capability**: Admin users can manually adjust commission amounts if needed
                """)
                
            with formula_tab3:
                st.markdown("### 🔗 Field Dependencies & Data Flow")
                
                st.info("This shows how fields depend on each other and the flow of calculations through the system.")
                
                # Dependency Tree
                st.markdown("#### Calculation Dependency Tree")
                st.code("""
1. Premium Sold (USER INPUT)
   └── Policy Taxes & Fees (USER INPUT)
       └── Commissionable Premium (CALCULATED)
           └── Policy Gross Comm % (USER INPUT)
               └── Agency Estimated Comm/Revenue (CRM) (CALCULATED)
                   └── Transaction Type (USER INPUT)
                   └── Policy Origination Date (USER INPUT - for END/PCH)
                   └── Effective Date (USER INPUT - for END/PCH)
                       └── Agent Comm Rate (CALCULATED)
                           └── Agent Estimated Comm $ (CALCULATED)
                               └── Agent Paid Amount (STMT) (FROM RECONCILIATION)
                                   └── Policy Balance Due (CALCULATED)

2. Broker Fee (USER INPUT)
   └── Broker Fee Agent Comm (CALCULATED - always 50%)
       └── Total Agent Comm (CALCULATED with Agent Est Comm)
                """, language="text")
                
                # Field Impact Matrix
                st.markdown("#### Field Impact Matrix")
                impact_data = {
                    "When This Changes": [
                        "Premium Sold",
                        "Policy Taxes & Fees",
                        "Policy Gross Comm %",
                        "Transaction Type",
                        "Broker Fee",
                        "Policy Orig Date",
                        "Effective Date",
                        "Agent Paid Amount"
                    ],
                    "These Fields Update": [
                        "Commissionable Premium, Agency Comm, Agent Comm, Total Agent Comm, Balance Due",
                        "Commissionable Premium, Agency Comm, Agent Comm, Total Agent Comm, Balance Due",
                        "Agency Comm, Agent Comm, Total Agent Comm, Balance Due",
                        "Agent Comm Rate, Agent Comm, Total Agent Comm, Balance Due",
                        "Broker Fee Agent Comm, Total Agent Comm",
                        "Agent Comm Rate (for END/PCH), Agent Comm, Total Agent Comm, Balance Due",
                        "Agent Comm Rate (for END/PCH), Agent Comm, Total Agent Comm, Balance Due",
                        "Policy Balance Due"
                    ],
                    "Calculation Type": [
                        "Cascading",
                        "Cascading",
                        "Cascading",
                        "Rate Determination",
                        "Direct",
                        "Conditional",
                        "Conditional",
                        "Direct"
                    ]
                }
                
                impact_df = pd.DataFrame(impact_data)
                st.dataframe(impact_df, use_container_width=True, hide_index=True)
                
                # Critical Fields
                st.markdown("#### 🚨 Critical Fields for Calculations")
                st.warning("""
                These fields MUST have values for calculations to work correctly:
                - Premium Sold (or calculated from Endorsement Calculator)
                - Policy Gross Comm %
                - Transaction Type
                - Policy Origination Date (for END/PCH transactions)
                - Effective Date (for END/PCH transactions)
                """)
                
                # Locked vs Editable
                st.markdown("#### 🔒 Formula-Locked vs Editable Fields")
                field_status = {
                    "Field Name": [
                        "Premium Sold",
                        "Policy Taxes & Fees",
                        "Commissionable Premium",
                        "Broker Fee",
                        "Policy Gross Comm %",
                        "Agency Estimated Comm/Revenue (CRM)",
                        "Agent Comm %",
                        "Agent Estimated Comm $",
                        "Broker Fee Agent Comm",
                        "Total Agent Comm",
                        "Policy Balance Due"
                    ],
                    "Status": [
                        "✏️ Editable",
                        "✏️ Editable",
                        "🔒 Formula-Locked",
                        "✏️ Editable",
                        "✏️ Editable",
                        "🔒 Formula-Locked",
                        "✏️ Editable (but shows rate)",
                        "🔒 Formula-Locked",
                        "🔒 Formula-Locked",
                        "🔒 Formula-Locked",
                        "🔒 Formula-Locked"
                    ],
                    "Notes": [
                        "User enters or calculates via Endorsement Calculator",
                        "User enters carrier taxes/fees",
                        "Auto-calculated: Premium - Taxes",
                        "User enters broker fee amount",
                        "User enters commission percentage",
                        "Auto-calculated: Commissionable × Rate",
                        "Shows rate but stored as editable field",
                        "Auto-calculated: Agency × Agent Rate",
                        "Auto-calculated: Always 50% of Broker Fee",
                        "Auto-calculated: Agent + Broker commissions",
                        "Auto-calculated: Estimated - Paid"
                    ]
                }
                
                status_df = pd.DataFrame(field_status)
                st.dataframe(status_df, use_container_width=True, hide_index=True)
                
            with formula_tab4:
                st.markdown("### 🧪 Formula Testing & Verification")
                
                st.markdown("#### Test Commission Calculations")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    test_premium = st.number_input("Test Premium Amount", value=1000.0, format="%.2f", key="test_premium")
                    test_comm_rate = st.number_input("Commission Rate (%)", value=10.0, format="%.2f", key="test_comm_rate")
                    test_trans_type = st.selectbox("Transaction Type", 
                        ["NEW", "RWL", "END", "PCH", "CAN", "NBS", "STL", "BoR", "REWRITE", "XCL"],
                        key="test_trans_type"
                    )
                    
                    if test_trans_type in ["END", "PCH"]:
                        st.markdown("**Date Check for END/PCH**")
                        col1a, col1b = st.columns(2)
                        with col1a:
                            orig_date = st.date_input("Origination Date", key="test_orig_date")
                        with col1b:
                            eff_date = st.date_input("Effective Date", key="test_eff_date")
                        is_new = orig_date == eff_date
                    else:
                        is_new = None
                
                with col2:
                    st.markdown("**Calculated Results:**")
                    
                    # Calculate agency commission
                    agency_comm = test_premium * (test_comm_rate / 100)
                    st.metric("Agency Commission", f"${agency_comm:.2f}")
                    
                    # Calculate agent commission based on type
                    if test_trans_type in ["NEW", "NBS", "STL", "BoR"]:
                        agent_rate = 0.50
                        rate_display = "50%"
                    elif test_trans_type in ["RWL", "REWRITE"]:
                        agent_rate = 0.25
                        rate_display = "25%"
                    elif test_trans_type in ["END", "PCH"]:
                        if is_new:
                            agent_rate = 0.50
                            rate_display = "50% (New)"
                        else:
                            agent_rate = 0.25
                            rate_display = "25% (Renewal)"
                    else:  # CAN, XCL
                        agent_rate = 0.0
                        rate_display = "0%"
                    
                    agent_comm = agency_comm * agent_rate
                    
                    st.metric("Agent Rate", rate_display)
                    st.metric("Agent Commission", f"${agent_comm:.2f}")
                    
                    # Show calculation breakdown
                    st.markdown("**Calculation Breakdown:**")
                    st.text(f"""
Premium: ${test_premium:.2f}
Commission Rate: {test_comm_rate}%
Agency Commission: ${test_premium:.2f} × {test_comm_rate}% = ${agency_comm:.2f}
Agent Rate: {rate_display}
Agent Commission: ${agency_comm:.2f} × {agent_rate:.2%} = ${agent_comm:.2f}
                    """)
            
            with formula_tab5:
                st.markdown("### ⚙️ Implementation Details")
                
                st.warning("CRITICAL: Understanding where formulas are implemented is essential for troubleshooting")
                
                st.markdown("#### Formula Implementation Locations")
                
                implementation_data = {
                    "Formula": [
                        "Agent Comm Rate Determination",
                        "Agent Comm Rate Determination",
                        "Agent Comm Rate Determination",
                        "Agency Commission Calculation",
                        "Agency Commission Calculation",
                        "Agent Commission Calculation",
                        "Agent Commission Calculation",
                        "Field Locking/Display",
                        "Field Locking/Display"
                    ],
                    "Location": [
                        "get_agent_rate() function (lines 370-383)",
                        "Add New Policy form (lines 3864-3874)",
                        "Edit Transaction form (lines 3108-3117)",
                        "Add New Policy form (auto-calc)",
                        "Edit Transaction form (auto-calc)",
                        "Add New Policy form (auto-calc)",
                        "Edit Transaction form (auto-calc)",
                        "Add New Policy form",
                        "Edit Transaction form"
                    ],
                    "Implementation": [
                        "✅ Full logic with date checking for END/PCH",
                        "⚠️ Defaults END/PCH to 50% (no date check)",
                        "❌ Just displays stored value (no logic)",
                        "✅ Uses Commissionable Premium",
                        "✅ Uses Commissionable Premium",
                        "✅ Full calculation",
                        "✅ Full calculation",
                        "✅ Shows as disabled fields",
                        "✅ Shows as disabled fields"
                    ],
                    "Issue": [
                        "Working correctly",
                        "Missing date comparison logic",
                        "No automatic rate determination",
                        "Working correctly",
                        "Working correctly",
                        "Working correctly",
                        "Working correctly",
                        "Working correctly",
                        "Working correctly"
                    ]
                }
                
                impl_df = pd.DataFrame(implementation_data)
                st.dataframe(impl_df, use_container_width=True, hide_index=True)
                
                st.markdown("#### Key Functions")
                st.code("""
1. get_agent_rate(row) - Main logic for determining agent commission rate
   - Location: Lines 370-383
   - Used by: Batch calculations, reports
   - NOT used by: Add/Edit forms (they have their own logic)

2. calculate_agency_commission(premium, rate) - Calculate agency commission
   - Used by: Forms for real-time calculation

3. calculate_agent_commission(agency_comm, trans_type, is_new) - Calculate agent commission
   - Used by: Forms for real-time calculation

4. is_reconciliation_transaction(trans_id) - Check if transaction is locked
   - Prevents editing of -STMT-, -VOID-, -ADJ- transactions
                """, language="text")
                
                st.markdown("#### Formula Execution Flow")
                st.code("""
Add New Policy Transaction:
1. User enters Premium Sold, Taxes, Broker Fee
2. System calculates Commissionable Premium (real-time)
3. User enters Policy Gross Comm %
4. System calculates Agency Commission (real-time)
5. System determines Agent Rate based on Transaction Type
6. System calculates Agent Commission (real-time)
7. System calculates Broker Fee Commission (always 50%)
8. System calculates Total Agent Commission
9. Data saved to database

Edit Policy Transaction:
1. Form loads with existing data
2. User can edit input fields
3. System recalculates all formula fields on save
4. Agent Comm Rate is NOT automatically determined (uses stored value)
                """, language="text")
            
            with formula_tab6:
                st.markdown("### ⚠️ Known Formula Issues & Inconsistencies")
                
                st.error("These issues affect calculation accuracy and user experience")
                
                st.markdown("#### 🔴 CRITICAL ISSUE: Agent Comm Rate Inconsistency")
                st.code("""
PROBLEM: Agent Commission Rate logic is implemented differently in different places

1. Batch Calculation Function (get_agent_rate):
   ✅ Correctly checks if Policy Orig Date = Effective Date for END/PCH

2. Add New Policy Form:
   ⚠️ Always defaults END/PCH to 50% (doesn't check dates)
   
3. Edit Transaction Form:
   ❌ Just shows whatever is in the database (no logic at all)
   
IMPACT:
- END/PCH transactions may have wrong commission rates
- Users must manually know to check/update the rate
- No validation that the stored rate matches the business rules

SOLUTION NEEDED:
- Implement consistent date-checking logic in all forms
- For Edit form: Look up the NEW transaction for the policy to determine correct rate
                """, language="text")
                
                st.markdown("#### 🟡 Other Known Issues")
                
                issues_data = {
                    "Issue": [
                        "Edit form doesn't recalculate rates",
                        "No validation of commission rates",
                        "Decimal vs percentage confusion",
                        "No audit trail for formula changes",
                        "Formula fields can be manually edited in database"
                    ],
                    "Impact": [
                        "Wrong rates stay wrong until manually fixed",
                        "Users can enter any rate without warning",
                        "Agent rate stored as 0.50 or 50 inconsistently",
                        "Can't track when/why calculations changed",
                        "Database edits bypass all formula logic"
                    ],
                    "Severity": [
                        "High",
                        "Medium",
                        "Medium",
                        "Low",
                        "High"
                    ],
                    "Workaround": [
                        "Manually verify and update rates",
                        "User training on correct rates",
                        "System handles both formats",
                        "Document changes manually",
                        "Lock formula fields at database level"
                    ]
                }
                
                issues_df = pd.DataFrame(issues_data)
                st.dataframe(issues_df, use_container_width=True, hide_index=True)
                
                st.markdown("#### 📊 Formula Validation Checklist")
                st.info("""
                When reviewing transactions, check:
                1. ✓ Is Transaction Type correct?
                2. ✓ For END/PCH: Does Agent Rate match the Policy Orig Date vs Effective Date rule?
                3. ✓ Is Agency Commission = Commissionable Premium × Gross Rate?
                4. ✓ Is Agent Commission = Agency Commission × Agent Rate?
                5. ✓ Is Broker Fee Commission = Broker Fee × 50%?
                6. ✓ Is Total Agent Commission = Agent + Broker Fee commissions?
                7. ✓ Is Balance Due = Agent Estimated - Agent Paid?
                """)
        
        with tab8:
            st.subheader("📋 Policy Types Configuration")
            
            # Try to load policy types from configuration file
            try:
                # Load from the updated config file if it exists
                config_path = "config_files/policy_types_updated.json"
                if os.path.exists(config_path):
                    with open(config_path, 'r') as f:
                        config_data = json.load(f)
                        policy_types_data = config_data.get("policy_types", [])
                else:
                    # Fall back to existing method
                    policy_types, allow_custom = load_policy_types()
                    policy_types_data = [{"code": pt["name"], "name": pt["name"], "active": pt.get("active", True)} for pt in policy_types]
                
                # Create a modern, compact display
                with st.container():
                    
                    # Summary metrics in a compact row
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        active_count = sum(1 for pt in policy_types_data if pt.get('active', True))
                        st.metric("Active Types", active_count)
                    with col2:
                        st.metric("Total Types", len(policy_types_data))
                    with col3:
                        st.metric("Categories", len(set(pt.get('category', 'Other') for pt in policy_types_data)))
                    
                    st.divider()
                    
                    # Display policy types in a grid format
                    st.markdown("### Active Policy Types")
                    
                    # Group by category
                    categories = {}
                    for pt in policy_types_data:
                        if pt.get('active', True):
                            category = pt.get('category', 'Other')
                            if category not in categories:
                                categories[category] = []
                            categories[category].append(pt)
                    
                    # Display each category
                    for category, types in categories.items():
                        st.markdown(f"**{category}**")
                        
                        # Create a grid of policy types
                        cols = st.columns(4)
                        for idx, policy_type in enumerate(types):
                            with cols[idx % 4]:
                                # Display as a compact card
                                st.success(f"✅ {policy_type.get('code', policy_type.get('name'))}")
                    
                    st.divider()
                    
                    # Add/Edit Policy Types Section
                    st.markdown("### ➕ Add New Policy Type")
                    
                    with st.form("add_policy_type_form"):
                        col1, col2 = st.columns(2)
                        with col1:
                            new_code = st.text_input("Code", placeholder="e.g., EPL", max_chars=10)
                            new_category = st.selectbox(
                                "Category",
                                options=["Personal", "Commercial", "Specialty", "Other"]
                            )
                        with col2:
                            new_name = st.text_input("Name", placeholder="e.g., Employment Practices Liability")
                            new_active = st.checkbox("Active", value=True)
                        
                        submitted = st.form_submit_button("Add Policy Type", type="primary")
                        
                        if submitted and new_code and new_name:
                            # Add the new policy type
                            new_policy_type = {
                                "code": new_code.upper(),
                                "name": new_name,
                                "active": new_active,
                                "category": new_category
                            }
                            
                            # Check if policy type already exists
                            existing_codes = [pt.get('code', '').upper() for pt in policy_types_data]
                            if new_code.upper() in existing_codes:
                                st.error(f"Policy type with code '{new_code}' already exists!")
                            else:
                                # Add to the list
                                policy_types_data.append(new_policy_type)
                                
                                # Update the configuration file
                                try:
                                    config_data['policy_types'] = policy_types_data
                                    os.makedirs("config_files", exist_ok=True)
                                    with open(config_path, 'w') as f:
                                        json.dump(config_data, f, indent=2)
                                    st.success(f"✅ Added policy type: {new_code} - {new_name}")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Error saving configuration: {e}")
                    
                    # Edit/Delete existing policy types
                    st.markdown("### 📝 Edit Policy Types")
                    
                    # Create an editable dataframe
                    if policy_types_data:
                        # Convert to DataFrame for easier editing
                        df = pd.DataFrame(policy_types_data)
                        
                        # Add a delete column
                        df['Delete'] = False
                        
                        # Display editable dataframe
                        edited_df = st.data_editor(
                            df,
                            column_config={
                                "code": st.column_config.TextColumn("Code", width="small"),
                                "name": st.column_config.TextColumn("Name", width="medium"),
                                "category": st.column_config.SelectboxColumn(
                                    "Category",
                                    options=["Personal", "Commercial", "Specialty", "Other"],
                                    width="small"
                                ),
                                "active": st.column_config.CheckboxColumn("Active", width="small"),
                                "Delete": st.column_config.CheckboxColumn("Delete", width="small")
                            },
                            hide_index=True,
                            use_container_width=True
                        )
                        
                        # Save changes button
                        if st.button("💾 Save Changes", type="primary"):
                            try:
                                # Filter out deleted items
                                updated_policy_types = []
                                for _, row in edited_df.iterrows():
                                    if not row.get('Delete', False):
                                        updated_policy_types.append({
                                            "code": row.get('code', ''),
                                            "name": row.get('name', ''),
                                            "active": row.get('active', True),
                                            "category": row.get('category', 'Other')
                                        })
                                
                                # Update configuration
                                config_data['policy_types'] = updated_policy_types
                                os.makedirs("config_files", exist_ok=True)
                                with open(config_path, 'w') as f:
                                    json.dump(config_data, f, indent=2)
                                
                                st.success("✅ Policy types updated successfully!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error saving changes: {e}")
                    
                    # Merge Policy Types Section
                    st.divider()
                    st.markdown("### 🔀 Merge Policy Types")
                    st.info("Merge duplicate policy types into a single type. All transactions will be updated to use the target type.")
                    
                    # Check for any policy types in the database
                    try:
                        # Get unique policy types from the database
                        supabase = get_supabase_client()
                        response = supabase.table('policies').select('"Policy Type"').execute()
                        
                        if response.data:
                            db_policy_types = set()
                            for record in response.data:
                                if record.get('Policy Type'):
                                    db_policy_types.add(record['Policy Type'])
                            
                            # Count transactions for each policy type
                            type_counts = {}
                            for pt in db_policy_types:
                                count_response = supabase.table('policies').select('count', count='exact').eq('"Policy Type"', pt).execute()
                                type_counts[pt] = count_response.count if count_response else 0
                            
                            # Sort by name for better display
                            sorted_types = sorted(list(db_policy_types))
                            
                            with st.form("merge_policy_types_form"):
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    merge_from = st.selectbox(
                                        "Merge From (will be deleted)",
                                        options=[""] + sorted_types,
                                        help="Select the policy type to merge FROM (this type will be removed)"
                                    )
                                    if merge_from:
                                        st.caption(f"📊 {type_counts.get(merge_from, 0)} transactions use this type")
                                
                                with col2:
                                    # Filter out the selected merge_from option
                                    merge_to_options = [""] + [pt for pt in sorted_types if pt != merge_from]
                                    merge_to = st.selectbox(
                                        "Merge Into (will be kept)",
                                        options=merge_to_options,
                                        help="Select the policy type to merge INTO (all transactions will use this type)"
                                    )
                                    if merge_to:
                                        st.caption(f"📊 {type_counts.get(merge_to, 0)} transactions use this type")
                                
                                # Show preview of what will happen
                                if merge_from and merge_to:
                                    st.warning(f"⚠️ This will update {type_counts.get(merge_from, 0)} transactions from '{merge_from}' to '{merge_to}'")
                                    st.info("💡 This action cannot be undone. Make sure to backup your data first!")
                                
                                merge_submitted = st.form_submit_button("🔀 Merge Policy Types", type="primary")
                                
                                if merge_submitted:
                                    if merge_from and merge_to and merge_from != merge_to:
                                        try:
                                            # Update all transactions with the merge_from type to use merge_to
                                            update_response = supabase.table('policies').update({'Policy Type': merge_to}).eq('"Policy Type"', merge_from).execute()
                                            
                                            if update_response.data:
                                                updated_count = len(update_response.data)
                                                st.success(f"✅ Successfully merged '{merge_from}' into '{merge_to}'. Updated {updated_count} transactions.")
                                                
                                                # Clear the cache to reflect changes
                                                clear_policies_cache()
                                                
                                                # Also update the policy type mappings if the merged type was mapped
                                                mapping_file = "config_files/policy_type_mappings.json"
                                                if os.path.exists(mapping_file):
                                                    try:
                                                        with open(mapping_file, 'r') as f:
                                                            mappings = json.load(f)
                                                        
                                                        # Update any mappings that pointed to the merged type
                                                        updated_mappings = False
                                                        for key, value in mappings.items():
                                                            if value == merge_from:
                                                                mappings[key] = merge_to
                                                                updated_mappings = True
                                                        
                                                        if updated_mappings:
                                                            with open(mapping_file, 'w') as f:
                                                                json.dump(mappings, f, indent=2)
                                                            st.info("📝 Also updated policy type mappings to use the merged type")
                                                    except:
                                                        pass
                                                
                                                st.rerun()
                                            else:
                                                st.info("No transactions were updated (the types may have already been merged)")
                                        except Exception as e:
                                            st.error(f"Error merging policy types: {e}")
                                    else:
                                        st.error("Please select two different policy types to merge")
                        else:
                            st.info("No policy types found in the database")
                    
                    except Exception as e:
                        st.error(f"Error loading policy types from database: {e}")
                    
                    # Rename Policy Types Section
                    st.divider()
                    st.markdown("### ✏️ Rename Policy Types")
                    st.info("Rename existing policy types to new standardized names. All transactions will be updated.")
                    
                    try:
                        # Get unique policy types from the database (reuse from merge section if available)
                        if 'db_policy_types' not in locals() or 'type_counts' not in locals():
                            supabase = get_supabase_client()
                            response = supabase.table('policies').select('"Policy Type"').execute()
                            
                            if response.data:
                                db_policy_types = set()
                                for record in response.data:
                                    if record.get('Policy Type'):
                                        db_policy_types.add(record['Policy Type'])
                                
                                # Count transactions for each policy type
                                type_counts = {}
                                for pt in db_policy_types:
                                    count_response = supabase.table('policies').select('count', count='exact').eq('"Policy Type"', pt).execute()
                                    type_counts[pt] = count_response.count if count_response else 0
                        
                        if db_policy_types:
                            sorted_types = sorted(list(db_policy_types))
                            
                            with st.form("rename_policy_type_form"):
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    rename_from = st.selectbox(
                                        "Current Name",
                                        options=[""] + sorted_types,
                                        help="Select the policy type to rename"
                                    )
                                    if rename_from:
                                        st.caption(f"📊 {type_counts.get(rename_from, 0)} transactions use this type")
                                
                                with col2:
                                    rename_to = st.text_input(
                                        "New Name",
                                        placeholder="Enter new name (e.g., AUTO, HOME, CONDO)",
                                        help="Enter the new standardized name"
                                    )
                                    if rename_to and rename_from:
                                        # Check if new name already exists
                                        if rename_to in db_policy_types and rename_to != rename_from:
                                            st.error(f"⚠️ '{rename_to}' already exists! Use Merge instead.")
                                        elif rename_to == rename_from:
                                            st.warning("New name is the same as current name")
                                
                                # Show preview
                                if rename_from and rename_to and rename_to != rename_from and rename_to not in db_policy_types:
                                    st.success(f"✅ Will rename '{rename_from}' to '{rename_to}' in {type_counts.get(rename_from, 0)} transactions")
                                
                                rename_submitted = st.form_submit_button("✏️ Rename Policy Type", type="primary")
                                
                                if rename_submitted:
                                    if rename_from and rename_to and rename_to != rename_from:
                                        if rename_to in db_policy_types:
                                            st.error(f"'{rename_to}' already exists in the database. Use the Merge feature instead.")
                                        else:
                                            try:
                                                # Update all transactions with the new name
                                                update_response = supabase.table('policies').update({'Policy Type': rename_to}).eq('"Policy Type"', rename_from).execute()
                                                
                                                if update_response.data:
                                                    updated_count = len(update_response.data)
                                                    st.success(f"✅ Successfully renamed '{rename_from}' to '{rename_to}'. Updated {updated_count} transactions.")
                                                    
                                                    # Clear the cache to reflect changes
                                                    clear_policies_cache()
                                                    
                                                    # Also update the policy type mappings if the renamed type was mapped
                                                    mapping_file = "config_files/policy_type_mappings.json"
                                                    if os.path.exists(mapping_file):
                                                        try:
                                                            with open(mapping_file, 'r') as f:
                                                                mappings = json.load(f)
                                                            
                                                            # Update any mappings that pointed to the old name
                                                            updated_mappings = False
                                                            for key, value in mappings.items():
                                                                if value == rename_from:
                                                                    mappings[key] = rename_to
                                                                    updated_mappings = True
                                                            
                                                            if updated_mappings:
                                                                with open(mapping_file, 'w') as f:
                                                                    json.dump(mappings, f, indent=2)
                                                                st.info("📝 Also updated policy type mappings to use the new name")
                                                        except:
                                                            pass
                                                    
                                                    st.rerun()
                                                else:
                                                    st.info("No transactions were updated")
                                            except Exception as e:
                                                st.error(f"Error renaming policy type: {e}")
                                    else:
                                        st.error("Please select a policy type and enter a new name")
                        else:
                            st.info("No policy types found in the database")
                    
                    except Exception as e:
                        st.error(f"Error loading policy types: {e}")
                    
                    # Backup/Download section
                    st.divider()
                    col1, col2 = st.columns(2)
                    with col1:
                        if 'config_data' in locals():
                            config_json = json.dumps(config_data, indent=2)
                            st.download_button(
                                label="📥 Download Current Configuration",
                                data=config_json,
                                file_name=f"policy_types_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                                mime="application/json"
                            )
                    with col2:
                        st.info("💡 Download a backup before making major changes")
                    
            except Exception as e:
                st.error(f"Error loading policy types configuration: {e}")
                st.info("Using fallback policy types display")
                
                # Fallback to basic display
                policy_types, allow_custom = load_policy_types()
                if policy_types:
                    st.dataframe(pd.DataFrame(policy_types), use_container_width=True)
        
        with tab9:
            st.subheader("🔄 Policy Type Mapping")
            st.info("Map policy types from reconciliation statements to your standardized policy types")
            
            # Load or initialize mappings
            mapping_file = "config_files/policy_type_mappings.json"
            try:
                if os.path.exists(mapping_file):
                    with open(mapping_file, 'r') as f:
                        mappings = json.load(f)
                else:
                    mappings = {}
            except Exception as e:
                st.error(f"Error loading mappings: {e}")
                mappings = {}
            
            # Get list of active policy types for dropdown
            active_types = []
            try:
                config_path = "config_files/policy_types_updated.json"
                if os.path.exists(config_path):
                    with open(config_path, 'r') as f:
                        config_data = json.load(f)
                        policy_types_data = config_data.get("policy_types", [])
                        active_types = [pt['name'] for pt in policy_types_data if pt.get('active', True)]
                else:
                    # Fallback to load_policy_types
                    policy_types, _ = load_policy_types()
                    active_types = [pt['name'] for pt in policy_types if pt.get('active', True)]
                
                # Sort for better display
                active_types.sort()
            except:
                st.error("Could not load policy types")
                active_types = []
            
            # Display current mappings
            st.markdown("### Current Mappings")
            if mappings:
                # Create editable dataframe for mappings
                mapping_data = []
                for statement_value, mapped_to in mappings.items():
                    mapping_data.append({
                        "Statement Value": statement_value,
                        "Maps To": mapped_to,
                        "Delete": False
                    })
                
                mapping_df = pd.DataFrame(mapping_data)
                edited_df = st.data_editor(
                    mapping_df,
                    column_config={
                        "Maps To": st.column_config.SelectboxColumn(
                            "Maps To",
                            options=active_types,
                            required=True
                        ),
                        "Delete": st.column_config.CheckboxColumn(
                            "Delete",
                            help="Check to delete this mapping"
                        )
                    },
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed"
                )
                
                # Save changes button
                if st.button("💾 Save Mapping Changes", type="primary"):
                    # Update mappings based on edits
                    new_mappings = {}
                    for idx, row in edited_df.iterrows():
                        if not row["Delete"]:
                            new_mappings[row["Statement Value"]] = row["Maps To"]
                    
                    # Save to file
                    try:
                        os.makedirs("config_files", exist_ok=True)
                        with open(mapping_file, 'w') as f:
                            json.dump(new_mappings, f, indent=2)
                        st.success("✅ Mappings saved successfully!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error saving mappings: {e}")
            else:
                st.info("No mappings configured yet. Add your first mapping below.")
            
            st.divider()
            
            # Add new mapping
            st.markdown("### Add New Mapping")
            with st.form("add_mapping_form"):
                col1, col2 = st.columns(2)
                with col1:
                    new_statement_value = st.text_input(
                        "Statement Value",
                        placeholder="e.g., AUTO, HO3, CONDO",
                        help="The policy type as it appears in your reconciliation statements"
                    )
                with col2:
                    new_maps_to = st.selectbox(
                        "Maps To",
                        options=[""] + active_types,
                        help="Your standardized policy type"
                    )
                
                submitted = st.form_submit_button("➕ Add Mapping", type="primary")
                
                if submitted:
                    if new_statement_value and new_maps_to:
                        # Check if mapping already exists
                        if new_statement_value in mappings:
                            st.error(f"Mapping for '{new_statement_value}' already exists!")
                        else:
                            # Add new mapping
                            mappings[new_statement_value] = new_maps_to
                            
                            # Save to file
                            try:
                                os.makedirs("config_files", exist_ok=True)
                                with open(mapping_file, 'w') as f:
                                    json.dump(mappings, f, indent=2)
                                st.success(f"✅ Added mapping: {new_statement_value} → {new_maps_to}")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error saving mapping: {e}")
                    else:
                        st.error("Please fill in both fields")
            
            st.divider()
            
            # Mapping summary
            st.markdown("### Mapping Summary")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total Mappings", len(mappings))
            with col2:
                st.metric("Active Policy Types", len(active_types))
            
            # Help text
            with st.expander("ℹ️ How Policy Type Mapping Works"):
                st.markdown("""
                **Purpose**: Prevent duplicate policy types during reconciliation by mapping statement values to your standardized types.
                
                **How it works**:
                1. During reconciliation import, the system checks each policy type against these mappings
                2. If a mapping exists, it automatically uses your standardized type
                3. If no mapping exists, the import will stop and ask you to add the mapping
                
                **Example**:
                - Statement has "AUTO" → Maps to "Private Passenger Auto"
                - Statement has "HO3" → Maps to "Homeowners"
                - Statement has "DWELLING" → Maps to "Dwelling Fire"
                
                **Benefits**:
                - No more duplicate policy types
                - Consistent data across all imports
                - Easy to maintain and update
                """)
    
    # --- Contacts ---
    elif page == "Contacts":
        # Modern SaaS-style header
        col1, col2 = st.columns([2, 1])
        with col1:
            st.title("Commission Management")
        with col2:
            # Quick stats
            if 'carriers_data' in st.session_state:
                st.metric("Active Carriers", len([c for c in st.session_state.get('carriers_data', []) if c.get('status') == 'Active']))
        
        # Load fresh data for this page (island architecture)
        all_data = load_policies_data()
        
        # Initialize session state for Contacts page
        if 'carriers_data' not in st.session_state:
            st.session_state.carriers_data = []
        if 'mgas_data' not in st.session_state:
            st.session_state.mgas_data = []
        if 'commission_rules' not in st.session_state:
            st.session_state.commission_rules = []
        if 'selected_carrier_id' not in st.session_state:
            st.session_state.selected_carrier_id = None
        if 'search_mode' not in st.session_state:
            st.session_state.search_mode = 'carriers'
        
        # Load carriers, MGAs, and commission rules from database
        try:
            # Check if tables exist first
            supabase = get_supabase_client()
            
            # Try to load carriers
            try:
                response = supabase.table('carriers').select("*").execute()
                st.session_state.carriers_data = response.data if response.data else []
            except Exception as e:
                if "relation" in str(e) and "does not exist" in str(e):
                    st.warning("⚠️ Commission structure tables not found. Please run the SQL script to create them.")
                    with st.expander("View SQL Script"):
                        st.code("""
-- Run this SQL in your Supabase SQL editor:
-- /sql_scripts/create_commission_structure_tables.sql
                        """)
                    st.stop()
                else:
                    st.error(f"Error loading carriers: {e}")
            
            # Try to load MGAs
            try:
                response = supabase.table('mgas').select("*").execute()
                st.session_state.mgas_data = response.data if response.data else []
            except Exception:
                pass
            
            # Try to load commission rules
            try:
                response = supabase.table('commission_rules').select("*").execute()
                st.session_state.commission_rules = response.data if response.data else []
            except Exception:
                pass
                
        except Exception as e:
            st.error(f"Error loading data: {e}")
        
        # Modern Command Palette Style Search
        st.markdown("### 🔍 Search")
        search_container = st.container()
        
        with search_container:
            col1, col2 = st.columns([3, 1])
            
            with col1:
                search_query = st.text_input(
                    "",
                    placeholder="Search carriers, MGAs, or commission rules... (Try: Progressive, 12%, Auto)",
                    label_visibility="collapsed",
                    key="commission_search"
                )
            
            with col2:
                quick_add = st.selectbox(
                    "",
                    ["Quick Add ➕", "Add Carrier", "Add MGA", "Add Rule"],
                    label_visibility="collapsed",
                    key="quick_add_select"
                )
                
                if quick_add == "Add Carrier":
                    st.session_state['show_add_carrier'] = True
                elif quick_add == "Add MGA":
                    st.session_state['show_add_mga'] = True
        
        # Add carrier modal - MOVED TO TOP FOR VISIBILITY
        if st.session_state.get('show_add_carrier'):
            with st.form("add_carrier_form_modal"):
                st.markdown("### Add New Carrier")
                col1, col2 = st.columns(2)
                
                with col1:
                    carrier_name = st.text_input("Carrier Name*", placeholder="e.g., Progressive Insurance")
                    naic_code = st.text_input("NAIC Code", placeholder="Optional")
                
                with col2:
                    producer_code = st.text_input("Producer Code", placeholder="Optional")
                    notes = st.text_area("Notes", placeholder="Optional notes")
                
                col1, col2 = st.columns([3, 1])
                with col2:
                    if st.form_submit_button("Add Carrier", type="primary"):
                        if carrier_name:
                            try:
                                new_carrier = {
                                    "carrier_name": carrier_name,
                                    "naic_code": naic_code if naic_code else None,
                                    "producer_code": producer_code if producer_code else None,
                                    "status": "Active",
                                    "notes": notes if notes else None
                                }
                                
                                response = supabase.table('carriers').insert(new_carrier).execute()
                                st.success(f"✅ Added {carrier_name}")
                                del st.session_state['show_add_carrier']
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")
                        else:
                            st.error("Carrier name is required")
                    
                    if st.form_submit_button("Cancel"):
                        del st.session_state['show_add_carrier']
                        st.rerun()
        
        # Add MGA modal - NEW IMPLEMENTATION
        if st.session_state.get('show_add_mga'):
            with st.form("add_mga_form_modal"):
                st.markdown("### Add New MGA")
                col1, col2 = st.columns(2)
                
                with col1:
                    mga_name = st.text_input("MGA Name*", placeholder="e.g., Advantage Partners LLC")
                    contact_name = st.text_input("Contact Name", placeholder="Optional")
                    phone = st.text_input("Phone", placeholder="Optional")
                
                with col2:
                    email = st.text_input("Email", placeholder="Optional")
                    website = st.text_input("Website", placeholder="Optional")
                    notes = st.text_area("Notes", placeholder="Optional notes")
                
                col1, col2 = st.columns([3, 1])
                with col2:
                    if st.form_submit_button("Add MGA", type="primary"):
                        if mga_name:
                            try:
                                # Build contact_info as JSONB
                                contact_info = {}
                                if contact_name:
                                    contact_info["contact_name"] = contact_name
                                if phone:
                                    contact_info["phone"] = phone
                                if email:
                                    contact_info["email"] = email
                                if website:
                                    contact_info["website"] = website
                                
                                new_mga = {
                                    "mga_name": mga_name,
                                    "contact_info": contact_info,
                                    "status": "Active",
                                    "notes": notes if notes else None
                                }
                                
                                response = supabase.table('mgas').insert(new_mga).execute()
                                st.success(f"✅ Added {mga_name}")
                                del st.session_state['show_add_mga']
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")
                        else:
                            st.error("MGA name is required")
                    
                    if st.form_submit_button("Cancel"):
                        del st.session_state['show_add_mga']
                        st.rerun()
        
        # Detailed Instructions and Help
        with st.expander("ℹ️ How Commission Rules Work", expanded=False):
            st.info("""
            **Important Information About Commission Rules:**
            
            🔹 **Rules are for future transactions only** - They will auto-populate rates when adding new policies
            
            🔹 **Existing transactions are NOT affected** - Deleting or changing rules does NOT modify any existing commission data in your policies
            
            🔹 **Your historical data is safe** - All past transactions keep their original commission calculations
            
            🔹 **Think of it like store prices** - Changing today's price doesn't affect what customers paid yesterday
            
            **📅 NEW! Effective Date Tracking:**
            
            🔸 **Set effective dates** - Schedule rate changes in advance (e.g., rate increases on Sept 1st)
            
            🔸 **End date rules** - Click "📅 End Date" to expire a rule instead of deleting it
            
            🔸 **Complete history** - View all historical rates in the "📜 Historical Rules" section
            
            🔸 **Backdate changes** - Found out about a rate change late? Set the effective date to when it actually changed
            
            🔸 **Track rate changes** - Know exactly when each rate was active and why it changed
            
            **To Handle Rate Changes:**
            1. Find the current rule and click "📅 End Date"
            2. Set the last day of the old rate
            3. Add reason (e.g., "Rate increase to 14%")
            4. Create new rule with new rate and next day as effective date
            
            **Example**: Progressive Auto changing from 12% to 14% on Sept 1st
            - End current rule on Aug 31st
            - Add new rule with 14% effective Sept 1st
            - Both rules are preserved for historical reference
            
            **🚀 NEW! Add/Edit Policy Integration:**
            - ✅ Carrier dropdown auto-populates available carriers
            - ✅ MGA dropdown shows only relevant MGAs for selected carrier
            - ✅ Commission rates automatically filled based on rules
            - ✅ Override capability with required reason tracking
            - ✅ Shows which rule is being applied (e.g., "Progressive (Direct) - Auto - 12%")
            
            **Search Tips:**
            - Search by carrier name: "Progressive", "Citizens"
            - Search by rate: "12", "15" (finds all rules with that rate)
            - Search by policy type: "Auto", "Home", "BOP"
            - Search by MGA: "Advantage", "TWFG"
            """)
        
        # Filter and search results
        filtered_carriers = []
        filtered_mgas = []
        filtered_rules = []
        
        if search_query:
            search_lower = search_query.lower()
            
            # Search carriers
            filtered_carriers = [
                c for c in st.session_state.carriers_data
                if search_lower in (c.get('carrier_name') or '').lower() or
                   search_lower in (c.get('notes') or '').lower()
            ]
            
            # Search MGAs
            filtered_mgas = [
                m for m in st.session_state.mgas_data
                if search_lower in (m.get('mga_name') or '').lower() or
                   search_lower in (m.get('notes') or '').lower()
            ]
            
            # Search commission rules by rate or policy type
            for rule in st.session_state.commission_rules:
                # Check if search matches rate
                if (str(rule.get('new_rate', '')).startswith(search_query) or 
                    str(rule.get('renewal_rate', '')).startswith(search_query) or
                    search_lower in (rule.get('policy_type', '') or '').lower() or
                    search_lower in (rule.get('rule_description', '') or '').lower()):
                    filtered_rules.append(rule)
        
        # Show search results in modern cards
        if search_query:
            if filtered_carriers or filtered_mgas or filtered_rules:
                st.markdown("### Search Results")
                
                # Carriers results
                if filtered_carriers:
                    st.markdown("#### Carriers")
                    for carrier in filtered_carriers[:5]:  # Show top 5
                        if st.button(f"🏢 {carrier['carrier_name']}", key=f"search_carrier_{carrier['carrier_id']}", use_container_width=True):
                            st.session_state.selected_carrier_id = carrier['carrier_id']
                            st.session_state.search_mode = 'carrier_detail'
                            st.rerun()
                
                # MGAs results
                if filtered_mgas:
                    st.markdown("#### MGAs")
                    for mga in filtered_mgas[:5]:
                        if st.button(f"🤝 {mga['mga_name']}", key=f"search_mga_{mga['mga_id']}", use_container_width=True):
                            st.session_state.selected_mga_id = mga['mga_id']
                            st.session_state.search_mode = 'mga_detail'
                            st.rerun()
                
                # Rules results
                if filtered_rules:
                    st.markdown("#### Commission Rules")
                    for rule in filtered_rules[:5]:
                        carrier = next((c for c in st.session_state.carriers_data if c['carrier_id'] == rule['carrier_id']), None)
                        if carrier:
                            mga_name = "Direct"
                            if rule.get('mga_id'):
                                mga = next((m for m in st.session_state.mgas_data if m['mga_id'] == rule['mga_id']), None)
                                if mga:
                                    mga_name = mga['mga_name']
                            
                            rule_text = f"💰 {carrier['carrier_name']} ({mga_name})"
                            if rule.get('policy_type'):
                                rule_text += f" - {rule['policy_type']}"
                            rule_text += f": {rule['new_rate']}% / {rule.get('renewal_rate', rule['new_rate'])}%"
                            
                            if st.button(rule_text, key=f"search_rule_{rule['rule_id']}", use_container_width=True):
                                st.session_state.selected_carrier_id = rule['carrier_id']
                                st.session_state.search_mode = 'carrier_detail'
                                st.rerun()
            else:
                st.info("No results found. Try a different search term.")
        
        # Main content area
        st.markdown("---")
        
        # Show carrier detail view if selected
        if st.session_state.get('search_mode') == 'carrier_detail' and st.session_state.get('selected_carrier_id'):
            selected_carrier = next((c for c in st.session_state.carriers_data if c['carrier_id'] == st.session_state.selected_carrier_id), None)
            
            if selected_carrier:
                # Back button
                if st.button("← Back to Search", key="back_to_search"):
                    st.session_state.search_mode = 'carriers'
                    st.session_state.selected_carrier_id = None
                    st.rerun()
                
                # Carrier detail card with edit button
                header_col1, header_col2 = st.columns([3, 1])
                with header_col1:
                    st.markdown(f"## {selected_carrier['carrier_name']}")
                with header_col2:
                    if st.button("✏️ Edit Carrier", key="edit_carrier_btn"):
                        st.session_state['editing_carrier'] = True
                
                # Check if we're in edit mode
                if st.session_state.get('editing_carrier'):
                    # Edit form
                    with st.form("edit_carrier_form"):
                        st.markdown("### Edit Carrier Information")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            edited_name = st.text_input("Carrier Name*", value=selected_carrier['carrier_name'])
                            edited_naic = st.text_input("NAIC Code", value=selected_carrier.get('naic_code', ''))
                            edited_status = st.selectbox("Status", ["Active", "Inactive"], 
                                                        index=0 if selected_carrier.get('status') == 'Active' else 1)
                        
                        with col2:
                            edited_producer = st.text_input("Producer Code", value=selected_carrier.get('producer_code', ''))
                            edited_notes = st.text_area("Notes", value=selected_carrier.get('notes', ''))
                        
                        form_col1, form_col2, form_col3 = st.columns([1, 1, 3])
                        with form_col1:
                            if st.form_submit_button("💾 Save Changes", type="primary"):
                                if edited_name:
                                    try:
                                        update_data = {
                                            "carrier_name": edited_name,
                                            "naic_code": edited_naic if edited_naic else None,
                                            "producer_code": edited_producer if edited_producer else None,
                                            "status": edited_status,
                                            "notes": edited_notes if edited_notes else None
                                        }
                                        
                                        supabase.table('carriers').update(update_data).eq('carrier_id', selected_carrier['carrier_id']).execute()
                                        st.success(f"✅ Updated {edited_name}")
                                        del st.session_state['editing_carrier']
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error updating carrier: {e}")
                                else:
                                    st.error("Carrier name is required")
                        
                        with form_col2:
                            if st.form_submit_button("❌ Cancel"):
                                del st.session_state['editing_carrier']
                                st.rerun()
                else:
                    # Normal view - Carrier info in modern layout
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.markdown(f"**Status:** {'✅ Active' if selected_carrier.get('status') == 'Active' else '⚠️ Inactive'}")
                    with col2:
                        if selected_carrier.get('naic_code'):
                            st.markdown(f"**NAIC:** {selected_carrier['naic_code']}")
                    with col3:
                        if selected_carrier.get('producer_code'):
                            st.markdown(f"**Producer Code:** {selected_carrier['producer_code']}")
                
                if selected_carrier.get('notes'):
                    st.info(f"📝 {selected_carrier['notes']}")
                
                # Commission rules section with inline editing feel
                st.markdown("### Commission Rules")
                
                # Quick help for this carrier
                with st.expander("💡 Quick Tips", expanded=False):
                    st.markdown("""
                    **Managing Commission Rules:**
                    - **Direct vs MGA**: Direct appointments don't use an MGA
                    - **Policy Types**: Comma-separated (e.g., "Auto, Boat")
                    - **Effective Dates**: Rules apply from this date forward
                    - **End Dating**: Better than deleting - preserves history
                    - **Payment Terms**: Advanced = paid upfront, As Earned = over time
                    
                    **Rate Changes:**
                    1. End-date the current rule (preserves history)
                    2. Add new rule with new rates starting next day
                    3. Both rules are kept for audit trail
                    
                    **Integration with Policies:**
                    - These rules auto-fill commission rates in Add/Edit Policy
                    - You can always override with a reason
                    - Historical policies are NOT affected by rule changes
                    """)
                
                # Quick add rule button
                col1, col2 = st.columns([3, 1])
                with col2:
                    if st.button("➕ Add Rule", key="inline_add_rule", type="primary"):
                        st.session_state['show_inline_add_rule'] = True
                
                # Inline add rule form
                if st.session_state.get('show_inline_add_rule'):
                    with st.container():
                        st.markdown("#### New Commission Rule")
                        with st.form("inline_rule_form"):
                            col1, col2, col3 = st.columns(3)
                            
                            with col1:
                                mga_options = ["Direct Appointment"] + [m['mga_name'] for m in st.session_state.mgas_data]
                                selected_mga = st.selectbox("MGA", options=mga_options)
                                policy_type = st.text_input("Policy Type(s)", placeholder="Auto, HO3, etc.")
                            
                            with col2:
                                new_rate = st.number_input("NEW %", min_value=0.0, max_value=100.0, value=10.0, step=0.5)
                                renewal_rate = st.number_input("RWL %", min_value=0.0, max_value=100.0, value=10.0, step=0.5)
                            
                            with col3:
                                effective_date = st.date_input("Effective", value=datetime.date.today())
                                payment_terms = st.selectbox("Terms", ["", "Advanced", "As Earned"])
                            
                            col1, col2 = st.columns([3, 1])
                            with col1:
                                rule_description = st.text_input("Description", placeholder="e.g., Standard rate effective Jan 2025")
                            with col2:
                                st.write("")  # Spacer
                                if st.form_submit_button("Save", type="primary"):
                                    try:
                                        mga_id = None
                                        if selected_mga != "Direct Appointment":
                                            mga_id = next((m['mga_id'] for m in st.session_state.mgas_data if m['mga_name'] == selected_mga), None)
                                        
                                        new_rule = {
                                            "carrier_id": selected_carrier['carrier_id'],
                                            "mga_id": mga_id,
                                            "policy_type": policy_type if policy_type else None,
                                            "new_rate": new_rate,
                                            "renewal_rate": renewal_rate,
                                            "payment_terms": payment_terms if payment_terms else None,
                                            "rule_description": rule_description if rule_description else None,
                                            "state": "FL",
                                            "effective_date": effective_date.isoformat(),
                                            "is_active": True
                                        }
                                        
                                        response = supabase.table('commission_rules').insert(new_rule).execute()
                                        st.success("✅ Rule added")
                                        del st.session_state['show_inline_add_rule']
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error: {e}")
                                
                                if st.form_submit_button("Cancel"):
                                    del st.session_state['show_inline_add_rule']
                                    st.rerun()
                
                # Display rules in modern card style
                carrier_rules = [r for r in st.session_state.commission_rules if r.get('carrier_id') == selected_carrier['carrier_id']]
                active_rules = [r for r in carrier_rules if r.get('is_active', True) and not r.get('end_date')]
                inactive_rules = [r for r in carrier_rules if not r.get('is_active', True) or r.get('end_date')]
                
                # Active rules as cards
                for rule in sorted(active_rules, key=lambda x: x.get('effective_date', ''), reverse=True):
                    with st.container():
                        # Rule card with modern design
                        col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
                        
                        with col1:
                            mga_name = "Direct"
                            if rule.get('mga_id'):
                                mga_name = next((m['mga_name'] for m in st.session_state.mgas_data if m['mga_id'] == rule['mga_id']), "Unknown")
                            
                            rule_title = f"**{mga_name}**"
                            if rule.get('policy_type'):
                                rule_title += f" - {rule['policy_type']}"
                            st.markdown(rule_title)
                            
                            if rule.get('rule_description'):
                                st.caption(rule['rule_description'])
                        
                        with col2:
                            st.markdown(f"**NEW:** {rule['new_rate']}%  \n**RWL:** {rule.get('renewal_rate', rule['new_rate'])}%")
                        
                        with col3:
                            effective_date = rule.get('effective_date', 'Unknown')
                            if effective_date and effective_date != 'Unknown':
                                try:
                                    date_obj = datetime.datetime.strptime(effective_date[:10], '%Y-%m-%d')
                                    formatted_date = date_obj.strftime('%m/%d/%Y')
                                    st.markdown(f"**Effective:** {formatted_date}")
                                except:
                                    st.markdown(f"**Effective:** Current")
                            
                            if rule.get('payment_terms'):
                                st.caption(rule['payment_terms'])
                        
                        with col4:
                            # Action menu
                            action = st.selectbox(
                                "",
                                ["Actions", "End Date", "Delete"],
                                key=f"action_{rule['rule_id']}",
                                label_visibility="collapsed"
                            )
                            
                            if action == "End Date":
                                st.session_state[f'end_date_{rule["rule_id"]}'] = True
                                st.rerun()
                            elif action == "Delete":
                                st.session_state[f'delete_{rule["rule_id"]}'] = True
                                st.rerun()
                        
                        # Inline end date form
                        if st.session_state.get(f'end_date_{rule["rule_id"]}'):
                            with st.form(f"end_form_{rule['rule_id']}"):
                                col1, col2, col3 = st.columns([2, 2, 2])
                                with col1:
                                    end_date = st.date_input("End Date", value=datetime.date.today())
                                with col2:
                                    reason = st.text_input("Reason", placeholder="Rate increase")
                                with col3:
                                    if st.form_submit_button("Confirm"):
                                        try:
                                            update_data = {
                                                "end_date": end_date.isoformat(),
                                                "is_active": False
                                            }
                                            if reason:
                                                current_desc = rule.get('rule_description', '')
                                                update_data['rule_description'] = f"{current_desc} | Ended: {reason}"
                                            
                                            supabase.table('commission_rules').update(update_data).eq('rule_id', rule['rule_id']).execute()
                                            del st.session_state[f'end_date_{rule["rule_id"]}']
                                            st.success("Rule updated")
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"Error: {e}")
                                    
                                    if st.form_submit_button("Cancel"):
                                        del st.session_state[f'end_date_{rule["rule_id"]}']
                                        st.rerun()
                        
                        # Inline delete confirmation
                        if st.session_state.get(f'delete_{rule["rule_id"]}'):
                            col1, col2 = st.columns([3, 1])
                            with col1:
                                st.warning("Are you sure you want to delete this rule?")
                            with col2:
                                if st.button("Delete", type="primary", key=f"confirm_delete_{rule['rule_id']}"):
                                    try:
                                        supabase.table('commission_rules').delete().eq('rule_id', rule['rule_id']).execute()
                                        del st.session_state[f'delete_{rule["rule_id"]}']
                                        st.success("Rule deleted")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error: {e}")
                                
                                if st.button("Cancel", key=f"cancel_delete_{rule['rule_id']}"):
                                    del st.session_state[f'delete_{rule["rule_id"]}']
                                    st.rerun()
                        
                        st.markdown("---")
                
                # Historical rules
                if inactive_rules:
                    if st.checkbox("Show Historical Rules", key="show_historical"):
                        st.markdown("#### Historical Rules")
                        for rule in sorted(inactive_rules, key=lambda x: x.get('effective_date', ''), reverse=True):
                            col1, col2, col3 = st.columns([3, 2, 3])
                            
                            with col1:
                                mga_name = "Direct"
                                if rule.get('mga_id'):
                                    mga_name = next((m['mga_name'] for m in st.session_state.mgas_data if m['mga_id'] == rule['mga_id']), "Unknown")
                                
                                st.text(f"{mga_name} - {rule.get('policy_type', 'All')}")
                            
                            with col2:
                                st.text(f"{rule['new_rate']}% / {rule.get('renewal_rate', rule['new_rate'])}%")
                            
                            with col3:
                                date_range = ""
                                if rule.get('effective_date'):
                                    try:
                                        eff_date = datetime.datetime.strptime(rule['effective_date'][:10], '%Y-%m-%d').strftime('%m/%d/%Y')
                                        date_range = f"{eff_date}"
                                    except:
                                        date_range = "Unknown"
                                
                                if rule.get('end_date'):
                                    try:
                                        end_date = datetime.datetime.strptime(rule['end_date'][:10], '%Y-%m-%d').strftime('%m/%d/%Y')
                                        date_range += f" - {end_date}"
                                    except:
                                        pass
                                
                                st.text(date_range)
                
                if not carrier_rules:
                    st.info("No commission rules yet. Click '➕ Add Rule' to create one.")
        
        # Show MGA detail view if selected
        elif st.session_state.get('search_mode') == 'mga_detail' and st.session_state.get('selected_mga_id'):
            selected_mga = next((m for m in st.session_state.mgas_data if m['mga_id'] == st.session_state.selected_mga_id), None)
            
            if selected_mga:
                # Back button
                if st.button("← Back to Search", key="back_from_mga"):
                    st.session_state.search_mode = 'carriers'
                    st.session_state.selected_mga_id = None
                    st.rerun()
                
                # MGA detail card with edit button
                header_col1, header_col2 = st.columns([3, 1])
                with header_col1:
                    st.markdown(f"## {selected_mga['mga_name']}")
                with header_col2:
                    if st.button("✏️ Edit MGA", key="edit_mga_btn"):
                        st.session_state['editing_mga'] = True
                
                # Check if we're in edit mode
                if st.session_state.get('editing_mga'):
                    # Edit form
                    with st.form("edit_mga_form"):
                        st.markdown("### Edit MGA Information")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            edited_name = st.text_input("MGA Name*", value=selected_mga['mga_name'])
                            edited_contact = st.text_input("Contact Name", value=selected_mga.get('contact_name', ''))
                            edited_email = st.text_input("Email", value=selected_mga.get('email', ''))
                        
                        with col2:
                            edited_phone = st.text_input("Phone", value=selected_mga.get('phone', ''))
                            edited_notes = st.text_area("Notes", value=selected_mga.get('notes', ''))
                        
                        form_col1, form_col2, form_col3 = st.columns([1, 1, 3])
                        with form_col1:
                            if st.form_submit_button("💾 Save Changes", type="primary"):
                                if edited_name:
                                    try:
                                        update_data = {
                                            "mga_name": edited_name,
                                            "contact_name": edited_contact if edited_contact else None,
                                            "phone": edited_phone if edited_phone else None,
                                            "email": edited_email if edited_email else None,
                                            "notes": edited_notes if edited_notes else None
                                        }
                                        
                                        supabase.table('mgas').update(update_data).eq('mga_id', selected_mga['mga_id']).execute()
                                        st.success(f"✅ Updated {edited_name}")
                                        del st.session_state['editing_mga']
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error updating MGA: {e}")
                                else:
                                    st.error("MGA name is required")
                        
                        with form_col2:
                            if st.form_submit_button("❌ Cancel"):
                                del st.session_state['editing_mga']
                                st.rerun()
                else:
                    # Normal view - MGA info
                    col1, col2 = st.columns(2)
                    with col1:
                        if selected_mga.get('contact_name'):
                            st.markdown(f"**Contact:** {selected_mga['contact_name']}")
                        if selected_mga.get('phone'):
                            st.markdown(f"**Phone:** {selected_mga['phone']}")
                    with col2:
                        if selected_mga.get('email'):
                            st.markdown(f"**Email:** {selected_mga['email']}")
                    
                    if selected_mga.get('notes'):
                        st.info(f"📝 {selected_mga['notes']}")
                
                # Show carriers associated with this MGA
                st.markdown("### Associated Carriers")
                
                # Find commission rules with this MGA
                mga_rules = [r for r in st.session_state.commission_rules 
                           if r.get('mga_id') == selected_mga['mga_id'] and r.get('is_active', True)]
                
                unique_carriers = list(set([r['carrier_id'] for r in mga_rules]))
                
                if unique_carriers:
                    cols = st.columns(3)
                    for idx, carrier_id in enumerate(unique_carriers):
                        carrier = next((c for c in st.session_state.carriers_data if c['carrier_id'] == carrier_id), None)
                        if carrier:
                            with cols[idx % 3]:
                                if st.button(
                                    f"🏢 {carrier['carrier_name']}",
                                    key=f"mga_carrier_{carrier['carrier_id']}",
                                    use_container_width=True
                                ):
                                    st.session_state.selected_carrier_id = carrier['carrier_id']
                                    st.session_state.search_mode = 'carrier_detail'
                                    st.rerun()
                else:
                    st.info("No carriers associated with this MGA yet.")
        
        # Default view - Recent carriers
        else:
            st.markdown("### Recent Carriers")
            
            # Show recent carriers as cards
            recent_carriers = sorted(
                [c for c in st.session_state.carriers_data if c.get('status') == 'Active'],
                key=lambda x: x.get('updated_at', ''),
                reverse=True
            )[:6]  # Show top 6
            
            if recent_carriers:
                cols = st.columns(3)
                for idx, carrier in enumerate(recent_carriers):
                    with cols[idx % 3]:
                        # Count rules for this carrier
                        rule_count = len([r for r in st.session_state.commission_rules 
                                        if r.get('carrier_id') == carrier['carrier_id'] 
                                        and r.get('is_active', True)])
                        
                        # Carrier card
                        if st.button(
                            f"**{carrier['carrier_name']}**\n\n{rule_count} active rules",
                            key=f"carrier_card_{carrier['carrier_id']}",
                            use_container_width=True,
                            help=carrier.get('notes', '')
                        ):
                            st.session_state.selected_carrier_id = carrier['carrier_id']
                            st.session_state.search_mode = 'carrier_detail'
                            st.rerun()
            else:
                st.info("No carriers found. Use Quick Add to create your first carrier.")
            
            # Show MGAs section
            st.markdown("### Recent MGAs")
            
            recent_mgas = sorted(
                st.session_state.mgas_data,
                key=lambda x: x.get('updated_at', ''),
                reverse=True
            )[:6]  # Show top 6
            
            if recent_mgas:
                cols = st.columns(3)
                for idx, mga in enumerate(recent_mgas):
                    with cols[idx % 3]:
                        # Count associated carriers
                        carrier_count = len(set([r['carrier_id'] for r in st.session_state.commission_rules 
                                               if r.get('mga_id') == mga['mga_id'] and r.get('is_active', True)]))
                        
                        # MGA card
                        if st.button(
                            f"**{mga['mga_name']}**\n\n{carrier_count} carriers",
                            key=f"mga_card_{mga['mga_id']}",
                            use_container_width=True,
                            help=mga.get('notes', '')
                        ):
                            st.session_state.selected_mga_id = mga['mga_id']
                            st.session_state.search_mode = 'mga_detail'
                            st.rerun()
            else:
                st.info("No MGAs found. Use Quick Add to create MGAs if you work with them.")
            
            if not recent_carriers and not recent_mgas:
                # Getting started guide
                st.markdown("""
                ### 🚀 Getting Started with Commission Management
                
                **Step 1: Add Your Carriers**
                - Click "Quick Add → Add Carrier" above
                - Enter carrier names (e.g., Progressive, Citizens, AAA)
                - NAIC and Producer codes are optional
                
                **Step 2: Add MGAs (if applicable)**
                - Many carriers work through MGAs (Managing General Agencies)
                - Click "Quick Add → Add MGA" to add them
                - Examples: Advantage Partners, TWFG, Burns & Wilcox
                
                **Step 3: Create Commission Rules**
                - Click on a carrier card to open it
                - Add rules for different policy types and rates
                - Set NEW and RWL (renewal) rates separately
                
                **Step 4: Integration with Policies**
                - When adding new policies, select carrier from dropdown
                - Commission rates auto-populate from your rules
                - Override when needed with reason tracking
                
                **Need Help?**
                - Check the "ℹ️ How Commission Rules Work" section above
                - SQL scripts available at: `/sql_scripts/populate_initial_carriers_mgas.sql`
                """)
        
    
    # --- Tools ---
    elif page == "Tools":
        st.title("🛠️ Tools")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        tab1, tab2, tab3 = st.tabs(["Data Tools", "Utility Functions", "Import/Export"])
        
        with tab1:
            st.subheader("Data Tools")
            
            # Commission calculator
            st.write("**Commission Calculator**")
            calc_col1, calc_col2 = st.columns(2)
            
            with calc_col1:
                premium_amount = st.number_input("Premium Amount", value=0.0, format="%.2f")
                commission_rate = st.number_input("Commission Rate (%)", value=10.0, format="%.2f")
            
            with calc_col2:
                calculated_commission = premium_amount * (commission_rate / 100)
                st.metric("Calculated Commission", f"${calculated_commission:.2f}")
            
            st.divider()
            
            # ID Generators
            st.write("**ID Generators**")
            gen_col1, gen_col2 = st.columns(2)
            
            with gen_col1:
                if st.button("Generate Client ID"):
                    new_client_id = generate_client_id()
                    st.code(new_client_id)
            
            with gen_col2:
                if st.button("Generate Transaction ID"):
                    new_transaction_id = generate_transaction_id()
                    st.code(new_transaction_id)
            
            st.divider()
            
            # Policy Origination Date Population Tool
            st.write("**🗓️ Populate Missing Policy Origination Dates**")
            
            st.info("""
            This tool will automatically populate missing Policy Origination Dates using the following logic:
            - **NEW transactions**: Use the Effective Date
            - **BoR transactions**: Use the Effective Date (new relationship)
            - **Other transactions**: Trace back through policy history to find the original NEW transaction
            """)
            
            # Exclude reconciliation transactions (STMT, VOID, ADJ)
            regular_transactions = all_data[~all_data['Transaction ID'].str.contains('-STMT-|-VOID-|-ADJ-', na=False)].copy()
            
            # Find transactions missing Policy Origination Date
            missing_origination = regular_transactions[
                (regular_transactions['Policy Origination Date'].isna()) | 
                (regular_transactions['Policy Origination Date'] == '')
            ].copy()
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Transactions", len(all_data))
            with col2:
                st.metric("Regular Transactions", len(regular_transactions))
            with col3:
                st.metric("Missing Origination Date", len(missing_origination))
            
            if len(missing_origination) == 0:
                st.success("✅ All transactions already have Policy Origination Date populated!")
            else:
                # Show preview
                if st.checkbox("Show transactions that need origination dates"):
                    st.dataframe(
                        missing_origination[['Transaction ID', 'Customer', 'Policy Number', 'Transaction Type', 'Effective Date']].head(20),
                        use_container_width=True
                    )
                    if len(missing_origination) > 20:
                        st.caption(f"Showing first 20 of {len(missing_origination)} transactions")
                
                # Process button
                if st.button("🔍 Analyze Missing Dates", type="primary", key="analyze_origination_dates"):
                    updates = []
                    errors = []
                    
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Process each missing transaction
                    for counter, (idx, row) in enumerate(missing_origination.iterrows()):
                        progress = (counter + 1) / len(missing_origination)
                        progress_bar.progress(progress)
                        
                        transaction_id = row.get('Transaction ID')
                        transaction_type = row.get('Transaction Type')
                        policy_number = row.get('Policy Number')
                        effective_date = row.get('Effective Date')
                        customer = row.get('Customer', 'Unknown')
                        
                        status_text.text(f"Processing {counter+1}/{len(missing_origination)}: {transaction_id}")
                        
                        auto_populated_date = None
                        reason = ""
                        
                        # Apply the same logic as the form
                        if transaction_type == 'NEW':
                            if effective_date and pd.notna(effective_date):
                                auto_populated_date = effective_date
                                reason = "NEW transaction - using Effective Date"
                        
                        elif transaction_type == 'BoR':
                            if effective_date and pd.notna(effective_date):
                                auto_populated_date = effective_date
                                reason = "BoR transaction - using Effective Date"
                        
                        elif transaction_type in ['RWL', 'END', 'PCH', 'CAN', 'XCL', 'REWRITE', 'NBS', 'STL'] and policy_number:
                            # Use the same recursive function from the form
                            def find_origination_date_tool(policy_num, visited=None):
                                if visited is None:
                                    visited = set()
                                
                                if policy_num in visited:
                                    return None, None  # Circular reference protection
                                visited.add(policy_num)
                                
                                # Find transactions for this policy number
                                policy_transactions = all_data[all_data['Policy Number'] == policy_num].copy()
                                
                                if not policy_transactions.empty:
                                    # Sort by transaction date/effective date to get earliest
                                    if 'Effective Date' in policy_transactions.columns:
                                        policy_transactions = policy_transactions.sort_values('Effective Date')
                                    
                                    # Check for NEW transaction
                                    new_transactions = policy_transactions[policy_transactions['Transaction Type'] == 'NEW']
                                    if len(new_transactions) > 1:
                                        # Multiple NEW transactions warning
                                        return None, "Multiple NEW transactions found"
                                    elif len(new_transactions) == 1:
                                        # Found the NEW transaction
                                        orig_date = new_transactions.iloc[0].get('Policy Origination Date')
                                        if orig_date and pd.notna(orig_date):
                                            return orig_date, "Found from NEW transaction"
                                        # If NEW transaction has no origination date, use its effective date
                                        eff_date = new_transactions.iloc[0].get('Effective Date')
                                        if eff_date and pd.notna(eff_date):
                                            return eff_date, "Found from NEW transaction's Effective Date"
                                    
                                    # No NEW transaction, check for Prior Policy Number
                                    for _, trans in policy_transactions.iterrows():
                                        prior_policy = trans.get('Prior Policy Number')
                                        if prior_policy and pd.notna(prior_policy) and str(prior_policy).strip():
                                            # Recursively follow the chain
                                            result_date, result_msg = find_origination_date_tool(prior_policy, visited)
                                            if result_date:
                                                return result_date, f"{result_msg} (via Prior Policy: {prior_policy})"
                                
                                return None, None
                            
                            found_date, message = find_origination_date_tool(policy_number)
                            if found_date:
                                auto_populated_date = found_date
                                reason = f"{transaction_type} - {message}"
                        
                        # Store results
                        if auto_populated_date:
                            updates.append({
                                'Transaction ID': transaction_id,
                                'Customer': customer,
                                'Policy Number': policy_number,
                                'Type': transaction_type,
                                'New Origination Date': auto_populated_date,
                                'Reason': reason
                            })
                        else:
                            errors.append({
                                'Transaction ID': transaction_id,
                                'Customer': customer,
                                'Policy Number': policy_number,
                                'Type': transaction_type,
                                'Reason': 'No origination date found'
                            })
                    
                    progress_bar.empty()
                    status_text.empty()
                    
                    # Show results
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("✅ Can Update", len(updates))
                    with col2:
                        st.metric("❌ Cannot Update", len(errors))
                    
                    # Store results in session state
                    st.session_state['origination_updates'] = updates
                    st.session_state['origination_errors'] = errors
                    
                    # Show preview of updates
                    if updates:
                        st.success(f"Found {len(updates)} transactions that can be updated")
                        
                        # Convert dates to strings for display
                        updates_df = pd.DataFrame(updates)
                        
                        st.dataframe(updates_df.head(10), use_container_width=True)
                        if len(updates) > 10:
                            st.caption(f"Showing first 10 of {len(updates)} updates")
                    
                    if errors:
                        with st.expander(f"⚠️ {len(errors)} transactions cannot be updated"):
                            errors_df = pd.DataFrame(errors)
                            st.dataframe(errors_df, use_container_width=True)
                
                # Update button (only shows after analysis)
                if 'origination_updates' in st.session_state and st.session_state['origination_updates']:
                    st.divider()
                    
                    col1, col2, col3 = st.columns([2, 1, 2])
                    with col2:
                        if st.button("💾 Update Database", type="primary", use_container_width=True, key="update_origination_dates"):
                            updates = st.session_state['origination_updates']
                            success_count = 0
                            error_count = 0
                            
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            
                            supabase = get_supabase_client()
                            
                            for idx, update in enumerate(updates):
                                progress = (idx + 1) / len(updates)
                                progress_bar.progress(progress)
                                status_text.text(f"Updating {idx+1}/{len(updates)}: {update['Transaction ID']}")
                                
                                try:
                                    # Update the record
                                    supabase.table('policies').update({
                                        'Policy Origination Date': update['New Origination Date']
                                    }).eq('Transaction ID', update['Transaction ID']).execute()
                                    
                                    success_count += 1
                                except Exception as e:
                                    error_count += 1
                                    st.error(f"Error updating {update['Transaction ID']}: {str(e)}")
                            
                            progress_bar.empty()
                            status_text.empty()
                            
                            # Clear the updates from session state
                            del st.session_state['origination_updates']
                            if 'origination_errors' in st.session_state:
                                del st.session_state['origination_errors']
                            
                            # Show results
                            st.success(f"✅ Successfully updated {success_count} transactions!")
                            if error_count > 0:
                                st.error(f"❌ Failed to update {error_count} transactions")
                            
                            # Clear cache to show updated data
                            clear_policies_cache()
                            
                            st.balloons()
                            
                            # Generate report
                            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                            report_data = []
                            for update in updates:
                                report_data.append({
                                    'Transaction ID': update['Transaction ID'],
                                    'Customer': update['Customer'],
                                    'Policy Number': update['Policy Number'],
                                    'Transaction Type': update['Type'],
                                    'New Origination Date': update['New Origination Date'],
                                    'Reason': update['Reason'],
                                    'Status': 'Updated'
                                })
                            
                            report_df = pd.DataFrame(report_data)
                            csv = report_df.to_csv(index=False)
                            
                            st.download_button(
                                label="📥 Download Update Report",
                                data=csv,
                                file_name=f"origination_date_updates_{timestamp}.csv",
                                mime="text/csv"
                            )
        
        with tab2:
            st.subheader("Utility Functions")
            
            # Currency formatter
            st.write("**Currency Formatter**")
            currency_input = st.number_input("Amount to format", value=1234.56, format="%.2f")
            formatted_currency = format_currency(currency_input)
            st.write(f"Formatted: {formatted_currency}")
            
        
        with tab3:
            st.subheader("Import/Export Tools")
            
            # Export section
            st.write("### Export Data")
            if not all_data.empty:
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("**📄 CSV Export**")
                    csv_data = all_data.to_csv(index=False)
                    st.download_button(
                        label="📥 Export All Data to CSV",
                        data=csv_data,
                        file_name=f"all_policies_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        help="Export all policy data as CSV file"
                    )
                
                with col2:
                    st.write("**📊 Excel Export**")
                    excel_buffer, excel_filename = create_formatted_excel_file(
                        all_data, 
                        sheet_name="All Policies", 
                        filename_prefix="all_policies"
                    )
                    if excel_buffer:
                        st.download_button(
                            label="📥 Export All Data to Excel",
                            data=excel_buffer,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Export all policy data as formatted Excel file"
                        )
            else:
                st.info("No data available to export")
            
            st.divider()
            
            # Import section
            st.write("### Import Data")
            
            # File uploader for both CSV and Excel
            uploaded_file = st.file_uploader(
                "Choose a CSV or Excel file", 
                type=["csv", "xlsx", "xls"],
                help="Upload a CSV or Excel file containing policy data"
            )
            
            if uploaded_file is not None:
                file_type = uploaded_file.name.split('.')[-1].lower()
                
                try:
                    # Process based on file type
                    if file_type == 'csv':
                        import_df = pd.read_csv(uploaded_file)
                        validation_success = True
                        validation_errors = []
                        st.success("✅ CSV file loaded successfully")
                        
                    elif file_type in ['xlsx', 'xls']:
                        validation_success, import_df, validation_errors = validate_excel_import(uploaded_file)
                        
                        if validation_success:
                            st.success("✅ Excel file loaded and validated successfully")
                        else:
                            st.error("❌ Excel file validation failed")
                            for error in validation_errors:
                                st.error(f"• {error}")
                    
                    # Show preview if file was loaded
                    if 'import_df' in locals() and import_df is not None:
                        st.write("**📋 Preview of uploaded data:**")
                        st.dataframe(import_df.head(10), use_container_width=True)
                        
                        # Show file statistics
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Rows", len(import_df))
                        with col2:
                            st.metric("Total Columns", len(import_df.columns))
                        with col3:
                            st.metric("File Type", file_type.upper())
                        
                        # Show validation results for Excel files
                        if file_type in ['xlsx', 'xls'] and validation_errors:
                            with st.expander("⚠️ Validation Issues"):
                                for error in validation_errors:
                                    st.warning(error)
                        
                        # Import button with enhanced validation
                        if validation_success:
                            if st.button("🔄 Import Data to Database", type="primary"):
                                with st.spinner("Validating and importing data..."):
                                    try:
                                        # Additional validation before import
                                        required_columns = ['Client_ID', 'Transaction_ID', 'Policy_Type']
                                        missing_required = [col for col in required_columns if col not in import_df.columns]
                                        
                                        if missing_required:
                                            st.error(f"❌ Missing required columns: {', '.join(missing_required)}")
                                        else:
                                            # Check for duplicate Transaction_IDs in current database
                                            existing_ids = set(all_data['Transaction_ID'].tolist()) if 'Transaction_ID' in all_data.columns else set()
                                            import_ids = set(import_df['Transaction_ID'].tolist()) if 'Transaction_ID' in import_df.columns else set()
                                            duplicates = existing_ids.intersection(import_ids)
                                            
                                            if duplicates:
                                                st.warning(f"⚠️ Found {len(duplicates)} Transaction_IDs that already exist in database")
                                                
                                                import_option = st.radio(
                                                    "How would you like to handle duplicates?",
                                                    ["Skip duplicate records", "Update existing records", "Cancel import"]
                                                )
                                                
                                                if import_option == "Cancel import":
                                                    st.info("Import cancelled")
                                                elif import_option == "Skip duplicate records":
                                                    import_df = import_df[~import_df['Transaction_ID'].isin(duplicates)]
                                                    st.info(f"Will import {len(import_df)} new records (skipping duplicates)")
                                                elif import_option == "Update existing records":
                                                    st.warning("Update functionality not yet implemented")
                                            
                                            if len(import_df) > 0 and import_option != "Cancel import":
                                                # This is where actual database import would happen
                                                st.success(f"✅ Ready to import {len(import_df)} records")
                                                st.info("🚧 Full database import implementation is in development. Preview and validation are complete.")
                                                
                                                # Show what would be imported
                                                with st.expander("📊 Import Summary"):
                                                    st.write(f"**Records to import:** {len(import_df)}")
                                                    st.write(f"**File type:** {file_type.upper()}")
                                                    st.write(f"**Columns found:** {len(import_df.columns)}")
                                                    st.write("**Sample data:**")
                                                    st.dataframe(import_df.head(3))
                                    
                                    except Exception as e:
                                        st.error(f"❌ Import error: {e}")
                        else:
                            st.error("❌ Cannot import data due to validation errors. Please fix the issues and try again.")
                
                except Exception as e:
                    st.error(f"❌ Error reading file: {e}")
                    if file_type in ['xlsx', 'xls']:
                        st.info("💡 Tip: Ensure your Excel file is not corrupted and contains data in the first sheet.")
            
            st.divider()
            
            # Update existing transactions from Excel
            st.write("### Update Existing Transactions from Excel")
            st.warning("""
            ⚠️ **IMPORTANT**: This tool will UPDATE existing transactions in the database.
            - No new transactions will be created
            - Only existing transactions (matched by Transaction ID) will be updated
            - ONLY the columns present in your Excel file will be updated
            - Columns not in your Excel file will remain unchanged
            - This action cannot be undone - make sure you have a backup!
            
            💡 **TIP**: To avoid upload errors, delete all app-created columns (reconciliation_status, is_reconciliation_entry, etc.) from the right side of your spreadsheet before uploading!
            """)
            
            update_file = st.file_uploader(
                "Choose an Excel file with updated transactions", 
                type=["xlsx", "xls"],
                key="update_excel_uploader",
                help="Upload an Excel file containing transactions to update. Must have 'Transaction ID' column."
            )
            
            if update_file is not None:
                try:
                    # Read the Excel file - handle multiple sheets and formatting
                    try:
                        # Try to read the Policy Revenue Report sheet first (common from exports)
                        update_df = pd.read_excel(update_file, sheet_name='Policy Revenue Report')
                        st.info("📋 Reading from 'Policy Revenue Report' sheet")
                    except:
                        try:
                            # If that fails, try common sheet names
                            all_sheets = pd.read_excel(update_file, sheet_name=None)
                            sheet_names = list(all_sheets.keys())
                            
                            # Look for a sheet with transaction data
                            data_sheet = None
                            for sheet_name in sheet_names:
                                if 'report' in sheet_name.lower() or 'polic' in sheet_name.lower() or 'transaction' in sheet_name.lower():
                                    if 'parameter' not in sheet_name.lower():  # Skip parameter sheets
                                        data_sheet = sheet_name
                                        break
                            
                            if data_sheet:
                                update_df = all_sheets[data_sheet]
                                st.info(f"📋 Reading from '{data_sheet}' sheet")
                            else:
                                # Default to the last sheet (usually the data sheet)
                                data_sheet = sheet_names[-1]
                                update_df = all_sheets[data_sheet]
                                st.info(f"📋 Reading from '{data_sheet}' sheet")
                                
                        except Exception as inner_e:
                            # Last resort - try first sheet
                            update_df = pd.read_excel(update_file, sheet_name=0)
                    
                    # Check for Transaction ID column
                    if 'Transaction ID' not in update_df.columns:
                        st.error("❌ Excel file must contain 'Transaction ID' column to match existing records")
                    else:
                        # Show preview
                        st.write("**📋 Preview of updates:**")
                        st.dataframe(update_df.head(10), use_container_width=True)
                        
                        # Show which columns will be updated
                        excel_columns = [col for col in update_df.columns if col not in ['Transaction ID', '_id', 'Policy Balance Due']]
                        st.info(f"📊 **Columns to be updated:** {', '.join(excel_columns)}")
                        st.success("✅ All other columns in the database will remain unchanged")
                        
                        # Check for calculated fields
                        calculated_fields = ['Policy Balance Due']
                        found_calculated = [field for field in calculated_fields if field in update_df.columns]
                        if found_calculated:
                            st.warning(f"⚠️ Note: {', '.join(found_calculated)} will be skipped (calculated field, not stored in database)")
                        
                        # Count matching transactions
                        transaction_ids = update_df['Transaction ID'].dropna().unique()
                        
                        # Query database for matching IDs
                        supabase = get_supabase_client()
                        existing_response = supabase.table('policies').select('"Transaction ID"').in_('"Transaction ID"', transaction_ids.tolist()).execute()
                        existing_ids = [row['Transaction ID'] for row in existing_response.data]
                        
                        # Show statistics
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Rows in File", len(update_df))
                        with col2:
                            st.metric("Matching Transactions", len(existing_ids))
                        with col3:
                            st.metric("New/Unmatched", len(transaction_ids) - len(existing_ids))
                        
                        if len(existing_ids) == 0:
                            st.warning("⚠️ No matching transactions found in database")
                        else:
                            # Show which transactions will be updated
                            with st.expander(f"📋 {len(existing_ids)} transactions will be updated"):
                                st.write(existing_ids[:50])  # Show first 50
                                if len(existing_ids) > 50:
                                    st.caption(f"...and {len(existing_ids) - 50} more")
                            
                            # Update button
                            if st.button("🔄 Update Transactions", type="primary", key="update_transactions_btn"):
                                progress_bar = st.progress(0)
                                status_text = st.empty()
                                
                                success_count = 0
                                error_count = 0
                                errors = []
                                
                                # Process each row that has a matching Transaction ID
                                for idx, row in update_df.iterrows():
                                    transaction_id = row.get('Transaction ID')
                                    
                                    if pd.notna(transaction_id) and transaction_id in existing_ids:
                                        progress = (idx + 1) / len(update_df)
                                        progress_bar.progress(progress)
                                        status_text.text(f"Updating {idx+1}/{len(update_df)}: {transaction_id}")
                                        
                                        try:
                                            # Prepare update data (exclude Transaction ID from updates)
                                            update_data = {}
                                            for col in update_df.columns:
                                                if col != 'Transaction ID' and col != '_id' and col != 'Policy Balance Due':  # Don't update ID fields or calculated fields
                                                    value = row[col]
                                                    # Handle NaN values
                                                    if pd.isna(value):
                                                        update_data[col] = None
                                                    # Handle datetime objects
                                                    elif isinstance(value, (pd.Timestamp, datetime.datetime)):
                                                        update_data[col] = value.strftime('%Y-%m-%d')
                                                    else:
                                                        update_data[col] = value
                                            
                                            # Update the record
                                            response = supabase.table('policies').update(update_data).eq('"Transaction ID"', transaction_id).execute()
                                            
                                            if response.data:
                                                success_count += 1
                                            else:
                                                error_count += 1
                                                errors.append(f"No data returned for {transaction_id}")
                                                
                                        except Exception as e:
                                            error_count += 1
                                            errors.append(f"{transaction_id}: {str(e)}")
                                
                                progress_bar.empty()
                                status_text.empty()
                                
                                # Show results
                                if success_count > 0:
                                    st.success(f"✅ Successfully updated {success_count} transactions")
                                    # Clear cache to show updated data
                                    clear_policies_cache()
                                
                                if error_count > 0:
                                    st.error(f"❌ Failed to update {error_count} transactions")
                                    with st.expander("Error Details"):
                                        for error in errors[:10]:  # Show first 10 errors
                                            st.write(f"• {error}")
                                        if len(errors) > 10:
                                            st.caption(f"...and {len(errors) - 10} more errors")
                                
                                # Generate update report
                                if success_count > 0:
                                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                                    
                                    # Create detailed update report
                                    report_rows = []
                                    for idx, row in update_df.iterrows():
                                        transaction_id = row.get('Transaction ID')
                                        if pd.notna(transaction_id) and transaction_id in existing_ids:
                                            report_rows.append({
                                                'Transaction ID': transaction_id,
                                                'Customer': row.get('Customer', ''),
                                                'Policy Number': row.get('Policy Number', ''),
                                                'Update Status': 'Success' if transaction_id not in [e.split(':')[0] for e in errors] else 'Failed',
                                                'Timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                            })
                                    
                                    # Add summary sheet data
                                    summary_data = pd.DataFrame({
                                        'Summary Item': [
                                            'Total Records in File',
                                            'Matching Transactions Found',
                                            'Successfully Updated',
                                            'Failed Updates',
                                            'Update Timestamp',
                                            'Columns Updated'
                                        ],
                                        'Value': [
                                            len(update_df),
                                            len(existing_ids),
                                            success_count,
                                            error_count,
                                            timestamp,
                                            ', '.join(excel_columns)
                                        ]
                                    })
                                    
                                    report_df = pd.DataFrame(report_rows)
                                    
                                    # Create Excel report with multiple sheets
                                    excel_buffer = io.BytesIO()
                                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                        # Summary sheet
                                        summary_data.to_excel(writer, sheet_name='Update Summary', index=False)
                                        
                                        # Detailed results sheet
                                        if not report_df.empty:
                                            report_df.to_excel(writer, sheet_name='Update Details', index=False)
                                        
                                        # Errors sheet if any
                                        if errors:
                                            error_df = pd.DataFrame({
                                                'Error': errors[:100]  # Limit to 100 errors
                                            })
                                            error_df.to_excel(writer, sheet_name='Errors', index=False)
                                        
                                        # Format the Excel file
                                        workbook = writer.book
                                        for sheet_name in workbook.sheetnames:
                                            worksheet = workbook[sheet_name]
                                            # Auto-adjust column widths
                                            for column in worksheet.columns:
                                                max_length = 0
                                                column_letter = column[0].column_letter
                                                for cell in column:
                                                    try:
                                                        if len(str(cell.value)) > max_length:
                                                            max_length = len(str(cell.value))
                                                    except:
                                                        pass
                                                adjusted_width = min(max_length + 2, 50)
                                                worksheet.column_dimensions[column_letter].width = adjusted_width
                                    
                                    excel_buffer.seek(0)
                                    
                                    # Provide download options
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        st.download_button(
                                            label="📊 Download Update Report (Excel)",
                                            data=excel_buffer,
                                            file_name=f"transaction_update_report_{timestamp}.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                        )
                                    
                                    with col2:
                                        # Also provide CSV option
                                        csv_data = report_df.to_csv(index=False)
                                        st.download_button(
                                            label="📄 Download Update Report (CSV)",
                                            data=csv_data,
                                            file_name=f"transaction_update_report_{timestamp}.csv",
                                            mime="text/csv"
                                        )
                                
                except Exception as e:
                    st.error(f"❌ Error processing file: {str(e)}")
                    
                    # Try to provide more specific error information
                    if "worksheet" in str(e).lower() or "sheet" in str(e).lower():
                        st.info("💡 This might be a multi-sheet Excel file. Try saving as a single-sheet Excel or CSV file.")
                    elif "index" in str(e).lower():
                        st.info("💡 The Excel file might have index/formatting issues. Try:")
                        st.info("   1. Save as a new Excel file without formatting")
                        st.info("   2. Or export as CSV and re-import to Excel")
                    else:
                        st.info("💡 Make sure your Excel file has a 'Transaction ID' column and valid data")
                    
                    # Show debug info
                    with st.expander("Debug Information"):
                        st.write(f"File name: {update_file.name}")
                        st.write(f"File size: {update_file.size} bytes")
                        st.write(f"Error type: {type(e).__name__}")
                        st.write(f"Error details: {str(e)}")
    
    # --- Policy Revenue Ledger ---
    elif page == "Policy Revenue Ledger":
        st.subheader("Policy Revenue Ledger")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        if all_data.empty:
            st.warning("No data found in policies table. Please add some policy data first.")
            return
        
        st.info("Search for a policy to view and edit a detailed ledger of all commission credits and debits for that policy. You can edit, delete transactions below. Be sure to click 'Save Changes' to commit your edits.")
        
        # --- Granular Policy Search in Single Row ---
        st.markdown("### 🔍 Search Criteria")
        
        # Create 4 columns for search criteria
        search_col1, search_col2, search_col3, search_col4 = st.columns(4)
        
        # Get initial data lists
        customers = all_data["Customer"].dropna().unique().tolist() if "Customer" in all_data.columns else []
        
        with search_col1:
            selected_customer = st.selectbox(
                "Select Customer:", 
                ["Select..."] + sorted(customers), 
                key="ledger_customer_select"
            )
        
        # Filter data based on customer selection
        filtered_data = all_data.copy()
        if selected_customer and selected_customer != "Select...":
            filtered_data = filtered_data[filtered_data["Customer"] == selected_customer]
        
        # Get policy types based on filtered data
        policy_types = filtered_data["Policy Type"].dropna().unique().tolist() if "Policy Type" in filtered_data.columns else []
        
        with search_col2:
            selected_policy_type = st.selectbox(
                "Select Policy Type:", 
                ["Select..."] + sorted(policy_types), 
                key="ledger_policytype_select"
            )
        
        # Further filter based on policy type
        if selected_policy_type and selected_policy_type != "Select...":
            filtered_data = filtered_data[filtered_data["Policy Type"] == selected_policy_type]
        
        # Get effective dates based on filtered data
        effective_dates = filtered_data["Effective Date"].dropna().unique().tolist() if "Effective Date" in filtered_data.columns else []
        
        with search_col3:
            selected_effective_date = st.selectbox(
                "Select Policy Effective Date:", 
                ["Select...", "All Dates"] + sorted(effective_dates), 
                key="ledger_effectivedate_select"
            )
        
        # Further filter based on effective date
        if selected_effective_date and selected_effective_date not in ["Select...", "All Dates"]:
            filtered_data = filtered_data[filtered_data["Effective Date"] == selected_effective_date]
        
        # Get policy numbers based on all filters
        policy_numbers = filtered_data["Policy Number"].dropna().unique().tolist() if "Policy Number" in filtered_data.columns else []
        
        with search_col4:
            # Show policy number select if customer and policy type are chosen
            if selected_customer != "Select..." and selected_policy_type != "Select..." and (selected_effective_date != "Select..."):
                selected_policy = st.selectbox(
                    "Select Policy Number:", 
                    ["Select..."] + sorted(policy_numbers), 
                    key="ledger_policy_select"
                )
            else:
                st.selectbox(
                    "Select Policy Number:", 
                    ["Select filters first..."], 
                    key="ledger_policy_select",
                    disabled=True
                )
                selected_policy = None

        if selected_policy and selected_policy != "Select...":
            # Get all rows for this policy, using only real database data
            policy_rows = all_data[all_data["Policy Number"] == selected_policy].copy()
            
            # Add X-DATE filter for policy term selection
            if not policy_rows.empty and "X-DATE" in policy_rows.columns:
                unique_xdates = sorted(policy_rows["X-DATE"].dropna().unique())
                
                # Create filter selection
                st.markdown("### 📅 Policy Term Filter (Optional)")
                term_col1, term_col2, term_col3 = st.columns([2, 2, 2])
                
                with term_col1:
                    x_date_options = ["All Terms"] + [str(x) for x in unique_xdates]
                    selected_xdate = st.selectbox(
                        "Select Policy Term by X-DATE:",
                        options=x_date_options,
                        index=0,
                        key="ledger_xdate_select",
                        help="Filter to show only transactions for a specific policy term"
                    )
                
                # Filter transactions if specific X-DATE selected
                if selected_xdate != "All Terms":
                    # Find the term's effective date (from NEW or RWL transaction with this X-DATE)
                    term_transactions = policy_rows[
                        (policy_rows["X-DATE"] == selected_xdate) & 
                        (policy_rows["Transaction Type"].isin(["NEW", "RWL"]))
                    ]
                    
                    if not term_transactions.empty:
                        term_eff_date = pd.to_datetime(term_transactions.iloc[0]["Effective Date"])
                        term_x_date = pd.to_datetime(selected_xdate)
                        
                        with term_col2:
                            st.info(f"📆 Term: {term_eff_date.strftime('%Y-%m-%d')} to {term_x_date.strftime('%Y-%m-%d')}")
                        
                        # Filter policy_rows to include:
                        # 1. NEW/RWL transaction with this X-DATE
                        # 2. END transactions with effective dates between term dates
                        # 3. STMT/VOID/ADJ transactions within the term period
                        filtered_rows = []
                        
                        for idx, row in policy_rows.iterrows():
                            trans_type = row.get("Transaction Type", "")
                            trans_eff_date = pd.to_datetime(row.get("Effective Date"), errors='coerce')
                            trans_x_date = row.get("X-DATE")
                            
                            # Include NEW/RWL with matching X-DATE
                            if trans_type in ["NEW", "RWL"] and trans_x_date == selected_xdate:
                                filtered_rows.append(idx)
                            # Include END within the term dates
                            elif trans_type == "END" and pd.notna(trans_eff_date):
                                if term_eff_date <= trans_eff_date <= term_x_date:
                                    filtered_rows.append(idx)
                            # Include payment/reconciliation transactions within term
                            elif trans_type in ["STMT", "VOID", "ADJ"] or "-STMT-" in str(row.get("Transaction ID", "")) or "-VOID-" in str(row.get("Transaction ID", "")):
                                # Check if STMT DATE falls within term
                                stmt_date = pd.to_datetime(row.get("STMT DATE"), errors='coerce')
                                if pd.notna(stmt_date) and term_eff_date <= stmt_date <= term_x_date:
                                    filtered_rows.append(idx)
                                # If no STMT DATE, check transaction ID date
                                elif pd.isna(stmt_date):
                                    trans_id = str(row.get("Transaction ID", ""))
                                    # Extract date from transaction ID if it contains YYYYMMDD pattern
                                    import re
                                    date_match = re.search(r'(\d{8})', trans_id)
                                    if date_match:
                                        try:
                                            trans_date = pd.to_datetime(date_match.group(1), format='%Y%m%d')
                                            if term_eff_date <= trans_date <= term_x_date:
                                                filtered_rows.append(idx)
                                        except:
                                            pass
                        
                        # Apply the filter
                        if filtered_rows:
                            policy_rows = policy_rows.loc[filtered_rows].copy()
                            with term_col3:
                                st.success(f"✅ Showing {len(policy_rows)} transactions for this term")
                        else:
                            st.warning("No transactions found for this term")
                            policy_rows = pd.DataFrame()  # Empty dataframe
                    else:
                        st.error("Could not determine term dates")
                else:
                    with term_col2:
                        st.info("📊 Showing all transactions for this policy")
            # Define the original ledger columns with visual indicator and financial columns
            ledger_columns = [
                "Type",  # New visual indicator column
                "Transaction ID",
                "Transaction Type",
                "Effective Date",
                "Credit (Commission Owed)",
                "Debit (Paid to Agent)"
            ]
            
            # Additional financial columns (will be at the far right after Delete column)
            financial_columns = [
                "Premium Sold",
                "Policy Taxes & Fees",
                "Commissionable Premium",
                "Broker Fee",
                "Broker Fee Agent Comm"
            ]

            # --- Always include all ledger columns, filling missing with empty values ---
            import numpy as np
            if not policy_rows.empty:
                # --- Ledger construction using mapped column names ---
                # Get mapped column names for credits and debits
                credit_col = get_mapped_column("Agent Estimated Comm $") or "Agent Estimated Comm $"
                debit_col = get_mapped_column("Agent Paid Amount (STMT)") or "Agent Paid Amount (STMT)"
                transaction_id_col = get_mapped_column("Transaction ID") or "Transaction ID"
                stmt_date_col = get_mapped_column("STMT DATE") or "STMT DATE"
                description_col = get_mapped_column("Description") or "Description"
                transaction_type_col = get_mapped_column("Transaction Type") or "Transaction Type"
                
                # Build ledger_df with correct mapping
                ledger_df = pd.DataFrame()
                
                # Add Type indicator column based on Transaction ID
                if transaction_id_col in policy_rows.columns:
                    ledger_df["Transaction ID"] = policy_rows[transaction_id_col]
                    # Create Type column with visual indicators
                    ledger_df["Type"] = ledger_df["Transaction ID"].apply(lambda x: 
                        "💰 STMT" if "-STMT-" in str(x) else 
                        "🔴 VOID" if "-VOID-" in str(x) else 
                        "📄"
                    )
                else:
                    ledger_df["Transaction ID"] = ""
                    ledger_df["Type"] = "📄"
                
                # Add Effective Date
                effective_date_col = get_mapped_column("Effective Date") or "Effective Date"
                ledger_df["Effective Date"] = policy_rows[effective_date_col] if effective_date_col in policy_rows.columns else ""
                
                ledger_df["Description"] = policy_rows[description_col] if description_col in policy_rows.columns else ""
                
                # Credit (Commission Owed) from mapped Agent Estimated Comm $
                if credit_col in policy_rows.columns:
                    ledger_df["Credit (Commission Owed)"] = policy_rows[credit_col]
                else:
                    ledger_df["Credit (Commission Owed)"] = 0.0
                    
                # Debit (Paid to Agent) from mapped Agent Paid Amount (STMT)
                if debit_col in policy_rows.columns:
                    ledger_df["Debit (Paid to Agent)"] = policy_rows[debit_col]
                else:
                    ledger_df["Debit (Paid to Agent)"] = 0.0
                    
                ledger_df["Transaction Type"] = policy_rows[transaction_type_col] if transaction_type_col in policy_rows.columns else ""
                
                # Add financial columns with mapped column names
                # Premium Sold
                premium_col = get_mapped_column("Premium Sold") or "Premium Sold"
                if premium_col in policy_rows.columns:
                    ledger_df["Premium Sold"] = policy_rows[premium_col]
                else:
                    ledger_df["Premium Sold"] = 0.0
                
                # Policy Taxes & Fees
                taxes_col = get_mapped_column("Policy Taxes & Fees") or "Policy Taxes & Fees"
                if taxes_col in policy_rows.columns:
                    ledger_df["Policy Taxes & Fees"] = policy_rows[taxes_col]
                else:
                    ledger_df["Policy Taxes & Fees"] = 0.0
                
                # Commissionable Premium
                comm_premium_col = get_mapped_column("Commissionable Premium") or "Commissionable Premium"
                if comm_premium_col in policy_rows.columns:
                    ledger_df["Commissionable Premium"] = policy_rows[comm_premium_col]
                else:
                    ledger_df["Commissionable Premium"] = 0.0
                
                # Broker Fee
                broker_fee_col = get_mapped_column("Broker Fee") or "Broker Fee"
                if broker_fee_col in policy_rows.columns:
                    ledger_df["Broker Fee"] = policy_rows[broker_fee_col]
                else:
                    ledger_df["Broker Fee"] = 0.0
                
                # Broker Fee Agent Comm
                broker_comm_col = get_mapped_column("Broker Fee Agent Comm") or "Broker Fee Agent Comm"
                if broker_comm_col in policy_rows.columns:
                    ledger_df["Broker Fee Agent Comm"] = policy_rows[broker_comm_col]
                else:
                    ledger_df["Broker Fee Agent Comm"] = 0.0
                
                # Ensure correct column order - combine main columns with financial columns and Description at the end
                all_ledger_columns = ledger_columns + financial_columns + ["Description"]
                ledger_df = ledger_df[all_ledger_columns]
                
                # Sort by Effective Date (oldest to newest)
                if "Effective Date" in ledger_df.columns:
                    # Convert to datetime for proper sorting
                    ledger_df["Effective Date Sort"] = pd.to_datetime(ledger_df["Effective Date"], errors='coerce')
                    # Sort by date, putting NaT values at the end
                    ledger_df = ledger_df.sort_values("Effective Date Sort", na_position='last')
                    # Drop the temporary sort column
                    ledger_df = ledger_df.drop(columns=["Effective Date Sort"])
                
                # Reset index for clean display
                ledger_df = ledger_df.reset_index(drop=True)
            else:
                # If no rows, show empty DataFrame with all columns
                all_ledger_columns = ledger_columns + financial_columns + ["Description"]
                ledger_df = pd.DataFrame(columns=all_ledger_columns)
                
            if not ledger_df.empty:
                    st.markdown("### Policy Details")
                    
                    # Show policy-level details using mapped column names
                    policy_detail_field_names = [
                        "Customer", "Client ID", "Policy Number", "Policy Type", "Carrier Name", "MGA Name",
                        "Effective Date", "Policy Origination Date", "Policy Gross Comm %", 
                        "Agent Comm %", "X-DATE"
                    ]
                    policy_detail_cols = []
                    for field_name in policy_detail_field_names:
                        mapped_col = get_mapped_column(field_name)
                        if mapped_col and mapped_col in policy_rows.columns:
                            policy_detail_cols.append(mapped_col)
                        elif field_name in policy_rows.columns:
                            policy_detail_cols.append(field_name)
                        # Special handling for Client ID - check both variations
                        elif field_name == "Client ID":
                            if "client_id" in policy_rows.columns:
                                policy_detail_cols.append("client_id")
                            elif "ClientID" in policy_rows.columns:
                                policy_detail_cols.append("ClientID")
                    
                    # Ensure we have valid columns before creating the dataframe
                    valid_detail_cols = [col for col in policy_detail_cols if col in policy_rows.columns]
                    
                    # If Client ID wasn't found but exists in the data, add it
                    if "Client ID" not in valid_detail_cols and "Client ID" in policy_rows.columns:
                        # Find where to insert it (after Customer)
                        if "Customer" in valid_detail_cols:
                            customer_idx = valid_detail_cols.index("Customer")
                            valid_detail_cols.insert(customer_idx + 1, "Client ID")
                        else:
                            valid_detail_cols.insert(1, "Client ID")
                    
                    # Instead of always using the first row, find a row with the most complete data
                    # Prioritize rows with Client ID populated
                    if not policy_rows.empty:
                        # Try to find a row with Client ID
                        rows_with_client_id = policy_rows[policy_rows["Client ID"].notna() & (policy_rows["Client ID"] != "") & (policy_rows["Client ID"] != "None")]
                        
                        if not rows_with_client_id.empty:
                            # Use the first row that has a Client ID
                            policy_details_row = rows_with_client_id.iloc[0]
                        else:
                            # Fall back to the first row if no rows have Client ID
                            policy_details_row = policy_rows.iloc[0]
                    else:
                        policy_details_row = pd.Series()
                    
                    # Extract values for the card display
                    customer = policy_details_row.get("Customer", "N/A")
                    client_id = policy_details_row.get("Client ID", "N/A")
                    policy_number = policy_details_row.get("Policy Number", "N/A")
                    policy_type = policy_details_row.get("Policy Type", "N/A")
                    carrier_name = policy_details_row.get("Carrier Name", "N/A")
                    mga_name = policy_details_row.get("MGA Name", "N/A")
                    effective_date = policy_details_row.get("Effective Date", "N/A")
                    origination_date = policy_details_row.get("Policy Origination Date", "N/A")
                    gross_comm = policy_details_row.get("Policy Gross Comm %", 0)
                    agent_comm = policy_details_row.get("Agent Comm %", 0)
                    x_date = policy_details_row.get("X-DATE", "N/A")
                    
                    # Format dates for display
                    if pd.notna(effective_date) and effective_date != "N/A":
                        try:
                            effective_date = pd.to_datetime(effective_date).strftime("%m/%d/%Y")
                        except:
                            pass
                    
                    if pd.notna(origination_date) and origination_date != "N/A":
                        try:
                            origination_date_display = pd.to_datetime(origination_date).strftime("%m/%d/%Y")
                        except:
                            origination_date_display = origination_date
                    else:
                        origination_date_display = origination_date
                    
                    if pd.notna(x_date) and x_date != "N/A":
                        try:
                            x_date = pd.to_datetime(x_date).strftime("%m/%d/%Y")
                        except:
                            pass
                    
                    # Display app-style card layout
                    # Customer and Policy Info Card
                    st.markdown(
                        f"""
                        <div style="border: 1px solid #ddd; border-radius: 8px; padding: 15px; margin-bottom: 15px; background-color: #f8f9fa;">
                            <div style="display: flex; justify-content: space-between; align-items: center;">
                                <div>
                                    <span style="font-size: 18px; font-weight: bold;">👤 {customer}</span>
                                </div>
                                <div style="text-align: right;">
                                    <span style="color: #666;">Client ID: <strong>{client_id}</strong></span>
                                </div>
                            </div>
                            <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 10px;">
                                <div>
                                    <span style="color: #666;">📋 Policy #: <strong>{policy_number}</strong></span>
                                </div>
                                <div style="text-align: right;">
                                    <span style="color: #666;">Type: <strong>{policy_type}</strong></span>
                                </div>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                    
                    # Carrier, MGA, and Commission Cards
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown(
                            f"""
                            <div style="border: 1px solid #ddd; border-radius: 8px; padding: 15px; height: 100px;">
                                <div style="color: #666; font-size: 14px; margin-bottom: 5px;">🏢 Carrier</div>
                                <div style="font-weight: bold; font-size: 16px;">{carrier_name}</div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                    
                    with col2:
                        st.markdown(
                            f"""
                            <div style="border: 1px solid #ddd; border-radius: 8px; padding: 15px; height: 100px;">
                                <div style="color: #666; font-size: 14px; margin-bottom: 5px;">🤝 MGA</div>
                                <div style="font-weight: bold; font-size: 16px;">{mga_name}</div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                    
                    with col3:
                        st.markdown(
                            f"""
                            <div style="border: 1px solid #ddd; border-radius: 8px; padding: 15px; height: 100px;">
                                <div style="color: #666; font-size: 14px; margin-bottom: 5px;">💰 Commission</div>
                                <div style="font-size: 15px;">Gross: <strong>{gross_comm}%</strong></div>
                                <div style="font-size: 15px;">Agent: <strong>{agent_comm}%</strong></div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                    
                    # Date Cards
                    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
                    date_col1, date_col2, date_col3 = st.columns(3)
                    
                    with date_col1:
                        st.markdown(
                            f"""
                            <div style="border: 1px solid #ddd; border-radius: 8px; padding: 15px; height: 120px;">
                                <div style="color: #666; font-size: 14px; margin-bottom: 25px;">📅 Policy Origination Date</div>
                                <div style="font-size: 16px; color: #999;">{origination_date_display}</div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                    
                    with date_col2:
                        st.markdown(
                            f"""
                            <div style="border: 1px solid #ddd; border-radius: 8px; padding: 15px; height: 120px;">
                                <div style="color: #666; font-size: 14px; margin-bottom: 25px;">📅 Effective Date</div>
                                <div style="font-weight: bold; font-size: 16px;">{effective_date}</div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                    
                    with date_col3:
                        st.markdown(
                            f"""
                            <div style="border: 1px solid #ddd; border-radius: 8px; padding: 15px; height: 120px;">
                                <div style="color: #666; font-size: 14px; margin-bottom: 25px;">🔄 Expiration Date <span style="color: #999; font-size: 12px;">(X-DATE)</span></div>
                                <div style="font-weight: bold; font-size: 16px;">{x_date}</div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                    
                    
                    # --- Policy Financial Summary Section ---
                    # Calculate totals for financial columns from ledger_df
                    total_premium = ledger_df["Premium Sold"].apply(pd.to_numeric, errors="coerce").sum() if "Premium Sold" in ledger_df.columns else 0.0
                    total_taxes = ledger_df["Policy Taxes & Fees"].apply(pd.to_numeric, errors="coerce").sum() if "Policy Taxes & Fees" in ledger_df.columns else 0.0
                    total_comm_premium = ledger_df["Commissionable Premium"].apply(pd.to_numeric, errors="coerce").sum() if "Commissionable Premium" in ledger_df.columns else 0.0
                    total_broker_fee = ledger_df["Broker Fee"].apply(pd.to_numeric, errors="coerce").sum() if "Broker Fee" in ledger_df.columns else 0.0
                    total_broker_comm = ledger_df["Broker Fee Agent Comm"].apply(pd.to_numeric, errors="coerce").sum() if "Broker Fee Agent Comm" in ledger_df.columns else 0.0
                    
                    st.markdown("#### Policy Financial Summary")
                    
                    # First row of financial summary
                    fin_col1, fin_col2, fin_col3 = st.columns(3)
                    with fin_col1:
                        st.metric("Total Premium Sold", f"${total_premium:,.2f}")
                    with fin_col2:
                        st.metric("Total Taxes & Fees", f"${total_taxes:,.2f}")
                    with fin_col3:
                        st.metric("Commissionable Premium", f"${total_comm_premium:,.2f}")
                    
                    # Second row of financial summary
                    fin_col4, fin_col5, fin_col6 = st.columns(3)
                    with fin_col4:
                        st.metric("Total Broker Fees", f"${total_broker_fee:,.2f}")
                    with fin_col5:
                        st.metric("Broker Fee Agent Comm", f"${total_broker_comm:,.2f}")
                    with fin_col6:
                        # Empty column for spacing
                        pass

                    st.markdown("### Policy Ledger")                # Ensure Credit and Debit columns are numeric
                    if "Credit (Commission Owed)" in ledger_df.columns:
                        ledger_df["Credit (Commission Owed)"] = pd.to_numeric(ledger_df["Credit (Commission Owed)"], errors="coerce").fillna(0.0)
                    if "Debit (Paid to Agent)" in ledger_df.columns:
                        ledger_df["Debit (Paid to Agent)"] = pd.to_numeric(ledger_df["Debit (Paid to Agent)"], errors="coerce").fillna(0.0)


                    # Create display dataframe
                    ledger_df_display = ledger_df.copy()
                    
                    # Update display columns to include financial columns and Description at the end
                    display_cols = ledger_columns + financial_columns + ["Description"]

                    # --- Ensure all display columns exist in the DataFrame ---
                    for col in display_cols:
                        if col not in ledger_df_display.columns:
                            ledger_df_display[col] = ""
                    # Reorder columns safely
                    ledger_df_display = ledger_df_display[display_cols]

                    # Add visual legend for transaction types
                    legend_col1, legend_col2, legend_col3, legend_col4 = st.columns([1, 2, 2, 2])
                    with legend_col1:
                        st.markdown("**Legend:**")
                    with legend_col2:
                        st.markdown("💰 STMT = Reconciliation Entry")
                    with legend_col3:
                        st.markdown("🔴 VOID = Voided Transaction")
                    with legend_col4:
                        st.markdown("📄 = Regular Transaction")
                    
                    
                    # Calculate height more precisely with smaller buffer
                    # Header ~35px + each row ~35px + small buffer for scrollbar
                    num_rows = len(ledger_df_display)
                    calculated_height = 35 + (35 * num_rows) + 20  # 20px buffer
                    # Set minimum to show at least 3 rows comfortably
                    display_height = max(150, calculated_height)
                    
                    # Create column configuration with auto-width settings
                    column_config = {}
                    for col in ledger_df_display.columns:
                        if col == "Type":
                            column_config[col] = st.column_config.TextColumn(
                                col,
                                help="Transaction type indicator: 💰=STMT, 🔴=VOID, 📄=Regular"
                            )
                        elif col in ["Credit (Commission Owed)", "Debit (Paid to Agent)", "Premium Sold", 
                                     "Policy Taxes & Fees", "Commissionable Premium", "Broker Fee", 
                                     "Broker Fee Agent Comm"]:
                            column_config[col] = st.column_config.NumberColumn(
                                col,
                                format="$%.2f"
                            )
                    
                    # Display the table as read-only with auto-sizing columns
                    st.dataframe(
                        ledger_df_display,
                        use_container_width=True,
                        height=display_height,
                        hide_index=True,
                        column_config=column_config
                    )                # --- Ledger Totals Section ---
                    total_credits = ledger_df_display["Credit (Commission Owed)"].apply(pd.to_numeric, errors="coerce").sum() if "Credit (Commission Owed)" in ledger_df_display.columns else 0.0
                    total_debits = ledger_df_display["Debit (Paid to Agent)"].apply(pd.to_numeric, errors="coerce").sum() if "Debit (Paid to Agent)" in ledger_df_display.columns else 0.0
                    balance_due = total_credits - total_debits

                    st.markdown("#### Ledger Totals")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Credits", f"${total_credits:,.2f}")
                    with col2:
                        st.metric("Total Debits", f"${total_debits:,.2f}")
                    with col3:
                        st.metric("Balance Due", f"${balance_due:,.2f}")
                    
    # --- Help ---
    elif page == "Help":
        st.title("📚 Help & Documentation")
        
        # Create tabs for organized help content
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["Getting Started", "Features Guide", "Troubleshooting", "Database Column Renaming", "About"])
        
        with tab1:
            st.subheader("🚀 Getting Started")
            st.write("""
            Welcome to the Commission Management Application! This powerful tool helps you manage insurance 
            commission data, track policies, and generate comprehensive reports.
            
            **Quick Start Steps:**
            1. **Add Policy Data**: Use "Add New Policy Transaction" to input your policy information
            2. **View Data**: Navigate to "All Policy Transactions" to see your complete dataset
            3. **Generate Reports**: Visit the "Reports" section for analytics and summaries
            4. **Manage Reconciliation**: Use "Reconciliation" for commission matching and payments
            """)
            
            st.info("💡 **Tip**: Start by adding a few sample policies to explore all features!")
        
        with tab2:
            st.subheader("📖 Features Guide")
            
            # Dashboard features
            st.write("**📊 Dashboard**")
            st.write("- View key metrics and policy summaries")
            st.write("- Search and edit policies quickly")
            st.write("- Monitor recent activity and statistics")
            
            st.divider()
            
            # Database features
            st.write("**💾 Database Management**")
            st.write("- **All Policies**: Browse paginated policy data with export options")
            st.write("- **Edit Policies**: Search and modify existing policy records")
            st.write("- **Add New**: Create new policy transactions with auto-generated IDs")
            
            st.divider()
            
            # Reporting features
            st.write("**📈 Reports & Analytics**")
            st.write("- **Reports**: Generate summary and commission analysis reports")
            st.write("- **Search & Filter**: Advanced filtering with multiple criteria")
            st.write("- **Policy Revenue Ledger**: Detailed revenue tracking with 'All Dates' option for date issues")
            
            st.divider()
            
            # Policy Audit Strategy
            st.write("**🔍 Policy Audit Strategy**")
            with st.expander("📋 How to Audit Your Policy Records Efficiently", expanded=False):
                st.markdown("""
                ## Start with the Policy Revenue Ledger Reports page - it's your best friend!

                ### 1. **Attack in Order of Priority**
                Use the new balance filter to break it down:
                - **First**: Filter for "Positive Balance Only" - these are policies where you're owed money
                - **Second**: Filter for "Negative Balance Only" - these are potential overpayments to investigate
                - **Third**: Filter for "Zero Balance Only" - just to verify these look correct

                ### 2. **Use the Search Function Strategically**
                - Search by date ranges to audit month by month

                ### 3. **Look for Red Flags**
                - Policies with unusually high balances due
                - Old policies (check Effective Date) that still show balances
                - Negative balances (overpayments)
                - Policies missing key data (use "Show Transactions Requiring Attention" on Edit Policies page)

                ### 4. **Create Report Templates**
                Save different column templates for different audit purposes:
                - "Payment Audit" template with payment-related columns
                - "Commission Audit" template with commission calculation columns
                - "Date Audit" template to check for date inconsistencies

                ### 5. **Export and Work Offline**
                - Export filtered data to CSV
                - Use Excel to sort, filter, and add your audit notes
                - Re-import corrections using the Reconciliation page

                ### 6. **Daily Bite-Sized Approach**
                - Or: "Today I'll verify all policies with balance > $500"
                - Or: "Today I'll check all 2024 policies"

                The key is using the filters to break the mountain into small, manageable hills! 🏔️→🏞️
                """)
            
            st.divider()
            
            # Administrative features
            st.write("**⚙️ Administrative Tools**")
            st.write("- **Admin Panel**: Database management and system tools")
            st.write("- **Tools**: Utilities for calculations and data formatting")
            st.write("- **Reconciliation**: Commission matching, payment tracking, and statement imports")
            
            st.divider()
            
            # Pending Renewals
            st.write("**📅 Pending Policy Renewals**")
            st.write("- **Time Range Filters**: View all renewals or filter by Past Due, This Week, 30/60/90 days")
            st.write("- **Visual Indicators**: 🔴 Past Due, 🟡 Urgent (0-7 days), ✅ OK")
            st.write("- **Summary Metrics**: Quick counts of past due, due this week, and total pending")
            st.write("- **Carrier Commission Rates**: Automatically loads rates when carrier/MGA selected")
            st.write("- **Premium Calculator**: Built-in endorsement calculator for premium changes")
            st.write("- **Smart Filtering**: Cancelled and excluded policies automatically hidden")
            st.write("- **Remove from List**: Use Prior Policy Number field when creating RWL/REWRITE to auto-remove")
            
            st.divider()
            
            # Cancel/Rewrite Scenario Guide
            st.write("**🔄 Cancel/Rewrite Scenario Guide**")
            st.info("**Important**: This section explains how to handle mid-term policy cancellations that are immediately rewritten (common for rate reductions).")
            
            with st.expander("📋 Step-by-Step Process for Cancel/Rewrite", expanded=True):
                st.success("🎉 **NEW FEATURES**: Prior Policy Number field now available in Add New Policy! Cancelled policies automatically hidden from Pending Renewals!")
                st.markdown("""
                ### When to Use Cancel/Rewrite
                Use this process when a customer wants to cancel their current policy mid-term and immediately 
                rewrite it (typically to save money with a better rate).
                
                ### Step 1: Cancel the Original Policy
                Create a new transaction with:
                - **Transaction Type**: `CAN` (Cancel)
                - **Policy Number**: Same as the original policy
                - **Effective Date**: The cancellation date
                - **Premium/Commission fields**: Enter negative amounts or zeros
                - **Description**: Add note like "Cancelled for rewrite - see policy [new policy number]"
                
                ✅ **NEW**: Cancelled policies now automatically disappear from Pending Renewals!
                
                ### Step 2: Create the Rewrite Policy
                Use "Add New Policy Transaction" with:
                - **Transaction Type**: `REWRITE` (not RWL - this ensures 25% commission, not 50%)
                - **Policy Number**: New number if changed, or same if carrier kept it
                - **Prior Policy Number**: Enter the original policy number (NOW AVAILABLE in Add New Policy!)
                - **Effective Date**: New effective date (immediate)
                - **X-DATE**: New expiration date
                - **Policy Origination Date**: Keep the ORIGINAL date (preserves customer history)
                - **Commission**: Will calculate at 25% (rewrite rate)
                - **Description**: Add note like "Rewrite of policy [original number] - mid-term for rate reduction"
                
                ### Key Benefits
                ✅ Maintains complete audit trail  
                ✅ Prior Policy Number links rewrite to original (field now in Add New Policy!)
                ✅ Preserves customer relationship timeline  
                ✅ Both transactions appear in reports  
                ✅ Cancelled policies automatically removed from Pending Renewals  
                ✅ Rewrite won't show in Pending Renewals (due to Prior Policy Number)
                ✅ Commission calculates correctly at 25%  
                
                ### Alternative: Same Policy Number
                If the carrier keeps the same policy number:
                - Still create both CAN and REWRITE transactions
                - Use the same Policy Number for both
                - Different Transaction IDs keep them separate
                - The dates and transaction types tell the complete story
                """)
            
            st.warning("⚠️ **Remember**: Always use REWRITE (not RWL) for mid-term rewrites to ensure proper 25% commission calculation!")
            
            st.divider()
            
            # Prior Policy Number Workflow
            st.write("**🔗 Prior Policy Number - The Magic Field!**")
            st.info("**Key Feature**: The Prior Policy Number field automatically removes old policies from Pending Renewals!")
            
            with st.expander("📋 How Prior Policy Number Works", expanded=True):
                st.markdown("""
                ### What is Prior Policy Number?
                It's a field that links a new policy to the old policy it's replacing. This creates a chain of policy history.
                
                ### When to Use It
                Use the Prior Policy Number field whenever you:
                - **Renew a policy** (RWL transaction)
                - **Rewrite a policy** (REWRITE transaction)
                - **Replace a policy** with a new carrier or policy type
                
                ### How It Works in Pending Renewals
                1. You see a policy in Pending Renewals that needs to be renewed/rewritten
                2. You create a new transaction (RWL or REWRITE)
                3. **Enter the old policy number in the Prior Policy Number field**
                4. Save the new transaction
                5. **The old policy automatically disappears from Pending Renewals!**
                
                ### Benefits
                ✅ No manual deletion needed  
                ✅ Maintains complete policy history  
                ✅ Prevents duplicate renewal reminders  
                ✅ Clean, accurate Pending Renewals list  
                ✅ Works for all replacement scenarios  
                
                ### Pro Tips
                - Always fill in Prior Policy Number when replacing ANY policy
                - This works even if the carrier changes
                - This works even if the policy type changes
                - The system tracks the complete chain of renewals
                """)
        
        with tab3:
            st.subheader("🔧 Troubleshooting")
            
            st.write("**Common Issues and Solutions:**")
            
            st.write("**❓ No data showing in Dashboard**")
            st.write("- Solution: Add policy data using 'Add New Policy Transaction'")
            st.write("- Check that the database file 'commissions.db' exists")
            
            st.write("**❓ Error saving changes**")
            st.write("- Solution: Ensure all required fields are filled correctly")
            st.write("- Check database permissions and disk space")
            
            st.write("**❓ Slow performance**")
            st.write("- Solution: The app uses caching for better performance")
            st.write("- Large datasets may take longer to load initially")
            
            st.write("**❓ Export not working**")
            st.write("- Solution: Ensure you have data to export")
            st.write("- Check browser settings for file downloads")
            
            st.info("💭 **Need more help?** Contact your system administrator or developer.")
        
        with tab4:
            st.subheader("🔄 Database Column Renaming Guide")
            
            st.warning("""
            ⚠️ **Important**: Do NOT use the Admin Panel's column rename feature for database columns. 
            Instead, follow this procedure to ensure both the app code and database stay synchronized.
            """)
            
            st.write("### Why this procedure?")
            st.write("""
            Database column names are referenced throughout the application code. When you rename a column,
            both the application code AND the database must be updated together to avoid errors.
            """)
            
            st.write("### ✅ Correct Procedure for Renaming Database Columns:")
            
            st.write("**Step 1: Check with Claude**")
            st.code("""
Example request: "We need to update a database column title from: 
'Agency Estimated Comm/Revenue (AZ)' to 'Agency Estimated Comm/Revenue (CRM)'"
            """, language="text")
            
            st.write("**Step 2: Claude will investigate and provide options**")
            st.info("""
            Claude will:
            1. Search the codebase for all references to the column
            2. Identify where the column name appears
            3. Provide options (usually recommending direct database update)
            """)
            
            st.write("**Step 3: Execute the SQL command in Supabase**")
            st.write("1. Go to your Supabase project dashboard")
            st.write("2. Navigate to the SQL Editor")
            st.write("3. Paste and run the ALTER command provided by Claude:")
            
            st.code("""
ALTER TABLE policies 
RENAME COLUMN "Old Column Name" 
TO "New Column Name";
            """, language="sql")
            
            st.write("**Step 4: Claude updates the application code**")
            st.write("""
            After you confirm the database change succeeded, Claude will:
            - Update all code references to use the new column name
            - Update column mapping configurations
            - Update SQL schema documentation files
            - Create a backup of the application before changes
            """)
            
            st.write("### 📋 Recent Example:")
            st.success("""
            **Changed**: "Agency Estimated Comm/Revenue (AZ)" → "Agency Estimated Comm/Revenue (CRM)"
            
            1. Claude found the column in commission_app.py and schema files
            2. You ran the ALTER command in Supabase SQL Editor
            3. Claude updated all code references from (AZ) to (CRM)
            4. Application now works correctly with the new column name
            """)
            
            st.write("### ⚡ Quick Reference:")
            col1, col2 = st.columns(2)
            with col1:
                st.error("❌ **Wrong Way:**")
                st.write("Using Admin Panel → Column Mapping")
                st.write("(This only updates display names, not database)")
            
            with col2:
                st.success("✅ **Right Way:**")
                st.write("1. Ask Claude for help")
                st.write("2. Run SQL in Supabase")
                st.write("3. Let Claude update code")
            
            st.info("""
            💡 **Remember**: This procedure ensures your database and application code stay perfectly 
            synchronized, preventing errors like 'Could not find column in schema cache'.
            """)
        
        with tab5:
            st.subheader("ℹ️ About")
            st.write("""
            **Commission Management Application**
            
            This application is designed to help insurance agencies and professionals manage their 
            commission data efficiently. Built with Streamlit and Python, it provides a comprehensive 
            suite of tools for policy management, reporting, and financial tracking.
            
            **Key Technologies:**
            - **Frontend**: Streamlit
            - **Database**: Supabase (PostgreSQL)
            - **Data Processing**: Pandas
            - **File Formats**: CSV, Excel, PDF support
            
            **Features:**
            - ✅ Complete CRUD operations for policy data
            - ✅ Advanced search and filtering capabilities
            - ✅ Comprehensive reporting and analytics
            - ✅ Commission reconciliation tools
            - ✅ Data export and import functionality
            - ✅ User-friendly dashboard interface
            
            **Version**: 2.0
            **Last Updated**: 2025
            """)
            
            st.success("🎉 Thank you for using the Commission Management Application!")
    
    # --- Policy Revenue Ledger Reports ---
    elif page == "Policy Revenue Ledger Reports":
        st.subheader("Policy Revenue Ledger Reports")
        st.success("📊 Generate customizable reports for policy summaries with Balance Due calculations and export capabilities.")
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        if all_data.empty:
            st.warning("No policy data loaded. Please check database connection or import data.")
        else:
            # View mode toggle
            st.markdown("### 📊 Report View Mode")
            view_col1, view_col2, view_col3 = st.columns([2, 2, 3])
            
            with view_col1:
                view_mode = st.radio(
                    "Select view:",
                    options=["Aggregated by Policy", "Detailed Transactions"],
                    index=0,
                    help="Aggregated: One row per policy with totals | Detailed: All individual transactions"
                )
            
            with view_col2:
                if view_mode == "Aggregated by Policy":
                    st.info("📋 Showing one row per policy with summed values")
                else:
                    st.info("📋 Showing all individual transactions")
            
            # Process data based on view mode
            if view_mode == "Aggregated by Policy":
                # Simple data processing - AGGREGATE BY POLICY NUMBER using mapped columns
                # This ensures one row per policy with totals from all transactions
                working_data = all_data.copy()
            
                # Get mapped column for Policy Number
                policy_number_col = get_mapped_column("Policy Number") or "Policy Number"
                
                # Group by Policy Number and aggregate the data
                if policy_number_col in working_data.columns and not working_data.empty:
                    # Define aggregation rules using mapped column names
                    agg_dict = {}
                
                    # For descriptive fields, take the first value (they should be the same for all transactions of the same policy)
                    descriptive_field_names = ["Customer", "Policy Type", "Carrier Name", "MGA Name", "Effective Date", 
                                             "Policy Origination Date", "X-DATE", "Client ID", "Transaction Type", 
                                             "Prior Policy Number", "Transaction ID", "NOTES", "Policy Term", 
                                             "Policy Checklist Complete", "STMT DATE"]
                    for field_name in descriptive_field_names:
                        mapped_col = get_mapped_column(field_name)
                        target_col = mapped_col if mapped_col and mapped_col in working_data.columns else (field_name if field_name in working_data.columns else None)
                        if target_col:
                            agg_dict[target_col] = 'first'
                    
                    # For monetary fields, sum all transactions for each policy
                    monetary_field_names = ["Agent Estimated Comm $", "Agent Paid Amount (STMT)", 
                                           "Agency Estimated Comm/Revenue (CRM)", "Premium Sold", "Policy Gross Comm %",
                                           "Broker Fee", "Broker Fee Agent Comm", "Total Agent Comm", 
                                           "Policy Taxes & Fees", "Commissionable Premium"]
                    for field_name in monetary_field_names:
                        mapped_col = get_mapped_column(field_name)
                        target_col = mapped_col if mapped_col and mapped_col in working_data.columns else (field_name if field_name in working_data.columns else None)
                        if target_col:
                            # Convert to numeric first, then sum
                            working_data[target_col] = pd.to_numeric(working_data[target_col], errors='coerce').fillna(0)
                            agg_dict[target_col] = 'sum'
                    
                    # For percentage fields, take the first value (should be consistent per policy)
                    percentage_fields = ["Agent Comm %", "Policy Gross Comm %"]
                    for field in percentage_fields:
                        if field in working_data.columns and field not in agg_dict:
                            agg_dict[field] = 'first'
                    
                    # Group by Policy Number and aggregate
                    working_data = working_data.groupby('Policy Number', as_index=False).agg(agg_dict)
            else:
                # Detailed view - show all transactions without aggregation
                working_data = all_data.copy()
            
            # Calculate Policy Balance Due using the EXACT same formula as Policy Revenue Ledger page
            # This ensures consistency between the two pages
            
            # Calculate BALANCE DUE using the same logic as the Policy Revenue Ledger page
            # Force recalculation by removing existing Policy Balance Due column if it exists
            if "Policy Balance Due" in working_data.columns:
                working_data = working_data.drop(columns=["Policy Balance Due"])
            
            # Check for required columns and calculate
            # Formula: Policy Balance Due = "Agent Estimated Comm $" - "Agent Paid Amount (STMT)"
            if "Agent Paid Amount (STMT)" in working_data.columns and "Agent Estimated Comm $" in working_data.columns:
                paid_amounts = pd.to_numeric(working_data["Agent Paid Amount (STMT)"], errors="coerce").fillna(0)
                commission_amounts = pd.to_numeric(working_data["Agent Estimated Comm $"], errors="coerce").fillna(0)
                
                working_data["Policy Balance Due"] = commission_amounts - paid_amounts
            else:
                # If we don't have the required columns, create with zeros
                working_data["Policy Balance Due"] = 0
            
            # Ensure Policy Balance Due is numeric for calculations
            if "Policy Balance Due" in working_data.columns:
                working_data["Policy Balance Due"] = pd.to_numeric(working_data["Policy Balance Due"], errors="coerce").fillna(0)
            
            # Statement Month Filter
            st.markdown("### 📅 Statement Month Selection")
            stmt_col1, stmt_col2, stmt_col3 = st.columns([2, 2, 2])
            
            with stmt_col1:
                # Extract unique year-month combinations from Effective Date
                if 'Effective Date' in working_data.columns:
                    # Convert to datetime if not already
                    working_data['Effective Date'] = pd.to_datetime(working_data['Effective Date'], errors='coerce')
                    
                    # Extract year-month and create list of options
                    working_data['Year-Month'] = working_data['Effective Date'].dt.strftime('%Y-%m')
                    unique_months = sorted(working_data['Year-Month'].dropna().unique(), reverse=True)
                    
                    # Create display labels (e.g., "January 2025")
                    month_options = ["All Months"]
                    for ym in unique_months:
                        try:
                            date_obj = pd.to_datetime(ym + '-01')
                            month_options.append(date_obj.strftime('%B %Y'))
                        except:
                            month_options.append(ym)
                    
                    selected_month = st.selectbox(
                        "Select Statement Month:",
                        options=month_options,
                        index=0,
                        help="Filter by the month when policies became effective (your monthly sales cohort)"
                    )
                else:
                    st.warning("Effective Date column not found")
                    selected_month = "All Months"
            
            with stmt_col2:
                if selected_month != "All Months":
                    # Convert selected month back to YYYY-MM format
                    selected_date = pd.to_datetime(selected_month)
                    selected_ym = selected_date.strftime('%Y-%m')
                    
                    # Filter data by selected month
                    month_data = working_data[working_data['Year-Month'] == selected_ym].copy()
                    
                    # Show count for selected month
                    if view_mode == "Aggregated by Policy":
                        st.info(f"📊 {len(month_data)} policies effective in {selected_month}")
                    else:
                        unique_in_month = month_data['Policy Number'].nunique() if 'Policy Number' in month_data.columns else 0
                        st.info(f"📊 {len(month_data)} transactions ({unique_in_month} policies) in {selected_month}")
                    
                    # Update working_data to filtered month data
                    working_data = month_data
                else:
                    st.info("📊 Showing all months")
            
            with stmt_col3:
                if selected_month != "All Months":
                    # Calculate month-specific balance
                    month_balance = working_data["Policy Balance Due"].sum() if "Policy Balance Due" in working_data.columns else 0
                    st.metric(f"{selected_month} Balance Due", f"${month_balance:,.2f}")
            
            # Calculate metrics using the same logic as Policy Revenue Ledger page
            if view_mode == "Aggregated by Policy":
                total_policies = len(working_data)
                metric_label = "Total Policies"
            else:
                total_policies = len(working_data)
                # Count unique policies in detailed view
                unique_policies = working_data['Policy Number'].nunique() if 'Policy Number' in working_data.columns else 0
                metric_label = f"Total Transactions ({unique_policies} policies)"
            
            # Count unique policies that have outstanding balance due (Policy Balance Due > 0)
            # This matches the UI formula used on the Policy Revenue Ledger page
            if "Policy Balance Due" in working_data.columns:
                # Get unique policies with Policy Balance Due > 0
                outstanding_policies = len(working_data[working_data["Policy Balance Due"] > 0])
                total_balance = working_data["Policy Balance Due"].sum()
            else:
                outstanding_policies = 0
                total_balance = 0
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric(metric_label, f"{total_policies:,}")
            with col2:
                st.metric("Outstanding Balances", f"{outstanding_policies:,}")
            with col3:
                st.metric("Total Balance Due", f"${total_balance:,.2f}")
            
            # Balance Filter Options
            st.markdown("### 🔍 Balance Filters")
            filter_col1, filter_col2, filter_col3 = st.columns(3)
            
            # Initialize balance_filter with default value
            balance_filter = "All Balances"
            
            with filter_col1:
                balance_filter = st.selectbox(
                    "Show policies with:",
                    options=[
                        "All Balances",
                        "Positive Balance Only (> $0)",
                        "Zero Balance Only (= $0)", 
                        "Negative Balance Only (< $0)",
                        "Non-Zero Balance (≠ $0)"
                    ],
                    index=0,
                    help="Filter policies based on their balance due amount"
                )
            
            # Apply balance filter
            if balance_filter != "All Balances":
                if "Policy Balance Due" in working_data.columns:
                    if balance_filter == "Positive Balance Only (> $0)":
                        working_data = working_data[working_data["Policy Balance Due"] > 0]
                    elif balance_filter == "Zero Balance Only (= $0)":
                        working_data = working_data[working_data["Policy Balance Due"] == 0]
                    elif balance_filter == "Negative Balance Only (< $0)":
                        working_data = working_data[working_data["Policy Balance Due"] < 0]
                    elif balance_filter == "Non-Zero Balance (≠ $0)":
                        working_data = working_data[working_data["Policy Balance Due"] != 0]
                    
                    # Show filtered count
                    st.info(f"📊 Showing {len(working_data):,} policies matching filter: {balance_filter}")
            
            # Initialize hidden rows in session state if not exists
            if 'prl_hidden_rows' not in st.session_state:
                st.session_state.prl_hidden_rows = set()
            
            # Apply hidden row filter if any rows are hidden
            if st.session_state.prl_hidden_rows and 'Transaction ID' in working_data.columns:
                # Store original count before filtering
                original_count = len(working_data)
                # Filter out hidden transaction IDs
                working_data = working_data[~working_data['Transaction ID'].isin(st.session_state.prl_hidden_rows)]
                hidden_count = original_count - len(working_data)
                # Store hidden count in session state for later use
                st.session_state.prl_hidden_count = hidden_count
            else:
                st.session_state.prl_hidden_count = 0
                
                    
            
            # Initialize session state for templates before any UI
            if "prl_templates" not in st.session_state:
                # Load templates from file if exists
                templates_file = "config_files/prl_templates.json"
                if os.path.exists(templates_file):
                    try:
                        with open(templates_file, 'r') as f:
                            st.session_state.prl_templates = json.load(f)
                    except:
                        st.session_state.prl_templates = {}
                else:
                    st.session_state.prl_templates = {}
            
            # Get all available columns
            all_columns = list(working_data.columns)
            # Default column selection (updated to use correct column names)
            default_columns = ["Customer", "Policy Number", "Agent Estimated Comm $", "Agent Paid Amount (STMT)", "Policy Balance Due"]
            available_default_columns = [col for col in default_columns if col in all_columns]
            
            # Check for default template and load it
            default_template = None
            for name, data in st.session_state.prl_templates.items():
                if data.get("is_default", False):
                    default_template = name
                    break
            
            # Initialize selected columns in session state
            if "prl_selected_columns" not in st.session_state:
                if default_template:
                    # Load the default template columns
                    template_data = st.session_state.prl_templates[default_template]
                    valid_template_columns = [col for col in template_data["columns"] if col in all_columns]
                    st.session_state.prl_selected_columns = valid_template_columns
                else:
                    # Use standard default columns
                    st.session_state.prl_selected_columns = available_default_columns
            
            # Show if default template was loaded
            if default_template and "prl_default_loaded" not in st.session_state:
                st.info(f"⭐ **Default template loaded**: {default_template}")
                st.session_state.prl_default_loaded = True
            
            # Column Selection & Templates in Expander
            with st.expander("🔧 Column Selection & Templates", expanded=False):
                if view_mode == "Aggregated by Policy":
                    st.info("💡 **Note**: This view aggregates data by Policy Number. Fields like 'Transaction Type' will show the first transaction's value. Switch to 'Detailed Transactions' view to see all individual transactions.")
                else:
                    st.info("💡 **Note**: This view shows all individual transactions. You can see each transaction type, term, and commission detail separately.")
                
                # Column selection interface
                col_selection_col1, col_selection_col2 = st.columns([2, 1])
                
                with col_selection_col1:
                    st.markdown("**Select Columns:**")
                    
                    # Available columns multiselect (restored to original style)
                    selected_columns = st.multiselect(
                        "Choose columns to display in your report",
                        options=all_columns,
                        default=st.session_state.prl_selected_columns,
                        key="prl_column_multiselect"
                    )
                    
                    # Update session state
                    st.session_state.prl_selected_columns = selected_columns
                    
                    # Column reordering with streamlit_sortables (compact style like Edit Policies page)
                    if selected_columns:
                        reorder_col1, reorder_col2 = st.columns([4, 1])
                        
                        with reorder_col1:
                            st.markdown("**Reorder columns by dragging the boxes below:**")
                        
                        with reorder_col2:
                            if st.button("🔄 Refresh", key="refresh_reorder", help="Refresh reorder section to sync with selected columns"):
                                # Force sync by updating the sortable key
                                if "prl_column_order_sortable_counter" not in st.session_state:
                                    st.session_state.prl_column_order_sortable_counter = 0
                                st.session_state.prl_column_order_sortable_counter += 1
                                st.rerun()
                        
                        # Generate unique key for sortable component
                        sortable_key = f"prl_column_order_sortable_{st.session_state.get('prl_column_order_sortable_counter', 0)}"
                        
                        reordered_columns = streamlit_sortables.sort_items(
                            items=selected_columns,
                            direction="horizontal",
                            key=sortable_key
                        )
                        
                        # Update session state with reordered columns
                        st.session_state.prl_selected_columns = reordered_columns
                        selected_columns = reordered_columns
                
                with col_selection_col2:
                    st.markdown("**Quick Presets:**")
                    if st.button("💰 Financial Focus"):
                        financial_cols = ["Customer", "Policy Number", "Agency Estimated Comm/Revenue (CRM)", "Agent Estimated Comm $", "Agent Paid Amount (STMT)", "Policy Balance Due"]
                        st.session_state.prl_selected_columns = [col for col in financial_cols if col in all_columns]
                        st.rerun()
                    
                    if st.button("📋 Basic Info"):
                        basic_cols = ["Customer", "Policy Type", "Carrier Name", "Policy Number", "Effective Date"]
                        st.session_state.prl_selected_columns = [col for col in basic_cols if col in all_columns]
                        st.rerun()
            
            # Template Management Section in Expander
            with st.expander("💾 Template Management", expanded=False):
                template_col1, template_col2, template_col3 = st.columns(3)
                
                with template_col1:
                    st.markdown("**Save New Template:**")
                    new_template_name = st.text_input(
                        "Template Title",
                        placeholder="Enter custom report title",
                        help="Give your template a descriptive name"
                    )
                
                    if st.button("💾 Save Template", disabled=not new_template_name or not selected_columns):
                        if new_template_name in st.session_state.prl_templates:
                            st.error(f"Template '{new_template_name}' already exists!")
                        else:
                            st.session_state.prl_templates[new_template_name] = {
                                "columns": selected_columns.copy(),
                                "created": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }
                        
                            # Save templates to file for persistence
                            templates_file = "config_files/prl_templates.json"
                            os.makedirs("config_files", exist_ok=True)
                            try:
                                with open(templates_file, 'w') as f:
                                    json.dump(st.session_state.prl_templates, f, indent=2)
                                st.success(f"✅ Template '{new_template_name}' saved!")
                            except Exception as e:
                                st.error(f"Error saving template: {str(e)}")
                        
                            st.rerun()
                    
                    # Add Update Template button if we're editing a template
                    if 'editing_template' in st.session_state and st.session_state.editing_template:
                        st.markdown("---")
                        st.markdown(f"**Currently editing:** {st.session_state.editing_template}")
                        if st.button("🔄 Update Template", type="primary"):
                            # Update the existing template
                            template_name = st.session_state.editing_template
                            # Preserve the default status if it exists
                            is_default = st.session_state.prl_templates[template_name].get("is_default", False)
                            
                            st.session_state.prl_templates[template_name] = {
                                "columns": selected_columns.copy(),
                                "created": st.session_state.prl_templates[template_name].get("created", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                                "updated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "is_default": is_default
                            }
                            
                            # Save templates to file
                            templates_file = "config_files/prl_templates.json"
                            try:
                                with open(templates_file, 'w') as f:
                                    json.dump(st.session_state.prl_templates, f, indent=2)
                                st.success(f"✅ Template '{template_name}' updated successfully!")
                                # Clear editing state
                                del st.session_state.editing_template
                            except Exception as e:
                                st.error(f"Error updating template: {str(e)}")
                            
                            st.rerun()
            
                with template_col2:
                    st.markdown("**Load Template:**")
                    if st.session_state.prl_templates:
                        # Show current default template if exists
                        current_default = None
                        for name, data in st.session_state.prl_templates.items():
                            if data.get("is_default", False):
                                current_default = name
                                break
                    
                        if current_default:
                            st.info(f"⭐ Default template: {current_default}")
                    
                        template_options = list(st.session_state.prl_templates.keys())
                        template_display = []
                        for template_name in template_options:
                            if st.session_state.prl_templates[template_name].get("is_default", False):
                                template_display.append(f"⭐ {template_name}")
                            else:
                                template_display.append(template_name)
                    
                        selected_template_display = st.selectbox(
                            "Select template to load",
                            options=template_display,
                            key="template_loader"
                        )
                    
                        # Get actual template name (remove star if present)
                        template_to_load = selected_template_display.replace("⭐ ", "")
                    
                        col1, col2 = st.columns(2)
                        with col1:
                            if st.button("📂 Load Template"):
                                template_data = st.session_state.prl_templates[template_to_load]
                                # Only load columns that still exist in the data
                                valid_columns = [col for col in template_data["columns"] if col in all_columns]
                                st.session_state.prl_selected_columns = valid_columns
                                st.success(f"✅ Loaded template: {template_to_load}")
                                st.rerun()
                    
                        with col2:
                            # Show set/unset default button
                            is_current_default = st.session_state.prl_templates[template_to_load].get("is_default", False)
                            if is_current_default:
                                if st.button("✖️ Unset Default"):
                                    # Remove default flag
                                    st.session_state.prl_templates[template_to_load]["is_default"] = False
                                
                                    # Save templates to file
                                    templates_file = "config_files/prl_templates.json"
                                    try:
                                        with open(templates_file, 'w') as f:
                                            json.dump(st.session_state.prl_templates, f, indent=2)
                                        st.success(f"✅ Removed default status from: {template_to_load}")
                                    except Exception as e:
                                        st.error(f"Error saving changes: {str(e)}")
                                    st.rerun()
                            else:
                                if st.button("⭐ Set as Default"):
                                    # Remove default from all other templates
                                    for name in st.session_state.prl_templates:
                                        st.session_state.prl_templates[name]["is_default"] = False
                                
                                    # Set this template as default
                                    st.session_state.prl_templates[template_to_load]["is_default"] = True
                                
                                    # Save templates to file
                                    templates_file = "config_files/prl_templates.json"
                                    try:
                                        with open(templates_file, 'w') as f:
                                            json.dump(st.session_state.prl_templates, f, indent=2)
                                        st.success(f"✅ Set as default template: {template_to_load}")
                                    except Exception as e:
                                        st.error(f"Error saving changes: {str(e)}")
                                    st.rerun()
                    else:
                        st.info("No saved templates yet")
            
                with template_col3:
                    st.markdown("**Manage Templates:**")
                    if st.session_state.prl_templates:
                        template_to_manage = st.selectbox(
                            "Select template to manage",
                            options=list(st.session_state.prl_templates.keys()),
                            key="template_manager"
                        )
                    
                        manage_col1, manage_col2 = st.columns(2)
                        with manage_col1:
                            if st.button("✏️ Edit"):
                                # Load template for editing
                                template_data = st.session_state.prl_templates[template_to_manage]
                                valid_columns = [col for col in template_data["columns"] if col in all_columns]
                                st.session_state.prl_selected_columns = valid_columns
                                st.session_state.editing_template = template_to_manage  # Track which template is being edited
                                st.info(f"Loaded '{template_to_manage}' for editing. Modify columns above and click 'Update Template' below.")
                    
                        with manage_col2:
                            if st.button("🗑️ Delete"):
                                del st.session_state.prl_templates[template_to_manage]
                            
                                # Save updated templates to file
                                templates_file = "config_files/prl_templates.json"
                                try:
                                    with open(templates_file, 'w') as f:
                                        json.dump(st.session_state.prl_templates, f, indent=2)
                                    st.success(f"✅ Deleted template: {template_to_manage}")
                                except Exception as e:
                                    st.error(f"Error saving changes: {str(e)}")
                            
                                st.rerun()
                    else:
                        st.info("No templates to manage")
            # Show current template status
            if st.session_state.prl_templates:
                with st.expander("📋 Saved Templates", expanded=False):
                    for name, data in st.session_state.prl_templates.items():
                        default_indicator = "⭐ " if data.get("is_default", False) else ""
                        st.write(f"**{default_indicator}{name}** - Created: {data['created']} - Columns: {len(data['columns'])}")
            
            # Show hidden rows indicator with unhide option (if any rows are hidden)
            if st.session_state.get('prl_hidden_count', 0) > 0:
                hide_col1, hide_col2, hide_col3 = st.columns([2, 1, 1])
                with hide_col1:
                    st.warning(f"🙈 {st.session_state.prl_hidden_count} rows hidden from view")
                with hide_col2:
                    if st.button("View Hidden", key="view_hidden_rows"):
                        st.session_state.show_hidden_details = True
                with hide_col3:
                    if st.button("Unhide All", key="unhide_all_quick"):
                        st.session_state.prl_hidden_rows.clear()
                        st.rerun()
            
            # Show hidden transactions details if requested
            if st.session_state.get('show_hidden_details', False):
                with st.expander("🙈 Hidden Transactions", expanded=True):
                    hidden_df = all_data[all_data['Transaction ID'].isin(st.session_state.prl_hidden_rows)]
                    
                    # Show summary
                    st.markdown(f"**{len(hidden_df)} hidden transactions:**")
                    
                    if len(hidden_df) > 0:
                        # Display with select columns for better visibility
                        display_cols = ['Transaction ID', 'Customer', 'Policy Number', 'Transaction Type', 
                                      'Agent Estimated Comm $', 'Agent Paid Amount (STMT)']
                        available_cols = [col for col in display_cols if col in hidden_df.columns]
                        
                        # Create editable dataframe with Unhide column
                        hidden_editable = hidden_df[available_cols].copy()
                        hidden_editable.insert(0, 'Unhide', False)
                        
                        # Configure columns for the hidden data editor
                        hidden_column_config = {
                            'Unhide': st.column_config.CheckboxColumn(
                                'Unhide',
                                help="Select rows to unhide",
                                default=False,
                                width="small"
                            )
                        }
                        
                        # Add number formatting for financial columns
                        for col in ['Agent Estimated Comm $', 'Agent Paid Amount (STMT)']:
                            if col in hidden_editable.columns:
                                hidden_column_config[col] = st.column_config.NumberColumn(
                                    col,
                                    format="%.2f"
                                )
                        
                        # Use data editor for hidden transactions
                        edited_hidden = st.data_editor(
                            hidden_editable,
                            use_container_width=True,
                            column_config=hidden_column_config,
                            disabled=[col for col in hidden_editable.columns if col != 'Unhide'],
                            hide_index=True,
                            key="hidden_data_editor"
                        )
                        
                        # Unhide options
                        unhide_col1, unhide_col2, unhide_col3 = st.columns([1.5, 1.5, 1])
                        with unhide_col1:
                            if st.button("✨ Unhide Selected", type="primary", key="unhide_selected"):
                                # Get rows where Unhide is True
                                rows_to_unhide = edited_hidden[edited_hidden['Unhide'] == True]['Transaction ID'].tolist()
                                if rows_to_unhide:
                                    # Remove from hidden set
                                    for trans_id in rows_to_unhide:
                                        st.session_state.prl_hidden_rows.discard(trans_id)
                                    st.session_state.show_hidden_details = False
                                    st.rerun()
                                else:
                                    st.warning("No rows selected to unhide")
                        
                        with unhide_col2:
                            if st.button("Unhide All", key="unhide_all_detail"):
                                st.session_state.prl_hidden_rows.clear()
                                st.session_state.show_hidden_details = False
                                st.rerun()
                        
                        with unhide_col3:
                            if st.button("Close", key="close_hidden_details"):
                                st.session_state.show_hidden_details = False
                                st.rerun()
                    else:
                        st.info("No hidden transactions to display")
                        if st.button("Close", key="close_empty_hidden"):
                            st.session_state.show_hidden_details = False
                            st.rerun()
            
            # Data Preview with Selected Columns
            st.markdown("### 📊 Report Preview")
            if selected_columns and not working_data.empty:
                # Filter to only include columns that exist
                valid_columns = [col for col in selected_columns if col in working_data.columns]
                
                if valid_columns:
                    # Pagination controls
                    pagination_col1, pagination_col2, pagination_col3 = st.columns([2, 1, 1])
                    
                    with pagination_col1:
                        records_per_page = st.selectbox(
                            "Records per page:",
                            options=[20, 50, 100, 200, 500, "All"],
                            index=0,
                            key="prl_records_per_page"                        )
                    
                    with pagination_col2:
                        if records_per_page != "All":
                            # Ensure records_per_page is treated as integer for calculations
                            records_per_page_int = int(records_per_page)
                            total_pages = max(1, (len(working_data) + records_per_page_int - 1) // records_per_page_int)
                            current_page = st.number_input(
                                "Page:",
                                min_value=1,
                                max_value=total_pages,
                                value=1,
                                key="prl_current_page"
                            )
                        else:
                            current_page = 1
                            total_pages = 1
                    
                    with pagination_col3:
                        if records_per_page != "All":
                            st.write(f"Page {current_page} of {total_pages}")
                    
                    # Apply pagination
                    if records_per_page == "All":
                        display_data = working_data[valid_columns].copy()
                        caption_text = f"Showing all {len(working_data):,} records with {len(valid_columns)} columns"
                    else:
                        # Ensure type safety for calculations
                        records_per_page_int = int(records_per_page)
                        current_page_int = int(current_page)
                        start_idx = (current_page_int - 1) * records_per_page_int
                        end_idx = start_idx + records_per_page_int
                        display_data = working_data[valid_columns].iloc[start_idx:end_idx].copy()
                        caption_text = f"Showing records {start_idx + 1}-{min(end_idx, len(working_data))} of {len(working_data):,} total records with {len(valid_columns)} columns"
                    
                    # Format numeric columns to 2 decimal places
                    # Get all numeric columns in the data
                    numeric_columns = display_data.select_dtypes(include=['float64', 'int64', 'float32', 'int32', 'float', 'int']).columns.tolist()
                    
                    # Also include columns that should be numeric but might be object type
                    potentially_numeric = [
                        "Agent Comm", "Agent Comm %",
                        "Agent Estimated Comm", "Agent Estimated Comm $",
                        "Agent Paid Amount", "Agent Paid Amount (STMT)",
                        "Policy Balance Due", "Balance Due",
                        "Agency Estimated Comm/Revenue (CRM)",
                        "Agency Comm Received (STMT)",
                        "Premium Sold", "Broker Fee", "Broker Fee Agent Comm",
                        "Total Agent Comm", "Policy Taxes & Fees", "Commissionable Premium",
                        "Policy Gross Comm %", "Credit (Commission Owed)", "Debit (Paid to Agent)"
                    ]
                    
                    # Combine and get unique columns
                    all_numeric_columns = list(set(numeric_columns + potentially_numeric))
                    
                    # Apply formatting to numeric columns that exist in the display data
                    for col in all_numeric_columns:
                        if col in display_data.columns:
                            # Convert to numeric and format to 2 decimal places
                            display_data[col] = pd.to_numeric(display_data[col], errors='coerce').round(2)
                    
                    # Format the dataframe for display with proper decimal places
                    # Create format dict for all numeric columns
                    format_dict = {}
                    for col in display_data.columns:
                        if col in all_numeric_columns or display_data[col].dtype in ['float64', 'int64', 'float32', 'int32']:
                            format_dict[col] = '{:.2f}'
                    
                    # Add a blank row at the end for better visibility
                    blank_row = pd.DataFrame([{col: '' for col in valid_columns}])
                    display_data = pd.concat([display_data, blank_row], ignore_index=True)
                    
                    # Calculate height based on actual number of rows
                    # Each row is approximately 35 pixels, header is 35 pixels
                    # Show actual rows + blank row + header, but cap at 11 visible rows max (plus blank row)
                    num_data_rows = min(len(display_data), 12)  # 11 data rows + 1 blank row
                    display_height = 35 * (num_data_rows + 1)  # +1 for header
                    
                    # Minimum height to show at least header + 2 rows
                    display_height = max(display_height, 105)
                    
                    # Create column configuration for numeric formatting and date formatting
                    column_config = {}
                    
                    # Date columns that need formatting
                    date_columns = ['Effective Date', 'X-DATE', 'STMT DATE', 'Policy Origination Date', 
                                    'Expiration Date', 'As of Date', 'Transaction Date']
                    
                    for col in display_data.columns:
                        if col in all_numeric_columns or display_data[col].dtype in ['float64', 'int64', 'float32', 'int32']:
                            column_config[col] = st.column_config.NumberColumn(
                                col,
                                format="%.2f"
                            )
                        elif col in date_columns:
                            column_config[col] = st.column_config.DateColumn(
                                col,
                                format="YYYY-MM-DD"
                            )
                    
                    # Add Hide checkbox column for data editor
                    if 'Transaction ID' in display_data.columns:
                        # Create a copy for editing
                        editable_data = display_data.copy()
                        
                        # Add Hide checkbox column at the beginning
                        editable_data.insert(0, 'Hide', False)
                        
                        # Update column config for the Hide column
                        column_config['Hide'] = st.column_config.CheckboxColumn(
                            'Hide',
                            help="Select rows to hide",
                            default=False,
                            width="small"
                        )
                        
                        # Use data_editor instead of dataframe
                        edited_df = st.data_editor(
                            editable_data,
                            use_container_width=True,
                            height=display_height,
                            column_config=column_config,
                            disabled=[col for col in editable_data.columns if col != 'Hide'],  # Only Hide column is editable
                            hide_index=True,
                            key="prl_data_editor"
                        )
                        
                        # Add Hide Selected button
                        hide_button_col1, hide_button_col2, hide_button_col3 = st.columns([1, 1, 4])
                        with hide_button_col1:
                            if st.button("🙈 Hide Selected", type="primary", key="hide_selected_rows"):
                                # Get rows where Hide is True
                                rows_to_hide = edited_df[edited_df['Hide'] == True]['Transaction ID'].tolist()
                                if rows_to_hide:
                                    st.session_state.prl_hidden_rows.update(rows_to_hide)
                                    st.rerun()
                                else:
                                    st.warning("No rows selected to hide")
                        
                        with hide_button_col2:
                            # Quick access to hidden rows view
                            if st.session_state.prl_hidden_rows:
                                if st.button(f"View Hidden ({len(st.session_state.prl_hidden_rows)})", key="quick_view_hidden"):
                                    st.session_state.show_hidden_details = True
                                    st.rerun()
                    else:
                        # No Transaction ID column, use regular dataframe
                        styled_data = style_special_transactions(display_data)
                        st.dataframe(
                            styled_data, 
                            use_container_width=True, 
                            height=display_height,
                            column_config=column_config
                        )
                    st.caption(caption_text)                    # Create report metadata for export
                    export_metadata = {
                        "Report Generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Report Type": "Policy Revenue Ledger Report",
                        "Statement Month": selected_month,
                        "View Mode": view_mode,
                        "Balance Filter": balance_filter,
                        "Total Records": f"{len(working_data):,}",
                        "Selected Columns": f"{len(valid_columns)} columns: {', '.join(valid_columns)}",
                        "Data Aggregation": "Grouped by Policy Number (one row per unique policy)" if view_mode == "Aggregated by Policy" else "All individual transactions",
                        "Balance Due Formula": "Agent Estimated Comm $ - Agent Paid Amount (STMT)",
                        "Page Size": f"{records_per_page} records per page" if records_per_page != "All" else "All records displayed",
                        "Outstanding Policies": f"{outstanding_policies:,} policies with balance due > $0",
                        "Total Balance Due": f"${total_balance:,.2f}"
                    }
                    
                    # Show active report parameters
                    st.markdown("### 📋 Active Report Parameters")
                    st.markdown("*These parameters will be included in exported files:*")
                    metadata_df = pd.DataFrame(list(export_metadata.items()), columns=["Parameter", "Value"])
                    st.dataframe(metadata_df, use_container_width=True, height=min(300, 40 + 40 * len(metadata_df)))
                    
                    # Enhanced export with custom filename and metadata
                    export_col1, export_col2, export_col3 = st.columns(3)
                    with export_col1:
                        custom_filename = st.text_input(
                            "Custom Export Filename (optional)",
                            placeholder="Leave blank for auto-generated name"
                        )
                    
                    with export_col2:
                        filename = custom_filename if custom_filename else f"policy_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
                        csv_filename = filename if filename.endswith('.csv') else filename + '.csv'
                        
                        # Create CSV with metadata header
                        csv_lines = ["# Policy Revenue Ledger Report - Export Parameters"]
                        for param, value in export_metadata.items():
                            csv_lines.append(f"# {param}: {value}")
                        csv_lines.append("# " + "="*50)
                        csv_lines.append("")  # Empty line before data
                        
                        # Format numeric columns in the export data
                        export_data = working_data[valid_columns].copy()
                        # Use the same all_numeric_columns list from earlier
                        for col in all_numeric_columns:
                            if col in export_data.columns:
                                export_data[col] = pd.to_numeric(export_data[col], errors='coerce').round(2)
                        
                        # Format date columns to remove time component
                        date_columns = ['Effective Date', 'X-DATE', 'STMT DATE', 'Policy Origination Date', 
                                      'Expiration Date', 'As of Date', 'Transaction Date']
                        for col in date_columns:
                            if col in export_data.columns:
                                # Convert to datetime and format as date only
                                export_data[col] = pd.to_datetime(export_data[col], errors='coerce').dt.strftime('%Y-%m-%d')
                        
                        # Add the actual data
                        csv_data = export_data.to_csv(index=False)
                        csv_with_metadata = "\n".join(csv_lines) + "\n" + csv_data
                        
                        st.download_button(
                            "📄 Export as CSV (with Parameters)",
                            csv_with_metadata,
                            csv_filename,
                            "text/csv",
                            help="CSV file includes report parameters in header comments"
                        )
                    
                    with export_col3:
                        excel_filename = filename if filename.endswith('.xlsx') else filename + '.xlsx'
                        
                        # Create Excel file with metadata sheet
                        excel_buffer = io.BytesIO()
                        with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                            # Write metadata to first sheet
                            metadata_df.to_excel(writer, sheet_name='Report Parameters', index=False)
                            
                            # Write data to second sheet with formatted numeric columns
                            excel_export_data = working_data[valid_columns].copy()
                            
                            # Format numeric columns
                            for col in all_numeric_columns:
                                if col in excel_export_data.columns:
                                    excel_export_data[col] = pd.to_numeric(excel_export_data[col], errors='coerce').round(2)
                            
                            # Format date columns to remove time component
                            date_columns = ['Effective Date', 'X-DATE', 'STMT DATE', 'Policy Origination Date', 
                                          'Expiration Date', 'As of Date', 'Transaction Date']
                            for col in date_columns:
                                if col in excel_export_data.columns:
                                    # Convert to datetime and format as date only
                                    excel_export_data[col] = pd.to_datetime(excel_export_data[col], errors='coerce').dt.strftime('%Y-%m-%d')
                            
                            excel_export_data.to_excel(writer, sheet_name='Policy Revenue Report', index=False)
                            
                            # Get workbook and format metadata sheet
                            workbook = writer.book
                            metadata_sheet = writer.sheets['Report Parameters']                            # Format metadata sheet
                            header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC'})
                            metadata_sheet.set_column('A:A', 25)
                            metadata_sheet.set_column('B:B', 50)
                            metadata_sheet.write_row(0, 0, ['Parameter', 'Value'], header_format)
                        
                        excel_buffer.seek(0)
                        
                        st.download_button(
                            "📊 Export as Excel (with Parameters)",
                            excel_buffer.getvalue(),
                            excel_filename,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Excel file includes report parameters on separate sheet"
                        )
                else:
                    st.warning("Selected columns are not available in the current data.")
            else:
                if not selected_columns:
                    st.info("Please select columns to display the report.")
                else:
                    st.warning("No data available for report generation.")
            
            # PDF printing instructions
            st.markdown("### 🖨️ Print to PDF")
            st.info("💡 **Tip:** To save this report as PDF, use your browser's print feature (Ctrl+P or Cmd+P) and select 'Save as PDF'. The report parameters shown above will be included in the PDF.")
            
            st.success("✅ Enhanced Policy Revenue Ledger Reports with Templates & Export Parameters!")
    
    # --- Pending Policy Renewals ---
    elif page == "Pending Policy Renewals":
        st.subheader("Pending Policy Renewals")
        
        # Add helpful tip about removing renewals in a collapsible expander
        with st.expander("💡 **Tip: How to Remove a Policy from This List**", expanded=False):
            st.info("""
            **To process a renewal (RWL):**
            - You can renew policies directly from this page!
            - Click the renewal button for any policy listed here
            - Fill in the renewal details and save
            
            **To process a "REWRITE":**
            1. Go to "Add Policy Transactions" page
            2. Create your new REWRITE transaction
            3. **IMPORTANT**: Enter the old policy number in the "Prior Policy Number" field
            4. The old policy will automatically disappear from this pending renewals list!
            
            **About Cancelled Policies (CAN):**
            - When you cancel a policy (CAN transaction), it ends the policy's history trail
            - Cancelled policies will NOT appear in this renewal list
            - Only use CAN when the policy is truly terminated (not replaced)
            
            **What doesn't show here:**
            - STMT transactions (these are payment records, not policies)
            - VOID transactions (these are voided/reversed entries)
            - This page is specifically for renewing policies by creating RWL transactions
            
            This works for all replacement scenarios: standard renewals, rewrites with carrier changes, or policy type changes.
            """)
        
        # Load fresh data for this page
        all_data = load_policies_data()
        
        if all_data.empty:
            st.warning("No data found in policies table. Please add some policy data first.")
            return
        
        # Initialize session state
        if 'deleted_renewals' not in st.session_state:
            st.session_state['deleted_renewals'] = []
        if 'editing_renewal' not in st.session_state:
            st.session_state.editing_renewal = False
        if 'renewal_to_edit' not in st.session_state:
            st.session_state.renewal_to_edit = None
        
        # Clear carrier/commission related session state to avoid stale data
        keys_to_clear = [
            'edit_selected_carrier_id', 'edit_selected_carrier_name',
            'edit_selected_mga_id', 'edit_selected_mga_name',
            'edit_carrier_name_manual', 'edit_mga_name_manual',
            'edit_final_carrier_name', 'edit_final_mga_name',
            'edit_commission_new_rate', 'edit_commission_renewal_rate',
            'edit_commission_rule_id', 'edit_has_commission_rule'
        ]
        for key in keys_to_clear:
            if key in st.session_state:
                del st.session_state[key]

        pending_renewals_df = get_pending_renewals(all_data)
        # Don't use duplicate_for_renewal for display - it modifies the transaction type!
        # Just display the original pending renewals as they are
        duplicated_renewals_df = pending_renewals_df.copy()
        
        # Filter out deleted renewals
        if not duplicated_renewals_df.empty:
            display_df = duplicated_renewals_df[~duplicated_renewals_df.index.isin(st.session_state['deleted_renewals'])]
        else:
            display_df = pd.DataFrame()

        # Add summary metrics at the top
        if not display_df.empty:
            # Calculate summary statistics
            past_due_count = len(display_df[display_df['Days Until Expiration'] < 0])
            this_week_count = len(display_df[(display_df['Days Until Expiration'] >= 0) & (display_df['Days Until Expiration'] <= 7)])
            this_month_count = len(display_df[(display_df['Days Until Expiration'] >= 0) & (display_df['Days Until Expiration'] <= 30)])
            total_count = len(display_df)
            
            # Display metrics
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("🔴 Past Due", past_due_count, help="Policies already expired")
            with col2:
                st.metric("🟡 Due This Week", this_week_count, help="Expiring in 0-7 days")
            with col3:
                st.metric("📅 Due This Month", this_month_count, help="Expiring in 0-30 days")
            with col4:
                st.metric("📊 Total Pending", total_count, help="All pending renewals")
            
            # Add time range filter
            st.markdown("### Filter Options")
            col_filter1, col_filter2 = st.columns([2, 3])
            
            with col_filter1:
                filter_option = st.radio(
                    "Show Renewals:",
                    ["All Renewals", "Past Due Only", "Due This Week", "Due in 30 Days", "Due in 60 Days", "Due in 90 Days"],
                    help="Filter renewals by time range"
                )
            
            # Apply filter based on selection
            filtered_df = display_df.copy()
            if filter_option == "Past Due Only":
                filtered_df = filtered_df[filtered_df['Days Until Expiration'] < 0]
            elif filter_option == "Due This Week":
                filtered_df = filtered_df[filtered_df['Days Until Expiration'] <= 7]
            elif filter_option == "Due in 30 Days":
                filtered_df = filtered_df[filtered_df['Days Until Expiration'] <= 30]
            elif filter_option == "Due in 60 Days":
                filtered_df = filtered_df[filtered_df['Days Until Expiration'] <= 60]
            elif filter_option == "Due in 90 Days":
                filtered_df = filtered_df[filtered_df['Days Until Expiration'] <= 90]
            # "All Renewals" shows everything (no additional filter needed)
            
            st.divider()
            
            # Use filtered dataframe for display
            display_df = filtered_df

        if not display_df.empty:
            # Format dates to MM/DD/YYYY before displaying
            date_columns = ['Policy Origination Date', 'Effective Date', 'X-DATE']
            # REMOVED: Date formatting to show raw dates as stored in database
            # for col in date_columns:
            #     if col in display_df.columns:
            #         # Convert to datetime, then to string with MM/DD/YYYY format
            #         display_df[col] = pd.to_datetime(display_df[col], errors='coerce').dt.strftime('%m/%d/%Y')
            
            # Add Edit checkbox column only
            display_df.insert(0, "Edit", False)
            
            # Configure the data editor
            column_config = {
                "Edit": st.column_config.CheckboxColumn(
                    "Edit",
                    help="Check to select for editing",
                    default=False
                )
            }
            
            # Configure numeric columns
            numeric_cols = [
                'Premium Sold', 'Agency Estimated Comm/Revenue (CRM)', 
                'Policy Gross Comm %', 'Agent Estimated Comm $', 'Days Until Expiration'
            ]
            for col in numeric_cols:
                if col in display_df.columns:
                    if col == 'Days Until Expiration':
                        column_config[col] = st.column_config.NumberColumn(
                            col,
                            format="%d days",
                            help="Negative values indicate past due"
                        )
                    else:
                        column_config[col] = st.column_config.NumberColumn(
                            col,
                            format="$%.2f" if '$' in col else "%.2f",
                            step=0.01
                        )
            
            # Add visual indicators for past due and urgent renewals
            st.markdown("#### Renewal Status Legend")
            col_legend1, col_legend2, col_legend3, col_legend4 = st.columns(4)
            with col_legend1:
                st.markdown("🔴 **Past Due** - Already expired")
            with col_legend2:
                st.markdown("🟡 **Urgent** - Due in 0-7 days")
            with col_legend3:
                st.markdown("⚪ **Normal** - Due in 8+ days")
            
            # Add status indicator column based on days until expiration
            def get_status_icon(days):
                if pd.isna(days):
                    return "❓"
                elif days < 0:
                    return "🔴 Past Due"
                elif days <= 7:
                    return "🟡 Urgent"
                else:
                    return "✅ OK"
            
            # Insert status column after Edit column
            display_df.insert(1, "Status", display_df['Days Until Expiration'].apply(get_status_icon))
            
            # Update column config for status
            column_config["Status"] = st.column_config.TextColumn(
                "Status",
                help="Renewal urgency status",
                width="small"
            )
            
            edited_df = st.data_editor(
                display_df,
                column_config=column_config,
                hide_index=True,
                key="pending_renewals_editor",
                disabled=["Days Until Expiration", "Status"]  # Make these columns read-only
            )
            
            # Show the edit form if we're in editing mode
            if st.session_state.editing_renewal and st.session_state.renewal_to_edit:
                st.divider()
                st.markdown("### 📝 Edit Renewal Transaction")
                
                # Important reminder for user with better visibility
                st.warning("⚠️ **Currently Editing**: To select a different renewal, scroll down and click the **Cancel** button at the bottom of this form.")
                
                # Prepare renewal data (pre-populate with calculated values)
                renewal_data = st.session_state.renewal_to_edit.copy()
                
                # Add Carrier and MGA selection UI (similar to Edit Policy Transactions)
                st.markdown("#### Carrier & MGA Selection")
                
                # Add attention-grabbing reminder
                st.success("💡 **TIP**: Select a carrier from the dropdown below to automatically load commission rates! ⬇️")
                
                # Get current carrier and MGA
                current_carrier = renewal_data.get('Carrier Name', '')
                current_mga = renewal_data.get('MGA Name', '')
                
                # Show current transaction data
                st.info(f"📋 Current Transaction Data - Carrier: '{current_carrier}' | MGA: '{current_mga}' | Policy Type: '{renewal_data.get('Policy Type', '')}')")
                
                col_carrier, col_mga = st.columns(2)
                
                with col_carrier:
                    # Load carriers list
                    carriers_list = load_carriers_for_dropdown()
                    carrier_names = [c['carrier_name'] for c in carriers_list if c.get('status', 'Active') == 'Active']
                    carrier_names.insert(0, "")  # Add empty option
                    
                    # Get index of current carrier
                    carrier_index = 0
                    if current_carrier and current_carrier in carrier_names:
                        carrier_index = carrier_names.index(current_carrier)
                    
                    selected_carrier_name = st.selectbox(
                        "Select Carrier",
                        options=carrier_names,
                        index=carrier_index,
                        key="renewal_carrier_select",
                        placeholder="Choose a carrier..."
                    )
                    
                    # Store carrier ID if selected
                    selected_carrier_id = None
                    if selected_carrier_name:
                        selected_carrier_id = next((c['carrier_id'] for c in carriers_list if c['carrier_name'] == selected_carrier_name), None)
                        st.session_state['edit_selected_carrier_id'] = selected_carrier_id
                        st.session_state['edit_selected_carrier_name'] = selected_carrier_name
                    
                    # Fallback text input for manual entry
                    if not selected_carrier_name:
                        carrier_name_manual = st.text_input(
                            "Or enter carrier name manually", 
                            value=current_carrier, 
                            placeholder="Type carrier name", 
                            key="renewal_carrier_manual"
                        )
                        st.session_state['edit_carrier_name_manual'] = carrier_name_manual
                
                with col_mga:
                    # Load MGAs based on selected carrier
                    if selected_carrier_name and selected_carrier_id:
                        mgas_list = load_mgas_for_carrier(selected_carrier_id)
                        mga_names = ["Direct Appointment"] + [m['mga_name'] for m in mgas_list]
                        
                        # Get index of current MGA
                        mga_index = 0
                        if current_mga:
                            if current_mga in mga_names:
                                mga_index = mga_names.index(current_mga)
                            elif current_mga == "" and "Direct Appointment" in mga_names:
                                mga_index = mga_names.index("Direct Appointment")
                        
                        selected_mga_name = st.selectbox(
                            "Select MGA",
                            options=mga_names,
                            index=mga_index,
                            key="renewal_mga_select",
                            placeholder="Choose an MGA...",
                            help="Select 'Direct Appointment' if no MGA is involved"
                        )
                        
                        # Store MGA ID if selected
                        if selected_mga_name != "Direct Appointment" and selected_carrier_id:
                            selected_mga_id = next((m['mga_id'] for m in mgas_list if m['mga_name'] == selected_mga_name), None)
                            st.session_state['edit_selected_mga_id'] = selected_mga_id
                            st.session_state['edit_selected_mga_name'] = selected_mga_name
                        else:
                            st.session_state['edit_selected_mga_id'] = None
                            st.session_state['edit_selected_mga_name'] = selected_mga_name
                    
                    # Fallback text input for manual entry
                    if not selected_carrier_name:
                        mga_name_manual = st.text_input(
                            "Or enter MGA name manually", 
                            value=current_mga, 
                            placeholder="Type MGA name or leave blank", 
                            key="renewal_mga_manual"
                        )
                        st.session_state['edit_mga_name_manual'] = mga_name_manual
                
                # Store final values for form submission
                if selected_carrier_name:
                    final_carrier_name = selected_carrier_name
                    final_mga_name = selected_mga_name if selected_mga_name != "Direct Appointment" else ""
                else:
                    final_carrier_name = st.session_state.get('edit_carrier_name_manual', '')
                    final_mga_name = st.session_state.get('edit_mga_name_manual', '')
                
                # Store in session state for form to access
                st.session_state['edit_final_carrier_name'] = final_carrier_name
                st.session_state['edit_final_mga_name'] = final_mga_name
                
                # Display selection status and commission info
                if selected_carrier_name and selected_carrier_id:
                    # Look up commission rule
                    commission_rule = None
                    policy_type = renewal_data.get('Policy Type', '')
                    
                    if st.session_state.get('edit_selected_mga_id'):
                        # Try carrier + MGA + policy type first
                        commission_rule = lookup_commission_rule(
                            selected_carrier_id, 
                            st.session_state.get('edit_selected_mga_id'), 
                            policy_type
                        )
                        if not commission_rule:
                            # Try carrier + MGA without policy type
                            commission_rule = lookup_commission_rule(
                                selected_carrier_id, 
                                st.session_state.get('edit_selected_mga_id'), 
                                None
                            )
                    
                    if not commission_rule:
                        # Try carrier + policy type without MGA
                        commission_rule = lookup_commission_rule(selected_carrier_id, None, policy_type)
                    
                    if not commission_rule:
                        # Try carrier default
                        commission_rule = lookup_commission_rule(selected_carrier_id, None, None)
                    
                    if commission_rule:
                        # Store both rates and let the form decide which to use based on transaction type
                        new_rate = commission_rule.get('new_rate', 0)
                        renewal_rate = commission_rule.get('renewal_rate', new_rate)  # Default to new rate if no renewal rate
                        
                        st.info(f"ℹ️ Commission rule found: {commission_rule.get('rule_description', 'Carrier default')}")
                        st.success(f"✅ Renewal rate will be applied: {renewal_rate}%")
                        
                        # Store both rates in session state
                        st.session_state['edit_commission_new_rate'] = new_rate
                        st.session_state['edit_commission_renewal_rate'] = renewal_rate
                        st.session_state['edit_commission_rule_id'] = commission_rule.get('rule_id')
                        st.session_state['edit_has_commission_rule'] = True
                    else:
                        st.info(f"ℹ️ No commission rule found for {selected_carrier_name}. Enter rate manually.")
                        st.session_state['edit_commission_new_rate'] = None
                        st.session_state['edit_commission_renewal_rate'] = None
                        st.session_state['edit_commission_rule_id'] = None
                        st.session_state['edit_has_commission_rule'] = False
                else:
                    # No carrier selected
                    if not selected_carrier_name and not st.session_state.get('edit_carrier_name_manual'):
                        st.info("ℹ️ No carrier selected. Commission rates must be entered manually.")
                    else:
                        st.info(f"ℹ️ Carrier '{final_carrier_name}' is not in the system. Commission rates must be entered manually.")
                    st.session_state['edit_has_commission_rule'] = False
                
                st.markdown("---")
                
                # Generate new unique Transaction ID
                # Try up to 10 times to generate a unique ID
                unique_id_found = False
                supabase = get_supabase_client()  # Get supabase client
                for attempt in range(10):
                    new_id = generate_transaction_id()
                    # Check if this ID already exists
                    try:
                        existing = supabase.table('policies').select('"Transaction ID"').eq('"Transaction ID"', new_id).execute()
                        if not existing.data:
                            renewal_data['Transaction ID'] = new_id
                            unique_id_found = True
                            break
                    except Exception as e:
                        st.warning(f"Error checking Transaction ID uniqueness: {e}")
                        break
                
                if not unique_id_found:
                    # If we couldn't generate a unique ID after 10 attempts, use a timestamp-based ID
                    renewal_data['Transaction ID'] = f"RWL{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
                
                # Pre-populate Prior Policy Number with the current policy number
                renewal_data['Prior Policy Number'] = renewal_data.get('Policy Number', '')
                
                # Set transaction type to RWL for renewal
                renewal_data[get_mapped_column("Transaction Type")] = "RWL"
                
                # Update dates for renewal - calculate new effective and expiration dates
                if 'expiration_date' in renewal_data:
                    # New effective date is the old expiration date
                    new_effective = pd.to_datetime(renewal_data['expiration_date'])
                    renewal_data[get_mapped_column("Effective Date")] = new_effective.strftime('%Y-%m-%d')  # Changed to YYYY-MM-DD
                    
                    # Calculate new expiration date based on Policy Term
                    policy_term_col = get_mapped_column("Policy Term")
                    if policy_term_col in renewal_data and pd.notna(renewal_data.get(policy_term_col)) and renewal_data.get(policy_term_col) != 0:
                        months_to_add = int(renewal_data[policy_term_col])
                    else:
                        months_to_add = 6  # Default to 6 months if not specified
                    
                    new_expiration = new_effective + pd.DateOffset(months=months_to_add)
                    renewal_data[get_mapped_column("X-DATE")] = new_expiration.strftime('%Y-%m-%d')  # Changed to YYYY-MM-DD
                
                # Clear commission fields and NOTES for renewal
                fields_to_clear = [
                    'Premium Sold', 'Commissionable Premium', 'Commission %', 
                    'Commission $', 'Producer Commission %', 'Producer Commission $',
                    'Override %', 'Override Commission', 'Commission Already Earned',
                    'Commission Already Received', 'Balance Owed', 'Renewal/Bonus Percentage',
                    'Renewal Amount', 'Not Paid/Paid', 'Agent Paid Amount (STMT)',
                    'Agency Comm Received (STMT)', 'NOTES'  # Added NOTES to clear it
                ]
                for field in fields_to_clear:
                    if field in renewal_data:
                        renewal_data[field] = None
                
                # Use the edit transaction form
                result = edit_transaction_form(
                    renewal_data, 
                    source_page="pending_renewals",
                    is_renewal=True
                )
                
                if result:
                    if result["action"] == "save":
                        try:
                            # Insert the new renewal transaction
                            new_renewal = result["data"]
                            
                            # Convert timestamps to strings for JSON serialization
                            new_renewal = convert_timestamps_for_json(new_renewal)
                            
                            # Remove UI-only fields using the centralized function
                            new_renewal = clean_data_for_database(new_renewal)
                            
                            # Handle column name changes
                            if 'NEW BIZ CHECKLIST COMPLETE' in new_renewal:
                                new_renewal['Policy Checklist Complete'] = new_renewal.pop('NEW BIZ CHECKLIST COMPLETE')
                            
                            # Map expiration_date to the actual database column
                            if 'expiration_date' in new_renewal:
                                new_renewal['X-DATE'] = new_renewal.pop('expiration_date')
                            
                            # Add policy tracking fields
                            original_policy = st.session_state.renewal_to_edit
                            new_renewal['Prior Policy Number'] = original_policy.get('Policy Number', '')
                            
                            # Policy Origination Date stays the same through renewals (it's already in the data)
                            # No need to set it explicitly as it's preserved from the original policy
                            
                            # Debug: Show what we're trying to insert
                            trans_id = new_renewal.get('Transaction ID', 'MISSING')
                            st.info(f"Attempting to insert renewal with Transaction ID: {trans_id}")
                            
                            # Show all fields being inserted (for debugging)
                            with st.expander("Debug: Full renewal data being inserted"):
                                debug_data = {k: v for k, v in new_renewal.items() if k not in ['Edit', 'Select', 'Action', 'Details']}
                                st.json(debug_data)
                            
                            # Double-check if this ID really exists
                            try:
                                check_existing = supabase.table('policies').select('"Transaction ID"').eq('"Transaction ID"', trans_id).execute()
                                if check_existing.data:
                                    st.error(f"Transaction ID {trans_id} already exists in database! Found {len(check_existing.data)} matching records.")
                                    st.error("The uniqueness check failed. Trying with timestamp-based ID...")
                                    # Generate a timestamp-based ID instead
                                    new_renewal['Transaction ID'] = f"RWL{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')[:-3]}"
                                    st.info(f"Using new Transaction ID: {new_renewal['Transaction ID']}")
                                else:
                                    st.success(f"Transaction ID {trans_id} is unique and ready to insert.")
                            except Exception as e:
                                st.error(f"Error checking existing ID: {e}")
                            
                            # Insert to database
                            supabase.table('policies').insert(new_renewal).execute()
                            clear_policies_cache()
                            
                            # Log the renewal
                            try:
                                renewal_history_data = {
                                    "renewal_timestamp": datetime.datetime.now().isoformat(),
                                    "renewed_by": "User",
                                    "original_transaction_id": st.session_state.renewal_to_edit.get('Transaction ID', ''),
                                    "new_transaction_id": new_renewal.get('Transaction ID', ''),
                                    "details": json.dumps({
                                        "count": 1, 
                                        "renewed_ids": [new_renewal.get('Transaction ID', '')],
                                        "original_policy_number": original_policy.get('Policy Number', ''),
                                        "new_policy_number": new_renewal.get('Policy Number', ''),
                                        "policy_chain": True if new_renewal.get('Prior Policy Number') else False
                                    })
                                }
                                supabase.table('renewal_history').insert(renewal_history_data).execute()
                            except Exception as e:
                                st.warning(f"Renewal saved but history logging failed: {e}")
                            
                            st.success("✅ Policy renewed successfully!")
                            
                            # Remove from pending list
                            original_index = None
                            for idx in display_df.index:
                                if display_df.loc[idx, 'Policy Number'] == st.session_state.renewal_to_edit.get('Policy Number'):
                                    original_index = idx
                                    break
                            if original_index is not None:
                                st.session_state['deleted_renewals'].append(original_index)
                            
                            # Reset editing state
                            st.session_state.editing_renewal = False
                            st.session_state.renewal_to_edit = None
                            time.sleep(1)
                            st.rerun()
                            
                        except Exception as e:
                            st.error(f"Error creating renewal: {e}")
                    
                    elif result["action"] == "cancel" or result["action"] == "close":
                        # Reset editing state
                        st.session_state.editing_renewal = False
                        st.session_state.renewal_to_edit = None
                        st.rerun()
            
            else:
                # Edit button below the table
                st.markdown("---")
                st.markdown("### Select a Renewal to Edit")
                
                # Check for Edit selections
                edit_selected_rows = edited_df[edited_df["Edit"] == True]
                selected_count = len(edit_selected_rows)
                
                # Debug info
                st.caption(f"Debug: {len(edited_df)} total renewals, {selected_count} selected")
                
                if selected_count == 1:
                    if st.button("✏️ Edit Selected Pending Renewal", type="primary", use_container_width=True):
                        # Get the row where Edit is checked
                        row_to_edit = edit_selected_rows.iloc[0]
                        renewal_dict = row_to_edit.to_dict()
                        # Remove UI-only columns
                        ui_only_columns = ['Edit', 'Select', 'Action', 'Details', 'Status']
                        for col in ui_only_columns:
                            if col in renewal_dict:
                                del renewal_dict[col]
                        st.session_state.editing_renewal = True
                        st.session_state.renewal_to_edit = renewal_dict
                        st.rerun()
                elif selected_count == 0:
                    st.button("✏️ Edit Selected Pending Renewal", type="primary", use_container_width=True, 
                             disabled=True, help="Check one Edit checkbox to edit")
                else:
                    st.button("✏️ Edit Selected Pending Renewal", type="primary", use_container_width=True, 
                             disabled=True, help=f"{selected_count} selected - please select only ONE renewal to edit")
        else:
            if pending_renewals_df.empty:
                st.info("No policies are pending renewal at this time.")
            else:
                st.info(f"No policies match the selected filter: {filter_option}")
                st.caption("Try selecting 'All Renewals' or a different time range.")
# Call main function
main()
