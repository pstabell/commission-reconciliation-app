import streamlit as st
st.set_page_config(layout="wide")
import traceback
import string
import random
import pandas as pd
import numpy as np
from supabase import create_client, Client
from dotenv import load_dotenv
import hashlib
import re

# Set pandas display options to avoid scientific notation and show 2 decimal places
pd.options.display.float_format = '{:.2f}'.format

# Load environment variables
load_dotenv()
import datetime
import time
import io
import streamlit_sortables
import os
import json
import pdfplumber
import shutil
import uuid
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from column_mapping_config import (
    column_mapper, get_mapped_column, get_ui_field_name, 
    is_calculated_field, safe_column_reference
)

def check_password():
    """Returns True if the user had the correct password."""
    
    def password_entered():
        """Checks whether a password entered by the user is correct."""
        # For initial setup, we'll use a simple password check
        # You should change this password and consider using environment variables
        correct_password = os.getenv("APP_PASSWORD", "CommissionApp2025!")  # Change this!
        
        # Check if password key exists before accessing it
        if "password" in st.session_state and st.session_state["password"] == correct_password:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # Don't store password
        else:
            st.session_state["password_correct"] = False

    # Return True if the password is validated
    if st.session_state.get("password_correct", False):
        return True

    # Show input for password
    st.title("🔐 Sales Commission App - Login")
    st.text_input(
        "Password", 
        type="password", 
        on_change=password_entered, 
        key="password",
        help="Contact administrator for password"
    )
    
    if "password_correct" in st.session_state and not st.session_state["password_correct"]:
        st.error("😕 Password incorrect. Please try again.")
    
    st.info("This application contains sensitive commission data. Authentication is required.")
    return False

@st.cache_resource
def get_supabase_client():
    """Get cached Supabase client."""
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_ANON_KEY")
    if not url or not key:
        st.error("Missing Supabase credentials. Please check your .env file.")
        st.stop()
    return create_client(url, key)

@st.cache_data(ttl=300)  # Cache for 5 minutes
def load_policies_data():
    """Load policies data from Supabase with caching."""
    try:
        supabase = get_supabase_client()
        response = supabase.table('policies').select("*").execute()
        if response.data:
            df = pd.DataFrame(response.data)
            # Ensure numeric columns are properly typed
            numeric_cols = [
                'Agent Estimated Comm $',
                'Policy Gross Comm %',
                'Agency Estimated Comm/Revenue (CRM)',
                'Agency Comm Received (STMT)',
                'Premium Sold',
                'Agent Paid Amount (STMT)',
                'Agency Comm Received (STMT)'
            ]
            
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Round all numeric columns to 2 decimal places
            df = round_numeric_columns(df)
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading data from Supabase: {e}")
        return pd.DataFrame()

def clear_policies_cache():
    """Clear the policies data cache."""
    load_policies_data.clear()

def log_debug(message, level="INFO", error_obj=None):
    """Add a debug log entry to session state."""
    if "debug_logs" not in st.session_state:
        st.session_state.debug_logs = []
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    log_entry = {
        "timestamp": timestamp,
        "level": level,
        "message": message
    }
    
    if error_obj:
        log_entry["error_details"] = str(error_obj)
        if hasattr(error_obj, '__dict__'):
            log_entry["error_attrs"] = str(error_obj.__dict__)
    
    st.session_state.debug_logs.append(log_entry)
    
    # Keep only last 500 entries
    if len(st.session_state.debug_logs) > 500:
        st.session_state.debug_logs = st.session_state.debug_logs[-500:]

def clear_debug_logs():
    """Clear all debug logs."""
    st.session_state.debug_logs = []

def is_reconciliation_transaction(transaction_id):
    """
    Check if transaction is a reconciliation entry that should be locked.
    Returns True for -STMT-, -VOID-, -ADJ- transactions.
    """
    if not transaction_id:
        return False
    
    transaction_id_str = str(transaction_id)
    reconciliation_types = ['-STMT-', '-VOID-', '-ADJ-']
    
    return any(suffix in transaction_id_str for suffix in reconciliation_types)

def apply_formula_display(df, show_formulas=True):
    """
    Apply formula calculations to existing columns with indicators.
    Optionally shows formulas or actual values based on toggle.
    """
    if df.empty:
        return df
    
    df = df.copy()
    
    # Store original values for comparison
    df['_original_agency'] = df['Agency Estimated Comm/Revenue (CRM)'].copy()
    df['_original_agent'] = df['Agent Estimated Comm $'].copy()
    
    if show_formulas:
        # Calculate formula values
        df['_formula_agency'] = df.apply(
            lambda row: (
                0.0 if pd.isna(row.get('Premium Sold', 0)) or pd.isna(row.get('Policy Gross Comm %', 0))
                else float(row.get('Premium Sold', 0) or 0) * float(row.get('Policy Gross Comm %', 0) or 0) / 100
            ),
            axis=1
        )
        
        # Get agent commission rate based on transaction type
        def get_agent_rate(row):
            trans_type = row.get('Transaction Type', '')
            if trans_type in ['NEW', 'NBS', 'STL', 'BoR']:
                return 50.0
            elif trans_type in ['RWL', 'REWRITE']:
                return 25.0
            elif trans_type in ['CAN', 'XCL']:
                return 0.0
            elif trans_type in ['END', 'PCH']:
                # Check if new business or renewal based on dates
                if row.get('Policy Origination Date') == row.get('Effective Date'):
                    return 50.0
                else:
                    return 25.0
            else:
                # Default to agent comm rate if available
                agent_rate = row.get('Agent Comm (NEW 50% RWL 25%)', 0)
                if agent_rate and agent_rate < 1:
                    return agent_rate * 100  # Convert decimal to percentage
                return agent_rate or 0
        
        # Calculate Agent formula
        df['_formula_agent'] = df.apply(
            lambda row: (
                row['_formula_agency'] * get_agent_rate(row) / 100
            ),
            axis=1
        )
        
        # Apply formulas to display columns
        df['Agency Estimated Comm/Revenue (CRM)'] = df['_formula_agency'].round(2)
        df['Agent Estimated Comm $'] = df['_formula_agent'].round(2)
        
        # Add indicators based on variance
        df['_agency_indicator'] = df.apply(
            lambda row: (
                '🔒' if row.get('Transaction ID', '').find('-STMT-') >= 0 or 
                       row.get('Transaction ID', '').find('-VOID-') >= 0 or
                       row.get('Transaction ID', '').find('-ADJ-') >= 0
                else '⚠️' if pd.isna(row.get('Premium Sold')) or pd.isna(row.get('Policy Gross Comm %')) or
                             row.get('Premium Sold', 0) == 0 or row.get('Policy Gross Comm %', 0) == 0
                else '✏️' if abs(float(row.get('_original_agency', 0) or 0) - row['_formula_agency']) > 0.01
                else '✓'
            ),
            axis=1
        )
        
        df['_agent_indicator'] = df.apply(
            lambda row: (
                '🔒' if row.get('Transaction ID', '').find('-STMT-') >= 0 or 
                       row.get('Transaction ID', '').find('-VOID-') >= 0 or
                       row.get('Transaction ID', '').find('-ADJ-') >= 0
                else '⚠️' if pd.isna(row.get('Premium Sold')) or pd.isna(row.get('Policy Gross Comm %')) or
                             row.get('Premium Sold', 0) == 0 or row.get('Policy Gross Comm %', 0) == 0
                else '✏️' if abs(float(row.get('_original_agent', 0) or 0) - row['_formula_agent']) > 0.01
                else '✓'
            ),
            axis=1
        )
        
        # Format display with indicators
        df['Agency Estimated Comm/Revenue (CRM)'] = df.apply(
            lambda row: f"${row['Agency Estimated Comm/Revenue (CRM)']:.2f} {row['_agency_indicator']}",
            axis=1
        )
        
        df['Agent Estimated Comm $'] = df.apply(
            lambda row: f"${row['Agent Estimated Comm $']:.2f} {row['_agent_indicator']}",
            axis=1
        )
    
    # Clean up temporary columns
    temp_cols = ['_original_agency', '_original_agent', '_formula_agency', '_formula_agent', 
                 '_agency_indicator', '_agent_indicator']
    for col in temp_cols:
        if col in df.columns:
            df = df.drop(columns=[col])
    
    return df

@st.cache_data
def get_custom_css():
    """Get cached CSS for better performance."""
    return """<style>        /* Remove default padding and maximize main block width */
        .main .block-container {
            padding-top: 1rem;
            padding-bottom: 0rem;
            padding-left: 1.5rem;
            padding-right: 1.5rem;
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
            max-width: calc(100vw - 240px) !important;
        }
        
        /* Ensure main content area starts after sidebar */
        .main {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
            max-width: calc(100vw - 240px) !important;
        }
        
        /* Additional targeting for content wrapper */
        [data-testid="stAppViewContainer"] > .main {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
        }
          /* Force all main content to respect sidebar space */
        section.main {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
        }
        
        /* Target the root app container - this is critical! */
        [data-testid="stAppViewContainer"] {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
            max-width: calc(100vw - 240px) !important;
        }
        
        /* Also target the main header area */
        [data-testid="stHeader"] {
            margin-left: 240px !important;
            width: calc(100vw - 240px) !important;
        }
        
        /* Force sidebar to be permanently visible and disable all collapse functionality */
        section[data-testid="stSidebar"] {
            min-width: 240px !important;
            max-width: 240px !important;
            width: 240px !important;
            display: block !important;
            visibility: visible !important;
            position: fixed !important;
            left: 0 !important;
            top: 0 !important;
            height: 100vh !important;
            z-index: 999 !important;
            transform: translateX(0px) !important;
            margin-left: 0px !important;
        }
        
        /* Force sidebar to stay expanded even when marked as collapsed */
        section[data-testid="stSidebar"][aria-expanded="false"] {
            transform: translateX(0px) !important;
            margin-left: 0px !important;
            left: 0 !important;
        }
        
        /* Hide ALL possible collapse/hide buttons with comprehensive selectors */
        button[data-testid="collapsedControl"],
        button[kind="header"],
        [data-testid="stSidebarNav"] button[kind="header"],
        [data-testid="stSidebar"] button[kind="header"],
        [data-testid="stSidebar"] [data-testid="collapsedControl"],
        [data-testid="stSidebar"] button[title*="collapse"],
        [data-testid="stSidebar"] button[title*="hide"],
        [data-testid="stSidebar"] button[aria-label*="collapse"],
        [data-testid="stSidebar"] button[aria-label*="hide"],
        .stSidebar button[kind="header"],
        section[data-testid="stSidebar"] > div > button,
        section[data-testid="stSidebar"] button:first-child {
            display: none !important;
            visibility: hidden !important;
            opacity: 0 !important;
            pointer-events: none !important;
            width: 0 !important;
            height: 0 !important;
        }
        
        /* Ensure sidebar content is always visible */
        [data-testid="stSidebar"] > div {
            display: block !important;
            visibility: visible !important;
            opacity: 1 !important;
        }
        
        /* Force sidebar radio buttons to be visible */
        [data-testid="stSidebar"] .stRadio {
            display: block !important;
            visibility: visible !important;
        }
        
        /* Override any CSS that might hide the sidebar */
        [data-testid="stSidebar"] {
            opacity: 1 !important;
            pointer-events: auto !important;
        }
        /* Remove extra margin from header/title */
        .main .block-container h1 {
            margin-bottom: 0.5rem;
        }
        /* Highlight all interactive input fields in Add New Policy Transaction form and Admin Panel rename headers */
        .stForm input:not([disabled]), .stForm select:not([disabled]), .stForm textarea:not([disabled]),
        .stTextInput > div > input:not([disabled]), .stNumberInput > div > input:not([disabled]), .stDateInput > div > input:not([disabled]) {
            background-color: #fff3b0 !important; /* Darker yellow */
            border: 2px solid #e6a800 !important; /* Darker yellow border */
            border-radius: 6px !important;
        }
        /* Make selectboxes match other fields and remove focus ring */
        .stSelectbox > div[data-baseweb="select"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        .stSelectbox > div[data-baseweb="select"]:focus,
        .stSelectbox > div[data-baseweb="select"]:active,
        .stSelectbox > div[data-baseweb="select"]:focus-visible {
            border: 2px solid #e6a800 !important;
            box-shadow: none !important;
            outline: none !important;
        }
        /* Add yellow border to the client name search input at the top of Add New Policy Transaction */
        input[type="text"][aria-label="Type client name to search:"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        /* Add yellow border to Premium Sold Calculator inputs */
        input[type="number"][aria-label="Existing Premium"],
        input[type="number"][aria-label="New/Revised Premium"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        /* Add yellow border to Enter Premium Sold and see Agency Estimated Comm/Revenue (CRM) input */
        input[type="number"][aria-label="Premium Sold"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        /* Custom fatter scrollbars for all tables and editors */
        .stDataFrame ::-webkit-scrollbar, .stDataEditor ::-webkit-scrollbar {
            height: 18px;
            width: 18px;
        }
        .stDataFrame ::-webkit-scrollbar-thumb, .stDataEditor ::-webkit-scrollbar-thumb {
            background: #888888;
            border-radius: 8px;
            border: 3px solid #b0b0b0;
        }
        .stDataFrame ::-webkit-scrollbar-track, .stDataEditor ::-webkit-scrollbar-track {
            background: #b0b0b0;
            border-radius: 8px;
        }
        /* For Firefox */
        .stDataFrame, .stDataEditor {
            scrollbar-width: thick;
            scrollbar-color: #888888 #b0b0b0;
        }
        /* Highlight selectboxes in Admin Panel reorder columns section */
        [id^="reorder_col_"] > div[data-baseweb="select"] {
            background-color: #fff3b0 !important;
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        [id^="reorder_col_"] > div[data-baseweb="select"]:focus,
        [id^="reorder_col_"] > div[data-baseweb="select"]:active,
        [id^="reorder_col_"] > div[data-baseweb="select"]:focus-visible {
            border: 2px solid #e6a800 !important;
            box-shadow: none !important;
            outline: none !important;
        }
        /* Highlight selectboxes in Admin Panel reorder columns section (force all backgrounds) */
        [data-testid^="stSelectbox"][id^="reorder_col_"] > div[data-baseweb="select"],
        [id^="reorder_col_"] [data-baseweb="select"],
        [id^="reorder_col_"] [class*="css-1wa3eu0-placeholder"],
        [id^="reorder_col_"] [class*="css-1uccc91-singleValue"],
        [id^="reorder_col_"] [class*="css-1okebmr-indicatorSeparator"],
        [id^="reorder_col_"] [class*="css-1pahdxg-control"],
        [id^="reorder_col_"] [class*="css-1s2u09g-control"],
        [id^="reorder_col_"] [class*="css-1n7v3ny-option"],
        [id^="reorder_col_"] [class*="css-9gakcf-option"],
        [id^="reorder_col_"] [class*="css-1n6sfyn-MenuList"],
        [id^="reorder_col_"] [class*="css-1n6sfyn-MenuList"] * {
            background-color: #fff3b0 !important;
            border-color: #e6a800 !important;
            color: #222 !important;
        }        /* Force highlight for all selectboxes labeled 'Transaction Type' everywhere */
        label:has(> div[data-baseweb="select"]) {
            background-color: #fff3b0 !important;
            border-radius: 6px !important;
            padding: 2px 4px !important;
        }
        /* Also target selectbox input and dropdown for 'Transaction Type' */
        [aria-label="Transaction Type"] > div[data-baseweb="select"],
        [aria-label="Transaction Type"] [data-baseweb="select"],
        [aria-label="Transaction Type"] [class*="css-1wa3eu0-placeholder"],
        [aria-label="Transaction Type"] [class*="css-1uccc91-singleValue"],
        [aria-label="Transaction Type"] [class*="css-1okebmr-indicatorSeparator"],
        [aria-label="Transaction Type"] [class*="css-1pahdxg-control"],
        [aria-label="Transaction Type"] [class*="css-1s2u09g-control"],
        [aria-label="Transaction Type"] [class*="css-1n7v3ny-option"],
        [aria-label="Transaction Type"] [class*="css-9gakcf-option"],
        [aria-label="Transaction Type"] [class*="css-1n6sfyn-MenuList"],
        [aria-label="Transaction Type"] [class*="css-1n6sfyn-MenuList"] * {
            background-color: #fff3b0 !important;
            border-color: #e6a800 !important;
            color: #222 !important;
        }
        [aria-label="Transaction Type"] > div[data-baseweb="select"] {
            border: 2px solid #e6a800 !important;
            border-radius: 6px !important;
        }
        [aria-label="Transaction Type"] > div[data-baseweb="select"]:focus,
        [aria-label="Transaction Type"] > div[data-baseweb="select"]:active,
        [aria-label="Transaction Type"] > div[data-baseweb="select"]:focus-visible {
            border: 2px solid #e6a800 !important;
            box-shadow: none !important;
            outline: none !important;
        }
        
        /* Maintain scroll position in data editor */
        [data-testid="stDataFrame"] > div,
        [data-testid="stDataEditor"] > div {
            scroll-behavior: auto !important;
        }
        
        /* Prevent auto-scroll to start */
        .stDataEditor {
            overflow-x: auto !important;
            overflow-y: auto !important;
        }
        </style>"""

def apply_css():
    """Apply custom CSS styling to the Streamlit app."""
    with st.container():
        st.markdown(get_custom_css(), unsafe_allow_html=True)

def format_currency(val):
    """Format the value as currency for display."""
    if pd.isna(val) or val is None:
        return ""
    try:
        return f"${val:,.2f}"
    except Exception:
        return val

def round_numeric_columns(df):
    """Round all numeric columns to 2 decimal places to avoid floating point precision issues."""
    numeric_columns = df.select_dtypes(include=['float64', 'float32', 'float', 'number']).columns
    for col in numeric_columns:
        df[col] = df[col].round(2)
    
    # Also handle columns that might be stored as strings but contain numeric values
    percentage_columns = ['Policy Gross Comm %', 'Agent Comm (NEW 50% RWL 25%)', 'Agent Comm (New 50% RWL 25%)']
    for col in percentage_columns:
        if col in df.columns:
            # Convert to numeric if it's a string, then round
            df[col] = pd.to_numeric(df[col], errors='coerce').round(2)
    
    return df

def clean_numeric_value(value):
    """Clean and round a numeric value to 2 decimal places."""
    if pd.isna(value) or value is None:
        return None
    try:
        # Convert to float and round to 2 decimal places
        return round(float(value), 2)
    except (ValueError, TypeError):
        return value

def format_dates_mmddyyyy(df):
    """Format date columns in MM/DD/YYYY format using mapped column names."""
    date_field_names = ["Policy Origination Date", "Effective Date", "X-Date", "X-DATE", "Statement Date", "STMT DATE"]
    
    for ui_field in date_field_names:
        # Try mapped column first, then fallback to exact match
        mapped_col = get_mapped_column(ui_field)
        target_col = mapped_col if mapped_col and mapped_col in df.columns else ui_field
        
        if target_col in df.columns:
            df[target_col] = pd.to_datetime(df[target_col], errors="coerce")
            df[target_col] = df[target_col].dt.strftime("%m/%d/%Y")
    return df

CURRENCY_COLUMNS = [
    "Premium Sold",    "Agency Comm Received (STMT)",
    "Gross Premium Paid",
    "Agency Gross Comm",
    "Agent Paid Amount (STMT)",
    "Estimated Agent Comm",
    "Policy Balance Due",  # Updated to use the new calculated column
    "Agent Estimated Comm $",
    "Agency Estimated Comm/Revenue (CRM)",
    "Agency Estimated Comm/Revenue (CRM)"
]

def format_currency_columns(df):
    # Use mapped column names for currency formatting
    for ui_field in CURRENCY_COLUMNS:
        mapped_col = get_mapped_column(ui_field)
        if mapped_col and mapped_col in df.columns:
            df[mapped_col] = df[mapped_col].apply(format_currency)
        elif ui_field in df.columns:  # Fallback to original name
            df[ui_field] = df[ui_field].apply(format_currency)
    return df

# --- Helper Functions ---

# Column mapping persistence functions
def load_saved_column_mappings():
    """Load saved column mappings from JSON file."""
    mappings_file = os.path.join('config_files', 'saved_mappings.json')
    try:
        if os.path.exists(mappings_file):
            with open(mappings_file, 'r') as f:
                return json.load(f)
    except Exception as e:
        st.error(f"Error loading saved mappings: {str(e)}")
    return {}

def save_column_mappings_to_file(mappings):
    """Save column mappings to JSON file."""
    mappings_file = os.path.join('config_files', 'saved_mappings.json')
    try:
        # Ensure the config_files directory exists
        os.makedirs('config_files', exist_ok=True)
        
        with open(mappings_file, 'w') as f:
            json.dump(mappings, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving mappings: {str(e)}")
        return False

def generate_client_id(length=6):
    """Generate a unique Client ID with exactly 3 letters and 3 numbers."""
    # Generate exactly 3 letters and 3 numbers
    letters = string.ascii_uppercase
    digits = string.digits
    
    # Pick 3 random letters
    selected_letters = [random.choice(letters) for _ in range(3)]
    
    # Pick 3 random numbers
    selected_numbers = [random.choice(digits) for _ in range(3)]
    
    # Combine them
    result = selected_letters + selected_numbers
    
    # Shuffle to create random pattern
    random.shuffle(result)
    return ''.join(result)

def generate_transaction_id(length=7):
    """Generate a unique Transaction ID with at least 3 letters and 3 numbers."""
    # Ensure at least 3 letters and 3 numbers for 7-character ID
    letters = string.ascii_uppercase
    digits = string.digits
    
    # Pick 3 random letters
    selected_letters = [random.choice(letters) for _ in range(3)]
    
    # Pick 3 random numbers
    selected_numbers = [random.choice(digits) for _ in range(3)]
    
    # For the 7th character, randomly choose letter or number
    if random.choice([True, False]):
        extra_char = random.choice(letters)
    else:
        extra_char = random.choice(digits)
    
    # Combine all characters
    result = selected_letters + selected_numbers + [extra_char]
    
    # Shuffle to create random pattern
    random.shuffle(result)
    return ''.join(result)

def generate_reconciliation_transaction_id(transaction_type="STMT", date=None):
    """Generate a reconciliation transaction ID with format: XXXXXXX-TYPE-YYYYMMDD"""
    base_id = generate_transaction_id()
    
    if date is None:
        date = datetime.datetime.now()
    
    date_str = date.strftime("%Y%m%d")
    
    return f"{base_id}-{transaction_type}-{date_str}"

# --- Excel Utility Functions ---
def create_formatted_excel_file(data, sheet_name="Data", filename_prefix="export"):
    """
    Create a formatted Excel file from DataFrame with professional styling.
    
    Args:
        data (pd.DataFrame): Data to export
        sheet_name (str): Name for the Excel sheet
        filename_prefix (str): Prefix for the filename
    
    Returns:
        io.BytesIO: Excel file buffer
        str: Suggested filename
    """
    try:
        # Create filename with timestamp
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        
        # Create Excel buffer
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Write data to Excel
            data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with reasonable limits
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
            
            # Format currency columns
            currency_columns = ['Commission_Paid', 'Agency_Commission_Received', 'Balance_Due', 
                              'Agency Estimated Comm/Revenue (CRM)', 'Premium_Amount']
            
            for col_name in currency_columns:
                if col_name in data.columns:
                    col_letter = None
                    for col_num, column_name in enumerate(data.columns, 1):
                        if column_name == col_name:
                            col_letter = openpyxl.utils.get_column_letter(col_num)
                            break
                    
                    if col_letter:
                        for row_num in range(2, len(data) + 2):  # Start from row 2 (after header)
                            cell = worksheet[f"{col_letter}{row_num}"]
                            cell.number_format = '"$"#,##0.00'
        
        excel_buffer.seek(0)
        return excel_buffer, filename
        
    except Exception as e:
        st.error(f"Error creating Excel file: {e}")
        return None, None

def create_multi_sheet_excel(data_dict, filename_prefix="multi_sheet_export"):
    """
    Create Excel file with multiple sheets and metadata.
    
    Args:
        data_dict (dict): Dictionary with sheet_name: data pairs
        filename_prefix (str): Prefix for filename
    
    Returns:
        io.BytesIO: Excel file buffer
        str: Suggested filename
    """
    try:
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Create metadata sheet first
            metadata = pd.DataFrame([
                ['Export Date', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Application', 'Commission Management System'],
                ['Export Type', 'Multi-Sheet Report'],
                ['Total Sheets', len(data_dict)],
                ['Sheet Names', ', '.join(data_dict.keys())]
            ], columns=['Parameter', 'Value'])
            
            metadata.to_excel(writer, sheet_name='Export Info', index=False)
            
            # Format metadata sheet
            workbook = writer.book
            metadata_sheet = writer.sheets['Export Info']
            
            # Style metadata headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in metadata_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust metadata columns
            metadata_sheet.column_dimensions['A'].width = 25
            metadata_sheet.column_dimensions['B'].width = 50
            
            # Add data sheets
            for sheet_name, data in data_dict.items():
                if not data.empty:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Format data sheet
                    worksheet = writer.sheets[sheet_name]
                    
                    # Format headers
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Auto-adjust columns
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 30)
                        worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
        
        excel_buffer.seek(0)
        return excel_buffer, filename
        
    except Exception as e:
        st.error(f"Error creating multi-sheet Excel file: {e}")
        return None, None

def validate_excel_import(uploaded_file):
    """
    Validate uploaded Excel file and return DataFrame with validation results.
    
    Args:
        uploaded_file: Streamlit uploaded file object
        
    Returns:
        tuple: (success: bool, data: DataFrame or None, errors: list)
    """
    try:
        # Read Excel file
        df = pd.read_excel(uploaded_file, engine='openpyxl')
        
        validation_errors = []
        
        # Check if file is empty
        if df.empty:
            validation_errors.append("Excel file is empty")
            return False, None, validation_errors
        
        # Check for required columns (based on column mapping)
        required_columns = [
            'Client_ID', 'Transaction_ID', 'Policy_Type', 
            'Transaction_Type', 'Effective_Date'
        ]
        
        missing_columns = []
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)
        
        if missing_columns:
            validation_errors.append(f"Missing required columns: {', '.join(missing_columns)}")
        
        # Validate data types and formats
        if 'Effective_Date' in df.columns:
            try:
                pd.to_datetime(df['Effective_Date'])
            except:
                validation_errors.append("Invalid date format in Effective_Date column")
        
        # Check for duplicate Transaction_IDs
        if 'Transaction_ID' in df.columns:
            duplicates = df['Transaction_ID'].duplicated().sum()
            if duplicates > 0:
                validation_errors.append(f"Found {duplicates} duplicate Transaction_IDs")
        
        # Validate numeric columns
        numeric_columns = ['Commission_Paid', 'Agency_Commission_Received', 'Balance_Due']
        for col in numeric_columns:
            if col in df.columns:
                try:
                    pd.to_numeric(df[col], errors='coerce')
                except:
                    validation_errors.append(f"Invalid numeric data in {col} column")
        
        # Return validation results
        if validation_errors:
            return False, df, validation_errors
        else:
            return True, df, []
            
    except Exception as e:
        return False, None, [f"Error reading Excel file: {str(e)}"]

# --- Commission calculation function ---
def calculate_commission(row):
    try:
        agency_revenue_col = get_mapped_column("Agency Estimated Comm/Revenue (CRM)")
        revenue = float(row[agency_revenue_col]) if agency_revenue_col and row[agency_revenue_col] is not None else 0.0
    except (ValueError, TypeError, KeyError):
        revenue = 0.0
    
    transaction_type_col = get_mapped_column("Transaction Type")
    policy_orig_col = get_mapped_column("Policy Origination Date") 
    effective_date_col = get_mapped_column("Effective Date")
    
    try:
        transaction_type = row.get(transaction_type_col, "") if transaction_type_col else ""
        if transaction_type in ["NEW", "NBS", "STL", "BoR"]:
            return revenue * 0.50
        elif transaction_type in ["END", "PCH"]:
            policy_orig = row.get(policy_orig_col, "") if policy_orig_col else ""
            effective_date = row.get(effective_date_col, "") if effective_date_col else ""
            return revenue * 0.50 if policy_orig == effective_date else revenue * 0.25
        elif transaction_type in ["RWL", "REWRITE"]:
            return revenue * 0.25
        elif transaction_type in ["CAN", "XCL"]:
            return 0
        else:
            return revenue * 0.25
    except (KeyError, TypeError):
        return revenue * 0.25

def get_pending_renewals(df: pd.DataFrame) -> pd.DataFrame:
    """
    Identifies and generates a DataFrame of policies pending renewal.
    """
    # Filter for relevant transaction types
    renewal_candidates = df[df[get_mapped_column("Transaction Type")].isin(["NEW", "RWL"])].copy()
    
    # Convert date columns to datetime objects
    renewal_candidates['expiration_date'] = pd.to_datetime(renewal_candidates[get_mapped_column("X-DATE")], errors='coerce')
    
    # Sort by policy number and expiration date to find the latest transaction
    renewal_candidates = renewal_candidates.sort_values(by=["Policy Number", "expiration_date"], ascending=[True, False])
    
    # Get the most recent transaction for each policy
    latest_renewals = renewal_candidates.drop_duplicates(subset="Policy Number", keep="first")
    
    # Filter for policies that are expired or expiring soon (e.g., within 60 days)
    today = pd.to_datetime(datetime.date.today())
    pending_renewals = latest_renewals[latest_renewals['expiration_date'] < (today + pd.DateOffset(days=60))]
    
    return pending_renewals

def normalize_business_name(name):
    """
    Normalize business names by removing common suffixes and punctuation.
    Helps match "RCM Construction" to "RCM Construction of SWFL LLC"
    """
    if not name:
        return ""
    
    # Common business suffixes to remove
    suffixes = [
        'LLC', 'L.L.C.', 'L.L.C', 'Inc', 'Inc.', 'Incorporated',
        'Corp', 'Corp.', 'Corporation', 'Ltd', 'Ltd.', 'Limited',
        'PA', 'P.A.', 'PC', 'P.C.', 'PLLC', 'P.L.L.C.',
        'LLP', 'L.L.P.', 'LP', 'L.P.', 'Company', 'Co.', 'Co'
    ]
    
    # Convert to string and strip
    normalized = str(name).strip()
    
    # Remove suffixes (case-insensitive)
    for suffix in suffixes:
        # Check both with and without leading comma
        patterns = [
            rf',?\s+{re.escape(suffix)}\s*$',
            rf'\s+{re.escape(suffix)}\s*$'
        ]
        for pattern in patterns:
            normalized = re.sub(pattern, '', normalized, flags=re.IGNORECASE)
    
    # Remove "of [Location]" patterns
    normalized = re.sub(r'\s+of\s+[A-Z][A-Za-z\s]+(?:LLC|Inc|Corp)?$', '', normalized, flags=re.IGNORECASE)
    
    # Clean up extra whitespace and punctuation
    normalized = ' '.join(normalized.split())
    normalized = normalized.strip(' ,.-')
    
    return normalized

def find_potential_customer_matches(search_name, existing_customers):
    """
    Find potential customer matches using various strategies.
    Returns list of (customer_name, match_type, score) tuples.
    """
    if not search_name:
        return []
    
    search_name_lower = search_name.lower().strip()
    search_normalized = normalize_business_name(search_name).lower()
    search_first_word = search_name.split()[0].lower() if search_name else ""
    
    # Handle "Last, First" format
    search_name_reversed = ""
    if "," in search_name:
        parts = search_name.split(",", 1)
        if len(parts) == 2:
            search_name_reversed = f"{parts[1].strip()} {parts[0].strip()}".lower()
    
    matches = {}  # Use dict to avoid duplicates
    
    for customer in existing_customers:
        if not customer:
            continue
            
        customer_lower = customer.lower().strip()
        customer_normalized = normalize_business_name(customer).lower()
        customer_first_word = customer.split()[0].lower() if customer else ""
        
        # 1. Exact match (highest priority)
        if search_name_lower == customer_lower:
            matches[customer] = ('exact', 100)
            continue
        
        # 2. Reversed name match (Last, First -> First Last)
        if search_name_reversed and search_name_reversed == customer_lower:
            if customer not in matches or matches[customer][1] < 98:
                matches[customer] = ('name_reversed', 98)
            continue
        
        # 3. Normalized match (very high priority)
        if search_normalized and search_normalized == customer_normalized:
            if customer not in matches or matches[customer][1] < 95:
                matches[customer] = ('normalized', 95)
            continue
        
        # 4. First word match (e.g., "Barboun" matches "Barboun, Thomas")
        if search_first_word and customer_first_word == search_first_word:
            if customer not in matches or matches[customer][1] < 90:
                matches[customer] = ('first_word', 90)
            continue
        
        # 5. Contains match (e.g., "RCM" in "RCM Construction")
        if len(search_name) >= 3:
            if search_name_lower in customer_lower:
                if customer not in matches or matches[customer][1] < 85:
                    matches[customer] = ('contains', 85)
                continue
            
            # Check if search is contained in normalized version
            if search_normalized in customer_normalized:
                if customer not in matches or matches[customer][1] < 83:
                    matches[customer] = ('normalized_contains', 83)
                continue
        
        # 6. Customer contains search (e.g., searching "RCM Construction" finds "RCM Construction of SWFL LLC")
        if customer_lower in search_name_lower and len(customer) >= 3:
            if customer not in matches or matches[customer][1] < 80:
                matches[customer] = ('reverse_contains', 80)
            continue
        
        # 7. Starts with match
        if customer_lower.startswith(search_name_lower[:3]) and len(search_name) >= 3:
            if customer not in matches or matches[customer][1] < 75:
                matches[customer] = ('starts_with', 75)
    
    # Convert to sorted list
    result = [(name, match_type, score) for name, (match_type, score) in matches.items()]
    result.sort(key=lambda x: (-x[2], x[0]))  # Sort by score desc, then name
    
    return result

def calculate_transaction_balances(all_data):
    """
    Calculate outstanding balances for all transactions.
    Reuses the exact logic from Unreconciled Transactions tab.
    Returns DataFrame with _balance column added.
    """
    if all_data.empty:
        return pd.DataFrame()
    
    # Get original transactions only (exclude -STMT-, -ADJ-, -VOID-)
    original_trans = all_data[
        ~all_data['Transaction ID'].str.contains('-STMT-|-ADJ-|-VOID-', na=False)
    ].copy()
    
    if original_trans.empty:
        return pd.DataFrame()
    
    # Calculate balance for each transaction
    for idx, row in original_trans.iterrows():
        # Calculate credits (commission owed)
        credit = float(row.get('Agent Estimated Comm $', 0) or 0)
        
        # Calculate debits (total paid for this policy)
        policy_num = row['Policy Number']
        effective_date = row['Effective Date']
        
        # Get all STMT entries for this specific policy and date
        stmt_entries = all_data[
            (all_data['Policy Number'] == policy_num) &
            (all_data['Effective Date'] == effective_date) &
            (all_data['Transaction ID'].str.contains('-STMT-', na=False))
        ]
        
        debit = 0
        if not stmt_entries.empty:
            debit = stmt_entries['Agent Paid Amount (STMT)'].fillna(0).sum()
        
        # Calculate balance
        balance = credit - debit
        original_trans.at[idx, '_balance'] = balance
    
    # Return only transactions with outstanding balance
    return original_trans[original_trans['_balance'] > 0.01]

def match_statement_transactions(statement_df, column_mapping, existing_data, statement_date):
    """
    Match statement transactions to existing database records.
    Now uses the same balance calculation as Unreconciled Transactions tab.
    Returns: (matched_list, unmatched_list, can_create_list)
    """
    matched = []
    unmatched = []
    can_create = []
    
    # Get transactions with outstanding balances using the same logic as Unreconciled tab
    outstanding_trans = calculate_transaction_balances(existing_data)
    
    # Build lookup dictionaries
    existing_lookup = {}
    customer_trans_lookup = {}  # Customer -> list of transactions
    
    if not outstanding_trans.empty:
        for idx, trans in outstanding_trans.iterrows():
            trans_dict = trans.to_dict()
            trans_dict['balance'] = trans['_balance']
            
            # Store by policy key - normalize date format for matching
            eff_date_normalized = trans['Effective Date']
            if pd.notna(eff_date_normalized):
                try:
                    eff_date_normalized = pd.to_datetime(eff_date_normalized).strftime('%Y-%m-%d')
                except:
                    eff_date_normalized = str(eff_date_normalized)
            policy_key = f"{trans['Policy Number']}_{eff_date_normalized}"
            existing_lookup[policy_key] = trans_dict
            
            # Store by customer (for fuzzy matching)
            customer = trans['Customer']
            if pd.notna(customer):
                customer_key = customer.lower().strip()
                if customer_key not in customer_trans_lookup:
                    customer_trans_lookup[customer_key] = []
                customer_trans_lookup[customer_key].append(trans_dict)
    
    # Get unique customer list for fuzzy matching
    all_customers = []
    if not existing_data.empty:
        all_customers = existing_data['Customer'].dropna().unique().tolist()
    
    # Process each statement row
    for idx, row in statement_df.iterrows():
        # Extract mapped values
        customer = str(row[column_mapping['Customer']]).strip() if 'Customer' in column_mapping else ''
        policy_num = str(row[column_mapping['Policy Number']]).strip() if 'Policy Number' in column_mapping else ''
        eff_date = row[column_mapping['Effective Date']] if 'Effective Date' in column_mapping else None
        
        # Skip rows that appear to be totals
        # Check if customer name contains common total indicators
        customer_lower = customer.lower()
        if any(total_word in customer_lower for total_word in ['total', 'totals', 'subtotal', 'sub-total', 'grand total', 'sum']):
            continue
        
        # Also skip if policy number or customer is empty/missing and it looks like a summary row
        if (not customer or customer.lower() in ['', 'nan', 'none']) and (not policy_num or policy_num.lower() in ['', 'nan', 'none']):
            continue
        # Primary reconciliation amount is what the agent was paid
        amount = float(row[column_mapping['Agent Paid Amount (STMT)']]) if 'Agent Paid Amount (STMT)' in column_mapping else 0
        # Agency amount is optional for audit purposes
        agency_amount = float(row[column_mapping['Agency Comm Received (STMT)']]) if 'Agency Comm Received (STMT)' in column_mapping else 0
        
        # Convert effective date to string format
        if pd.notna(eff_date):
            try:
                eff_date = pd.to_datetime(eff_date).strftime('%Y-%m-%d')
            except:
                eff_date = str(eff_date)
        
        match_result = {
            'row_index': idx,
            'customer': customer,
            'policy_number': policy_num,
            'effective_date': eff_date,
            'amount': amount,  # Agent Paid Amount (primary)
            'agency_amount': agency_amount,  # Agency Comm Received (audit)
            'statement_data': row.to_dict()
        }
        
        # Try to match: Policy Number + Effective Date (highest confidence)
        policy_key = f"{policy_num}_{eff_date}"
        if policy_key in existing_lookup:
            match_result['match'] = existing_lookup[policy_key]
            match_result['confidence'] = 100
            match_result['match_type'] = 'Policy + Date'
            matched.append(match_result)
            continue
        
        # Try enhanced customer matching
        potential_customers = find_potential_customer_matches(customer, all_customers)
        
        if potential_customers:
            # Check if we have a single high-confidence match
            if len(potential_customers) == 1 and potential_customers[0][2] >= 90:
                # Single high-confidence match - try to find transaction
                matched_customer = potential_customers[0][0]
                customer_key = matched_customer.lower().strip()
                
                if customer_key in customer_trans_lookup:
                    customer_trans = customer_trans_lookup[customer_key]
                    # Check for amount match among transactions
                    amount_matched = False
                    for trans in customer_trans:
                        if amount > 0 and abs(trans['balance'] - amount) / amount <= 0.05:  # 5% tolerance
                            match_result['match'] = trans
                            match_result['confidence'] = 90
                            match_result['match_type'] = f'{potential_customers[0][1]} + Amount'
                            match_result['matched_customer'] = matched_customer
                            matched.append(match_result)
                            amount_matched = True
                            break
                    
                    if not amount_matched:
                        # Multiple transactions but no amount match
                        if len(customer_trans) == 1:
                            # Single transaction for this customer
                            match_result['match'] = customer_trans[0]
                            match_result['confidence'] = 85
                            match_result['match_type'] = potential_customers[0][1]
                            match_result['matched_customer'] = matched_customer
                            matched.append(match_result)
                        else:
                            # Multiple transactions - needs selection
                            match_result['potential_matches'] = customer_trans
                            match_result['potential_customers'] = potential_customers
                            match_result['needs_selection'] = True
                            unmatched.append(match_result)
                else:
                    # Customer found but no transactions with balance
                    match_result['potential_customers'] = potential_customers
                    match_result['no_balance'] = True
                    unmatched.append(match_result)
            else:
                # Multiple potential matches or low confidence - needs manual selection
                match_result['potential_customers'] = potential_customers
                match_result['needs_selection'] = True
                
                # Collect all transactions for all potential customers
                all_potential_trans = []
                for potential_customer, match_type, score in potential_customers[:5]:  # Limit to top 5
                    customer_key = potential_customer.lower().strip()
                    if customer_key in customer_trans_lookup:
                        trans_list = customer_trans_lookup[customer_key]
                        for trans in trans_list:
                            trans_copy = trans.copy()  # Avoid modifying original
                            trans_copy['_customer_match'] = potential_customer
                            trans_copy['_match_type'] = match_type
                            trans_copy['_match_score'] = score
                            all_potential_trans.append(trans_copy)
                
                if all_potential_trans:
                    match_result['potential_matches'] = all_potential_trans
                    
                unmatched.append(match_result)
        else:
            # No fuzzy match found - check if exact customer exists
            customer_key = customer.lower().strip()
            if customer_key in customer_trans_lookup:
                # Exact match exists
                customer_matches = customer_trans_lookup[customer_key]
                # Check for amount match
                amount_matched = False
                for trans in customer_matches:
                    if amount > 0 and abs(trans['balance'] - amount) / amount <= 0.05:  # 5% tolerance
                        match_result['match'] = trans
                        match_result['confidence'] = 85
                        match_result['match_type'] = 'Customer + Amount'
                        matched.append(match_result)
                        amount_matched = True
                        break
                
                if not amount_matched:
                    # No amount match - needs selection
                    match_result['potential_matches'] = customer_matches
                    match_result['potential_customers'] = [(customer, 'exact', 100)]
                    match_result['needs_selection'] = True
                    unmatched.append(match_result)
            else:
                # No match found - can create new transaction
                match_result['can_create'] = True
                can_create.append(match_result)
    
    return matched, unmatched, can_create

def show_import_results(statement_date):
    """Display import results and allow user to review/confirm"""
    st.divider()
    st.markdown("### 📊 Import Preview")
    
    # Create tabs for different result types
    result_tabs = st.tabs([
        f"✅ Matched ({len(st.session_state.matched_transactions)})",
        f"❌ Unmatched ({len(st.session_state.unmatched_transactions)})",
        f"➕ Can Create ({len(st.session_state.transactions_to_create)})"
    ])
    
    with result_tabs[0]:  # Matched transactions
        if st.session_state.matched_transactions:
            matched_df = []
            for item in st.session_state.matched_transactions:
                # Show matched customer name if different from statement
                display_customer = item['customer']
                if 'matched_customer' in item and item['matched_customer'] != item['customer']:
                    display_customer = f"{item['customer']} → {item['matched_customer']}"
                
                matched_df.append({
                    'Status': '✅',
                    'Customer': display_customer,
                    'Policy': item['policy_number'],
                    'Eff Date': item['effective_date'],
                    'Statement Amt': item['amount'],
                    'DB Balance': item['match']['balance'],
                    'Confidence': f"{item['confidence']}%",
                    'Match Type': item['match_type']
                })
            
            df = pd.DataFrame(matched_df)
            st.dataframe(
                df,
                column_config={
                    "Statement Amt": st.column_config.NumberColumn(format="$%.2f"),
                    "DB Balance": st.column_config.NumberColumn(format="$%.2f")
                },
                use_container_width=True,
                hide_index=True
            )
            
            total_matched = sum(item['amount'] for item in st.session_state.matched_transactions)
            st.metric("Total Matched Amount", f"${total_matched:,.2f}")
        else:
            st.info("No matched transactions")
    
    with result_tabs[1]:  # Unmatched transactions
        if st.session_state.unmatched_transactions:
            st.warning("These transactions need manual review")
            
            # Create session state for manual matches if not exists
            if 'manual_matches' not in st.session_state:
                st.session_state.manual_matches = {}
            
            for idx, item in enumerate(st.session_state.unmatched_transactions):
                with st.expander(f"🔍 {item['customer']} - ${item['amount']:,.2f}", expanded=True):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.markdown(f"**Statement Details:**")
                        st.text(f"Policy: {item['policy_number']}")
                        st.text(f"Effective Date: {item['effective_date']}")
                        st.text(f"Amount: ${item['amount']:,.2f}")
                    
                    with col2:
                        # Check if we have potential customers
                        if 'potential_customers' in item:
                            st.markdown("**Potential Matches Found:**")
                            
                            customer_options = []
                            for cust, match_type, score in item['potential_customers'][:10]:
                                customer_options.append(f"{cust} ({match_type}: {score}%)")
                            
                            selected_customer_idx = st.selectbox(
                                "Select Customer",
                                range(len(customer_options)),
                                format_func=lambda x: customer_options[x],
                                key=f"customer_select_{idx}"
                            )
                            
                            if selected_customer_idx is not None:
                                selected_customer = item['potential_customers'][selected_customer_idx][0]
                                
                                # Show transactions for selected customer
                                if 'potential_matches' in item:
                                    customer_trans = [t for t in item['potential_matches'] 
                                                    if t.get('_customer_match') == selected_customer]
                                    
                                    if customer_trans:
                                        st.markdown("**Available Transactions:**")
                                        
                                        trans_options = []
                                        for trans in customer_trans:
                                            trans_desc = f"Policy: {trans['Policy Number']} | "
                                            trans_desc += f"Eff: {trans['Effective Date']} | "
                                            trans_desc += f"Balance: ${trans['balance']:,.2f}"
                                            trans_options.append(trans_desc)
                                        
                                        selected_trans_idx = st.selectbox(
                                            "Select Transaction",
                                            range(len(trans_options)),
                                            format_func=lambda x: trans_options[x],
                                            key=f"trans_select_{idx}"
                                        )
                                        
                                        if st.button("✅ Confirm Match", key=f"confirm_{idx}", type="primary"):
                                            # Add to manual matches
                                            st.session_state.manual_matches[idx] = {
                                                'statement_item': item,
                                                'matched_transaction': customer_trans[selected_trans_idx],
                                                'customer': selected_customer
                                            }
                                            st.success("Match confirmed!")
                                            st.rerun()
                                    else:
                                        st.info("No transactions with balance for this customer")
                        else:
                            st.info("No potential matches found")
                    
                    # Option to create new
                    if st.checkbox(f"Create as new transaction", key=f"create_new_{idx}"):
                        st.session_state.manual_matches[idx] = {
                            'statement_item': item,
                            'create_new': True
                        }
                        st.success("Will create new transaction")
            
            # Show confirmed matches
            if st.session_state.manual_matches:
                st.divider()
                st.markdown("### ✅ Confirmed Manual Matches")
                
                manual_match_count = len(st.session_state.manual_matches)
                st.metric("Manual Matches", manual_match_count)
                
                if st.button("Apply Manual Matches", type="primary"):
                    # Move manually matched items to appropriate lists
                    for idx, match_info in st.session_state.manual_matches.items():
                        if 'create_new' in match_info and match_info['create_new']:
                            # Move to create list
                            st.session_state.transactions_to_create.append(match_info['statement_item'])
                        else:
                            # Move to matched list
                            matched_item = match_info['statement_item'].copy()
                            matched_item['match'] = match_info['matched_transaction']
                            matched_item['confidence'] = 100
                            matched_item['match_type'] = 'Manual'
                            st.session_state.matched_transactions.append(matched_item)
                    
                    # Remove from unmatched
                    indices_to_remove = list(st.session_state.manual_matches.keys())
                    st.session_state.unmatched_transactions = [
                        item for i, item in enumerate(st.session_state.unmatched_transactions) 
                        if i not in indices_to_remove
                    ]
                    
                    # Clear manual matches
                    st.session_state.manual_matches = {}
                    st.success("Manual matches applied!")
                    st.rerun()
        else:
            st.success("All transactions were matched!")
    
    with result_tabs[2]:  # Can create
        if st.session_state.transactions_to_create:
            st.info("These transactions don't exist in the database but can be created")
            
            create_df = []
            for item in st.session_state.transactions_to_create:
                create_df.append({
                    'Create': True,
                    'Customer': item['customer'],
                    'Policy': item['policy_number'],
                    'Eff Date': item['effective_date'],
                    'Amount': item['amount']
                })
            
            df = pd.DataFrame(create_df)
            edited_df = st.data_editor(
                df,
                column_config={
                    "Create": st.column_config.CheckboxColumn("Create", help="Check to create this transaction"),
                    "Amount": st.column_config.NumberColumn(format="$%.2f")
                },
                use_container_width=True,
                hide_index=True,
                key="create_selector"
            )
            
            # Option to create all or selected
            col1, col2 = st.columns(2)
            with col1:
                create_selected = st.checkbox(
                    "Create missing transactions before reconciling", 
                    value=True,
                    key="create_selected"
                )
            with col2:
                selected_count = edited_df['Create'].sum()
                st.metric("Selected to Create", selected_count)
        else:
            st.info("No new transactions to create")
    
    # Import confirmation section
    st.divider()
    st.markdown("### 🚀 Import Confirmation")
    
    # Get create_selected value from session state or default
    create_selected = st.session_state.get('create_selected', True)
    
    # Calculate totals
    total_matched = sum(item['amount'] for item in st.session_state.matched_transactions)
    total_unmatched = sum(item['amount'] for item in st.session_state.unmatched_transactions)
    total_to_create = sum(item['amount'] for item in st.session_state.transactions_to_create) if create_selected else 0
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Matched", f"${total_matched:,.2f}")
    with col2:
        st.metric("Unmatched", f"${total_unmatched:,.2f}")
    with col3:
        st.metric("To Create", f"${total_to_create:,.2f}")
    with col4:
        st.metric("Total", f"${total_matched + total_to_create:,.2f}")
    
    # Show statement total for verification if available
    if 'statement_file_total' in st.session_state and st.session_state.statement_file_total > 0:
        st.divider()
        st.markdown("### ✅ Verification Check")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Statement Total", f"${st.session_state.statement_file_total:,.2f}", 
                     help="Total from your commission statement file")
        with col2:
            reconcile_total = total_matched + total_to_create
            st.metric("Ready to Reconcile", f"${reconcile_total:,.2f}",
                     help="Sum of matched + to be created transactions")
        with col3:
            difference = st.session_state.statement_file_total - reconcile_total
            if abs(difference) < 0.01:  # Less than a penny
                st.metric("Difference", "$0.00", delta_color="off")
                st.success("✓ Perfectly balanced!")
            else:
                st.metric("Difference", f"${abs(difference):,.2f}", 
                         delta=f"{'Over' if difference < 0 else 'Under'} by ${abs(difference):,.2f}",
                         delta_color="inverse")
                if difference > 0:
                    st.warning(f"⚠️ Missing ${difference:,.2f} from statement")
                else:
                    st.error(f"❌ Exceeding statement by ${abs(difference):,.2f}")
    
    # Import button
    if st.button("🔄 Proceed with Import", type="primary", disabled=len(st.session_state.matched_transactions) == 0):
        with st.spinner("Importing transactions..."):
            try:
                # Generate batch ID
                batch_id = f"IMPORT-{statement_date.strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
                created_count = 0
                reconciled_count = 0
                
                # Step 1: Create missing transactions if selected
                if create_selected and st.session_state.transactions_to_create:
                    for idx, item in enumerate(st.session_state.transactions_to_create):
                        # Check if this transaction should be created (default to True if no edit)
                        should_create = True
                        if 'create_selector' in st.session_state:
                            # Get the dataframe from session state if it exists
                            create_df = pd.DataFrame([{
                                'Create': True,
                                'Customer': t['customer'],
                                'Policy': t['policy_number'],
                                'Eff Date': t['effective_date'],
                                'Amount': t['amount']
                            } for t in st.session_state.transactions_to_create])
                            should_create = create_df.loc[idx, 'Create'] if idx < len(create_df) else True
                        
                        if should_create:
                            # Generate new transaction ID
                            new_trans_id = generate_transaction_id()
                            
                            # Create new transaction
                            new_trans = {
                                'Transaction ID': new_trans_id,
                                'Customer': item['customer'],
                                'Policy Number': item['policy_number'],
                                'Effective Date': item['effective_date'],
                                'Transaction Type': item['statement_data'].get(st.session_state.column_mapping.get('Transaction Type', ''), 'NEW'),
                                'Premium Sold': item['statement_data'].get(st.session_state.column_mapping.get('Premium Sold', ''), 0),
                                'Agent Estimated Comm $': item['amount'],  # Use statement amount as estimated
                                'Agency Estimated Comm/Revenue (CRM)': item['amount'],
                                'NOTES': f"Created from statement import {batch_id}"
                            }
                            
                            # Add other mapped fields
                            for sys_field, stmt_field in st.session_state.column_mapping.items():
                                if sys_field not in new_trans and stmt_field in item['statement_data']:
                                    new_trans[sys_field] = item['statement_data'][stmt_field]
                            
                            # Insert to database
                            supabase = get_supabase_client()
                            result = supabase.table('policies').insert(new_trans).execute()
                            if result.data:
                                created_count += 1
                                # Add to matched transactions for reconciliation
                                st.session_state.matched_transactions.append({
                                    'match': result.data[0],
                                    'amount': item['amount'],
                                    'customer': item['customer'],
                                    'policy_number': item['policy_number']
                                })
                
                # Step 2: Create reconciliation entries for all matched transactions
                for item in st.session_state.matched_transactions:
                    recon_id = generate_reconciliation_transaction_id("STMT", statement_date)
                    
                    # Create reconciliation entry
                    recon_entry = {
                        'Transaction ID': recon_id,
                        'Customer': item['match']['Customer'],
                        'Policy Number': item['match']['Policy Number'],
                        'Effective Date': item['match']['Effective Date'],
                        'Transaction Type': item['match'].get('Transaction Type', ''),
                        'Premium Sold': 0,
                        'Agency Comm Received (STMT)': item.get('agency_amount', 0),  # Agency amount for audit
                        'Agent Paid Amount (STMT)': item['amount'],  # Agent amount (primary)
                        'STMT DATE': statement_date.strftime('%Y-%m-%d'),
                        'reconciliation_status': 'reconciled',
                        'reconciliation_id': batch_id,
                        'is_reconciliation_entry': True,
                        'NOTES': f"Import batch {batch_id}"
                    }
                    
                    # Insert reconciliation entry
                    supabase = get_supabase_client()
                    supabase.table('policies').insert(recon_entry).execute()
                    reconciled_count += 1
                
                # Clear session state
                st.session_state.import_data = None
                st.session_state.matched_transactions = []
                st.session_state.unmatched_transactions = []
                st.session_state.transactions_to_create = []
                st.session_state.column_mapping = {}
                if 'statement_file_total' in st.session_state:
                    del st.session_state.statement_file_total
                
                st.success(f"""
                ✅ Import completed successfully!
                - Created {created_count} new transactions
                - Reconciled {reconciled_count} transactions
                - Batch ID: {batch_id}
                """)
                
                # Clear cache and refresh
                st.cache_data.clear()
                time.sleep(2)
                st.rerun()
                
            except Exception as e:
                st.error(f"Error during import: {str(e)}")
                st.exception(e)

def duplicate_for_renewal(df: pd.DataFrame) -> pd.DataFrame:
    """
    Duplicates the given policies and updates their dates for renewal.
    """
    if df.empty:
        return pd.DataFrame()

    renewed_df = df.copy()
    
    # Calculate new term dates
    renewed_df['new_effective_date'] = renewed_df['expiration_date']
    renewed_df['new_expiration_date'] = renewed_df['new_effective_date'] + pd.DateOffset(months=6) # Assuming 6-month terms for now
    
    # Update the relevant columns
    renewed_df[get_mapped_column("Effective Date")] = renewed_df['new_effective_date'].dt.strftime('%m/%d/%Y')
    renewed_df[get_mapped_column("X-DATE")] = renewed_df['new_expiration_date'].dt.strftime('%m/%d/%Y')
    renewed_df[get_mapped_column("Transaction Type")] = "RWL"
    
    return renewed_df

def main():
    # Check password before showing any content
    if not check_password():
        st.stop()
    
    # Apply CSS styling
    apply_css()
      # --- Page Selection ---
    page = st.sidebar.radio(
        "Navigation",        [
            "Dashboard",
            "Reports",
            "All Policies in Database",
            "Edit Policies in Database",
            "Add New Policy Transaction",
            "Search & Filter",
            "Reconciliation",
            "Admin Panel",
            "Tools",
            "Accounting",
            "Help",
            "Policy Revenue Ledger",
            "Policy Revenue Ledger Reports",
            "Pending Policy Renewals"
        ]
    )
    
    # Add logout button to sidebar
    st.sidebar.divider()
    if st.sidebar.button("🚪 Logout", type="secondary", use_container_width=True):
        st.session_state["password_correct"] = False
        st.rerun()
    
    # --- Load data with caching for better performance ---
    all_data = load_policies_data()
    supabase = get_supabase_client()
    
    if all_data.empty:
        st.warning("No data found in policies table. Please add some policy data first.")

    # --- Page Navigation ---
    if page == "Dashboard":
        st.title("📊 Commission Dashboard")
        
        if not all_data.empty:
            # --- Dashboard Metrics ---
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_policies = len(all_data)
                st.metric("Total Policies", f"{total_policies:,}")
            
            with col2:
                if 'Commission_Paid' in all_data.columns:
                    total_commission = all_data['Commission_Paid'].sum()
                    st.metric("Total Commission", f"${total_commission:,.2f}")
                else:
                    st.metric("Total Commission", "N/A")
            
            with col3:
                if 'Agency_Commission_Received' in all_data.columns:
                    agency_commission = all_data['Agency_Commission_Received'].sum()
                    st.metric("Agency Commission", f"${agency_commission:,.2f}")
                else:
                    st.metric("Agency Commission", "N/A")
            
            # Balance Due column removed - no longer needed
            
            st.divider()
            
            # --- Client Search and Quick Actions ---
            st.subheader("🔍 Quick Client Search")
            
            search_col1, search_col2 = st.columns([3, 1])
            
            with search_col1:
                search_term = st.text_input("Search by Client Name, Policy Number, or Transaction ID", 
                                          placeholder="Enter search term...")
            
            with search_col2:
                st.write("")  # Spacing
                search_button = st.button("Search", type="primary")
            
            # Search functionality
            if search_term or search_button:
                if search_term:
                    # Search across multiple columns
                    mask = pd.Series(False, index=all_data.index)
                    search_columns = ['Customer', 'Policy Number', 'Transaction ID', 'Client ID']
                    
                    for col in search_columns:
                        if col in all_data.columns:
                            mask |= all_data[col].astype(str).str.contains(search_term, case=False, na=False)
                    
                    search_results = all_data[mask]
                    
                    if not search_results.empty:
                        st.success(f"Found {len(search_results)} matching records")
                        
                        # Configure column settings for proper numeric display
                        column_config = {}
                        numeric_cols = [
                            'Agent Estimated Comm $',
                                        'Policy Gross Comm %',
                            'Agency Estimated Comm/Revenue (CRM)',
                            'Agency Comm Received (STMT)',
                            'Premium Sold',
                            'Agent Paid Amount (STMT)',
                            'Agency Comm Received (STMT)'
                        ]
                        
                        for col in numeric_cols:
                            if col in search_results.columns:
                                column_config[col] = st.column_config.NumberColumn(
                                    col,
                                    format="%.2f",
                                    step=0.01
                                )
                        
                        # Display search results in an editable table
                        edited_data = st.data_editor(
                            search_results,
                            use_container_width=True,
                            height=400,
                            key="dashboard_search_editor",
                            column_config=column_config
                        )
                        
                        # Update button for search results
                        if st.button("Save Changes", key="dashboard_save"):
                            try:
                                # Update the database with changes
                                for idx, row in edited_data.iterrows():
                                    # Update via Supabase
                                    update_dict = {}
                                    for col in edited_data.columns:
                                        if col not in ['_id', 'Transaction ID']:  # Don't update ID fields
                                            update_dict[col] = row[col]
                                    
                                    # Find the record by Transaction ID
                                    transaction_id = row.get('Transaction ID', row.get('Transaction_ID'))
                                    if transaction_id:
                                        try:
                                            supabase.table('policies').update(update_dict).eq('Transaction ID', transaction_id).execute()
                                        except Exception as update_error:
                                            st.error(f"Error updating record {transaction_id}: {update_error}")
                                
                                st.success("Changes saved successfully!")
                                st.rerun()
                                
                            except Exception as e:
                                st.error(f"Error saving changes: {e}")
                    else:
                        st.warning("No records found matching your search term")
                else:
                    st.info("Enter a search term to find records")
            
            st.divider()
            
            # --- Recent Activity ---
            st.subheader("📈 Recent Activity")
            
            # Show last 10 records
            recent_data = all_data.head(10)
            
            # Configure column settings for proper numeric display
            column_config = {}
            numeric_cols = [
                'Agent Estimated Comm $',
                'Policy Gross Comm %',
                'Agency Estimated Comm/Revenue (CRM)',
                'Agency Comm Received (STMT)',
                'Premium Sold',
                'Agent Paid Amount (STMT)',
                'Agency Comm Received (STMT)'
            ]
            
            for col in numeric_cols:
                if col in recent_data.columns:
                    column_config[col] = st.column_config.NumberColumn(
                        col,
                        format="%.2f"
                    )
            
            st.dataframe(
                recent_data,
                use_container_width=True,
                height=300,
                column_config=column_config
            )
            
            # --- Quick Statistics ---
            st.subheader("📊 Quick Statistics")
            
            stats_col1, stats_col2 = st.columns(2)
            
            with stats_col1:
                if 'Policy_Type' in all_data.columns:
                    policy_type_counts = all_data['Policy_Type'].value_counts()
                    st.write("**Policies by Type:**")
                    for policy_type, count in policy_type_counts.items():
                        st.write(f"• {policy_type}: {count}")
                        
            with stats_col2:
                if 'Transaction_Type' in all_data.columns:
                    transaction_type_counts = all_data['Transaction_Type'].value_counts()
                    st.write("**Transactions by Type:**")
                    for trans_type, count in transaction_type_counts.items():
                        st.write(f"• {trans_type}: {count}")
            
        else:
            st.info("No data available. Please add some policy data to see dashboard metrics.")
    
    # --- Reports ---
    elif page == "Reports":
        st.title("📈 Reports")
        
        if not all_data.empty:
            tab1, tab2, tab3, tab4 = st.tabs(["Summary Reports", "Commission Analysis", "Policy Reports", "Custom Reports"])
            
            with tab1:
                st.subheader("Summary Reports")
                
                # Overall summary
                col1, col2 = st.columns(2)
                
                with col1:
                    st.metric("Total Policies", len(all_data))
                    if 'Commission_Paid' in all_data.columns:
                        st.metric("Total Commission Paid", f"${all_data['Commission_Paid'].sum():,.2f}")
                    if 'Agency_Commission_Received' in all_data.columns:
                        st.metric("Total Agency Commission", f"${all_data['Agency_Commission_Received'].sum():,.2f}")
                
                with col2:
                    if 'Balance_Due' in all_data.columns:
                        st.metric("Total Balance Due", f"${all_data['Balance_Due'].sum():,.2f}")
                    if 'Effective_Date' in all_data.columns:
                        latest_date = pd.to_datetime(all_data['Effective_Date'], errors='coerce').max()
                        st.metric("Latest Policy Date", latest_date.strftime('%m/%d/%Y') if pd.notnull(latest_date) else "N/A")
            
            with tab2:
                st.subheader("Commission Analysis")
                
                if 'Commission_Paid' in all_data.columns and 'Policy_Type' in all_data.columns:
                    commission_by_type = all_data.groupby('Policy_Type')['Commission_Paid'].sum().sort_values(ascending=False)
                    st.write("**Commission by Policy Type:**")
                    st.dataframe(commission_by_type.reset_index())
                    
                if 'Transaction_Type' in all_data.columns and 'Commission_Paid' in all_data.columns:
                    commission_by_transaction = all_data.groupby('Transaction_Type')['Commission_Paid'].sum().sort_values(ascending=False)
                    st.write("**Commission by Transaction Type:**")
                    st.dataframe(commission_by_transaction.reset_index())
            
            with tab3:
                st.subheader("Policy Reports")
                
                if 'Policy_Type' in all_data.columns:
                    policy_counts = all_data['Policy_Type'].value_counts()
                    st.write("**Policy Count by Type:**")
                    st.dataframe(policy_counts.reset_index())
                
                if 'Transaction_Type' in all_data.columns:
                    transaction_counts = all_data['Transaction_Type'].value_counts()
                    st.write("**Transaction Count by Type:**")
                    st.dataframe(transaction_counts.reset_index())
            
            with tab4:
                st.subheader("Custom Reports")
                st.info("Custom report builder - Select criteria to generate custom reports")
                
                # Date range filter
                if 'Effective_Date' in all_data.columns:
                    date_col = st.columns(2)
                    with date_col[0]:
                        start_date = st.date_input("Start Date")
                    with date_col[1]:
                        end_date = st.date_input("End Date")
                
                # Policy type filter
                if 'Policy_Type' in all_data.columns:
                    policy_types = st.multiselect("Policy Types", all_data['Policy_Type'].unique())
                
                if st.button("Generate Custom Report"):
                    filtered_data = all_data.copy()
                    
                    # Apply filters
                    if 'Effective_Date' in all_data.columns and start_date and end_date:
                        filtered_data['Effective_Date'] = pd.to_datetime(filtered_data['Effective_Date'], errors='coerce')
                        filtered_data = filtered_data[
                            (filtered_data['Effective_Date'] >= pd.Timestamp(start_date)) &
                            (filtered_data['Effective_Date'] <= pd.Timestamp(end_date))
                        ]
                    
                    if policy_types:
                        filtered_data = filtered_data[filtered_data['Policy_Type'].isin(policy_types)]
                    
                    st.write(f"**Custom Report Results ({len(filtered_data)} records):**")
                    st.dataframe(filtered_data, use_container_width=True)
                    
                    # Export custom report results
                    if not filtered_data.empty:
                        st.subheader("Export Custom Report")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write("**📄 CSV Export**")
                            csv_data = filtered_data.to_csv(index=False)
                            st.download_button(
                                label="📥 Download Custom Report CSV",
                                data=csv_data,
                                file_name=f"custom_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                                mime="text/csv",
                                help="Export custom report results as CSV file"
                            )
                        
                        with col2:
                            st.write("**📊 Excel Export**")
                            excel_buffer, excel_filename = create_formatted_excel_file(
                                filtered_data, 
                                sheet_name="Custom Report", 
                                filename_prefix="custom_report"
                            )
                            if excel_buffer:
                                st.download_button(
                                    label="📥 Download Custom Report Excel",
                                    data=excel_buffer,
                                    file_name=excel_filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    help="Export custom report results as formatted Excel file"
                                )
            
            # Export all reports data
            if not all_data.empty:
                st.divider()
                st.subheader("Export All Policy Data")
                st.write("Export complete policy database for comprehensive reporting:")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("**📄 CSV Export**")
                    csv_all_data = all_data.to_csv(index=False)
                    st.download_button(
                        label="📥 Download All Policies CSV",
                        data=csv_all_data,
                        file_name=f"all_policies_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        help="Export all policy data as CSV file"
                    )
                
                with col2:
                    st.write("**📊 Excel Export**")
                    
                    # Create multi-sheet Excel with different report views
                    reports_data = {
                        "All Policies": all_data,
                        "Summary Stats": pd.DataFrame([
                            ["Total Policies", len(all_data)],
                            ["Total Commission", all_data['Commission_Paid'].sum() if 'Commission_Paid' in all_data.columns else 0],
                            ["Total Balance Due", all_data['Balance_Due'].sum() if 'Balance_Due' in all_data.columns else 0],
                            ["Export Date", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                        ], columns=['Metric', 'Value'])
                    }
                    
                    # Add policy type breakdown if available
                    if 'Policy_Type' in all_data.columns:
                        policy_breakdown = all_data['Policy_Type'].value_counts().reset_index()
                        policy_breakdown.columns = ['Policy_Type', 'Count']
                        reports_data["Policy Type Breakdown"] = policy_breakdown
                    
                    excel_buffer_multi, excel_filename_multi = create_multi_sheet_excel(
                        reports_data, 
                        filename_prefix="complete_policy_report"
                    )
                    
                    if excel_buffer_multi:
                        st.download_button(
                            label="📥 Download Comprehensive Excel Report",
                            data=excel_buffer_multi,
                            file_name=excel_filename_multi,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Export comprehensive multi-sheet Excel report with all data and summaries"
                        )
        else:
            st.info("No data available for reports.")
    
    # --- All Policies in Database ---
    elif page == "All Policies in Database":
        st.title("📋 All Policies in Database")
        
        if not all_data.empty:
            st.write(f"**Total Records: {len(all_data)}**")
            
            # Add toggle for formula view
            col_toggle, col_space = st.columns([2, 8])
            with col_toggle:
                show_formulas = st.toggle("📊 Show Formulas", value=True, help="Toggle between formula calculations and actual values")
            
            # Pagination controls
            items_per_page = st.selectbox("Items per page", [25, 50, 100, 200], index=1)
            total_pages = max(1, len(all_data) // items_per_page + (1 if len(all_data) % items_per_page > 0 else 0))
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                page_num = st.number_input("Page", min_value=1, max_value=total_pages, value=1, step=1)
            
            # Calculate pagination
            start_idx = (page_num - 1) * items_per_page
            end_idx = min(start_idx + items_per_page, len(all_data))
            
            # Display paginated data
            paginated_data = all_data.iloc[start_idx:end_idx]
            
            # Apply formula display to existing columns
            paginated_data_display = apply_formula_display(paginated_data, show_formulas=show_formulas)
            
            st.write(f"Showing records {start_idx + 1} to {end_idx} of {len(all_data)}")
            
            # Show mode info
            if show_formulas:
                st.info("📊 **Formula Mode**: Showing calculated values. Icons: ✓ = matches manual entry, ✏️ = manual override, ⚠️ = missing data, 🔒 = reconciliation entry")
            else:
                st.info("📝 **Actual Values Mode**: Showing original manual entries from the database.")
            
            # Configure column settings for proper numeric display
            column_config = {}
            numeric_cols = [
                'Agent Estimated Comm $',
                'Policy Gross Comm %',
                'Agency Estimated Comm/Revenue (CRM)',
                'Agency Comm Received (STMT)',
                'Premium Sold',
                'Agent Paid Amount (STMT)',
                'Agency Comm Received (STMT)',
            ]
            
            for col in numeric_cols:
                if col in paginated_data_display.columns:
                    # Note: When in formula mode, these columns will have indicators appended
                    # Streamlit will treat them as text columns due to the emoji indicators
                    if show_formulas and col in ['Agency Estimated Comm/Revenue (CRM)', 'Agent Estimated Comm $']:
                        column_config[col] = st.column_config.TextColumn(
                            col,
                            help="Formula calculation with status indicator"
                        )
                    else:
                        column_config[col] = st.column_config.NumberColumn(
                            col,
                            format="%.2f"
                        )
            
            # Display the data in a scrollable table
            st.dataframe(
                paginated_data_display,
                use_container_width=True,
                height=600,
                column_config=column_config
            )
            
            # Export options
            st.subheader("Export Options")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.write("**📄 CSV Export**")
                # Export based on current view mode
                export_data = apply_formula_display(paginated_data, show_formulas=show_formulas) if show_formulas else paginated_data
                csv = export_data.to_csv(index=False)
                st.download_button(
                    label="📥 Current Page CSV",
                    data=csv,
                    file_name=f"policies_page_{page_num}.csv",
                    mime="text/csv",
                    help="Export current page as CSV file"
                )
            
            with col2:
                # Export all data based on current view mode
                all_data_export = apply_formula_display(all_data, show_formulas=show_formulas) if show_formulas else all_data
                csv_all = all_data_export.to_csv(index=False)
                st.write("**📄 CSV Export**")
                st.download_button(
                    label="📥 All Data CSV",
                    data=csv_all,
                    file_name="all_policies.csv",
                    mime="text/csv",
                    help="Export all policies as CSV file"
                )
            
            with col3:
                st.write("**📊 Excel Export**")
                excel_buffer, excel_filename = create_formatted_excel_file(
                    export_data, 
                    sheet_name="Policies Page", 
                    filename_prefix=f"policies_page_{page_num}"
                )
                if excel_buffer:
                    st.download_button(
                        label="📥 Current Page Excel",
                        data=excel_buffer,
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Export current page as formatted Excel file"
                    )
            
            with col4:
                st.write("**📊 Excel Export**")
                excel_buffer_all, excel_filename_all = create_formatted_excel_file(
                    all_data_export, 
                    sheet_name="All Policies", 
                    filename_prefix="all_policies"
                )
                if excel_buffer_all:
                    st.download_button(
                        label="📥 All Data Excel",
                        data=excel_buffer_all,
                        file_name=excel_filename_all,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Export all policies as formatted Excel file"
                    )
        else:
            st.info("No policies found in database.")
    
    # --- Edit Policies in Database ---
    elif page == "Edit Policies in Database":
        st.title("✏️ Edit Policies in Database")
        
        if not all_data.empty:
            st.warning("⚠️ Be careful when editing data. Changes are saved directly to the database.")
            
            # Search and filter options
            st.subheader("Find Policies to Edit")
            
            # Wrap search in a form to prevent accidental reruns
            with st.form("search_form"):
                search_col1, search_col2 = st.columns([3, 1])
                
                with search_col1:
                    edit_search_term = st.text_input("Search policies to edit", placeholder="Enter search term...")
                
                with search_col2:
                    st.write("")  # Spacing
                    edit_search_button = st.form_submit_button("Find Records", type="primary")
            
            # Show filtered data for editing
            if edit_search_term or edit_search_button:
                if edit_search_term:
                    # Search across multiple columns
                    mask = pd.Series(False, index=all_data.index)
                    search_columns = ['Customer', 'Policy Number', 'Transaction ID', 'Client ID']
                    
                    for col in search_columns:
                        if col in all_data.columns:
                            mask |= all_data[col].astype(str).str.contains(edit_search_term, case=False, na=False)
                    
                    edit_results = all_data[mask].copy()
                    
                    if not edit_results.empty:
                        # Format numeric columns to ensure 2 decimal places
                        edit_results = round_numeric_columns(edit_results)
                        
                        # Filter out reconciliation transactions
                        # First, find the transaction ID column using multiple detection methods
                        transaction_id_col = None
                        
                        # Method 1: Try using column mapper
                        transaction_id_col = get_mapped_column("Transaction ID")
                        
                        # Method 2: If mapper fails, check common variations
                        if not transaction_id_col:
                            possible_names = ["Transaction ID", "Transaction_ID", "TransactionID", 
                                            "transaction_id", "Transaction Id", "TRANSACTION_ID",
                                            "transaction id", "transactionid"]
                            for col in edit_results.columns:
                                if col in possible_names:
                                    transaction_id_col = col
                                    break
                        
                        # Method 3: If still not found, use flexible search
                        if not transaction_id_col:
                            for col in edit_results.columns:
                                # Remove spaces and underscores for comparison
                                col_normalized = col.lower().replace(" ", "").replace("_", "")
                                if col_normalized == "transactionid":
                                    transaction_id_col = col
                                    break
                        
                        if transaction_id_col:
                            # Count total before filtering
                            total_count = len(edit_results)
                            
                            # Filter out reconciliation transactions
                            edit_results_editable = edit_results[
                                ~edit_results[transaction_id_col].apply(is_reconciliation_transaction)
                            ].copy()
                            
                            recon_count = total_count - len(edit_results_editable)
                            
                            # Show counts
                            if recon_count > 0:
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.success(f"Found {len(edit_results_editable)} editable transactions")
                                with col2:
                                    st.info(f"🔒 {recon_count} reconciliation entries (view in Reconciliation page)")
                            else:
                                st.success(f"Found {len(edit_results_editable)} transactions for editing")
                            
                            # Update edit_results to only contain editable transactions
                            edit_results = edit_results_editable
                            
                            # Check if we have any editable transactions left
                            if edit_results.empty:
                                st.warning("No editable transactions found. All transactions for this customer are reconciliation entries.")
                                return  # Exit early
                        else:
                            st.success(f"Found {len(edit_results)} records for editing")
                        
                        # Only proceed if we have editable transactions
                        if not edit_results.empty:
                            # Find the actual column names dynamically
                            transaction_id_col = None
                            client_id_col = None
                            for col in edit_results.columns:
                                if 'transaction' in col.lower() and 'id' in col.lower():
                                    transaction_id_col = col
                                if 'client' in col.lower() and 'id' in col.lower():
                                    client_id_col = col
                            
                            # Add a selection column for deletion
                            edit_results_with_selection = edit_results.copy()
                            edit_results_with_selection.insert(0, 'Select', False)
                        
                            # Configure column settings for the data editor
                            column_config = {
                                "Select": st.column_config.CheckboxColumn(
                                    "Select",
                                    help="Select rows to delete",
                                    default=False,
                                )
                            }
                            
                            # Configure numeric columns to display with 2 decimal places
                            numeric_cols = [
                                'Agent Estimated Comm $',
                                'Policy Gross Comm %',
                                'Agency Estimated Comm/Revenue (CRM)',
                                'Agency Comm Received (STMT)',
                                'Premium Sold',
                                'Agent Paid Amount (STMT)',
                                'Agency Comm Received (STMT)'
                            ]
                            
                            for col in numeric_cols:
                                if col in edit_results.columns:
                                    column_config[col] = st.column_config.NumberColumn(
                                        col,
                                        format="%.2f",
                                        step=0.01
                                    )
                            
                            # If we have transaction ID column, make it clear it will be auto-generated
                            if transaction_id_col:
                                column_config[transaction_id_col] = st.column_config.TextColumn(
                                    transaction_id_col,
                                    help="Leave blank for new rows - will be auto-generated on save",
                                    disabled=False  # Allow editing but we'll generate if blank
                                )
                            
                            # Create a unique key for this search result to track edits
                            editor_key = "edit_policies_editor"
                            
                            # Initialize or reset session state for this editor
                            search_key = f"last_search_{editor_key}"
                            edit_position_key = f"edit_position_{editor_key}"
                            unsaved_changes_key = f"unsaved_changes_{editor_key}"
                            
                            # Initialize if not exists or reset if new search
                            if (editor_key not in st.session_state or 
                                search_key not in st.session_state or 
                                st.session_state[search_key] != edit_search_term):
                                st.session_state[editor_key] = edit_results_with_selection.copy()
                                st.session_state[search_key] = edit_search_term
                                # Clear position tracking on new search
                                if edit_position_key in st.session_state:
                                    del st.session_state[edit_position_key]
                                if unsaved_changes_key in st.session_state:
                                    del st.session_state[unsaved_changes_key]
                            
                            # Preserve edit position and unsaved changes
                            if unsaved_changes_key in st.session_state:
                                # Restore any unsaved changes from before the refresh
                                for row_idx, col_name, value in st.session_state[unsaved_changes_key]:
                                    if row_idx < len(st.session_state[editor_key]) and col_name in st.session_state[editor_key].columns:
                                        st.session_state[editor_key].loc[row_idx, col_name] = value
                            
                            # Auto-save functionality setup
                            auto_save_key = f"auto_save_{editor_key}"
                            if auto_save_key not in st.session_state:
                                st.session_state[auto_save_key] = True
                            
                            # Auto-save toggle at the top
                            col1, col2, col3 = st.columns([4, 1, 1])
                            with col1:
                                st.markdown("### Edit Policies")
                            with col3:
                                st.session_state[auto_save_key] = st.checkbox(
                                    "🔄 Auto-save", 
                                    value=st.session_state[auto_save_key],
                                    help="Automatically save changes as you type"
                                )
                            
                            # Add a container for status messages
                            status_container = st.empty()
                            
                            # Show editing tips
                            with st.expander("💡 Editing Tips", expanded=False):
                                st.markdown("""
                                **Best editing experience:**
                                - 🖥️ **Use Full Screen mode** (expand icon in top-right of table) to prevent screen jumping
                                - Auto-save is enabled by default - changes save automatically
                                - Full screen maintains your position while editing
                                
                                **To minimize data loss:**
                                - Edit in Full Screen mode for best stability
                                - For bulk edits: Export to Excel → Edit → Import back
                                
                                **Keyboard shortcuts:**
                                - Tab: Move to next cell (may be limited)
                                - Enter: Confirm edit and move down
                                - Shift+Enter: Confirm edit and move up
                                """)
                            
                            # Editable data grid with selection column
                            edited_data = st.data_editor(
                                st.session_state[editor_key],
                                use_container_width=True,
                                height=500,
                                key=f"{editor_key}_widget",
                                num_rows="fixed",  # Changed from dynamic since it's not working
                                column_config=column_config,
                                disabled=False
                            )
                                
                                
                            # Detect changes and auto-save
                            if not edited_data.equals(st.session_state[editor_key]):
                                changes_detected = []
                                for idx in edited_data.index:
                                    for col in edited_data.columns:
                                        if edited_data.loc[idx, col] != st.session_state[editor_key].loc[idx, col]:
                                            changes_detected.append((idx, col, edited_data.loc[idx, col]))
                                
                                if changes_detected and st.session_state[auto_save_key]:
                                    # Auto-save changes immediately
                                    status_container.info("💾 Auto-saving changes...")
                                    
                                    try:
                                        # Get transaction ID column
                                        transaction_id_col = None
                                        for col in edited_data.columns:
                                            if 'transaction' in col.lower() and 'id' in col.lower():
                                                transaction_id_col = col
                                                break
                                        
                                        saved_count = 0
                                        for idx, col_name, new_value in changes_detected:
                                            if col_name != 'Select' and transaction_id_col:
                                                transaction_id = edited_data.loc[idx, transaction_id_col]
                                                if pd.notna(transaction_id) and str(transaction_id).strip():
                                                    # Update single cell
                                                    update_dict = {col_name: new_value if pd.notna(new_value) else None}
                                                    supabase.table('policies').update(update_dict).eq(transaction_id_col, transaction_id).execute()
                                                    saved_count += 1
                                        
                                        if saved_count > 0:
                                            status_container.success(f"✅ Auto-saved {saved_count} changes")
                                            # Update the base data to reflect saved changes
                                            st.session_state[editor_key] = edited_data.copy()
                                            # Don't rerun - just update state
                                    
                                    except Exception as e:
                                        status_container.error(f"Auto-save error: {str(e)}")
                                        log_debug(f"Auto-save error: {str(e)}", "ERROR", e)
                                
                                elif changes_detected and not st.session_state[auto_save_key]:
                                    status_container.info(f"📝 {len(changes_detected)} unsaved changes")
                                    st.session_state[unsaved_changes_key] = changes_detected
                            
                            else:
                                # No changes
                                if st.session_state[auto_save_key]:
                                    status_container.success("✅ All changes auto-saved")
                            
                            # Update session state without rerun
                            st.session_state[editor_key] = edited_data
                            
                            # Get the client ID from the search results (if all rows have the same client)
                            existing_client_id = None
                            if client_id_col:
                                unique_client_ids = edit_results[client_id_col].dropna().unique()
                                if len(unique_client_ids) == 1:
                                    existing_client_id = unique_client_ids[0]
                                    st.info(f"💡 New rows will receive unique Transaction IDs and use Client ID: {existing_client_id}")
                                elif len(unique_client_ids) > 1:
                                    st.info("💡 New rows will receive unique Transaction IDs. Multiple Client IDs found - new rows will need Client ID specified.")
                            else:
                                st.info("💡 New rows will receive unique Transaction IDs and Client IDs when you save.")
                            
                            # Add buttons for adding new row and editing selected row
                            button_col1, button_col2 = st.columns(2)
                            
                            with button_col1:
                                if st.button("➕ Add New Transaction for This Client", type="secondary", use_container_width=True):
                                    # Ensure session state exists
                                    if editor_key in st.session_state:
                                        # Create a new empty row with generated IDs
                                        new_row = pd.Series(dtype='object')
                                        for col in edit_results_with_selection.columns:
                                            new_row[col] = None
                                        
                                        # Set default values
                                        new_row['Select'] = False
                                        if transaction_id_col:
                                            new_row[transaction_id_col] = generate_transaction_id()
                                        if client_id_col and existing_client_id:
                                            new_row[client_id_col] = existing_client_id
                                        elif client_id_col:
                                            new_row[client_id_col] = generate_client_id()
                                        
                                        # Add the new row to session state
                                        new_df = pd.concat([st.session_state[editor_key], pd.DataFrame([new_row])], ignore_index=True)
                                        st.session_state[editor_key] = new_df
                                        st.rerun()
                                    else:
                                        st.error("Session state not initialized. Please try searching again.")
                            
                            with button_col2:
                                # Check for selected rows for edit button
                                selected_rows = edited_data[edited_data['Select'] == True].copy()
                                selected_count = len(selected_rows)
                                
                                if selected_count == 1:
                                    if st.button("✏️ Edit Selected Transaction", type="primary", use_container_width=True):
                                        st.session_state['show_edit_modal'] = True
                                        st.session_state['edit_modal_data'] = selected_rows.iloc[0].to_dict()
                                elif selected_count == 0:
                                    st.button("✏️ Edit Selected Transaction", type="primary", use_container_width=True, disabled=True, help="Select one transaction to edit")
                                else:
                                    st.button("✏️ Edit Selected Transaction", type="primary", use_container_width=True, disabled=True, help=f"{selected_count} selected - please select only ONE transaction")
                            
                            # Save and Delete buttons with status
                            st.markdown("---")
                            col1, col2, col3 = st.columns([2, 2, 2])
                            
                            with col1:
                                # Show save status
                                if unsaved_changes_key in st.session_state and st.session_state[unsaved_changes_key]:
                                    st.warning(f"⚠️ {len(st.session_state[unsaved_changes_key])} unsaved changes")
                                else:
                                    st.success("✅ All changes saved")
                            
                            with col2:
                                save_button = st.button("💾 Save All Changes", type="primary", use_container_width=True)
                            
                            if save_button:
                                    try:
                                        updated_count = 0
                                        inserted_count = 0
                                        
                                        
                                        # Store original transaction IDs to track which rows are updates vs inserts
                                        original_transaction_ids = set()
                                        if transaction_id_col:
                                            original_transaction_ids = set(edit_results[transaction_id_col].dropna().astype(str))
                                        
                                        # Process each record in the edited data
                                        for idx, row in edited_data.iterrows():
                                            # Skip the selection column for save operations
                                            if 'Select' in row:
                                                row = row.drop('Select')
                                            
                                            # Get the transaction ID for this row
                                            transaction_id = row.get(transaction_id_col) if transaction_id_col else None
                                            
                                            # Check if this is a new row by seeing if the transaction ID exists in original data
                                            is_new_row = False
                                            if transaction_id_col:
                                                # If there's no transaction ID or it's not in the original set, it's new
                                                if pd.isna(transaction_id) or str(transaction_id).strip() == '' or str(transaction_id) not in original_transaction_ids:
                                                    is_new_row = True
                                            else:
                                                # Fallback to index-based check if no transaction ID column
                                                is_new_row = idx >= len(edit_results)
                                            
                                            if is_new_row:
                                                # For new rows, generate unique IDs if they're missing
                                                if transaction_id_col and (pd.isna(transaction_id) or str(transaction_id).strip() == ''):
                                                    transaction_id = generate_transaction_id()
                                                    row[transaction_id_col] = transaction_id
                                                if client_id_col and (pd.isna(row[client_id_col]) or str(row[client_id_col]).strip() == ''):
                                                    # Use existing client ID if searching for a specific client, otherwise generate new
                                                    if existing_client_id:
                                                        row[client_id_col] = existing_client_id
                                                    else:
                                                        row[client_id_col] = generate_client_id()
                                            
                                            # Process all rows
                                            if is_new_row:
                                                # INSERT new record
                                                insert_dict = {}
                                                for col in edited_data.columns:
                                                    if col not in ['_id', 'Select']:  # Exclude auto-generated fields and selection column
                                                        value = row[col]
                                                        # Clean numeric values
                                                        if pd.notna(value) and isinstance(value, (int, float)):
                                                            insert_dict[col] = clean_numeric_value(value)
                                                        else:
                                                            insert_dict[col] = value if pd.notna(value) else None
                                                
                                                try:
                                                    supabase.table('policies').insert(insert_dict).execute()
                                                    inserted_count += 1
                                                except Exception as insert_error:
                                                    st.error(f"Error inserting new record: {insert_error}")
                                            elif transaction_id:  # Only update if we have a transaction ID
                                                # UPDATE existing record
                                                update_dict = {}
                                                for col in edited_data.columns:
                                                    if col not in ['_id', 'Select', transaction_id_col]:  # Don't update ID field or selection column
                                                        value = row[col]
                                                        # Clean numeric values
                                                        if pd.notna(value) and isinstance(value, (int, float)):
                                                            update_dict[col] = clean_numeric_value(value)
                                                        else:
                                                            update_dict[col] = value if pd.notna(value) else None
                                                
                                                try:
                                                    supabase.table('policies').update(update_dict).eq(transaction_id_col, transaction_id).execute()
                                                    updated_count += 1
                                                except Exception as update_error:
                                                    st.error(f"Error updating record: {update_error}")
                                            else:
                                                # This shouldn't happen, but log it if it does
                                                st.warning(f"Skipped row {idx} - existing row but no transaction ID found")
                                        
                                        # Clear cache and show success message
                                        clear_policies_cache()
                                        
                                        if inserted_count > 0 and updated_count > 0:
                                            st.success(f"Successfully inserted {inserted_count} new records and updated {updated_count} existing records!")
                                        elif inserted_count > 0:
                                            st.success(f"Successfully inserted {inserted_count} new records!")
                                        elif updated_count > 0:
                                            st.success(f"Successfully updated {updated_count} records!")
                                        else:
                                            st.info("No changes were made.")
                                        
                                        st.rerun()
                                        
                                        # Clear unsaved changes after successful save
                                        if unsaved_changes_key in st.session_state:
                                            del st.session_state[unsaved_changes_key]
                                        
                                    except Exception as e:
                                        st.error(f"Error saving changes: {e}")
                            
                            with col3:
                                if st.button("🔄 Refresh Data", use_container_width=True):
                                    # Clear session state and refresh
                                    if unsaved_changes_key in st.session_state:
                                        del st.session_state[unsaved_changes_key]
                                    st.rerun()
                            
                            # Modal Form Implementation for Edit
                            if st.session_state.get('show_edit_modal', False):
                                # Create a modal-like overlay
                                st.markdown("---")
                                st.markdown("### 📝 Edit Transaction")
                                
                                modal_data = st.session_state.get('edit_modal_data', {})
                                
                                # Remove the Select column from modal data
                                if 'Select' in modal_data:
                                    del modal_data['Select']
                                
                                # Get transaction ID and customer name for header
                                transaction_id = modal_data.get(transaction_id_col, 'Unknown')
                                customer_name = modal_data.get('Customer', 'Unknown')
                                
                                # Check if this is a reconciliation transaction
                                if is_reconciliation_transaction(transaction_id):
                                    st.error("🔒 This is a reconciliation transaction and cannot be edited.")
                                    st.info("Reconciliation entries (-STMT-, -VOID-, -ADJ-) are permanent audit records. Use the Reconciliation page to create adjustments if needed.")
                                    if st.button("Close", type="primary"):
                                        st.session_state['show_edit_modal'] = False
                                        st.session_state['edit_modal_data'] = None
                                        st.rerun()
                                    return  # Exit early
                                
                                st.info(f"**Transaction ID:** {transaction_id} | **Customer:** {customer_name}")
                                
                                # Create form
                                with st.form("edit_transaction_form"):
                                    # Group fields logically
                                    st.markdown("#### Client Information")
                                    col1, col2 = st.columns(2)
                                    
                                    # Track updated values
                                    updated_data = {}
                                    
                                    # Define internal system fields that should be read-only
                                    internal_fields = [
                                        'reconciliation_status', 
                                        'reconciliation_id', 
                                        'reconciled_at', 
                                        'is_reconciliation_entry',
                                        '_id',
                                        'Client ID',
                                        'Client ID (CRM)',
                                        'STMT DATE',
                                        'Agency Comm Received (STMT)',
                                        'Agent Paid Amount (STMT)'
                                    ]
                                    
                                    # Define field groups for better organization
                                    client_fields = ['Client ID (CRM)', 'Client ID', 'Customer', 'Client Name', 'Agent Name']
                                    policy_fields = ['Writing Code', 'Policy #', 'Product', 'Carrier']
                                    date_fields = ['Policy Issue Date', 'Policy Effective Date', 'As of Date']
                                    commission_fields = [
                                        'Premium Sold', 'Agency Estimated Comm/Revenue (CRM)', 
                                        'Policy Gross Comm %', 'Agent Estimated Comm $',
                                        'Agency Comm Received (STMT)', 'Agent Paid Amount (STMT)',
                                        'Agent Comm (NEW 50% RWL 25%)'
                                    ]
                                    status_fields = ['Reconciliation Notes', 'Reconciled?', 'Cross-Reference Key']
                                    
                                    # Client Information
                                    field_counter = 0
                                    for field in modal_data.keys():
                                        # Skip internal fields - they'll be shown at the bottom
                                        if field in client_fields and field not in internal_fields:
                                            with col1 if field_counter % 2 == 0 else col2:
                                                updated_data[field] = st.text_input(
                                                    field, 
                                                    value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                    key=f"modal_{field}"
                                                )
                                            field_counter += 1
                                    
                                    # Policy Information
                                    st.markdown("#### Policy Information")
                                    col3, col4 = st.columns(2)
                                    field_counter = 0
                                    for field in modal_data.keys():
                                        if field in policy_fields:
                                            with col3 if field_counter % 2 == 0 else col4:
                                                updated_data[field] = st.text_input(
                                                    field, 
                                                    value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                    key=f"modal_{field}"
                                                )
                                            field_counter += 1
                                    
                                    # Date Fields
                                    st.markdown("#### Dates")
                                    col5, col6 = st.columns(2)
                                    field_counter = 0
                                    for field in modal_data.keys():
                                        if field in date_fields:
                                            with col5 if field_counter % 2 == 0 else col6:
                                                # For date fields, try to parse existing value
                                                date_value = modal_data.get(field)
                                                if date_value and pd.notna(date_value):
                                                    try:
                                                        # Try to parse the date
                                                        parsed_date = pd.to_datetime(date_value)
                                                        updated_data[field] = st.date_input(
                                                            field,
                                                            value=parsed_date.date(),
                                                            key=f"modal_{field}"
                                                        )
                                                    except:
                                                        updated_data[field] = st.text_input(
                                                            field,
                                                            value=str(date_value),
                                                            key=f"modal_{field}"
                                                        )
                                                else:
                                                    updated_data[field] = st.date_input(
                                                        field,
                                                        value=None,
                                                        key=f"modal_{field}"
                                                    )
                                            field_counter += 1
                                    
                                    # Commission Fields
                                    st.markdown("#### Commission Details")
                                    col7, col8 = st.columns(2)
                                    
                                    # Left column fields in specific order
                                    left_commission_fields = [
                                        'Premium Sold',
                                        'Policy Gross Comm %',
                                        'Agent Comm (NEW 50% RWL 25%)'
                                    ]
                                    
                                    # Right column fields in specific order
                                    right_commission_fields = [
                                        'Agency Estimated Comm/Revenue (CRM)',
                                        'Agent Estimated Comm $'
                                    ]
                                    
                                    # Display left column fields
                                    with col7:
                                        for field in left_commission_fields:
                                            if field in modal_data.keys():
                                                current_value = modal_data.get(field, 0)
                                                if pd.isna(current_value):
                                                    current_value = 0.0
                                                updated_data[field] = st.number_input(
                                                    field,
                                                    value=float(current_value),
                                                    format="%.2f",
                                                    key=f"modal_{field}"
                                                )
                                    
                                    # Display right column fields
                                    with col8:
                                        # Add empty space to align with commission percentages
                                        # Create an invisible number input to match the height of Premium Sold field
                                        st.number_input(
                                            " ",  # Single space as label
                                            value=0.0,
                                            format="%.2f",
                                            key="modal_spacer_commission",
                                            disabled=True,
                                            label_visibility="hidden"
                                        )
                                        
                                        # Calculate formula values
                                        premium_sold = updated_data.get('Premium Sold', modal_data.get('Premium Sold', 0))
                                        gross_comm_pct = updated_data.get('Policy Gross Comm %', modal_data.get('Policy Gross Comm %', 0))
                                        agent_comm_rate = updated_data.get('Agent Comm (NEW 50% RWL 25%)', modal_data.get('Agent Comm (NEW 50% RWL 25%)', 0))
                                        
                                        # Ensure numeric values
                                        try:
                                            premium_sold = float(premium_sold) if premium_sold else 0.0
                                            gross_comm_pct = float(gross_comm_pct) if gross_comm_pct else 0.0
                                            agent_comm_rate = float(agent_comm_rate) if agent_comm_rate else 0.0
                                        except:
                                            premium_sold = 0.0
                                            gross_comm_pct = 0.0
                                            agent_comm_rate = 0.0
                                        
                                        # Handle agent commission rate - it might be stored as decimal (0.50) or percentage (50)
                                        # If it's less than 1, assume it's a decimal and convert to percentage
                                        if agent_comm_rate > 0 and agent_comm_rate < 1:
                                            agent_comm_rate = agent_comm_rate * 100
                                        
                                        # Calculate Agency Estimated Comm/Revenue (CRM)
                                        agency_comm = premium_sold * (gross_comm_pct / 100) if gross_comm_pct else 0.0
                                        
                                        # Calculate Agent Estimated Comm $
                                        agent_comm = agency_comm * (agent_comm_rate / 100) if agent_comm_rate else 0.0
                                        
                                        # Ensure calculated values are valid numbers
                                        if pd.isna(agency_comm) or (isinstance(agency_comm, float) and (agency_comm == float('inf') or agency_comm == float('-inf'))):
                                            agency_comm = 0.0
                                        if pd.isna(agent_comm) or (isinstance(agent_comm, float) and (agent_comm == float('inf') or agent_comm == float('-inf'))):
                                            agent_comm = 0.0
                                        
                                        for field in right_commission_fields:
                                            if field in modal_data.keys():
                                                if field == 'Agency Estimated Comm/Revenue (CRM)':
                                                    # Display as read-only calculated field
                                                    st.number_input(
                                                        field,
                                                        value=agency_comm,
                                                        format="%.2f",
                                                        key=f"modal_{field}_display",
                                                        disabled=True,
                                                        help=f"Formula: Premium Sold × Policy Gross Comm % = ${premium_sold:.2f} × {gross_comm_pct:.1f}% = ${agency_comm:.2f}"
                                                    )
                                                    # Store the calculated value
                                                    updated_data[field] = agency_comm
                                                elif field == 'Agent Estimated Comm $':
                                                    # Display as read-only calculated field
                                                    st.number_input(
                                                        field,
                                                        value=agent_comm,
                                                        format="%.2f",
                                                        key=f"modal_{field}_display",
                                                        disabled=True,
                                                        help=f"Formula: Agency Comm × Agent Rate = ${agency_comm:.2f} × {agent_comm_rate:.1f}% = ${agent_comm:.2f}"
                                                    )
                                                    # Store the calculated value
                                                    updated_data[field] = agent_comm
                                                else:
                                                    # This shouldn't happen but handle it just in case
                                                    current_value = modal_data.get(field, 0)
                                                    if pd.isna(current_value):
                                                        current_value = 0.0
                                                    updated_data[field] = st.number_input(
                                                        field,
                                                        value=float(current_value),
                                                        format="%.2f",
                                                        key=f"modal_{field}"
                                                    )
                                    
                                    # Internal Commission Fields (Reconciliation)
                                    commission_internal_fields = ['Agency Comm Received (STMT)', 'Agent Paid Amount (STMT)', 'STMT DATE']
                                    has_commission_internal = any(field in modal_data.keys() for field in commission_internal_fields)
                                    
                                    if has_commission_internal:
                                        st.markdown("---")
                                        st.markdown("##### Internal Fields (Read-only)")
                                        col7a, col8a = st.columns(2)
                                        field_counter = 0
                                        for field in commission_internal_fields:
                                            if field in modal_data.keys():
                                                with col7a if field_counter % 2 == 0 else col8a:
                                                    if field == 'STMT DATE':
                                                        # STMT DATE is a text/date field
                                                        st.text_input(
                                                            field,
                                                            value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                            key=f"modal_{field}",
                                                            disabled=True,
                                                            help="Statement date - update via Reconciliation page"
                                                        )
                                                    else:
                                                        # Agency Comm Received and Agent Paid Amount are number fields
                                                        current_value = modal_data.get(field, 0)
                                                        if pd.isna(current_value):
                                                            current_value = 0.0
                                                        st.number_input(
                                                            field,
                                                            value=float(current_value),
                                                            format="%.2f",
                                                            key=f"modal_{field}",
                                                            disabled=True,
                                                            help="Reconciliation field - update via Reconciliation page"
                                                        )
                                                field_counter += 1
                                    
                                    # Status Fields
                                    st.markdown("#### Status & Notes")
                                    for field in modal_data.keys():
                                        if field in status_fields:
                                            if field == 'Reconciled?':
                                                current_val = str(modal_data.get(field, 'No')).upper() == 'YES'
                                                updated_data[field] = 'Yes' if st.checkbox(
                                                    "Reconciled?",
                                                    value=current_val,
                                                    key=f"modal_{field}"
                                                ) else 'No'
                                            elif field == 'Reconciliation Notes':
                                                updated_data[field] = st.text_area(
                                                    field,
                                                    value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                    key=f"modal_{field}",
                                                    height=100
                                                )
                                            else:
                                                updated_data[field] = st.text_input(
                                                    field,
                                                    value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                    key=f"modal_{field}"
                                                )
                                    
                                    # Any remaining fields not in our categories OR internal fields that need to be shown at bottom
                                    st.markdown("#### Other Fields")
                                    # Get fields not in any category
                                    uncategorized_fields = [f for f in modal_data.keys() if f not in 
                                                          client_fields + policy_fields + date_fields + 
                                                          commission_fields + status_fields + [transaction_id_col]]
                                    # Add internal fields that might be in other categories but should show here
                                    internal_fields_to_show = [f for f in modal_data.keys() if f in internal_fields]
                                    
                                    # Separate editable and internal fields
                                    editable_other_fields = [f for f in uncategorized_fields if f not in internal_fields]
                                    readonly_fields = list(set(internal_fields_to_show))
                                    
                                    # Display editable fields first in specific order
                                    # Define the specific field order for Other Fields
                                    left_other_fields = [
                                        'Policy Type',
                                        'Carrier Name', 
                                        'Effective Date',
                                        'Transaction Type',
                                        'NEW BIZ CHECKLIST COMPLETE'
                                    ]
                                    
                                    right_other_fields = [
                                        'Policy Origination Date',
                                        'Policy Number',
                                        'X-DATE',
                                        'FULL OR MONTHLY PMTS',
                                        'NOTES'
                                    ]
                                    
                                    # Get any fields not in our specific lists
                                    specified_fields = left_other_fields + right_other_fields
                                    unspecified_editable_fields = [f for f in editable_other_fields if f not in specified_fields]
                                    
                                    # Display fields in specific order
                                    col9, col10 = st.columns(2)
                                    
                                    # Left column
                                    with col9:
                                        for field in left_other_fields:
                                            if field in editable_other_fields and field in modal_data.keys():
                                                updated_data[field] = st.text_input(
                                                    field,
                                                    value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                    key=f"modal_{field}"
                                                )
                                    
                                    # Right column
                                    with col10:
                                        for field in right_other_fields:
                                            if field in editable_other_fields and field in modal_data.keys():
                                                if field == 'NOTES':
                                                    # Make NOTES a text area for better usability
                                                    updated_data[field] = st.text_area(
                                                        field,
                                                        value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                        key=f"modal_{field}",
                                                        height=80
                                                    )
                                                else:
                                                    updated_data[field] = st.text_input(
                                                        field,
                                                        value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                        key=f"modal_{field}"
                                                    )
                                    
                                    # Display any remaining unspecified editable fields below the organized ones
                                    if unspecified_editable_fields:
                                        st.markdown("---")
                                        st.markdown("##### Additional Fields")
                                        col9a, col10a = st.columns(2)
                                        field_counter = 0
                                        for field in sorted(unspecified_editable_fields):
                                            with col9a if field_counter % 2 == 0 else col10a:
                                                updated_data[field] = st.text_input(
                                                    field,
                                                    value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                    key=f"modal_{field}"
                                                )
                                            field_counter += 1
                                    
                                    # Display internal/readonly fields at the bottom
                                    # Exclude commission internal fields that are already shown in Commission Details
                                    commission_internal_fields = ['Agency Comm Received (STMT)', 'Agent Paid Amount (STMT)', 'STMT DATE']
                                    other_readonly_fields = [f for f in readonly_fields if f not in commission_internal_fields]
                                    
                                    if other_readonly_fields:
                                        if editable_other_fields:
                                            st.markdown("---")  # Separator between editable and readonly
                                        st.markdown("##### Internal Fields (Read-only)")
                                        col11, col12 = st.columns(2)
                                        field_counter = 0
                                        for field in sorted(other_readonly_fields):  # Sort for consistent ordering
                                            with col11 if field_counter % 2 == 0 else col12:
                                                st.text_input(
                                                    field,
                                                    value=str(modal_data.get(field, '')) if modal_data.get(field) is not None else '',
                                                    key=f"modal_{field}",
                                                    disabled=True,
                                                    help="Internal system field (read-only)"
                                                )
                                            field_counter += 1
                                    
                                    # Form buttons
                                    st.markdown("---")
                                    col_save, col_cancel = st.columns(2)
                                    
                                    with col_save:
                                        save_modal = st.form_submit_button("💾 Save Changes", type="primary", use_container_width=True)
                                    
                                    with col_cancel:
                                        cancel_modal = st.form_submit_button("❌ Cancel", use_container_width=True)
                                
                                # Handle form submission
                                if save_modal:
                                    try:
                                        # Build complete update dictionary with ALL fields
                                        update_dict = {}
                                        
                                        # First, include ALL original fields from modal_data (except Select and ID)
                                        for field, original_value in modal_data.items():
                                            if field != transaction_id_col and field != 'Select':
                                                # Clean NaN values from original data
                                                if pd.isna(original_value):
                                                    update_dict[field] = None
                                                elif isinstance(original_value, float) and (original_value == float('inf') or original_value == float('-inf')):
                                                    update_dict[field] = None
                                                else:
                                                    update_dict[field] = original_value
                                        
                                        # Then override with any fields that were updated in the form
                                        for field, value in updated_data.items():
                                            # Convert date objects to strings
                                            if isinstance(value, datetime.date):
                                                update_dict[field] = value.isoformat()
                                            # Handle empty strings as None
                                            elif value == '':
                                                update_dict[field] = None
                                            # Handle NaN and None values
                                            elif pd.isna(value):
                                                update_dict[field] = None
                                            # Handle numeric values
                                            elif isinstance(value, (int, float)):
                                                # Check for NaN or infinity
                                                if pd.isna(value) or (isinstance(value, float) and (value == float('inf') or value == float('-inf'))):
                                                    update_dict[field] = None
                                                else:
                                                    update_dict[field] = value
                                            else:
                                                update_dict[field] = value
                                        
                                        # Debug: Show what we're updating
                                        st.write(f"DEBUG: Updating {len(update_dict)} fields for transaction {transaction_id}")
                                        st.write("DEBUG: Sample fields being updated:", list(update_dict.keys())[:5])
                                        
                                        # Check for NaN values before updating
                                        nan_fields = []
                                        for field, value in update_dict.items():
                                            if isinstance(value, float) and pd.isna(value):
                                                nan_fields.append(f"{field}: {value}")
                                                update_dict[field] = None  # Replace NaN with None
                                        
                                        if nan_fields:
                                            st.warning(f"Found NaN values in fields: {nan_fields}. Converting to None.")
                                        
                                        # Update in database with ALL fields
                                        result = supabase.table('policies').update(update_dict).eq(transaction_id_col, transaction_id).execute()
                                        
                                        # Check if update actually happened
                                        if result.data:
                                            st.success(f"✅ Transaction updated successfully! Updated {len(update_dict)} fields.")
                                            
                                            # Verify the update by reading back
                                            verify = supabase.table('policies').select("*").eq(transaction_id_col, transaction_id).execute()
                                            if verify.data and len(verify.data) > 0:
                                                # Check a sample field
                                                sample_field = list(updated_data.keys())[0] if updated_data else None
                                                if sample_field and sample_field in verify.data[0]:
                                                    st.write(f"DEBUG: Verified - {sample_field} is now: {verify.data[0][sample_field]}")
                                            
                                            # Force clear the session state for the editor
                                            if 'edit_policies_editor' in st.session_state:
                                                del st.session_state['edit_policies_editor']
                                            if 'last_search_edit_policies_editor' in st.session_state:
                                                del st.session_state['last_search_edit_policies_editor']
                                        else:
                                            st.error("❌ Update may have failed - no data returned")
                                        
                                        # Clear modal state
                                        st.session_state['show_edit_modal'] = False
                                        st.session_state['edit_modal_data'] = None
                                        
                                        # Clear cache and refresh
                                        clear_policies_cache()
                                        time.sleep(1)
                                        st.rerun()
                                        
                                    except Exception as e:
                                        st.error(f"Error updating transaction: {str(e)}")
                                
                                if cancel_modal:
                                    # Clear modal state
                                    st.session_state['show_edit_modal'] = False
                                    st.session_state['edit_modal_data'] = None
                                    st.rerun()
                            
                            # Delete functionality moved to bottom
                            st.divider()
                            st.subheader("🗑️ Delete Selected Records")
                            
                            # Re-check selected rows for delete functionality
                            selected_rows_for_delete = edited_data[edited_data['Select'] == True].copy()
                            
                            # Collect transaction IDs to delete BEFORE any modifications
                            transaction_ids_to_delete = []
                            reconciliation_attempts = []
                            if not selected_rows_for_delete.empty and transaction_id_col:
                                for idx, row in selected_rows_for_delete.iterrows():
                                    tid = row[transaction_id_col]
                                    if tid and pd.notna(tid):
                                        # Check if this is a reconciliation transaction
                                        if is_reconciliation_transaction(tid):
                                            reconciliation_attempts.append(str(tid))
                                        else:
                                            transaction_ids_to_delete.append(str(tid))
                            
                            # Show error if trying to delete reconciliation transactions
                            if reconciliation_attempts:
                                st.error(f"🔒 Cannot delete {len(reconciliation_attempts)} reconciliation transaction(s):")
                                for tid in reconciliation_attempts:
                                    st.write(f"- {tid}")
                                st.info("Reconciliation entries (-STMT-, -VOID-, -ADJ-) are permanent audit records. Use the Reconciliation page to create adjustments if needed.")
                            
                            if transaction_ids_to_delete:
                                st.warning(f"⚠️ You have selected {len(transaction_ids_to_delete)} record(s) for deletion:")
                                # Show which records are selected
                                for tid in transaction_ids_to_delete:
                                    # Find the customer name for this transaction ID
                                    customer_row = edited_data[edited_data[transaction_id_col] == tid]
                                    if not customer_row.empty:
                                        customer = customer_row.iloc[0].get('Customer', 'Unknown')
                                        st.write(f"- {tid} - {customer}")
                                
                                col1, col2 = st.columns([1, 3])
                                with col1:
                                    if st.button("🗑️ Confirm Delete", type="secondary"):
                                        try:
                                            deleted_count = 0
                                            # Use the pre-collected transaction IDs
                                            for tid in transaction_ids_to_delete:
                                                # Get the full row data for archiving
                                                row_to_archive = edited_data[edited_data[transaction_id_col] == tid]
                                                if not row_to_archive.empty:
                                                    row = row_to_archive.iloc[0]
                                                    # First, archive the record to deleted_policies table
                                                    # Convert row to dict, handling NaN values and type conversion
                                                    policy_dict = {}
                                                    for col in row.index:
                                                        if col != 'Select':  # Skip the selection column
                                                            value = row[col]
                                                            if pd.notna(value):
                                                                # Convert to Python native types for JSON serialization
                                                                if hasattr(value, 'item'):  # numpy scalar
                                                                    policy_dict[col] = value.item()
                                                                elif isinstance(value, (int, float, bool, str)):
                                                                    policy_dict[col] = value
                                                                else:
                                                                    policy_dict[col] = str(value)
                                                            else:
                                                                policy_dict[col] = None
                                                    
                                                    # Create archive record with JSONB structure
                                                    archive_record = {
                                                        'transaction_id': tid,
                                                        'customer_name': row.get('Customer', 'Unknown'),
                                                        'policy_data': policy_dict
                                                    }
                                                    
                                                    try:
                                                        # Insert into deleted_policies table
                                                        supabase.table('deleted_policies').insert(archive_record).execute()
                                                        
                                                        # Then delete from main policies table
                                                        supabase.table('policies').delete().eq(transaction_id_col, tid).execute()
                                                        deleted_count += 1
                                                    except Exception as archive_error:
                                                        st.error(f"Error archiving record {tid}: {archive_error}")
                                            
                                            # Clear cache and session state before rerun
                                            clear_policies_cache()
                                            
                                            # Clear the editor session state to force refresh
                                            if 'edit_policies_editor' in st.session_state:
                                                del st.session_state['edit_policies_editor']
                                            if 'edit_policies_editor_widget' in st.session_state:
                                                del st.session_state['edit_policies_editor_widget']
                                            if 'last_search_edit_policies_editor' in st.session_state:
                                                del st.session_state['last_search_edit_policies_editor']
                                            
                                            st.success(f"Successfully deleted {deleted_count} records! (Archived for recovery)")
                                            time.sleep(1)  # Brief pause to show success message
                                            st.rerun()
                                            
                                        except Exception as e:
                                            st.error(f"Error deleting records: {e}")
                                with col2:
                                    st.info("Click 'Confirm Delete' to permanently remove the selected records.")
                            else:
                                if not selected_rows_for_delete.empty:
                                    st.error("Could not identify transaction IDs for selected rows. Make sure the Transaction ID column is properly identified.")
                                else:
                                    st.info("Check the 'Select' checkbox in the data editor above to select rows for deletion.")
                    else:
                        st.warning("No records found matching your search")
                else:
                    st.info("Enter a search term to find records to edit")
            else:
                st.info("Use the search box above to find policies to edit, or edit all policies below:")
                
                # Option to edit all data (with pagination for performance)
                show_all = st.checkbox("Show all policies for editing (use with caution)")
                
                if show_all:
                    st.warning("Editing all policies at once can be slow with large datasets")
                    
                    # Sort by Customer A-Z first, then limit to first 50 records for performance
                    if 'Customer' in all_data.columns:
                        all_data_sorted = all_data.sort_values(by='Customer', ascending=True)
                        edit_all_data = all_data_sorted.head(50)
                    else:
                        edit_all_data = all_data.head(50)
                    st.write(f"Showing first 50 records for editing out of {len(all_data)} total (sorted by Customer A-Z)")
                    
                    # Find the actual column names dynamically
                    transaction_id_col = None
                    client_id_col = None
                    for col in edit_all_data.columns:
                        if 'transaction' in col.lower() and 'id' in col.lower():
                            transaction_id_col = col
                        if 'client' in col.lower() and 'id' in col.lower():
                            client_id_col = col
                    
                    # Try minimal column config to see if it helps with sorting
                    edited_all_data = st.data_editor(
                        edit_all_data,
                        use_container_width=True,
                        height=500,
                        key="edit_all_policies_editor",
                        num_rows="dynamic",
                        disabled=['_id'] if '_id' in edit_all_data.columns else []  # Only disable _id column
                    )
                    
                    if st.button("💾 Save Changes", key="save_all_edits"):
                        try:
                            updated_count = 0
                            inserted_count = 0
                            
                            for idx, row in edited_all_data.iterrows():
                                # Check if this is a new row
                                is_new_row = idx >= len(edit_all_data)
                                
                                if is_new_row:
                                    # For new rows, generate unique IDs if they're missing
                                    if transaction_id_col and (pd.isna(row[transaction_id_col]) or str(row[transaction_id_col]).strip() == ''):
                                        row[transaction_id_col] = generate_transaction_id()
                                    if client_id_col and (pd.isna(row[client_id_col]) or str(row[client_id_col]).strip() == ''):
                                        row[client_id_col] = generate_client_id()
                                
                                # Get the transaction ID for processing
                                transaction_id = row.get(transaction_id_col) if transaction_id_col else None
                                
                                if transaction_id:
                                    
                                    if is_new_row:
                                        # INSERT new record
                                        insert_dict = {}
                                        for col in edited_all_data.columns:
                                            if col not in ['_id']:  # Exclude auto-generated fields
                                                insert_dict[col] = row[col] if pd.notna(row[col]) else None
                                        
                                        try:
                                            supabase.table('policies').insert(insert_dict).execute()
                                            inserted_count += 1
                                        except Exception as insert_error:
                                            st.error(f"Error inserting new record: {insert_error}")
                                    else:
                                        # UPDATE existing record
                                        update_dict = {}
                                        for col in edited_all_data.columns:
                                            if col not in ['_id', transaction_id_col]:  # Don't update the ID field itself
                                                update_dict[col] = row[col] if pd.notna(row[col]) else None
                                        
                                        try:
                                            supabase.table('policies').update(update_dict).eq(transaction_id_col, transaction_id).execute()
                                            updated_count += 1
                                        except Exception as update_error:
                                            st.error(f"Error updating record: {update_error}")
                            
                            # Clear cache and show success message
                            clear_policies_cache()
                            
                            if inserted_count > 0 and updated_count > 0:
                                st.success(f"Successfully inserted {inserted_count} new records and updated {updated_count} existing records!")
                            elif inserted_count > 0:
                                st.success(f"Successfully inserted {inserted_count} new records!")
                            elif updated_count > 0:
                                st.success(f"Successfully updated {updated_count} records!")
                            else:
                                st.info("No changes were made.")
                            
                            st.rerun()
                            
                        except Exception as e:
                            st.error(f"Error saving changes: {e}")
        else:
            st.info("No policies available to edit.")
    
    # --- Add New Policy Transaction ---
    elif page == "Add New Policy Transaction":
        st.title("➕ Add New Policy Transaction")
        
        # Search for Existing Client ID by Name
        st.subheader("Search for Existing Client ID by Name")
        client_search = st.text_input("Type client name to search:", key="client_search")
        
        selected_client_id = None
        if client_search:
            # Search for matching clients
            matching_clients = all_data[all_data['Customer'].str.contains(client_search, case=False, na=False)][['Client ID', 'Customer']].drop_duplicates()
            if not matching_clients.empty:
                st.info(f"Found {len(matching_clients)} matching clients:")
                # Display options
                for idx, client in matching_clients.iterrows():
                    if st.button(f"Use {client['Client ID']} - {client['Customer']}", key=f"select_client_{idx}"):
                        selected_client_id = client['Client ID']
                        st.session_state['selected_client_id'] = selected_client_id
                        st.session_state['selected_customer_name'] = client['Customer']
            else:
                st.warning("No matching clients found")
        
        # Get selected client ID and name from session state if available
        if 'selected_client_id' in st.session_state:
            selected_client_id = st.session_state['selected_client_id']
        
        selected_customer_name = None
        if 'selected_customer_name' in st.session_state:
            selected_customer_name = st.session_state['selected_customer_name']
        
        # Premium Sold Calculator
        st.markdown("---")
        st.subheader("Premium Sold Calculator (for Endorsements)")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            existing_premium = st.number_input("Existing Premium", value=-1000000.00, format="%.2f", step=100.0, key="existing_premium")
        with col2:
            new_premium = st.number_input("New/Revised Premium", value=-1000000.00, format="%.2f", step=100.0, key="new_premium")
        with col3:
            premium_sold_calc = new_premium - existing_premium
            st.metric("Premium Sold (New - Existing):", f"${premium_sold_calc:+,.2f}")
        
        # Agency Revenue Calculator
        st.markdown("---")
        st.subheader("Enter Premium Sold and see Agency Revenue:")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            premium_sold = st.number_input("Premium Sold", value=-1000000.00, format="%.2f", step=100.0, key="premium_sold_input")
        with col2:
            policy_gross_comm_percent = st.number_input("Policy Gross Comm %", value=0.0, format="%.2f", min_value=0.0, max_value=100.0, key="policy_gross_comm_calc")
        with col3:
            agency_revenue = premium_sold * (policy_gross_comm_percent / 100)
            st.metric("Agency Revenue", f"${agency_revenue:,.2f}")
        
        # Main Form
        st.markdown("---")
        with st.form("add_policy_form"):
            # Row 1: Client ID and Transaction ID
            col1, col2 = st.columns(2)
            with col1:
                if selected_client_id:
                    client_id = st.text_input("Client ID", value=selected_client_id, disabled=True)
                else:
                    client_id = st.text_input("Client ID", value=generate_client_id())
            with col2:
                transaction_id = st.text_input("Transaction ID", value=generate_transaction_id(), disabled=True)
            
            # Row 2: Customer
            if selected_customer_name:
                customer = st.text_input("Customer", value=selected_customer_name)
            else:
                customer = st.text_input("Customer", placeholder="Enter customer name")
            
            # Row 3: Carrier Name
            carrier_name = st.text_input("Carrier Name", placeholder="Enter carrier name")
            
            # Row 4: Policy Type
            policy_type = st.selectbox("Policy Type", ["Auto", "Home", "Life", "Health", "Commercial", "Umbrella", "Flood", "Other"])
            
            # Row 5: Policy Number
            policy_number = st.text_input("Policy Number", placeholder="Enter policy number")
            
            # Row 6: Transaction Type
            transaction_type = st.selectbox("Transaction Type", ["NEW", "RWL", "END", "PCH", "CAN", "XCL", "NBS", "STL", "BoR", "REWRITE"])
            
            # Row 7: Agent Commission
            agent_comm_label = st.text_input("Agent Comm (NEW 50% RWL 25%)", value="Calculated based on Transaction Type", disabled=True)
            
            # Date fields
            col1, col2, col3 = st.columns(3)
            with col1:
                policy_orig_date = st.date_input("Policy Origination Date", value=datetime.date.today(), format="MM/DD/YYYY")
            with col2:
                effective_date = st.date_input("Effective Date", value=datetime.date.today(), format="MM/DD/YYYY")
            with col3:
                x_date = st.date_input("X-DATE", value=datetime.date.today() + datetime.timedelta(days=180), format="MM/DD/YYYY")
            
            # Checklist
            new_biz_checklist = st.checkbox("NEW BIZ CHECKLIST COMPLETE")
            
            # Financial fields
            col1, col2 = st.columns(2)
            with col1:
                agency_est_comm = st.number_input("Agency Estimated Comm/Revenue (CRM)", value=agency_revenue if 'agency_revenue' in locals() else 0.0, format="%.2f")
                agency_gross_comm = st.number_input("Agency Gross Comm Received", value=0.0, format="%.2f")
            
            with col2:
                # Calculate agent commission based on transaction type
                if transaction_type in ["NEW", "NBS", "STL", "BoR"]:
                    agent_comm_rate = 0.50
                elif transaction_type in ["END", "PCH"]:
                    agent_comm_rate = 0.50 if policy_orig_date == effective_date else 0.25
                elif transaction_type in ["RWL", "REWRITE"]:
                    agent_comm_rate = 0.25
                else:
                    agent_comm_rate = 0.0
                
                agent_est_comm = agency_est_comm * agent_comm_rate
                agent_est_comm_input = st.number_input("Agent Estimated Comm $", value=agent_est_comm, format="%.2f")
                
                # BALANCE DUE field removed - no longer needed
                full_or_monthly = st.selectbox("FULL OR MONTHLY PMTS", ["FULL", "MONTHLY", ""])
            
            # Notes
            notes = st.text_area("NOTES", placeholder="Enter any additional notes...")
            
            # Submit button
            submitted = st.form_submit_button("💾 Save Policy Transaction", type="primary", use_container_width=True)
            
            if submitted:
                if customer and policy_number:
                    try:
                        # Prepare the new policy record
                        new_policy = {
                            "Client ID": client_id,
                            "Transaction ID": transaction_id,
                            "Customer": customer,
                            "Carrier Name": carrier_name,
                            "Policy Type": policy_type,
                            "Policy Number": policy_number,
                            "Transaction Type": transaction_type,
                            "Policy Origination Date": policy_orig_date.strftime('%m/%d/%Y'),
                            "Effective Date": effective_date.strftime('%m/%d/%Y'),
                            "X-DATE": x_date.strftime('%m/%d/%Y'),
                            "NEW BIZ CHECKLIST COMPLETE": "Yes" if new_biz_checklist else "No",
                            "Policy Gross Comm %": f"{policy_gross_comm_percent}%" if 'policy_gross_comm_percent' in locals() and policy_gross_comm_percent else None,
                            "Agency Estimated Comm/Revenue (CRM)": clean_numeric_value(agency_est_comm),
                            "Agency Gross Comm Received": clean_numeric_value(agency_gross_comm),
                            "Agent Estimated Comm $": clean_numeric_value(agent_est_comm_input),
                            # "BALANCE DUE" field removed - no longer needed
                            "FULL OR MONTHLY PMTS": full_or_monthly,
                            "NOTES": notes,
                            "Premium Sold": f"${premium_sold:,.2f}" if 'premium_sold' in locals() else None
                        }
                        
                        # Remove None values to avoid database issues
                        new_policy = {k: v for k, v in new_policy.items() if v is not None}
                        
                        # Insert into database
                        supabase.table('policies').insert(new_policy).execute()
                        clear_policies_cache()
                        
                        st.success(f"✅ Successfully added new policy transaction for {customer}")
                        st.balloons()
                        
                        # Clear session state
                        if 'selected_client_id' in st.session_state:
                            del st.session_state['selected_client_id']
                        if 'selected_customer_name' in st.session_state:
                            del st.session_state['selected_customer_name']
                        
                        # Set a flag to show the add another button
                        st.session_state['show_add_another'] = True
                        st.rerun()
                            
                    except Exception as e:
                        st.error(f"Error adding policy: {e}")
                        st.error("Please check that all required fields are filled correctly.")
                else:
                    st.error("Please fill in at least Customer Name and Policy Number")
        
        # Show "Add Another" button outside the form if a policy was just added
        if st.session_state.get('show_add_another', False):
            if st.button("➕ Add Another Policy", type="primary"):
                # Clear the flag
                st.session_state['show_add_another'] = False
                st.rerun()
    
    # --- Search & Filter ---
    elif page == "Search & Filter":
        st.title("🔍 Advanced Search & Filter")
        
        if not all_data.empty:
            st.subheader("Search Criteria")
            
            # Create filter form
            with st.form("advanced_search_form"):
                search_col1, search_col2 = st.columns(2)
                
                with search_col1:
                    # Text search fields
                    customer_search = st.text_input("Customer Name Contains")
                    policy_number_search = st.text_input("Policy Number Contains")
                    client_id_search = st.text_input("Client ID Contains")
                    transaction_id_search = st.text_input("Transaction ID Contains")
                
                with search_col2:
                    # Dropdown filters
                    if 'Policy_Type' in all_data.columns:
                        policy_type_filter = st.multiselect("Policy Type", all_data['Policy_Type'].unique())
                    
                    if 'Transaction_Type' in all_data.columns:
                        transaction_type_filter = st.multiselect("Transaction Type", all_data['Transaction_Type'].unique())
                    
                    # Date range
                    if 'Effective_Date' in all_data.columns:
                        date_from = st.date_input("Effective Date From", value=None)
                        date_to = st.date_input("Effective Date To", value=None)
                
                # Numeric filters
                st.subheader("Numeric Filters")
                numeric_col1, numeric_col2 = st.columns(2)
                
                with numeric_col1:
                    if 'Commission_Paid' in all_data.columns:
                        commission_min = st.number_input("Min Commission Paid", value=0.0, format="%.2f")
                        commission_max = st.number_input("Max Commission Paid", value=float(all_data['Commission_Paid'].max() if all_data['Commission_Paid'].max() > 0 else 999999), format="%.2f")
                
                with numeric_col2:
                    if 'Balance_Due' in all_data.columns:
                        balance_min = st.number_input("Min Balance Due", value=0.0, format="%.2f")
                        balance_max = st.number_input("Max Balance Due", value=float(all_data['Balance_Due'].max() if all_data['Balance_Due'].max() > 0 else 999999), format="%.2f")
                
                # Submit search
                search_submitted = st.form_submit_button("🔍 Apply Filters", type="primary")
            
            # Reset filters button
            if st.button("🔄 Clear All Filters"):
                st.rerun()
            
            # Apply filters and show results
            if search_submitted or any([customer_search, policy_number_search, client_id_search, transaction_id_search]):
                filtered_data = all_data.copy()
                
                # Apply text filters
                if customer_search:
                    filtered_data = filtered_data[filtered_data['Customer'].str.contains(customer_search, case=False, na=False)]
                
                if policy_number_search:
                    filtered_data = filtered_data[filtered_data['Policy_Number'].str.contains(policy_number_search, case=False, na=False)]
                
                if client_id_search:
                    filtered_data = filtered_data[filtered_data['Client_ID'].str.contains(client_id_search, case=False, na=False)]
                
                if transaction_id_search:
                    filtered_data = filtered_data[filtered_data['Transaction_ID'].str.contains(transaction_id_search, case=False, na=False)]
                
                # Apply dropdown filters
                if 'Policy_Type' in all_data.columns and policy_type_filter:
                    filtered_data = filtered_data[filtered_data['Policy_Type'].isin(policy_type_filter)]
                
                if 'Transaction_Type' in all_data.columns and transaction_type_filter:
                    filtered_data = filtered_data[filtered_data['Transaction_Type'].isin(transaction_type_filter)]
                
                # Apply date filters
                if 'Effective_Date' in all_data.columns and (date_from or date_to):
                    filtered_data['Effective_Date'] = pd.to_datetime(filtered_data['Effective_Date'], errors='coerce')
                    
                    if date_from:
                        filtered_data = filtered_data[filtered_data['Effective_Date'] >= pd.Timestamp(date_from)]
                    
                    if date_to:
                        filtered_data = filtered_data[filtered_data['Effective_Date'] <= pd.Timestamp(date_to)]
                
                # Apply numeric filters
                if 'Commission_Paid' in all_data.columns:
                    filtered_data = filtered_data[
                        (filtered_data['Commission_Paid'] >= commission_min) &
                        (filtered_data['Commission_Paid'] <= commission_max)
                    ]
                
                if 'Balance_Due' in all_data.columns:
                    filtered_data = filtered_data[
                        (filtered_data['Balance_Due'] >= balance_min) &
                        (filtered_data['Balance_Due'] <= balance_max)
                    ]
                
                # Show results
                st.subheader(f"Search Results ({len(filtered_data)} records found)")
                
                if not filtered_data.empty:
                    # Configure column settings for proper numeric display
                    column_config = {}
                    numeric_cols = [
                        'Agent Estimated Comm $',
                                'Policy Gross Comm %',
                        'Agency Estimated Comm/Revenue (CRM)',
                        'Agency Comm Received (STMT)',
                        'Premium Sold',
                        'Agent Paid Amount (STMT)',
                        'Agency Comm Received (STMT)'
                    ]
                    
                    for col in numeric_cols:
                        if col in filtered_data.columns:
                            column_config[col] = st.column_config.NumberColumn(
                                col,
                                format="%.2f"
                            )
                    
                    # Display filtered data
                    st.dataframe(
                        filtered_data,
                        use_container_width=True,
                        height=500,
                        column_config=column_config
                    )
                    
                    # Export filtered results
                    st.subheader("Export Filtered Results")
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write("**📄 CSV Export**")
                        csv = filtered_data.to_csv(index=False)
                        st.download_button(
                            label="📥 Download as CSV",
                            data=csv,
                            file_name="filtered_policies.csv",
                            mime="text/csv",
                            help="Export filtered results as CSV file"
                        )
                    
                    with col2:
                        st.write("**📊 Excel Export**")
                        excel_buffer, excel_filename = create_formatted_excel_file(
                            filtered_data, 
                            sheet_name="Filtered Results", 
                            filename_prefix="filtered_policies"
                        )
                        if excel_buffer:
                            st.download_button(
                                label="📥 Download as Excel",
                                data=excel_buffer,
                                file_name=excel_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                help="Export filtered results as formatted Excel file"
                            )
                else:
                    st.warning("No records match your search criteria")
            else:
                st.info("Use the form above to search and filter policies")
        else:
            st.info("No data available to search.")
    
    # --- Reconciliation ---
    elif page == "Reconciliation":
        st.title("💳 Commission Reconciliation")
        
        # Create tabs for different reconciliation functions
        rec_tab1, rec_tab2, rec_tab3, rec_tab4 = st.tabs([
            "Import Statement", 
            "Unreconciled Transactions", 
            "Reconciliation History",
            "Adjustments & Voids"
        ])
        
        with rec_tab1:
            st.subheader("📥 Import Commission Statement")
            
            # Add custom CSS for yellow border on Statement Date
            st.markdown("""
                <style>
                /* Target the Statement Date input specifically */
                div.stDateInput > div > div[data-baseweb="input"] {
                    border: 3px solid #ffd700 !important;
                    border-radius: 4px !important;
                }
                
                /* Alternative selector for broader compatibility */
                section[data-testid="stSidebar"] + div div.row-widget.stDateInput:first-of-type input {
                    border: 3px solid #ffd700 !important;
                    border-radius: 4px !important;
                    box-shadow: 0 0 0 2px #ffd700 !important;
                }
                
                /* Target the container div */
                div.row-widget.stDateInput:first-of-type > div {
                    border: 3px solid #ffd700 !important;
                    border-radius: 4px !important;
                    padding: 2px;
                }
                </style>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Create a container with yellow background as a visual highlight
                st.markdown("""
                    <div style="background-color: #fffacd; border: 3px solid #ffd700; border-radius: 8px; padding: 10px; margin-bottom: 10px;">
                        <p style="margin: 0; font-weight: bold; color: #333;">📅 Statement Date</p>
                    </div>
                """, unsafe_allow_html=True)
                
                statement_date = st.date_input(
                    "",  # Empty label since we have the custom header above
                    value=datetime.date.today(),
                    help="The date on the commission statement",
                    format="MM/DD/YYYY",
                    key="statement_date_input"
                )
            
            with col2:
                reconciliation_date = st.date_input(
                    "Date Reconciled",
                    value=datetime.date.today(),
                    disabled=True,
                    help="Today's date (when reconciliation is performed)",
                    format="MM/DD/YYYY"
                )
            
            st.divider()
            
            # Method selection
            import_method = st.radio(
                "Select Import Method",
                ["Manual Entry", "Upload CSV/Excel File"],
                horizontal=True
            )
            
            if import_method == "Manual Entry":
                st.subheader("Manual Statement Entry")
                
                # Initialize batch in session state
                if 'reconciliation_batch' not in st.session_state:
                    st.session_state.reconciliation_batch = []
                if 'statement_total' not in st.session_state:
                    st.session_state.statement_total = 0.0
                
                # Statement total input
                col1, col2 = st.columns([2, 1])
                with col1:
                    statement_total = st.number_input(
                        "Enter Total Commission from Statement",
                        min_value=0.0,
                        value=st.session_state.statement_total,
                        step=0.01,
                        format="%.2f",
                        help="Enter the total commission amount shown on your statement"
                    )
                    st.session_state.statement_total = statement_total
                
                with col2:
                    batch_total = sum(item['amount'] for item in st.session_state.reconciliation_batch)
                    if statement_total > 0:
                        progress = min(batch_total / statement_total, 1.0)
                        st.metric("Batch Progress", f"${batch_total:,.2f}")
                        st.progress(progress)
                        if batch_total == statement_total:
                            st.success("✓ Ready to reconcile!")
                        elif batch_total > statement_total:
                            st.error("Over statement amount!")
                        else:
                            st.info(f"${statement_total - batch_total:,.2f} remaining")
                
                # Show current batch if any
                if st.session_state.reconciliation_batch:
                    st.divider()
                    st.markdown("### Current Reconciliation Batch")
                    
                    # Show batch summary
                    batch_df = pd.DataFrame(st.session_state.reconciliation_batch)
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Items in Batch", len(batch_df))
                    with col2:
                        batch_total = batch_df['amount'].sum()
                        st.metric("Batch Total", f"${batch_total:,.2f}")
                    with col3:
                        if statement_total > 0:
                            variance = batch_total - statement_total
                            if variance == 0:
                                st.metric("Variance", "$0.00", delta_color="off")
                                st.success("✓ Perfectly balanced!")
                            else:
                                st.metric("Variance", f"${variance:,.2f}", delta=f"${abs(variance):,.2f}")
                    
                    # Enhanced batch display with more details
                    batch_df['Remove'] = False
                    
                    # Add more display columns
                    display_columns = ['Remove', 'transaction_id', 'customer', 'policy_number', 
                                     'policy_type', 'effective_date', 'balance', 'amount']
                    
                    # Only include columns that exist
                    available_columns = [col for col in display_columns if col in batch_df.columns]
                    
                    edited_batch = st.data_editor(
                        batch_df[available_columns],
                        column_config={
                            "Remove": st.column_config.CheckboxColumn("Remove", help="Check to remove from batch"),
                            "transaction_id": st.column_config.TextColumn("Transaction ID", disabled=True),
                            "customer": st.column_config.TextColumn("Customer", disabled=True),
                            "policy_number": st.column_config.TextColumn("Policy #", disabled=True),
                            "policy_type": st.column_config.TextColumn("Type", disabled=True),
                            "effective_date": st.column_config.TextColumn("Effective", disabled=True),
                            "balance": st.column_config.NumberColumn("Full Balance", disabled=True, format="$%.2f", help="Total outstanding balance"),
                            "amount": st.column_config.NumberColumn("To Reconcile", disabled=True, format="$%.2f", help="Amount to reconcile in this batch")
                        },
                        use_container_width=True,
                        hide_index=True,
                        key="batch_editor"
                    )
                    
                    # Remove checked items
                    if edited_batch['Remove'].any():
                        items_to_remove = edited_batch[edited_batch['Remove']]['transaction_id'].tolist()
                        st.session_state.reconciliation_batch = [
                            item for item in st.session_state.reconciliation_batch 
                            if item['transaction_id'] not in items_to_remove
                        ]
                        st.rerun()
                    
                    # Reconcile batch button
                    col1, col2, col3 = st.columns([1, 1, 2])
                    
                    with col1:
                        # Only enable reconcile button if totals match
                        can_reconcile = batch_total == statement_total and statement_total > 0 and len(st.session_state.reconciliation_batch) > 0
                        
                        if can_reconcile:
                            if st.button("🔄 Reconcile Batch", type="primary", key="reconcile_batch"):
                                try:
                                    # Generate batch reconciliation ID
                                    batch_id = f"REC-{statement_date.strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
                                    success_count = 0
                                    
                                    # Create reconciliation entries for each item in batch
                                    for item in st.session_state.reconciliation_batch:
                                        recon_id = generate_reconciliation_transaction_id("STMT", statement_date)
                                        
                                        # Create reconciliation entry
                                        recon_entry = {
                                            'Transaction ID': recon_id,
                                            'Client ID': item['client_id'],
                                            'Customer': item['customer'],
                                            'Carrier Name': item['carrier_name'],
                                            'Policy Type': item['policy_type'],
                                            'Policy Number': item['policy_number'],
                                            'Transaction Type': item['transaction_type'],
                                            'Effective Date': item['effective_date'],
                                            'X-DATE': item['x_date'],
                                            'Premium Sold': 0,
                                            'Policy Gross Comm %': 0,
                                            'Agency Estimated Comm/Revenue (CRM)': 0,
                                            'Agency Comm Received (STMT)': item['amount'],
                                            'Agent Estimated Comm $': 0,
                                            'Agent Paid Amount (STMT)': item['amount'],
                                            'STMT DATE': statement_date.strftime('%Y-%m-%d'),
                                            'reconciliation_status': 'reconciled',
                                            'reconciliation_id': batch_id,
                                            'reconciled_at': datetime.datetime.now().isoformat(),
                                            'is_reconciliation_entry': True
                                        }
                                        
                                        # Insert reconciliation entry
                                        supabase.table('policies').insert(recon_entry).execute()
                                        
                                        # Update original transaction
                                        supabase.table('policies').update({
                                            'reconciliation_status': 'reconciled',
                                            'reconciliation_id': batch_id,
                                            'reconciled_at': datetime.datetime.now().isoformat()
                                        }).eq('_id', item['_id']).execute()
                                        
                                        success_count += 1
                                    
                                    # Clear batch
                                    st.session_state.reconciliation_batch = []
                                    st.session_state.statement_total = 0.0
                                    
                                    st.success(f"✅ Batch reconciled successfully! {success_count} transactions processed. Batch ID: {batch_id}")
                                    st.cache_data.clear()
                                    time.sleep(2)
                                    st.rerun()
                                    
                                except Exception as e:
                                    st.error(f"Error reconciling batch: {str(e)}")
                        else:
                            # Show disabled button with explanation
                            st.button("🔄 Reconcile Batch", type="primary", disabled=True, key="reconcile_batch_disabled")
                            if batch_total != statement_total:
                                if batch_total > statement_total:
                                    st.error(f"Batch exceeds statement by ${batch_total - statement_total:,.2f}")
                                else:
                                    st.warning(f"Need ${statement_total - batch_total:,.2f} more to match statement")
                            elif statement_total == 0:
                                st.warning("Enter statement total first")
                            elif len(st.session_state.reconciliation_batch) == 0:
                                st.info("Add transactions to batch")
                    
                    with col2:
                        if st.button("Clear Batch", type="secondary"):
                            st.session_state.reconciliation_batch = []
                            st.rerun()
                
                st.divider()
                
                # Precise drill-down selection (from Accounting page)
                st.markdown("### Select Transaction to Add to Batch")
                
                # Initialize session state for selections if not exists
                if 'recon_selected_customer' not in st.session_state:
                    st.session_state.recon_selected_customer = None
                if 'recon_selected_policy_type' not in st.session_state:
                    st.session_state.recon_selected_policy_type = None
                if 'recon_selected_policy_number' not in st.session_state:
                    st.session_state.recon_selected_policy_number = None
                if 'recon_selected_effective_date' not in st.session_state:
                    st.session_state.recon_selected_effective_date = None
                
                # Customer selection
                customers = sorted(all_data["Customer"].dropna().unique().tolist()) if not all_data.empty else []
                selected_customer = st.selectbox(
                    "Select Customer", 
                    ["Select..."] + customers, 
                    key="recon_drill_customer",
                    index=0 if st.session_state.recon_selected_customer is None else (
                        customers.index(st.session_state.recon_selected_customer) + 1 
                        if st.session_state.recon_selected_customer in customers else 0
                    )
                )
                
                # Initialize variables
                policy_types = []
                policy_numbers = []
                effective_dates = []
                selected_policy_type = None
                selected_policy_number = None
                selected_effective_date = None
                client_id = None
                transaction_id = None
                
                # Policy Type selection
                if selected_customer and selected_customer != "Select...":
                    st.session_state.recon_selected_customer = selected_customer
                    
                    # Get all original transactions for this customer
                    customer_data = all_data[
                        (all_data["Customer"] == selected_customer) &
                        (~all_data['Transaction ID'].str.contains('-STMT-|-ADJ-|-VOID-', na=False))
                    ]
                    
                    # Calculate balance for each transaction
                    if not customer_data.empty:
                        # Get all reconciliation entries for balance calculation
                        for idx, row in customer_data.iterrows():
                            trans_id = row['Transaction ID']
                            policy_num = row['Policy Number']
                            
                            # Calculate credits (commission owed)
                            credit = float(row.get('Agent Estimated Comm $', 0) or 0)
                            
                            # Calculate debits (total paid for this transaction)
                            # Find all -STMT- entries for this original transaction
                            stmt_entries = all_data[
                                (all_data['Policy Number'] == policy_num) &
                                (all_data['Transaction ID'].str.contains('-STMT-', na=False))
                            ]
                            
                            debit = 0
                            if not stmt_entries.empty:
                                debit = stmt_entries['Agent Paid Amount (STMT)'].fillna(0).sum()
                            
                            # Calculate balance
                            balance = credit - debit
                            customer_data.at[idx, '_balance'] = balance
                        
                        # Filter to show only transactions with outstanding balance
                        customer_data = customer_data[customer_data['_balance'] > 0.01]  # Small threshold to avoid floating point issues
                    
                    if not customer_data.empty:
                        policy_types = sorted(customer_data["Policy Type"].dropna().unique().tolist())
                        selected_policy_type = st.selectbox(
                            "Select Policy Type", 
                            ["Select..."] + policy_types, 
                            key="recon_drill_policy_type"
                        )
                        
                        # Policy Number selection
                        if selected_policy_type and selected_policy_type != "Select...":
                            st.session_state.recon_selected_policy_type = selected_policy_type
                            
                            policy_data = customer_data[customer_data["Policy Type"] == selected_policy_type]
                            policy_numbers = sorted(policy_data["Policy Number"].dropna().unique().tolist())
                            selected_policy_number = st.selectbox(
                                "Select Policy Number", 
                                ["Select..."] + policy_numbers, 
                                key="recon_drill_policy_number"
                            )
                            
                            # Effective Date selection
                            if selected_policy_number and selected_policy_number != "Select...":
                                st.session_state.recon_selected_policy_number = selected_policy_number
                                
                                number_data = policy_data[policy_data["Policy Number"] == selected_policy_number]
                                effective_dates = sorted(number_data["Effective Date"].dropna().unique().tolist())
                                selected_effective_date = st.selectbox(
                                    "Select Effective Date", 
                                    ["Select..."] + effective_dates, 
                                    key="recon_drill_effective_date"
                                )
                                
                                # Get exact transaction when all fields are selected
                                if selected_effective_date and selected_effective_date != "Select...":
                                    st.session_state.recon_selected_effective_date = selected_effective_date
                                    
                                    exact_match = number_data[number_data["Effective Date"] == selected_effective_date]
                                    
                                    if not exact_match.empty:
                                        # Get the specific transaction details
                                        transaction = exact_match.iloc[0]
                                        client_id = transaction.get("Client ID", "")
                                        transaction_id = transaction.get("Transaction ID", "")
                                        
                                        # Display transaction details
                                        st.success(f"✅ Found Transaction: {transaction_id}")
                                        
                                        # Calculate outstanding balance
                                        credit = float(transaction.get('Agent Estimated Comm $', 0) or 0)
                                        balance = float(transaction.get('_balance', credit))
                                        
                                        # Show transaction summary with balance
                                        col1, col2, col3, col4 = st.columns(4)
                                        with col1:
                                            st.metric("Premium Sold", f"${transaction.get('Premium Sold', 0):,.2f}")
                                        with col2:
                                            st.metric("Commission Owed", f"${credit:,.2f}")
                                        with col3:
                                            previous_payments = credit - balance
                                            st.metric("Previous Payments", f"${previous_payments:,.2f}")
                                        with col4:
                                            st.metric("Outstanding Balance", f"${balance:,.2f}", 
                                                     help="Amount still owed for this transaction")
                                        
                                        # Show full transaction details
                                        with st.expander("📋 View Full Transaction Details", expanded=True):
                                            # Create two columns for organized display
                                            detail_col1, detail_col2 = st.columns(2)
                                            
                                            with detail_col1:
                                                st.markdown("**Policy Information**")
                                                details_dict = {
                                                    "Transaction ID": transaction.get('Transaction ID', ''),
                                                    "Client ID": transaction.get('Client ID', ''),
                                                    "Customer": transaction.get('Customer', ''),
                                                    "Policy Number": transaction.get('Policy Number', ''),
                                                    "Policy Type": transaction.get('Policy Type', ''),
                                                    "Carrier Name": transaction.get('Carrier Name', ''),
                                                    "Transaction Type": transaction.get('Transaction Type', ''),
                                                }
                                                for key, value in details_dict.items():
                                                    st.text(f"{key}: {value}")
                                            
                                            with detail_col2:
                                                st.markdown("**Dates & Amounts**")
                                                dates_dict = {
                                                    "Effective Date": transaction.get('Effective Date', ''),
                                                    "Policy Origination Date": transaction.get('Policy Origination Date', ''),
                                                    "X-DATE": transaction.get('X-DATE', ''),
                                                    "Premium Sold": f"${float(transaction.get('Premium Sold', 0) or 0):,.2f}",
                                                    "Policy Gross Comm %": f"{float(transaction.get('Policy Gross Comm %', 0) or 0):.2f}%",
                                                    "Agency Est. Comm": f"${float(transaction.get('Agency Estimated Comm/Revenue (CRM)', 0) or 0):,.2f}",
                                                    "Agent Comm Rate": transaction.get('Agent Comm (NEW 50% RWL 25%)', ''),
                                                }
                                                for key, value in dates_dict.items():
                                                    st.text(f"{key}: {value}")
                                            
                                            # Show payment history if any
                                            if previous_payments > 0:
                                                st.divider()
                                                st.markdown("**Payment History**")
                                                
                                                # Find all reconciliation entries for this policy
                                                policy_num = transaction.get('Policy Number', '')
                                                payment_history = all_data[
                                                    (all_data['Policy Number'] == policy_num) &
                                                    (all_data['Transaction ID'].str.contains('-STMT-', na=False))
                                                ]
                                                
                                                if not payment_history.empty:
                                                    payment_df = payment_history[['Transaction ID', 'STMT DATE', 'Agent Paid Amount (STMT)']].copy()
                                                    payment_df['STMT DATE'] = pd.to_datetime(payment_df['STMT DATE']).dt.strftime('%m/%d/%Y')
                                                    payment_df = payment_df.rename(columns={
                                                        'Transaction ID': 'Reconciliation ID',
                                                        'STMT DATE': 'Statement Date',
                                                        'Agent Paid Amount (STMT)': 'Amount Paid'
                                                    })
                                                    
                                                    st.dataframe(
                                                        payment_df,
                                                        column_config={
                                                            "Amount Paid": st.column_config.NumberColumn(format="$%.2f")
                                                        },
                                                        use_container_width=True,
                                                        hide_index=True
                                                    )
                                        
                                        st.divider()
                                        
                                        # Add to batch section
                                        st.markdown("### Add to Reconciliation Batch")
                                        
                                        # Amount to reconcile
                                        amount_to_reconcile = st.number_input(
                                            "Amount to Reconcile",
                                            min_value=0.01,
                                            max_value=balance,
                                            value=balance,
                                            step=0.01,
                                            format="%.2f",
                                            key="recon_amount",
                                            help=f"Enter amount from statement for this transaction (max: ${balance:,.2f})"
                                        )
                                        
                                        # Check if already in batch
                                        already_in_batch = any(
                                            item['transaction_id'] == transaction_id 
                                            for item in st.session_state.reconciliation_batch
                                        )
                                        
                                        col1, col2 = st.columns(2)
                                        
                                        with col1:
                                            if already_in_batch:
                                                st.warning("✓ Already in batch")
                                            else:
                                                if st.button("➕ Add to Batch", type="primary", key="add_to_batch"):
                                                    # Add to reconciliation batch
                                                    batch_item = {
                                                        '_id': transaction['_id'],
                                                        'transaction_id': transaction_id,
                                                        'client_id': client_id,
                                                        'customer': selected_customer,
                                                        'carrier_name': transaction.get('Carrier Name', ''),
                                                        'policy_type': selected_policy_type,
                                                        'policy_number': selected_policy_number,
                                                        'transaction_type': transaction.get('Transaction Type', ''),
                                                        'effective_date': selected_effective_date,
                                                        'x_date': transaction.get('X-DATE', ''),
                                                        'amount': amount_to_reconcile,
                                                        'balance': balance
                                                    }
                                                    
                                                    st.session_state.reconciliation_batch.append(batch_item)
                                                    st.success(f"✅ Added ${amount_to_reconcile:,.2f} to batch")
                                                    
                                                    # Clear selections
                                                    st.session_state.recon_selected_customer = None
                                                    st.session_state.recon_selected_policy_type = None
                                                    st.session_state.recon_selected_policy_number = None
                                                    st.session_state.recon_selected_effective_date = None
                                                    
                                                    time.sleep(1)
                                                    st.rerun()
                                        
                                        with col2:
                                            # Show how this affects batch total
                                            if not already_in_batch and st.session_state.statement_total > 0:
                                                new_batch_total = sum(item['amount'] for item in st.session_state.reconciliation_batch) + amount_to_reconcile
                                                if new_batch_total > st.session_state.statement_total:
                                                    st.error(f"Would exceed statement by ${new_batch_total - st.session_state.statement_total:,.2f}")
                                                else:
                                                    remaining = st.session_state.statement_total - new_batch_total
                                                    st.info(f"Would leave ${remaining:,.2f} to reconcile")
                                    else:
                                        st.warning("No transaction found for this selection")
                    else:
                        st.info(f"No transactions with outstanding balance found for {selected_customer}")
            
            else:  # Upload CSV/Excel File
                st.subheader("📤 Upload Statement File")
                
                # Initialize session state for import
                if 'import_data' not in st.session_state:
                    st.session_state.import_data = None
                if 'column_mapping' not in st.session_state:
                    st.session_state.column_mapping = {}
                if 'matched_transactions' not in st.session_state:
                    st.session_state.matched_transactions = []
                if 'unmatched_transactions' not in st.session_state:
                    st.session_state.unmatched_transactions = []
                if 'transactions_to_create' not in st.session_state:
                    st.session_state.transactions_to_create = []
                
                # Step 1: File Upload
                uploaded_file = st.file_uploader(
                    "Choose a CSV or Excel file",
                    type=['csv', 'xlsx', 'xls'],
                    help="Upload your commission statement file"
                )
                
                if uploaded_file is not None:
                    # Parse file
                    try:
                        if uploaded_file.name.endswith('.csv'):
                            df = pd.read_csv(uploaded_file)
                        else:
                            df = pd.read_excel(uploaded_file)
                        
                        st.session_state.import_data = df
                        st.success(f"✅ Loaded {len(df)} rows from {uploaded_file.name}")
                        
                        # Show preview
                        with st.expander("📊 File Preview", expanded=True):
                            st.dataframe(df.head(10), use_container_width=True)
                        
                        # Step 2: Column Mapping
                        st.divider()
                        st.markdown("### 🔗 Map Statement Columns to System Fields")
                        st.info(f"📅 Statement Date will be set to: {statement_date.strftime('%m/%d/%Y')}")
                        
                        # Required fields
                        required_fields = {
                            'Customer': 'Customer/Client Name',
                            'Policy Number': 'Policy Number',
                            'Effective Date': 'Policy Effective Date',
                            'Agent Paid Amount (STMT)': 'Agent Payment Amount (Required)',
                            'Agency Comm Received (STMT)': 'Agency Commission (for Audit)'
                        }
                        
                        # Optional fields
                        optional_fields = {
                            'Policy Type': 'Policy Type',
                            'Transaction Type': 'Transaction Type (NEW/RWL/END/CXL)',
                            'Premium Sold': 'Premium Amount',
                            'X-DATE': 'Expiration/Cancellation Date',
                            'NOTES': 'Notes/Description'
                        }
                        
                        # Create mapping interface
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown("**Required Fields**")
                            for sys_field, description in required_fields.items():
                                selected_col = st.selectbox(
                                    f"{sys_field} ({description})",
                                    options=[''] + list(df.columns),
                                    key=f"map_{sys_field}",
                                    help=f"Select the column that contains {description}"
                                )
                                if selected_col:
                                    st.session_state.column_mapping[sys_field] = selected_col
                        
                        with col2:
                            st.markdown("**Optional Fields**")
                            for sys_field, description in optional_fields.items():
                                selected_col = st.selectbox(
                                    f"{sys_field} ({description})",
                                    options=[''] + list(df.columns),
                                    key=f"map_{sys_field}",
                                    help=f"Select the column that contains {description}"
                                )
                                if selected_col:
                                    st.session_state.column_mapping[sys_field] = selected_col
                        
                        # Column Mapping Management
                        st.divider()
                        st.markdown("### 💾 Save/Load Column Mappings")
                        
                        # Initialize saved mappings in session state if not exists
                        if 'saved_column_mappings' not in st.session_state:
                            st.session_state.saved_column_mappings = load_saved_column_mappings()
                        
                        col1, col2, col3 = st.columns([2, 1, 1])
                        
                        with col1:
                            mapping_name = st.text_input(
                                "Mapping Name",
                                placeholder="e.g., ABC Insurance Statement",
                                help="Give this mapping a name to save it for future use"
                            )
                        
                        with col2:
                            # Save current mapping
                            if st.button("💾 Save Mapping", type="secondary", disabled=not mapping_name):
                                if st.session_state.column_mapping:
                                    st.session_state.saved_column_mappings[mapping_name] = {
                                        'mapping': st.session_state.column_mapping.copy(),
                                        'created': datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                                        'field_count': len(st.session_state.column_mapping)
                                    }
                                    # Save to file for persistence
                                    if save_column_mappings_to_file(st.session_state.saved_column_mappings):
                                        st.success(f"✅ Saved mapping: {mapping_name}")
                                    else:
                                        st.error("Failed to save mapping to file")
                                else:
                                    st.warning("No mapping to save")
                        
                        with col3:
                            # Load saved mapping
                            if st.session_state.saved_column_mappings:
                                selected_mapping = st.selectbox(
                                    "Load Saved Mapping",
                                    options=[''] + list(st.session_state.saved_column_mappings.keys()),
                                    format_func=lambda x: 'Select a mapping...' if x == '' else x
                                )
                                
                                if selected_mapping and st.button("📂 Load", type="secondary"):
                                    saved_map = st.session_state.saved_column_mappings[selected_mapping]['mapping']
                                    # Verify columns exist in current file
                                    valid_mapping = {}
                                    missing_cols = []
                                    
                                    for sys_field, stmt_col in saved_map.items():
                                        if stmt_col in df.columns:
                                            valid_mapping[sys_field] = stmt_col
                                        else:
                                            missing_cols.append(stmt_col)
                                    
                                    st.session_state.column_mapping = valid_mapping
                                    
                                    if missing_cols:
                                        st.warning(f"Some columns not found in current file: {', '.join(missing_cols)}")
                                    else:
                                        st.success(f"✅ Loaded mapping: {selected_mapping}")
                                    st.rerun()
                        
                        # Show saved mappings
                        if st.session_state.saved_column_mappings:
                            with st.expander("📋 Saved Mappings", expanded=False):
                                for name, info in st.session_state.saved_column_mappings.items():
                                    col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
                                    with col1:
                                        st.text(name)
                                    with col2:
                                        st.text(f"Created: {info['created']}")
                                    with col3:
                                        st.text(f"{info['field_count']} fields")
                                    with col4:
                                        if st.button("🗑️", key=f"delete_{name}", help=f"Delete {name}"):
                                            del st.session_state.saved_column_mappings[name]
                                            save_column_mappings_to_file(st.session_state.saved_column_mappings)
                                            st.rerun()
                        
                        # Check if all required fields are mapped
                        required_mapped = all(
                            field in st.session_state.column_mapping and st.session_state.column_mapping[field]
                            for field in required_fields.keys()
                        )
                        
                        if required_mapped:
                            st.divider()
                            
                            # Process and match transactions
                            if st.button("🔍 Process & Match Transactions", type="primary"):
                                with st.spinner("Matching transactions..."):
                                    # Calculate statement total from all rows (including totals row)
                                    # This gives us the check-and-balance figure
                                    statement_total_amount = 0
                                    if 'Agent Paid Amount (STMT)' in st.session_state.column_mapping:
                                        agent_col = st.session_state.column_mapping['Agent Paid Amount (STMT)']
                                        try:
                                            # Sum all rows including any totals
                                            all_amounts = pd.to_numeric(df[agent_col], errors='coerce').fillna(0)
                                            # Check if there's a totals row (last row or row with 'total' in customer name)
                                            customer_col = st.session_state.column_mapping.get('Customer', '')
                                            if customer_col:
                                                # Find rows that look like totals
                                                totals_mask = df[customer_col].astype(str).str.lower().str.contains('total|sum', na=False)
                                                if totals_mask.any():
                                                    # Use the totals row value
                                                    statement_total_amount = all_amounts[totals_mask].max()
                                                else:
                                                    # No totals row found, sum all non-total rows
                                                    statement_total_amount = all_amounts.sum()
                                            else:
                                                statement_total_amount = all_amounts.sum()
                                        except:
                                            statement_total_amount = 0
                                    
                                    st.session_state.statement_file_total = statement_total_amount
                                    
                                    matched, unmatched, to_create = match_statement_transactions(
                                        df, 
                                        st.session_state.column_mapping,
                                        all_data,
                                        statement_date
                                    )
                                    
                                    st.session_state.matched_transactions = matched
                                    st.session_state.unmatched_transactions = unmatched
                                    st.session_state.transactions_to_create = to_create
                                    
                                    # Show summary
                                    st.success("✅ Matching complete!")
                                    
                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        st.metric("Matched", len(matched))
                                    with col2:
                                        st.metric("Unmatched", len(unmatched))
                                    with col3:
                                        st.metric("Can Create", len(to_create))
                            
                            # Show match results
                            if st.session_state.matched_transactions or st.session_state.unmatched_transactions or st.session_state.transactions_to_create:
                                show_import_results(statement_date)
                        
                        else:
                            st.warning("⚠️ Please map all required fields before proceeding")
                            
                    except Exception as e:
                        st.error(f"Error reading file: {str(e)}")
                        st.exception(e)
        
        with rec_tab2:
            st.subheader("📋 Transactions with Outstanding Balance")
            
            # Get all transactions with outstanding balances using shared function
            if not all_data.empty:
                outstanding = calculate_transaction_balances(all_data)
                
                if not outstanding.empty:
                    # Summary metrics
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Transactions with Balance", len(outstanding))
                    
                    with col2:
                        total_outstanding = outstanding['_balance'].sum()
                        st.metric("Total Outstanding", f"${total_outstanding:,.2f}")
                    
                    with col3:
                        avg_balance = outstanding['_balance'].mean()
                        st.metric("Average Balance", f"${avg_balance:,.2f}")
                    
                    # Customer filter
                    st.divider()
                    
                    customer_filter = st.selectbox(
                        "Filter by Customer",
                        options=['All'] + sorted(outstanding['Customer'].dropna().unique().tolist()),
                        key="outstanding_customer_filter"
                    )
                    
                    if customer_filter != 'All':
                        display_df = outstanding[outstanding['Customer'] == customer_filter]
                    else:
                        display_df = outstanding
                
                    # Display the data with balance
                    display_df['Outstanding Balance'] = display_df['_balance']
                    st.dataframe(
                        display_df[[
                            'Transaction ID', 'Customer', 'Policy Number', 
                            'Effective Date', 'Agent Estimated Comm $', 
                            'Outstanding Balance'
                        ]],
                        column_config={
                            "Agent Estimated Comm $": st.column_config.NumberColumn(format="$%.2f"),
                            "Outstanding Balance": st.column_config.NumberColumn(format="$%.2f")
                        },
                        use_container_width=True,
                        hide_index=True
                    )
                else:
                    st.success("✅ All transactions have been fully reconciled!")
            else:
                st.warning("No data available")
        
        with rec_tab3:
            st.subheader("📜 Reconciliation History")
            
            # Show reconciliation entries
            if not all_data.empty:
                recon_entries = all_data[
                    all_data['Transaction ID'].str.contains('-STMT-', na=False)
                ]
                
                if not recon_entries.empty:
                    # Date range filter
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        start_date = st.date_input(
                            "From Date",
                            value=datetime.date.today() - datetime.timedelta(days=30),
                            format="MM/DD/YYYY"
                        )
                    
                    with col2:
                        end_date = st.date_input(
                            "To Date",
                            value=datetime.date.today(),
                            format="MM/DD/YYYY"
                        )
                    
                    # Filter by date range
                    if 'STMT DATE' in recon_entries.columns:
                        recon_entries['STMT DATE'] = pd.to_datetime(recon_entries['STMT DATE'])
                        mask = (recon_entries['STMT DATE'].dt.date >= start_date) & (recon_entries['STMT DATE'].dt.date <= end_date)
                        filtered_recon = recon_entries[mask]
                    else:
                        filtered_recon = recon_entries
                    
                    if not filtered_recon.empty:
                        # Summary metrics
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            st.metric("Total Entries", len(filtered_recon))
                        with col2:
                            # Use Agent Paid Amount as primary reconciliation field
                            if 'Agent Paid Amount (STMT)' in filtered_recon.columns:
                                total_agent_paid = filtered_recon['Agent Paid Amount (STMT)'].sum()
                            else:
                                # Fallback to Agency amount for backward compatibility
                                total_agent_paid = filtered_recon['Agency Comm Received (STMT)'].sum()
                            st.metric("Total Agent Payments", f"${total_agent_paid:,.2f}")
                        with col3:
                            # Count unique batch IDs
                            unique_batches = filtered_recon['reconciliation_id'].nunique() if 'reconciliation_id' in filtered_recon.columns else 0
                            st.metric("Reconciliation Batches", unique_batches)
                        with col4:
                            if 'Agent Paid Amount (STMT)' in filtered_recon.columns:
                                avg_per_batch = filtered_recon['Agent Paid Amount (STMT)'].sum() / unique_batches if unique_batches > 0 else 0
                            else:
                                avg_per_batch = filtered_recon['Agency Comm Received (STMT)'].sum() / unique_batches if unique_batches > 0 else 0
                            st.metric("Avg per Batch", f"${avg_per_batch:,.2f}")
                        
                        # Group by reconciliation batch
                        st.divider()
                        
                        view_mode = st.radio(
                            "View Mode",
                            ["By Batch", "All Transactions"],
                            horizontal=True,
                            key="history_view_mode"
                        )
                        
                        if view_mode == "By Batch" and 'reconciliation_id' in filtered_recon.columns:
                            # Show batch summary
                            # Use Agent Paid Amount for batch totals
                            agg_dict = {
                                'Transaction ID': 'count',
                                'STMT DATE': 'first'
                            }
                            if 'Agent Paid Amount (STMT)' in filtered_recon.columns:
                                agg_dict['Agent Paid Amount (STMT)'] = 'sum'
                            else:
                                agg_dict['Agency Comm Received (STMT)'] = 'sum'
                            
                            batch_summary = filtered_recon.groupby('reconciliation_id').agg(agg_dict).reset_index()
                            
                            rename_dict = {
                                'reconciliation_id': 'Batch ID',
                                'Transaction ID': 'Transaction Count',
                                'STMT DATE': 'Statement Date'
                            }
                            if 'Agent Paid Amount (STMT)' in batch_summary.columns:
                                rename_dict['Agent Paid Amount (STMT)'] = 'Agent Payment Total'
                            else:
                                rename_dict['Agency Comm Received (STMT)'] = 'Agent Payment Total'
                            
                            batch_summary = batch_summary.rename(columns=rename_dict)
                            
                            # Format date
                            batch_summary['Statement Date'] = pd.to_datetime(batch_summary['Statement Date']).dt.strftime('%m/%d/%Y')
                            
                            # Display batch summary
                            selected_batch = st.selectbox(
                                "Select Batch to View Details",
                                options=[''] + batch_summary['Batch ID'].tolist(),
                                format_func=lambda x: 'Select a batch...' if x == '' else f"{x} ({batch_summary[batch_summary['Batch ID']==x]['Statement Date'].iloc[0] if x != '' else ''})"
                            )
                            
                            st.dataframe(
                                batch_summary.sort_values('Statement Date', ascending=False),
                                column_config={
                                    "Agent Payment Total": st.column_config.NumberColumn(format="$%.2f")
                                },
                                use_container_width=True,
                                hide_index=True
                            )
                            
                            # Show batch details if selected
                            if selected_batch:
                                st.divider()
                                st.subheader(f"Batch Details: {selected_batch}")
                                
                                batch_details = filtered_recon[filtered_recon['reconciliation_id'] == selected_batch]
                                
                                # Format dates
                                if 'STMT DATE' in batch_details.columns:
                                    batch_details['STMT DATE'] = pd.to_datetime(batch_details['STMT DATE']).dt.strftime('%m/%d/%Y')
                                
                                st.dataframe(
                                    batch_details[[
                                        'Transaction ID', 'Customer', 'Policy Number', 
                                        'STMT DATE', 'Agency Comm Received (STMT)', 'Agent Paid Amount (STMT)'
                                    ]],
                                    column_config={
                                        "Agency Comm Received (STMT)": st.column_config.NumberColumn(format="$%.2f"),
                                        "Agent Paid Amount (STMT)": st.column_config.NumberColumn(format="$%.2f")
                                    },
                                    use_container_width=True,
                                    hide_index=True
                                )
                        else:
                            # Show all transactions
                            # Format dates
                            if 'STMT DATE' in filtered_recon.columns:
                                display_recon = filtered_recon.copy()
                                display_recon['STMT DATE'] = pd.to_datetime(display_recon['STMT DATE']).dt.strftime('%m/%d/%Y')
                            else:
                                display_recon = filtered_recon
                            
                            # Include both agent and agency amounts
                            display_columns = ['Transaction ID', 'Customer', 'Policy Number', 'STMT DATE']
                            if 'Agent Paid Amount (STMT)' in display_recon.columns:
                                display_columns.append('Agent Paid Amount (STMT)')
                            if 'Agency Comm Received (STMT)' in display_recon.columns:
                                display_columns.append('Agency Comm Received (STMT)')
                            
                            st.dataframe(
                                display_recon[display_columns].sort_values('STMT DATE', ascending=False),
                                column_config={
                                    "Agent Paid Amount (STMT)": st.column_config.NumberColumn(format="$%.2f"),
                                    "Agency Comm Received (STMT)": st.column_config.NumberColumn(format="$%.2f")
                                },
                                use_container_width=True,
                                hide_index=True
                            )
                    else:
                        st.info("No reconciliations found in the selected date range")
                else:
                    st.info("No reconciliation history found")
            else:
                st.warning("No data available")
        
        with rec_tab4:
            st.subheader("🔧 Adjustments & Voids")
            
            action = st.radio(
                "Select Action",
                ["Create Adjustment", "Void Reconciliation"],
                horizontal=True
            )
            
            if action == "Create Adjustment":
                st.info("Adjustments allow you to correct reconciliation errors without modifying original records")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Find transaction to adjust
                    trans_id = st.text_input(
                        "Transaction ID to Adjust",
                        placeholder="Enter the original transaction ID"
                    )
                    
                    adjustment_amount = st.number_input(
                        "Adjustment Amount",
                        help="Positive to increase, negative to decrease",
                        step=0.01
                    )
                
                with col2:
                    adjustment_reason = st.text_area(
                        "Reason for Adjustment",
                        placeholder="Explain why this adjustment is needed"
                    )
                
                if st.button("Create Adjustment", type="primary"):
                    if trans_id and adjustment_amount != 0 and adjustment_reason:
                        try:
                            # Verify the transaction exists
                            original_trans = all_data[all_data['Transaction ID'] == trans_id]
                            
                            if original_trans.empty:
                                st.error(f"Transaction ID {trans_id} not found")
                            else:
                                # Generate adjustment ID
                                adj_id = generate_reconciliation_transaction_id("ADJ")
                                adj_date = datetime.datetime.now()
                                
                                # Get original transaction details
                                orig_row = original_trans.iloc[0]
                                
                                # Create adjustment entry
                                adjustment_entry = {
                                    'Transaction ID': adj_id,
                                    'Customer': orig_row['Customer'],
                                    'Policy Number': orig_row.get('Policy Number', ''),
                                    'Policy Type': orig_row.get('Policy Type', ''),
                                    'Effective Date': orig_row.get('Effective Date', ''),
                                    'Transaction Type': 'ADJUSTMENT',
                                    'Premium Sold': 0,
                                    'Policy Gross Comm %': 0,
                                    'Agent Estimated Comm $': adjustment_amount,  # The adjustment amount
                                    'Agency Estimated Comm/Revenue (CRM)': 0,
                                    'Agency Comm Received (STMT)': 0,
                                    'Agent Paid Amount (STMT)': 0,
                                    'STMT DATE': adj_date.date(),
                                    'NEW BIZ CHECKLIST COMPLETE': orig_row.get('NEW BIZ CHECKLIST COMPLETE', ''),
                                    'NOTES': f"ADJUSTMENT for {trans_id}: {adjustment_reason}",
                                    'reconciliation_status': 'adjustment',
                                    'reconciliation_id': f"ADJ-{trans_id}",
                                    'is_reconciliation_entry': True,
                                    'Client ID': orig_row.get('Client ID', ''),
                                    'FULL OR MONTHLY PMTS': orig_row.get('FULL OR MONTHLY PMTS', ''),
                                    'X-DATE': orig_row.get('X-DATE', ''),
                                    'Description': f"Adjustment: {adjustment_reason}"
                                }
                                
                                # Save to database
                                result = supabase.table('policies').insert(adjustment_entry).execute()
                                
                                if result.data:
                                    st.success(f"✅ Adjustment {adj_id} created successfully!")
                                    st.info(f"Adjustment of ${adjustment_amount:,.2f} applied to transaction {trans_id}")
                                    
                                    # Clear the form
                                    if 'trans_id_input' in st.session_state:
                                        del st.session_state['trans_id_input']
                                    
                                    # Refresh data
                                    st.cache_data.clear()
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error("Failed to create adjustment entry")
                                    
                        except Exception as e:
                            st.error(f"Error creating adjustment: {str(e)}")
                            traceback.print_exc()
                    else:
                        st.error("Please fill in all fields")
            
            else:  # Void Reconciliation
                st.warning("⚠️ Voiding a reconciliation will reverse the entire statement batch")
                
                # Show recent reconciliation batches
                if not all_data.empty and 'reconciliation_id' in all_data.columns:
                    # Get unique reconciliation batches
                    reconciliation_entries = all_data[
                        (all_data['Transaction ID'].str.contains('-STMT-', na=False)) &
                        (all_data['reconciliation_id'].notna())
                    ]
                    
                    if not reconciliation_entries.empty:
                        # Group by batch to show batch summary
                        batch_summary = reconciliation_entries.groupby('reconciliation_id').agg({
                            'Transaction ID': 'count',
                            'Agency Comm Received (STMT)': 'sum',
                            'STMT DATE': 'first',
                            'reconciled_at': 'first'
                        }).reset_index()
                        
                        # Sort by date descending
                        batch_summary = batch_summary.sort_values('reconciled_at', ascending=False).head(20)
                        
                        st.subheader("Recent Reconciliation Batches")
                        
                        # Select batch to void
                        batch_options = batch_summary['reconciliation_id'].tolist()
                        
                        if batch_options:
                            selected_batch = st.selectbox(
                                "Select Batch to Void",
                                options=[''] + batch_options,
                                format_func=lambda x: 'Select a batch...' if x == '' else f"{x} - ${batch_summary[batch_summary['reconciliation_id']==x]['Agency Comm Received (STMT)'].iloc[0]:,.2f} ({batch_summary[batch_summary['reconciliation_id']==x]['Transaction ID'].iloc[0]} transactions)"
                            )
                            
                            if selected_batch:
                                # Show batch details
                                batch_details = reconciliation_entries[reconciliation_entries['reconciliation_id'] == selected_batch]
                                
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    st.metric("Transactions in Batch", len(batch_details))
                                with col2:
                                    batch_total = batch_details['Agency Comm Received (STMT)'].sum()
                                    st.metric("Batch Total", f"${batch_total:,.2f}")
                                with col3:
                                    stmt_date = pd.to_datetime(batch_details['STMT DATE'].iloc[0]).strftime('%m/%d/%Y')
                                    st.metric("Statement Date", stmt_date)
                                
                                # Show transactions that will be voided
                                with st.expander("View Transactions to be Voided", expanded=True):
                                    display_cols = ['Transaction ID', 'Customer', 'Policy Number', 'Agency Comm Received (STMT)']
                                    st.dataframe(
                                        batch_details[display_cols],
                                        column_config={
                                            "Agency Comm Received (STMT)": st.column_config.NumberColumn(format="$%.2f")
                                        },
                                        use_container_width=True,
                                        hide_index=True
                                    )
                                
                                # Void confirmation
                                st.divider()
                                
                                void_reason = st.text_area(
                                    "Reason for Voiding (Required)",
                                    placeholder="Explain why this batch needs to be voided",
                                    key="void_reason"
                                )
                                
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    confirm_void = st.checkbox(
                                        "I understand this will reverse all transactions in this batch",
                                        key="confirm_void"
                                    )
                                
                                with col2:
                                    if confirm_void and void_reason:
                                        if st.button("🗑️ Void Batch", type="secondary"):
                                            try:
                                                void_count = 0
                                                void_date = datetime.datetime.now()
                                                
                                                # Create void entries for each transaction in the batch
                                                for idx, row in batch_details.iterrows():
                                                    # Generate void transaction ID
                                                    void_id = generate_reconciliation_transaction_id("VOID", void_date.date())
                                                    
                                                    # Create void entry (negative amounts)
                                                    void_entry = {
                                                        'Transaction ID': void_id,
                                                        'Client ID': row.get('Client ID', ''),
                                                        'Customer': row.get('Customer', ''),
                                                        'Carrier Name': row.get('Carrier Name', ''),
                                                        'Policy Type': row.get('Policy Type', ''),
                                                        'Policy Number': row.get('Policy Number', ''),
                                                        'Transaction Type': row.get('Transaction Type', ''),
                                                        'Effective Date': row.get('Effective Date', ''),
                                                        'X-DATE': row.get('X-DATE', ''),
                                                        'Premium Sold': 0,
                                                        'Policy Gross Comm %': 0,
                                                        'Agency Estimated Comm/Revenue (CRM)': 0,
                                                        'Agency Comm Received (STMT)': -float(row.get('Agency Comm Received (STMT)', 0)),  # Negative
                                                        'Agent Estimated Comm $': 0,
                                                        'Agent Paid Amount (STMT)': -float(row.get('Agent Paid Amount (STMT)', 0)),  # Negative
                                                        'STMT DATE': void_date.strftime('%Y-%m-%d'),
                                                        'reconciliation_status': 'void',
                                                        'reconciliation_id': f"VOID-{selected_batch}",
                                                        'reconciled_at': void_date.isoformat(),
                                                        'is_reconciliation_entry': True,
                                                        'NOTES': f"VOID: {void_reason}"
                                                    }
                                                    
                                                    # Insert void entry
                                                    supabase.table('policies').insert(void_entry).execute()
                                                    void_count += 1
                                                
                                                # Update original transactions to mark as unreconciled again
                                                # Find original transactions that were reconciled in this batch
                                                original_trans = all_data[
                                                    (all_data['reconciliation_id'] == selected_batch) &
                                                    (~all_data['Transaction ID'].str.contains('-STMT-|-ADJ-|-VOID-', na=False))
                                                ]
                                                
                                                for idx, orig in original_trans.iterrows():
                                                    supabase.table('policies').update({
                                                        'reconciliation_status': 'unreconciled',
                                                        'reconciliation_id': None,
                                                        'reconciled_at': None
                                                    }).eq('_id', orig['_id']).execute()
                                                
                                                # Log the void in reconciliations table
                                                void_log = {
                                                    'reconciliation_date': void_date.date().isoformat(),
                                                    'statement_date': batch_details['STMT DATE'].iloc[0],
                                                    'carrier_name': 'VOID',
                                                    'total_amount': -batch_total,
                                                    'transaction_count': void_count,
                                                    'notes': f"VOIDED Batch {selected_batch}: {void_reason}"
                                                }
                                                supabase.table('reconciliations').insert(void_log).execute()
                                                
                                                st.success(f"✅ Successfully voided batch {selected_batch}")
                                                st.info(f"Created {void_count} void entries")
                                                
                                                # Clear cache and refresh
                                                st.cache_data.clear()
                                                time.sleep(2)
                                                st.rerun()
                                                
                                            except Exception as e:
                                                st.error(f"Error voiding batch: {str(e)}")
                                    else:
                                        if not void_reason:
                                            st.warning("Please provide a reason for voiding")
                                        if not confirm_void:
                                            st.info("Check the confirmation box to proceed")
                    else:
                        st.info("No reconciliation batches found to void")
                else:
                    st.info("No reconciliation data available")
    
    # --- Admin Panel ---
    elif page == "Admin Panel":
        st.title("⚙️ Admin Panel")
        
        tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["Database Info", "Column Mapping", "Data Management", "System Tools", "Deletion History", "Debug Logs", "Formulas & Calculations"])
        
        with tab1:
            st.subheader("Database Information")
            
            # Database stats
            if not all_data.empty:
                st.metric("Total Records", len(all_data))
                st.metric("Total Columns", len(all_data.columns))
                
                st.subheader("Column Information")
                col_info = pd.DataFrame({
                    'Column': all_data.columns,
                    'Data Type': [str(all_data[col].dtype) for col in all_data.columns],
                    'Non-Null Count': [all_data[col].count() for col in all_data.columns],
                    'Null Count': [all_data[col].isnull().sum() for col in all_data.columns]
                })
                st.dataframe(col_info, use_container_width=True)
                
                st.subheader("Sample Data")
                st.dataframe(all_data.head(), use_container_width=True)
            else:
                st.info("No data available")
        
        with tab2:
            st.subheader("Column Mapping Configuration")
            st.info("Column mapping settings are managed in column_mapping_config.py")
            
            if not all_data.empty:
                st.write("**Current Database Columns:**")
                for col in sorted(all_data.columns):
                    st.write(f"• {col}")
        
        with tab3:
            st.subheader("Data Management")
            
            st.warning("⚠️ These operations affect your database. Use with caution!")
            
            # Database backup
            if st.button("📁 Create Database Backup"):
                try:
                    import shutil
                    backup_name = f"commissions_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                    shutil.copy2("commissions.db", backup_name)
                    st.success(f"Database backed up as {backup_name}")
                except Exception as e:
                    st.error(f"Backup failed: {e}")
            
            # Data validation
            if st.button("🔍 Validate Data Integrity"):
                if not all_data.empty:
                    issues = []
                    
                    # Check for duplicates
                    if 'Transaction_ID' in all_data.columns:
                        duplicates = all_data['Transaction_ID'].duplicated().sum()
                        if duplicates > 0:
                            issues.append(f"Found {duplicates} duplicate Transaction IDs")
                    
                    # Check for missing critical data
                    if 'Customer' in all_data.columns:
                        missing_customers = all_data['Customer'].isnull().sum()
                        if missing_customers > 0:
                            issues.append(f"Found {missing_customers} records with missing customer names")
                    
                    if issues:
                        st.warning("Data integrity issues found:")
                        for issue in issues:
                            st.write(f"• {issue}")
                    else:
                        st.success("No data integrity issues found!")
                else:
                    st.info("No data to validate")
        
        with tab4:
            st.subheader("System Tools")
            
            # System information
            st.write("**System Information:**")
            st.write(f"• Python version: {pd.__version__}")
            st.write(f"• Pandas version: {pd.__version__}")
            st.write(f"• Database: Supabase Cloud")
            
            # Clear session state
            if st.button("🔄 Clear Session State"):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.success("Session state cleared!")
                st.rerun()
        
        with tab5:
            st.subheader("🗑️ Deletion History - Last 100 Deleted Policy Transactions")
            st.info("View and restore recently deleted policies. Records are kept for recovery purposes.")
            
            try:
                # Fetch deleted policies from Supabase
                deleted_response = supabase.table('deleted_policies').select("*").order('deleted_at', desc=True).limit(100).execute()
                
                if deleted_response.data:
                    # Extract policy data from JSONB structure
                    deleted_records = []
                    for record in deleted_response.data:
                        policy_info = {
                            'deletion_id': record['deletion_id'],
                            'deleted_at': record['deleted_at'],
                            'transaction_id': record['transaction_id'],
                            'customer_name': record['customer_name']
                        }
                        # Add the policy data fields
                        if 'policy_data' in record and record['policy_data']:
                            policy_info.update(record['policy_data'])
                        deleted_records.append(policy_info)
                    
                    deleted_df = pd.DataFrame(deleted_records)
                    
                    # Convert deleted_at to datetime
                    if 'deleted_at' in deleted_df.columns:
                        deleted_df['deleted_at'] = pd.to_datetime(deleted_df['deleted_at'])
                    
                    # Add a selection column for restoration
                    deleted_df.insert(0, 'Restore', False)
                    
                    # Display the deleted records
                    st.write(f"**Found {len(deleted_df)} deleted records:**")
                    
                    # Show key info at the top
                    edited_deleted = st.data_editor(
                        deleted_df,
                        use_container_width=True,
                        height=400,
                        key="deleted_policies_editor",
                        column_config={
                            "Restore": st.column_config.CheckboxColumn(
                                "Restore",
                                help="Select records to restore",
                                default=False,
                            ),
                            "deleted_at": st.column_config.DatetimeColumn(
                                "Deleted At",
                                format="DD/MM/YYYY HH:mm",
                                timezone="local"
                            )
                        }
                    )
                    
                    # Restore functionality
                    st.divider()
                    col1, col2 = st.columns([2, 3])
                    
                    with col1:
                        if st.button("♻️ Restore Selected Records", type="primary"):
                            # Find rows where Restore checkbox is True
                            selected_to_restore = edited_deleted[edited_deleted['Restore'] == True]
                            
                            if not selected_to_restore.empty:
                                try:
                                    restored_count = 0
                                    for idx, row in selected_to_restore.iterrows():
                                        # Prepare data for restoration (exclude deletion-specific columns)
                                        restore_data = {}
                                        for col in row.index:
                                            # Exclude metadata columns that aren't part of the policies table
                                            if col not in ['Restore', 'deletion_id', 'deleted_at', 'transaction_id', 'customer_name']:
                                                if pd.notna(row[col]):
                                                    value = row[col]
                                                    # Clean numeric values for proper data types
                                                    if isinstance(value, (int, float)):
                                                        # Check if it should be an integer (no decimal part)
                                                        if isinstance(value, float) and value.is_integer():
                                                            restore_data[col] = int(value)
                                                        else:
                                                            restore_data[col] = clean_numeric_value(value)
                                                    else:
                                                        restore_data[col] = value
                                        
                                        # Restore to policies table
                                        supabase.table('policies').insert(restore_data).execute()
                                        
                                        # Remove from deleted_policies table
                                        deletion_id = row['deletion_id']
                                        supabase.table('deleted_policies').delete().eq('deletion_id', deletion_id).execute()
                                        
                                        restored_count += 1
                                    
                                    # Clear cache and show success
                                    clear_policies_cache()
                                    st.success(f"Successfully restored {restored_count} records!")
                                    st.rerun()
                                    
                                except Exception as restore_error:
                                    st.error(f"Error restoring records: {restore_error}")
                            else:
                                st.warning("Please select records to restore using the checkboxes.")
                    
                    with col2:
                        if st.button("🗑️ Permanently Delete Selected", type="secondary"):
                            # Find rows where Restore checkbox is True (using same checkbox for selection)
                            selected_to_delete = edited_deleted[edited_deleted['Restore'] == True]
                            
                            if not selected_to_delete.empty:
                                st.warning(f"⚠️ This will permanently delete {len(selected_to_delete)} records from history!")
                                if st.button("Confirm Permanent Deletion", key="confirm_perm_delete"):
                                    try:
                                        for idx, row in selected_to_delete.iterrows():
                                            deletion_id = row['deletion_id']
                                            supabase.table('deleted_policies').delete().eq('deletion_id', deletion_id).execute()
                                        
                                        st.success(f"Permanently deleted {len(selected_to_delete)} records from history.")
                                        st.rerun()
                                    except Exception as perm_delete_error:
                                        st.error(f"Error permanently deleting records: {perm_delete_error}")
                    
                    # Export deleted records
                    st.divider()
                    st.write("**Export Deletion History:**")
                    csv_data = deleted_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Deletion History CSV",
                        data=csv_data,
                        file_name=f"deletion_history_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                else:
                    st.info("No deleted policies found. Deleted records will appear here for recovery.")
                    
            except Exception as e:
                if "relation \"deleted_policies\" does not exist" in str(e):
                    st.warning("The deleted_policies table doesn't exist yet. Please run the SQL script to create it:")
                    st.code("""
-- Run this in your Supabase SQL editor:
CREATE TABLE IF NOT EXISTS deleted_policies (
    deletion_id SERIAL PRIMARY KEY,
    deleted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    -- Copy all columns from policies table
    _id INTEGER,
    "Client ID" TEXT,
    "Transaction ID" TEXT,
    "Customer" TEXT,
    -- ... (see create_deleted_policies_table.sql for full schema)
);
                    """)
                else:
                    st.error(f"Error loading deletion history: {e}")
        
        with tab6:
            st.subheader("🐛 Debug Logs")
            st.info("This section captures all debug messages, errors, and system events to help diagnose issues.")
            
            # Initialize debug logs if not exists
            if "debug_logs" not in st.session_state:
                st.session_state.debug_logs = []
            
            # Control buttons
            col1, col2, col3 = st.columns([1, 1, 4])
            with col1:
                if st.button("🗑️ Clear Logs"):
                    clear_debug_logs()
                    st.success("Debug logs cleared!")
                    st.rerun()
            
            with col2:
                if st.button("🔄 Refresh"):
                    st.rerun()
            
            with col3:
                st.write(f"Total log entries: {len(st.session_state.debug_logs)}")
            
            # Filter options
            st.markdown("**Filter Options:**")
            col1, col2 = st.columns(2)
            with col1:
                log_level_filter = st.selectbox(
                    "Log Level",
                    ["All", "ERROR", "WARNING", "INFO", "DEBUG"],
                    key="log_level_filter"
                )
            
            with col2:
                search_term = st.text_input("Search in logs", key="log_search")
            
            # Display logs
            if st.session_state.debug_logs:
                # Filter logs
                filtered_logs = st.session_state.debug_logs
                
                if log_level_filter != "All":
                    filtered_logs = [log for log in filtered_logs if log.get("level") == log_level_filter]
                
                if search_term:
                    filtered_logs = [log for log in filtered_logs if search_term.lower() in str(log).lower()]
                
                # Display in reverse order (newest first)
                st.markdown(f"**Showing {len(filtered_logs)} log entries:**")
                
                for log in reversed(filtered_logs[-100:]):  # Show last 100 entries
                    level = log.get("level", "INFO")
                    timestamp = log.get("timestamp", "")
                    message = log.get("message", "")
                    
                    # Color code by level
                    if level == "ERROR":
                        st.error(f"**[{timestamp}]** {message}")
                        if "error_details" in log:
                            with st.expander("Error Details"):
                                st.code(log["error_details"])
                                if "error_attrs" in log:
                                    st.code(log["error_attrs"])
                    elif level == "WARNING":
                        st.warning(f"**[{timestamp}]** {message}")
                    elif level == "DEBUG":
                        st.info(f"**[{timestamp}]** {message}")
                    else:
                        st.write(f"**[{timestamp}]** {message}")
                
                # Export logs
                st.divider()
                if st.button("📥 Export Logs as JSON"):
                    logs_json = json.dumps(st.session_state.debug_logs, indent=2)
                    st.download_button(
                        label="Download Debug Logs",
                        data=logs_json,
                        file_name=f"debug_logs_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                        mime="application/json"
                    )
            else:
                st.info("No debug logs yet. Logs will appear here as you use the application.")
        
        with tab7:
            st.subheader("📊 Formulas & Calculations")
            st.info("This section displays all active formulas and calculations used throughout the application.")
            
            # Formula tabs
            formula_tab1, formula_tab2, formula_tab3, formula_tab4 = st.tabs(["Commission Formulas", "Rate Matrix", "Calculated Fields", "Formula Testing"])
            
            with formula_tab1:
                st.markdown("### 💰 Commission Calculation Formulas")
                
                # Agency Commission Formula
                st.markdown("#### 1. Agency Commission (Estimated)")
                st.code("""
Formula: Premium Sold × Policy Gross Comm %
Example: $1,000 × 10% = $100

Used in: Add New Policy, Edit Policies
                """, language="text")
                
                # Agent Commission Formula
                st.markdown("#### 2. Agent Commission Calculation")
                st.code("""
Base Formula: Agency Commission × Agent Rate

Agent Rates by Transaction Type:
- NEW, NBS, STL, BoR: 50% of Agency Commission
- RWL, REWRITE: 25% of Agency Commission
- END, PCH: 
  - If Policy Origination Date = Effective Date: 50% (New Business)
  - Otherwise: 25% (Renewal)
- CAN, XCL: 0% (No commission on cancellations)

Example (NEW): $100 Agency Comm × 50% = $50 Agent Comm
Example (RWL): $100 Agency Comm × 25% = $25 Agent Comm
                """, language="text")
                
                # Premium Calculator
                st.markdown("#### 3. Premium Calculator (Endorsements)")
                st.code("""
Formula: New Premium - Existing Premium = Additional Premium
Example: $1,200 - $1,000 = $200 Additional Premium

Used in: Add New Policy (for END/PCH transactions)
                """, language="text")
                
                # Balance Due
                st.markdown("#### 4. Balance Due Calculation")
                st.code("""
Formula: Agent Estimated Comm $ - Agent Paid Amount (STMT)
Example: $50 Estimated - $45 Paid = $5 Balance Due

Used in: Reports, Dashboard searches
                """, language="text")
                
            with formula_tab2:
                st.markdown("### 📋 Commission Rate Matrix")
                
                # Create rate matrix dataframe
                rate_data = {
                    "Transaction Type": ["NEW", "NBS", "STL", "BoR", "RWL", "REWRITE", "END (New)", "END (Renewal)", "PCH (New)", "PCH (Renewal)", "CAN", "XCL"],
                    "Full Name": [
                        "New Business", "New Business (Special)", "Still (Continuing)", "Book of Renewals",
                        "Renewal", "Rewrite", "Endorsement (New)", "Endorsement (Renewal)",
                        "Policy Change (New)", "Policy Change (Renewal)", "Cancellation", "Excluded"
                    ],
                    "Agent Rate": ["50%", "50%", "50%", "50%", "25%", "25%", "50%", "25%", "50%", "25%", "0%", "0%"],
                    "Condition": [
                        "Always", "Always", "Always", "Always", "Always", "Always",
                        "If Orig Date = Eff Date", "If Orig Date ≠ Eff Date",
                        "If Orig Date = Eff Date", "If Orig Date ≠ Eff Date",
                        "No commission", "No commission"
                    ]
                }
                
                rate_df = pd.DataFrame(rate_data)
                st.dataframe(rate_df, use_container_width=True, hide_index=True)
                
                st.markdown("#### Special Rules")
                st.markdown("""
                - **NEW vs RWL Detection**: Based on Policy Origination Date vs Effective Date
                - **Endorsements/Changes**: Commission rate depends on whether it's on a new or renewal policy
                - **Cancellations**: No agent commission paid on CAN or XCL transactions
                - **Override Capability**: Admin users can manually adjust commission amounts if needed
                """)
                
            with formula_tab3:
                st.markdown("### 🔢 Calculated Fields Reference")
                
                calculated_fields = {
                    "Field Name": [
                        "Agency Estimated Comm/Revenue (CRM)",
                        "Agent Estimated Comm $",
                        "Balance Due",
                        "Commission Difference",
                        "Year-to-Date Totals",
                        "Monthly Summaries"
                    ],
                    "Calculation": [
                        "Premium Sold × Policy Gross Comm %",
                        "Agency Commission × Agent Rate (varies by type)",
                        "Agent Estimated - Agent Paid",
                        "Agency Estimated - Agency Received",
                        "SUM of commissions for current year",
                        "SUM grouped by month"
                    ],
                    "Update Frequency": [
                        "On data entry",
                        "On data entry",
                        "Real-time in reports",
                        "Real-time in reports",
                        "On report generation",
                        "On report generation"
                    ],
                    "Used In": [
                        "All policy views",
                        "All policy views",
                        "Reports, Search filters",
                        "Reconciliation reports",
                        "Dashboard, Reports",
                        "Monthly reports"
                    ]
                }
                
                calc_df = pd.DataFrame(calculated_fields)
                st.dataframe(calc_df, use_container_width=True, hide_index=True)
                
            with formula_tab4:
                st.markdown("### 🧪 Formula Testing & Verification")
                
                st.markdown("#### Test Commission Calculations")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    test_premium = st.number_input("Test Premium Amount", value=1000.0, format="%.2f", key="test_premium")
                    test_comm_rate = st.number_input("Commission Rate (%)", value=10.0, format="%.2f", key="test_comm_rate")
                    test_trans_type = st.selectbox("Transaction Type", 
                        ["NEW", "RWL", "END", "PCH", "CAN", "NBS", "STL", "BoR", "REWRITE", "XCL"],
                        key="test_trans_type"
                    )
                    
                    if test_trans_type in ["END", "PCH"]:
                        st.markdown("**Date Check for END/PCH**")
                        col1a, col1b = st.columns(2)
                        with col1a:
                            orig_date = st.date_input("Origination Date", key="test_orig_date")
                        with col1b:
                            eff_date = st.date_input("Effective Date", key="test_eff_date")
                        is_new = orig_date == eff_date
                    else:
                        is_new = None
                
                with col2:
                    st.markdown("**Calculated Results:**")
                    
                    # Calculate agency commission
                    agency_comm = test_premium * (test_comm_rate / 100)
                    st.metric("Agency Commission", f"${agency_comm:.2f}")
                    
                    # Calculate agent commission based on type
                    if test_trans_type in ["NEW", "NBS", "STL", "BoR"]:
                        agent_rate = 0.50
                        rate_display = "50%"
                    elif test_trans_type in ["RWL", "REWRITE"]:
                        agent_rate = 0.25
                        rate_display = "25%"
                    elif test_trans_type in ["END", "PCH"]:
                        if is_new:
                            agent_rate = 0.50
                            rate_display = "50% (New)"
                        else:
                            agent_rate = 0.25
                            rate_display = "25% (Renewal)"
                    else:  # CAN, XCL
                        agent_rate = 0.0
                        rate_display = "0%"
                    
                    agent_comm = agency_comm * agent_rate
                    
                    st.metric("Agent Rate", rate_display)
                    st.metric("Agent Commission", f"${agent_comm:.2f}")
                    
                    # Show calculation breakdown
                    st.markdown("**Calculation Breakdown:**")
                    st.text(f"""
Premium: ${test_premium:.2f}
Commission Rate: {test_comm_rate}%
Agency Commission: ${test_premium:.2f} × {test_comm_rate}% = ${agency_comm:.2f}
Agent Rate: {rate_display}
Agent Commission: ${agency_comm:.2f} × {agent_rate:.0%} = ${agent_comm:.2f}
                    """)
    
    # --- Tools ---
    elif page == "Tools":
        st.title("🛠️ Tools")
        
        tab1, tab2, tab3 = st.tabs(["Data Tools", "Utility Functions", "Import/Export"])
        
        with tab1:
            st.subheader("Data Tools")
            
            # Commission calculator
            st.write("**Commission Calculator**")
            calc_col1, calc_col2 = st.columns(2)
            
            with calc_col1:
                premium_amount = st.number_input("Premium Amount", value=0.0, format="%.2f")
                commission_rate = st.number_input("Commission Rate (%)", value=10.0, format="%.2f")
            
            with calc_col2:
                calculated_commission = premium_amount * (commission_rate / 100)
                st.metric("Calculated Commission", f"${calculated_commission:.2f}")
            
            st.divider()
            
            # ID Generators
            st.write("**ID Generators**")
            gen_col1, gen_col2 = st.columns(2)
            
            with gen_col1:
                if st.button("Generate Client ID"):
                    new_client_id = generate_client_id()
                    st.code(new_client_id)
            
            with gen_col2:
                if st.button("Generate Transaction ID"):
                    new_transaction_id = generate_transaction_id()
                    st.code(new_transaction_id)
        
        with tab2:
            st.subheader("Utility Functions")
            
            # Currency formatter
            st.write("**Currency Formatter**")
            currency_input = st.number_input("Amount to format", value=1234.56, format="%.2f")
            formatted_currency = format_currency(currency_input)
            st.write(f"Formatted: {formatted_currency}")
            
            st.divider()
            
            # Date formatter
            st.write("**Date Formatter**")
            date_input = st.date_input("Date to format")
            formatted_date = date_input.strftime('%m/%d/%Y')
            st.write(f"Formatted (MM/DD/YYYY): {formatted_date}")
        
        with tab3:
            st.subheader("Import/Export Tools")
            
            # Export section
            st.write("### Export Data")
            if not all_data.empty:
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("**📄 CSV Export**")
                    csv_data = all_data.to_csv(index=False)
                    st.download_button(
                        label="📥 Export All Data to CSV",
                        data=csv_data,
                        file_name=f"all_policies_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        help="Export all policy data as CSV file"
                    )
                
                with col2:
                    st.write("**📊 Excel Export**")
                    excel_buffer, excel_filename = create_formatted_excel_file(
                        all_data, 
                        sheet_name="All Policies", 
                        filename_prefix="all_policies"
                    )
                    if excel_buffer:
                        st.download_button(
                            label="📥 Export All Data to Excel",
                            data=excel_buffer,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Export all policy data as formatted Excel file"
                        )
            else:
                st.info("No data available to export")
            
            st.divider()
            
            # Import section
            st.write("### Import Data")
            
            # File uploader for both CSV and Excel
            uploaded_file = st.file_uploader(
                "Choose a CSV or Excel file", 
                type=["csv", "xlsx", "xls"],
                help="Upload a CSV or Excel file containing policy data"
            )
            
            if uploaded_file is not None:
                file_type = uploaded_file.name.split('.')[-1].lower()
                
                try:
                    # Process based on file type
                    if file_type == 'csv':
                        import_df = pd.read_csv(uploaded_file)
                        validation_success = True
                        validation_errors = []
                        st.success("✅ CSV file loaded successfully")
                        
                    elif file_type in ['xlsx', 'xls']:
                        validation_success, import_df, validation_errors = validate_excel_import(uploaded_file)
                        
                        if validation_success:
                            st.success("✅ Excel file loaded and validated successfully")
                        else:
                            st.error("❌ Excel file validation failed")
                            for error in validation_errors:
                                st.error(f"• {error}")
                    
                    # Show preview if file was loaded
                    if 'import_df' in locals() and import_df is not None:
                        st.write("**📋 Preview of uploaded data:**")
                        st.dataframe(import_df.head(10), use_container_width=True)
                        
                        # Show file statistics
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Rows", len(import_df))
                        with col2:
                            st.metric("Total Columns", len(import_df.columns))
                        with col3:
                            st.metric("File Type", file_type.upper())
                        
                        # Show validation results for Excel files
                        if file_type in ['xlsx', 'xls'] and validation_errors:
                            with st.expander("⚠️ Validation Issues"):
                                for error in validation_errors:
                                    st.warning(error)
                        
                        # Import button with enhanced validation
                        if validation_success:
                            if st.button("🔄 Import Data to Database", type="primary"):
                                with st.spinner("Validating and importing data..."):
                                    try:
                                        # Additional validation before import
                                        required_columns = ['Client_ID', 'Transaction_ID', 'Policy_Type']
                                        missing_required = [col for col in required_columns if col not in import_df.columns]
                                        
                                        if missing_required:
                                            st.error(f"❌ Missing required columns: {', '.join(missing_required)}")
                                        else:
                                            # Check for duplicate Transaction_IDs in current database
                                            existing_ids = set(all_data['Transaction_ID'].tolist()) if 'Transaction_ID' in all_data.columns else set()
                                            import_ids = set(import_df['Transaction_ID'].tolist()) if 'Transaction_ID' in import_df.columns else set()
                                            duplicates = existing_ids.intersection(import_ids)
                                            
                                            if duplicates:
                                                st.warning(f"⚠️ Found {len(duplicates)} Transaction_IDs that already exist in database")
                                                
                                                import_option = st.radio(
                                                    "How would you like to handle duplicates?",
                                                    ["Skip duplicate records", "Update existing records", "Cancel import"]
                                                )
                                                
                                                if import_option == "Cancel import":
                                                    st.info("Import cancelled")
                                                elif import_option == "Skip duplicate records":
                                                    import_df = import_df[~import_df['Transaction_ID'].isin(duplicates)]
                                                    st.info(f"Will import {len(import_df)} new records (skipping duplicates)")
                                                elif import_option == "Update existing records":
                                                    st.warning("Update functionality not yet implemented")
                                            
                                            if len(import_df) > 0 and import_option != "Cancel import":
                                                # This is where actual database import would happen
                                                st.success(f"✅ Ready to import {len(import_df)} records")
                                                st.info("🚧 Full database import implementation is in development. Preview and validation are complete.")
                                                
                                                # Show what would be imported
                                                with st.expander("📊 Import Summary"):
                                                    st.write(f"**Records to import:** {len(import_df)}")
                                                    st.write(f"**File type:** {file_type.upper()}")
                                                    st.write(f"**Columns found:** {len(import_df.columns)}")
                                                    st.write("**Sample data:**")
                                                    st.dataframe(import_df.head(3))
                                    
                                    except Exception as e:
                                        st.error(f"❌ Import error: {e}")
                        else:
                            st.error("❌ Cannot import data due to validation errors. Please fix the issues and try again.")
                
                except Exception as e:
                    st.error(f"❌ Error reading file: {e}")
                    if file_type in ['xlsx', 'xls']:
                        st.info("💡 Tip: Ensure your Excel file is not corrupted and contains data in the first sheet.")
    
    # --- Accounting ---
    elif page == "Accounting":
        st.subheader("Accounting")
        st.info("This section provides accounting summaries, reconciliation tools, and export options. Use the reconciliation tool below to match your commission statement to your database and mark payments as received.")

        # --- Tables are already created in Supabase ---
        # No need to create tables as they were created during schema setup        # --- Load manual entries from DB if session state is empty ---
        if "manual_commission_rows" not in st.session_state:
            st.session_state["manual_commission_rows"] = []
            
        # Only reload from DB if session state is completely empty (not after deletions)
        if not st.session_state["manual_commission_rows"] and "deletion_performed" not in st.session_state:
            try:
                response = supabase.table('manual_commission_entries').select("*").execute()
                manual_entries_df = pd.DataFrame(response.data) if response.data else pd.DataFrame()
            except Exception as e:
                st.error(f"Error loading manual entries: {e}")
                manual_entries_df = pd.DataFrame()
            if not manual_entries_df.empty:
                # Convert DB rows to dicts for session state
                st.session_state["manual_commission_rows"] = [
                    {
                        "Customer": row["customer"],
                        "Policy Type": row["policy_type"],
                        "Policy Number": row["policy_number"],
                        "Effective Date": row["effective_date"],
                        "Transaction Type": row["transaction_type"],
                        "Commission Paid": row["commission_paid"],
                        "Agency Comm Received (STMT)": row["agency_commission_received"],
                        "Statement Date": row["statement_date"],
                        "Client ID": row.get("client_id", ""),
                        "Transaction ID": row.get("transaction_id", "")
                    }
                    for _, row in manual_entries_df.iterrows()
                ]

        # --- Accounting UI code continues here ---
        st.markdown("## Reconcile Commission Statement")
        entry_mode = st.radio("How would you like to enter your commission statement?", ["Manual Entry", "Upload File"], key="reconcile_entry_mode")
        statement_df = None
        if entry_mode == "Upload File":
            uploaded_statement = st.file_uploader(
                "Upload your commission statement (CSV, Excel, or PDF)",
                type=["csv", "xlsx", "xls", "pdf"],
                key="reconcile_statement_upload"
            )
            if uploaded_statement:
                if uploaded_statement.name.endswith(".csv"):
                    statement_df = pd.read_csv(uploaded_statement)
                elif uploaded_statement.name.endswith(".xlsx") or uploaded_statement.name.endswith(".xls"):
                    statement_df = pd.read_excel(uploaded_statement)
                elif uploaded_statement.name.endswith(".pdf"):
                    with pdfplumber.open(uploaded_statement) as pdf:
                        all_tables = []
                        for page in pdf.pages:
                            tables = page.extract_tables()
                            for table in tables:
                                temp_df = pd.DataFrame(table[1:], columns=table[0])
                                all_tables.append(temp_df)
                        if all_tables:
                            statement_df = pd.concat(all_tables, ignore_index=True)
                        else:
                            st.error("No tables found in PDF.")
                            st.stop()
        elif entry_mode == "Manual Entry":
            st.markdown("### Manually Enter Commission Statement Data")
            if "manual_commission_rows" not in st.session_state:
                st.session_state["manual_commission_rows"] = []
            customers = sorted(all_data["Customer"].dropna().unique().tolist())
            selected_customer = st.selectbox("Select Customer", ["Select..."] + customers, key="recon_customer_select")
            # Initialize variables for policy selections
            policy_types = []
            policy_numbers = []
            effective_dates = []
            selected_policy_type = None
            selected_policy_number = None
            selected_effective_date = None
            client_id = None
            
            if selected_customer and selected_customer != "Select...":
                policy_types = sorted(all_data[all_data["Customer"] == selected_customer]["Policy Type"].dropna().unique().tolist())
                selected_policy_type = st.selectbox("Select Policy Type", ["Select..."] + policy_types, key="recon_policy_type_select")
                if selected_policy_type and selected_policy_type != "Select...":
                    policy_numbers = sorted(all_data[(all_data["Customer"] == selected_customer) & (all_data["Policy Type"] == selected_policy_type)]["Policy Number"].dropna().unique().tolist())
                    selected_policy_number = st.selectbox("Select Policy Number", ["Select..."] + policy_numbers, key="recon_policy_number_select")
                    if selected_policy_number and selected_policy_number != "Select...":
                        effective_dates = sorted(all_data[(all_data["Customer"] == selected_customer) & (all_data["Policy Type"] == selected_policy_type) & (all_data["Policy Number"] == selected_policy_number)]["Effective Date"].dropna().unique().tolist())
                        selected_effective_date = st.selectbox("Select Effective Date", ["Select..."] + effective_dates, key="recon_effective_date_select")
                        
                        # Lookup Client ID using exact match of all selected fields
                        if selected_effective_date and selected_effective_date != "Select...":
                            exact_match = all_data[
                                (all_data["Customer"] == selected_customer) &
                                (all_data["Policy Type"] == selected_policy_type) &
                                (all_data["Policy Number"] == selected_policy_number) &
                                (all_data["Effective Date"] == selected_effective_date)
                            ]
                            if not exact_match.empty and "Client ID" in exact_match.columns:
                                client_id = exact_match.iloc[0]["Client ID"]
                                st.success(f"✅ Client ID found: {client_id}")
                            else:
                                st.warning("⚠️ Client ID not found for this exact combination")
            transaction_types = ["NEW", "NBS", "STL", "BoR", "END", "PCH", "RWL", "REWRITE", "CAN", "XCL"]
            transaction_type = st.selectbox("Transaction Type", transaction_types, key="recon_transaction_type_select")
            # Auto-calculate Agent Comm (New 50% RWL 25%) based on Transaction Type
            def calc_agent_comm_pct(tx_type):
                if tx_type == "NEW":
                    return 50.0
                elif tx_type in ["RWL", "RENEWAL"]:
                    return 25.0
                else:
                    return 0.0
            agent_comm_new_rwl = calc_agent_comm_pct(transaction_type)
            # --- Enter statement date ONCE for all entries in this session ---
            if "recon_stmt_date" not in st.session_state:
                st.session_state["recon_stmt_date"] = None
            st.session_state["recon_stmt_date"] = st.date_input(
                "Commission Statement Date (applies to all entries below)",
                value=st.session_state["recon_stmt_date"] or None,
                format="MM/DD/YYYY",
                key="recon_stmt_date_input_global"
            )
            statement_date = st.session_state["recon_stmt_date"]
            agency_comm_received = st.number_input("Agency Comm Received (STMT)", min_value=0.0, step=0.01, format="%.2f", key="recon_agency_comm_received_input")
            amount_paid = st.number_input("Agent Paid Amount (STMT)", min_value=0.0, step=0.01, format="%.2f", key="recon_amount_paid_input")
            if st.button("Add Entry", key="recon_add_entry_btn"):
                # Use the same transaction ID rules as new policy transaction: 7 chars, uppercase letters and digits
                transaction_id = generate_transaction_id()  # 7 chars, uppercase letters and digits
                # Use the global statement date for all entries
                stmt_date_str = ""
                if statement_date is None or statement_date == "":
                    stmt_date_str = ""
                elif isinstance(statement_date, datetime.date):
                    stmt_date_str = statement_date.strftime("%m/%d/%Y")
                else:
                    try:
                        stmt_date_str = pd.to_datetime(statement_date).strftime("%m/%d/%Y")
                    except Exception:
                        stmt_date_str = ""
                entry = {
                    "Client ID": client_id if client_id else "",
                    "Transaction ID": transaction_id,
                    "Customer": selected_customer,
                    "Policy Type": selected_policy_type if selected_policy_type else "",
                    "Policy Number": selected_policy_number if selected_policy_number else "",
                    "Effective Date": selected_effective_date if selected_effective_date else "",
                    "Transaction Type": transaction_type,
                    "Agent Comm (New 50% RWL 25%)": agent_comm_new_rwl,
                    "Agency Comm Received (STMT)": agency_comm_received,
                    "Agent Paid Amount (STMT)": amount_paid,
                    "Statement Date": stmt_date_str
                }
                # Insert new row into DB (non-destructive, always insert)
                # Insert into Supabase
                entry_data = {
                    "client_id": client_id if client_id else "",
                    "transaction_id": transaction_id,
                    "customer": selected_customer,
                    "policy_type": selected_policy_type if selected_policy_type else "",
                    "policy_number": selected_policy_number if selected_policy_number else "",
                    "effective_date": selected_effective_date if selected_effective_date else "",
                    "transaction_type": transaction_type,
                    "commission_paid": amount_paid,
                    "agency_commission_received": agency_comm_received,
                    "statement_date": stmt_date_str
                }
                try:
                    supabase.table('manual_commission_entries').insert(entry_data).execute()
                except Exception as e:
                    st.error(f"Error saving manual entry: {e}")
                st.session_state["manual_commission_rows"].append(entry)
                st.success("Entry added (non-destructive, unique 7-character Transaction ID). You can review and edit below.")        # Show manual entries for editing/management if any exist        if st.session_state.get("manual_commission_rows"):
            st.markdown("### Manual Commission Entries")
              # Show persistent deletion success message
            if "deletion_success_msg" in st.session_state:
                st.success(st.session_state["deletion_success_msg"])
                del st.session_state["deletion_success_msg"]  # Clear after showing
            
            st.info("Manual entries are non-destructive and saved to database. Use controls below to manage entries.")
              # Display summary of entries
            total_entries = len(st.session_state["manual_commission_rows"])
            total_paid = sum(float(row.get("Agent Paid Amount (STMT)", 0)) for row in st.session_state["manual_commission_rows"])
            total_received = sum(float(row.get("Agency Comm Received (STMT)", 0)) for row in st.session_state["manual_commission_rows"])
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Entries", total_entries)
            with col2:
                st.metric("Total Agency Comm", f"${total_received:,.2f}")
            with col3:
                st.metric("Total Agent Comm Paid", f"${total_paid:,.2f}")
            
            # Show entries in an editable table with selection capability
            df_display = pd.DataFrame(st.session_state["manual_commission_rows"])
            df_display.insert(0, "Select", False)  # Add selection column at the beginning
            
            edited_df = st.data_editor(
                df_display, 
                use_container_width=True, 
                height=max(400, 40 + 40 * len(df_display)),
                column_config={
                    "Select": st.column_config.CheckboxColumn("Select", help="Check to select rows for deletion")
                },
                disabled=[col for col in df_display.columns if col != "Select"],  # Only allow editing the Select column
                key="manual_entries_editor"            )
            
            # Delete selected rows functionality
            selected_indices = edited_df[edited_df["Select"] == True].index.tolist()
            if selected_indices:
                col1, col2 = st.columns([1, 4])
                with col1:
                    if st.button("🗑️ Delete Selected", type="secondary", key="delete_selected_entries"):
                        # Remove selected entries from session state
                        original_count = len(st.session_state["manual_commission_rows"])
                          # Get the transaction IDs of rows to delete BEFORE removing from session state
                        transaction_ids_to_delete = []
                        for idx in selected_indices:
                            if idx < len(st.session_state["manual_commission_rows"]):
                                entry = st.session_state["manual_commission_rows"][idx]
                                transaction_id = entry.get("Transaction ID", "")
                                transaction_ids_to_delete.append(transaction_id)
                        
                        # Remove from session state
                        st.session_state["manual_commission_rows"] = [
                            entry for i, entry in enumerate(st.session_state["manual_commission_rows"]) 
                            if i not in selected_indices
                        ]
                        deleted_count = original_count - len(st.session_state["manual_commission_rows"])
                          # Delete from database using the collected transaction IDs
                        db_deleted_count = 0
                        failed_deletions = []
                        for transaction_id in transaction_ids_to_delete:
                            if transaction_id and transaction_id.strip():  # Only if transaction_id is not empty
                                try:
                                    # Delete from Supabase
                                    result = supabase.table('manual_commission_entries').delete().eq('transaction_id', transaction_id).execute()
                                    if result.data:
                                        db_deleted_count += 1
                                    else:
                                        # Check if record exists
                                        check_result = supabase.table('manual_commission_entries').select("*").eq('transaction_id', transaction_id).execute()
                                        if not check_result.data:
                                            failed_deletions.append(f"No DB row found for ID: {transaction_id}")
                                        else:
                                            db_deleted_count += 1  # Delete was successful even if no data returned
                                except Exception as e:
                                    failed_deletions.append(f"Error deleting {transaction_id}: {str(e)}")
                            else:
                                failed_deletions.append("Empty transaction ID")
                         # Get mapped column names
                        customer_col = get_mapped_column("Customer")
                        policy_type_col = get_mapped_column("Policy Type")
                        policy_number_col = get_mapped_column("Policy Number")
                        effective_date_col = get_mapped_column("Effective Date")
                        transaction_type_col = get_mapped_column("Transaction Type")
                        agent_paid_col = get_mapped_column("Agent Paid Amount (STMT)")
                        agency_comm_col = get_mapped_column("Agency Comm Received (STMT)")
                        client_id_col = get_mapped_column("Client ID")
                        transaction_id_col = get_mapped_column("Transaction ID")
                        
                       
                        # Set flag to prevent automatic reload from database
                        st.session_state["deletion_performed"] = True
                          # Create success message
                        success_msg = f"Deleted {deleted_count} entries from session and {db_deleted_count} from database."
                        if failed_deletions:
                            success_msg += f" Issues: {'; '.join(failed_deletions[:3])}"  # Show first 3 issues
                        elif db_deleted_count == 0 and transaction_ids_to_delete:
                            success_msg += f" (Transaction IDs attempted: {transaction_ids_to_delete})"
                        
                        st.session_state["deletion_success_msg"] = success_msg
                        st.rerun()
                with col2:
                    st.info(f"📋 {len(selected_indices)} row(s) selected for deletion")
            else:
                st.info("💡 Tip: Check the 'Select' boxes next to entries you want to delete, then click the delete button.")
              # --- Reconcile Button to Save Statement to History ---
            st.markdown("---")
            if st.session_state["manual_commission_rows"]:
                st.markdown("#### Reconcile Commission Statement")
                col1, col2 = st.columns([2, 3])
                with col1:
                    statement_date = st.date_input(
                        "Statement Date",
                        value=datetime.date.today(),
                        format="MM/DD/YYYY",
                        key="reconcile_statement_date_v2"
                    )
                with col2:
                    statement_description = st.text_input(
                        "Statement Description (optional)",
                        placeholder="e.g., Monthly Commission Statement - December 2024",
                        key="reconcile_description"                    )
                  # Calculate totals for the statement
                total_commission_paid = sum(float(row.get("Agent Paid Amount (STMT)", 0)) for row in st.session_state["manual_commission_rows"])
                total_agency_received = sum(float(row.get("Agency Comm Received (STMT)", 0)) for row in st.session_state["manual_commission_rows"])
                
                st.markdown(f"**Statement Summary:** {len(st.session_state['manual_commission_rows'])} entries | Amount Paid: ${total_commission_paid:,.2f} | Agency Received: ${total_agency_received:,.2f}")
                
                if st.button("💾 Reconcile & Save to History", type="primary", key="reconcile_statement_btn"):
                    error_placeholder = st.empty()
                    try:
                        # Debug: Check statement date
                        log_debug(f"Starting reconciliation process with statement date: {statement_date}", "INFO")
                        st.info(f"Using statement date: {statement_date}")
                        if not statement_date:
                            log_debug("No statement date selected", "ERROR")
                            st.error("⚠️ No statement date selected! Please select a date above.")
                            st.stop()
                        
                        # Prepare statement details for JSON storage
                        statement_details = []
                        for row in st.session_state["manual_commission_rows"]:
                            statement_details.append({
                                "Customer": row.get("Customer", ""),
                                "Policy Type": row.get("Policy Type", ""),
                                "Policy Number": row.get("Policy Number", ""),
                                "Effective Date": row.get("Effective Date", ""),
                                "Transaction Type": row.get("Transaction Type", ""),
                                "Agent Paid Amount (STMT)": row.get("Agent Paid Amount (STMT)", 0),
                                "Agency Comm Received (STMT)": row.get("Agency Comm Received (STMT)", 0),
                                "Client ID": row.get("Client ID", ""),
                                "Transaction ID": row.get("Transaction ID", "")
                            })
                        
                        # PHASE 1: Existing audit/history functionality (preserved exactly)
                        # Save to commission_payments table
                        log_debug(f"Phase 1: Saving to commission_payments table. Total paid: {total_commission_paid}, Total received: {total_agency_received}", "INFO")
                        try:
                            # Prepare commission payment data
                            payment_data = {
                                "policy_number": "STMT-" + datetime.datetime.now().strftime('%Y%m%d'),
                                "customer": "Statement Reconciliation",
                                "payment_amount": float(total_commission_paid),  # Ensure it's a float
                                "statement_date": statement_date.strftime('%Y-%m-%d') if statement_date else None,
                                "payment_timestamp": datetime.datetime.now().isoformat(),
                                "statement_details": json.dumps({
                                    "description": statement_description,
                                    "total_commission_paid": float(total_commission_paid),
                                    "total_agency_received": float(total_agency_received),
                                    "entries": statement_details
                                })
                            }
                            # Debug: Show what we're trying to save
                            st.info("Attempting to save payment record...")
                            st.json(payment_data)
                            log_debug(f"Saving payment data: {json.dumps(payment_data, indent=2)}", "DEBUG")
                            
                            result = supabase.table('commission_payments_simple').insert(payment_data).execute()
                            log_debug("Payment record saved successfully to commission_payments_simple", "INFO")
                            st.success("Payment record saved successfully!")
                        except Exception as e:
                            log_debug(f"Error saving payment record: {str(e)}", "ERROR", e)
                            st.error(f"Error saving payment record: {e}")
                            st.error(f"Payment data that failed: {payment_data}")
                            raise  # Re-raise to trigger outer exception handler
                        
                        # Save individual entries to manual_commission_entries table as well
                        st.info("💾 Saving individual entries to manual_commission_entries table...")
                        for row in st.session_state["manual_commission_rows"]:
                            try:
                                entry_data = {
                                    "client_id": row.get("Client ID", ""),
                                    "transaction_id": row.get("Transaction ID", ""),
                                    "customer": row.get("Customer", ""),
                                    "policy_type": row.get("Policy Type", ""),
                                    "policy_number": row.get("Policy Number", ""),
                                    "effective_date": row.get("Effective Date", ""),
                                    "transaction_type": row.get("Transaction Type", ""),
                                    "commission_paid": float(row.get("Agent Paid Amount (STMT)", 0)),
                                    "agency_commission_received": float(row.get("Agency Comm Received (STMT)", 0)),
                                    "statement_date": statement_date.strftime('%Y-%m-%d')
                                }
                                # Check if entry already exists
                                existing = supabase.table('manual_commission_entries').select("*").eq('transaction_id', entry_data['transaction_id']).execute()
                                if existing.data:
                                    # Update existing entry
                                    supabase.table('manual_commission_entries').update(entry_data).eq('transaction_id', entry_data['transaction_id']).execute()
                                else:
                                    # Insert new entry
                                    supabase.table('manual_commission_entries').insert(entry_data).execute()
                            except Exception as e:
                                st.error(f"Error saving manual entry: {e}")
                        
                        # PHASE 2: NEW - Add reconciled transactions to main policies database
                        try:
                            log_debug("Phase 2: Starting to add transactions to policies database", "INFO")
                            st.info("📊 Phase 2: Adding transactions to main policies database...")
                            
                            # Debug: Show what columns are being mapped
                            with st.expander("Debug: Column Mapping", expanded=True):
                                st.write("Column mappings being used:")
                                debug_mappings = {
                                    "Transaction ID": get_mapped_column("Transaction ID") or "Transaction ID",
                                    "Customer": get_mapped_column("Customer") or "Customer",
                                    "Policy Type": get_mapped_column("Policy Type") or "Policy Type",
                                    "Policy Number": get_mapped_column("Policy Number") or "Policy Number",
                                    "STMT DATE": get_mapped_column("STMT DATE") or "STMT DATE"
                                }
                                st.json(debug_mappings)
                                log_debug(f"Column mappings: {json.dumps(debug_mappings, indent=2)}", "DEBUG")
                            
                            new_main_db_transactions = []
                            for idx, row in enumerate(st.session_state["manual_commission_rows"]):
                                # Create new transaction record for main policies table using centralized mapping
                                new_transaction = {}
                                
                                # Use centralized mapping to ensure field consistency
                                transaction_id_col = get_mapped_column("Transaction ID") or "Transaction ID"
                                customer_col = get_mapped_column("Customer") or "Customer"
                                policy_type_col = get_mapped_column("Policy Type") or "Policy Type"
                                policy_number_col = get_mapped_column("Policy Number") or "Policy Number"
                                effective_date_col = get_mapped_column("Effective Date") or "Effective Date"
                                transaction_type_col = get_mapped_column("Transaction Type") or "Transaction Type"
                                stmt_date_col = get_mapped_column("STMT DATE") or "STMT DATE"
                                agency_comm_col = get_mapped_column("Agency Comm Received (STMT)") or "Agency Comm Received (STMT)"
                                agent_paid_col = get_mapped_column("Agent Paid Amount (STMT)") or "Agent Paid Amount (STMT)"
                                client_id_col = get_mapped_column("Client ID") or "Client ID"
                                
                                # Debug: Show what we're building
                                st.write(f"Building transaction {idx + 1}...")
                                log_debug(f"Building transaction {idx + 1} for {row.get('Customer', 'Unknown')}", "DEBUG")
                                
                                # Populate with reconciliation data using flexible transaction types
                                if transaction_id_col:
                                    # Generate new unique transaction ID for main DB to avoid conflicts
                                    new_transaction[transaction_id_col] = generate_transaction_id()
                                    st.write(f"- {transaction_id_col}: {new_transaction[transaction_id_col]}")
                                if customer_col:
                                    new_transaction[customer_col] = row.get("Customer", "")
                                if policy_type_col:
                                    new_transaction[policy_type_col] = row.get("Policy Type", "")
                                if policy_number_col:
                                    new_transaction[policy_number_col] = row.get("Policy Number", "")
                                if effective_date_col:
                                    new_transaction[effective_date_col] = row.get("Effective Date", "")
                                if transaction_type_col:
                                    # PRESERVE USER'S TRANSACTION TYPE FLEXIBILITY - use exactly what they entered
                                    new_transaction[transaction_type_col] = row.get("Transaction Type", "")
                                if stmt_date_col:
                                    new_transaction[stmt_date_col] = statement_date.strftime('%m/%d/%Y')
                                if agency_comm_col:
                                    new_transaction[agency_comm_col] = float(row.get("Agency Comm Received (STMT)", 0))
                                if agent_paid_col:
                                    new_transaction[agent_paid_col] = float(row.get("Agent Paid Amount (STMT)", 0))
                                if client_id_col:
                                    new_transaction[client_id_col] = row.get("Client ID", "")
                                
                                # Add notes field to identify reconciled transactions
                                notes_col = get_mapped_column("NOTES") or "NOTES"
                                if notes_col:
                                    new_transaction[notes_col] = f"Reconciled Statement - {statement_date.strftime('%m/%d/%Y')}"
                                
                                new_main_db_transactions.append(new_transaction)
                                log_debug(f"Transaction {idx + 1} built: {json.dumps(new_transaction, default=str)}", "DEBUG")
                            
                            # Insert new transactions into main policies table
                            main_db_added_count = 0
                            failed_insertions = []
                            if new_main_db_transactions:
                                try:
                                    st.info(f"Attempting to add {len(new_main_db_transactions)} transactions to main database...")
                                    new_df = pd.DataFrame(new_main_db_transactions)
                                    
                                    # Show what we're trying to insert
                                    with st.expander("Debug: Data being added to policies table"):
                                        st.dataframe(new_df)
                                    
                                    # Insert via Supabase in batches
                                    for idx, row in new_df.iterrows():
                                        policy_data = row.to_dict()
                                        # Handle NaN values
                                        for key, value in policy_data.items():
                                            if pd.isna(value):
                                                policy_data[key] = None
                                        try:
                                            log_debug(f"Attempting to insert policy row {idx}: {json.dumps(policy_data, default=str)}", "DEBUG")
                                            result = supabase.table('policies').insert(policy_data).execute()
                                            main_db_added_count += 1
                                            log_debug(f"Successfully inserted policy row {idx}", "INFO")
                                        except Exception as e:
                                            error_msg = f"Row {idx}: {str(e)}"
                                            failed_insertions.append(error_msg)
                                            log_debug(f"Error inserting policy row {idx}: {str(e)}", "ERROR", e)
                                            st.error(f"Error inserting policy: {e}")
                                            # Show the problematic data
                                            st.error("Failed data:")
                                            st.json(policy_data)
                                    
                                    if failed_insertions:
                                        st.error(f"❌ Failed to insert {len(failed_insertions)} records to main database:")
                                        for err in failed_insertions:
                                            st.write(f"- {err}")
                                    
                                    clear_policies_cache()
                                except Exception as e:
                                    st.error(f"⚠️ Reconciliation saved to history, but could not add to main database: {str(e)}")
                                    st.error("Full error details:")
                                    st.code(str(e))
                                    # Make the error persistent
                                    st.stop()
                                
                        except Exception as phase2_error:
                            log_debug(f"Error in Phase 2: {str(phase2_error)}", "ERROR", phase2_error)
                            st.error(f"❌ Error in Phase 2 (Adding to policies table): {str(phase2_error)}")
                            st.error("This error occurred while trying to add transactions to the main database")
                            st.code(str(phase2_error))
                            # Don't clear entries on phase 2 error
                            st.stop()
                        
                        # Only clear entries and show success if we get here without errors
                        # Enhanced success message showing both operations
                        log_debug(f"Reconciliation completed successfully. Added {main_db_added_count} transactions to policies table", "INFO")
                        success_msg = f"✅ Commission statement reconciled and saved to history! Statement date: {statement_date.strftime('%m/%d/%Y')}"
                        if main_db_added_count > 0:
                            success_msg += f"\n💾 Added {main_db_added_count} new policy transactions to main policies database"
                            success_msg += f"\n🔍 View in 'All Policies in Database' or 'Policy Revenue Ledger' pages"
                        st.success(success_msg)
                        
                        # Clear the manual entries ONLY after showing success message
                        st.session_state["manual_commission_rows"] = []
                        st.rerun()
                        
                    except Exception as e:
                        log_debug(f"Unexpected error during reconciliation: {str(e)}", "ERROR", e)
                        # Create a persistent error container
                        with error_placeholder.container():
                            st.error(f"❌ Error saving commission statement: {str(e)}")
                            st.error("Full error details:")
                            st.code(str(e))
                            
                            # Show the statement date that was being used
                            st.error(f"Statement date being used: {statement_date}")
                            
                            # Don't clear the manual entries on error - preserve user's work
                            st.warning("⚠️ Your entries have been preserved. Please try again or contact support.")
                            
                            # Show the data that was attempted to be saved
                            with st.expander("Debug Information - Click to expand", expanded=True):
                                st.write("**Manual entries that were attempted to be saved:**")
                                st.json(st.session_state.get("manual_commission_rows", []))
                                st.write("**Statement date from session:**")
                                st.write(st.session_state.get("recon_stmt_date", "Not set"))
                        
                        # Stop execution to prevent the error from being cleared
                        st.stop()
            else:
                st.info("ℹ️ Add manual commission entries above to reconcile a statement.")

        # --- Payment/Reconciliation History Viewer ---
        st.markdown("---")
        st.markdown("### Payment/Reconciliation History")
        try:
            response = supabase.table('commission_payments_simple').select("*").order('payment_timestamp', desc=True).execute()
            payment_history = pd.DataFrame(response.data) if response.data else pd.DataFrame()
        except Exception as e:
            st.error(f"Error loading payment history: {e}")
            payment_history = pd.DataFrame()
        if not payment_history.empty:
            payment_history["statement_date"] = pd.to_datetime(payment_history["statement_date"], errors="coerce").dt.strftime("%m/%d/%Y").fillna("")
            payment_history["payment_timestamp"] = pd.to_datetime(payment_history["payment_timestamp"], errors="coerce").dt.strftime("%m/%d/%Y %I:%M %p").fillna("")

            for idx, row in payment_history.iterrows():
                col1, col2 = st.columns([0.9, 0.1])
                with col1:
                    with st.expander(f"Statement Date: {row['statement_date']} | Customer: {row['customer']} | Amount: ${row['payment_amount']:,.2f} | Time: {row['payment_timestamp']}"):
                        # Show the full commission statement as a locked table
                        try:
                            details = json.loads(row.get('statement_details', '[]'))
                            if details:
                                df_details = pd.DataFrame(details)
                                st.dataframe(df_details, use_container_width=True, height=min(400, 40 + 40 * len(df_details)))
                            else:
                                st.info("No statement details available for this record.")
                        except Exception:
                            st.info("No statement details available or could not parse.")
                with col2:
                    if 'id' in row and pd.notna(row['id']):
                        if st.button("Delete", key=f"delete_history_{row['id']}"):
                            st.session_state['pending_delete_history_id'] = row['id']
                            st.rerun()
                    else:
                        st.warning("Cannot delete row without a valid ID.")
        
        if 'pending_delete_history_id' in st.session_state:
            st.warning(f"Are you sure you want to delete this history record? This cannot be undone.")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Confirm Delete", key="confirm_delete_history"):
                    try:
                        supabase.table('commission_payments_simple').delete().eq('id', st.session_state['pending_delete_history_id']).execute()
                        st.success("History record deleted.")
                        del st.session_state['pending_delete_history_id']
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error deleting record: {e}")
            with col2:
                if st.button("Cancel", key="cancel_delete_history"):
                    del st.session_state['pending_delete_history_id']
                    st.rerun()
        else:
            st.info("No payment/reconciliation history found.")# --- Policy Revenue Ledger ---
    elif page == "Policy Revenue Ledger":
        st.subheader("Policy Revenue Ledger")
        st.info("Search for a policy to view and edit a detailed ledger of all commission credits and debits for that policy. You can edit, delete transactions below. Be sure to click 'Save Changes' to commit your edits.")

        # --- Granular Policy Search ---
        customers = all_data["Customer"].dropna().unique().tolist() if "Customer" in all_data.columns else []
        selected_customer = st.selectbox("Select Customer:", ["Select..."] + sorted(customers), key="ledger_customer_select")

        filtered_data = all_data.copy()
        if selected_customer and selected_customer != "Select...":
            filtered_data = filtered_data[filtered_data["Customer"] == selected_customer]

        policy_types = filtered_data["Policy Type"].dropna().unique().tolist() if "Policy Type" in filtered_data.columns else []
        selected_policy_type = st.selectbox("Select Policy Type:", ["Select..."] + sorted(policy_types), key="ledger_policytype_select")
        if selected_policy_type and selected_policy_type != "Select...":
            filtered_data = filtered_data[filtered_data["Policy Type"] == selected_policy_type]

        effective_dates = filtered_data["Effective Date"].dropna().unique().tolist() if "Effective Date" in filtered_data.columns else []
        selected_effective_date = st.selectbox("Select Policy Effective Date:", ["Select..."] + sorted(effective_dates), key="ledger_effectivedate_select")
        if selected_effective_date and selected_effective_date != "Select...":
            filtered_data = filtered_data[filtered_data["Effective Date"] == selected_effective_date]

        # Only show policy number select if all three are chosen
        policy_numbers = filtered_data["Policy Number"].dropna().unique().tolist() if "Policy Number" in filtered_data.columns else []
        selected_policy = None
        if selected_customer != "Select..." and selected_policy_type != "Select..." and selected_effective_date != "Select...":
            selected_policy = st.selectbox("Select Policy Number:", ["Select..."] + sorted(policy_numbers), key="ledger_policy_select")

        if selected_policy and selected_policy != "Select...":            # Get all rows for this policy, using only real database data
            policy_rows = all_data[all_data["Policy Number"] == selected_policy].copy()
            # Define the original ledger columns
            ledger_columns = [
                "Transaction ID",
                "Description",
                "Credit (Commission Owed)",
                "Debit (Paid to Agent)",
                "Transaction Type"
            ]

            # --- Always include all ledger columns, filling missing with empty values ---
            import numpy as np
            if not policy_rows.empty:
                # --- Ledger construction using mapped column names ---
                # Get mapped column names for credits and debits
                credit_col = get_mapped_column("Agent Estimated Comm $") or "Agent Estimated Comm $"
                debit_col = get_mapped_column("Agent Paid Amount (STMT)") or "Agent Paid Amount (STMT)"
                transaction_id_col = get_mapped_column("Transaction ID") or "Transaction ID"
                stmt_date_col = get_mapped_column("STMT DATE") or "STMT DATE"
                description_col = get_mapped_column("Description") or "Description"
                transaction_type_col = get_mapped_column("Transaction Type") or "Transaction Type"
                
                # Build ledger_df with correct mapping
                ledger_df = pd.DataFrame()
                ledger_df["Transaction ID"] = policy_rows[transaction_id_col] if transaction_id_col in policy_rows.columns else ""
                ledger_df["Description"] = policy_rows[description_col] if description_col in policy_rows.columns else ""
                
                # Credit (Commission Owed) from mapped Agent Estimated Comm $
                if credit_col in policy_rows.columns:
                    ledger_df["Credit (Commission Owed)"] = policy_rows[credit_col]
                else:
                    ledger_df["Credit (Commission Owed)"] = 0.0                # Debit (Paid to Agent) from mapped Agent Paid Amount (STMT)
                if debit_col in policy_rows.columns:
                    ledger_df["Debit (Paid to Agent)"] = policy_rows[debit_col]
                else:
                    ledger_df["Debit (Paid to Agent)"] = 0.0
                ledger_df["Transaction Type"] = policy_rows[transaction_type_col] if transaction_type_col in policy_rows.columns else ""
                # Ensure correct column order
                ledger_df = ledger_df[ledger_columns]
                # Reset index for clean display
                ledger_df = ledger_df.reset_index(drop=True)
            else:
                # If no rows, show empty DataFrame with correct columns
                ledger_df = pd.DataFrame(columns=ledger_columns)            
            if not ledger_df.empty:
                st.markdown("### Policy Details (Editable)")
                # Show policy-level details using mapped column names
                policy_detail_field_names = [
                    "Customer", "Client ID", "Policy Number", "Policy Type", "Carrier Name",
                    "Effective Date", "Policy Origination Date", "Policy Gross Comm %", 
                    "Agent Comm (NEW 50% RWL 25%)", "X-DATE"
                ]
                policy_detail_cols = []
                for field_name in policy_detail_field_names:
                    mapped_col = get_mapped_column(field_name)
                    if mapped_col and mapped_col in policy_rows.columns:
                        policy_detail_cols.append(mapped_col)
                    elif field_name in policy_rows.columns:
                        policy_detail_cols.append(field_name)
                
                policy_details_df = policy_rows.iloc[[0]][policy_detail_cols].copy() if not policy_rows.empty else pd.DataFrame(columns=policy_detail_cols)
                edited_details_df = st.data_editor(
                    policy_details_df,
                    use_container_width=True,
                    key="policy_details_editor",
                    num_rows="fixed",
                    height=100
                )
                if st.button("Test Mapping (Preview Policy Details)", key="test_mapping_policy_details_btn") or st.session_state.get("show_policy_details_mapping_preview", False):
                    st.session_state["show_policy_details_mapping_preview"] = True
                    
                    def get_policy_details_column_mapping(col, val):
                        mapping = {
                            "Customer": "customer",
                            "Client ID": "client_id",
                            "Policy Number": "policy_number",
                            "Policy Type": "policy_type",
                            "Carrier Name": "carrier_name",
                            "Effective Date": "policy_effective_date",
                            "Policy Origination Date": "policy_origination_date",
                            "Policy Gross Comm %": "policy_commission_pct",
                            "Agent Comm (NEW 50% RWL 25%)": "agent_commission_pct",
                            "X-DATE": "policy_expiration_date"
                        }
                        db_col = mapping.get(col, col)
                        return {
                            "UI Column Title": col,
                            "Database Column Name": db_col,
                            "Current Value": val
                        }
                    mapping_table = [get_policy_details_column_mapping(col, edited_details_df.iloc[0][col]) for col in policy_detail_cols] if not edited_details_df.empty else []
                    mapping_df = pd.DataFrame(mapping_table)
                    with st.expander("Preview Policy Details Mapping & Changes", expanded=True):
                        st.write("**Column Mapping (UI → Database):**")
                        st.dataframe(mapping_df, use_container_width=True)
                else:
                    st.session_state["show_policy_details_mapping_preview"] = False
                if st.button("Save Policy Details", key="save_policy_details_btn") and not edited_details_df.empty:
                    update_sql = "UPDATE policies SET " + ", ".join([f'"{col}" = :{col}' for col in policy_detail_cols if col != "Policy Number"]) + " WHERE \"Policy Number\" = :policy_number"
                    update_params = {col: edited_details_df.iloc[0][col] for col in policy_detail_cols if col != "Policy Number"}
                    update_params["policy_number"] = edited_details_df.iloc[0]["Policy Number"]
                    with engine.begin() as conn:
                        # Update via Supabase
                        if '_id' in selected_df.columns:
                            policy_id = selected_df.iloc[0]['_id']
                        else:
                            # Get the record ID first
                            search_key = 'Policy Number' if 'Policy Number' in update_params else 'Transaction ID'
                            search_value = update_params.get('policy_number', update_params.get('transaction_id'))
                            if search_value:
                                result = supabase.table('policies').select('_id').eq(search_key, search_value).execute()
                                if result.data:
                                    policy_id = result.data[0]['_id']
                                else:
                                    policy_id = None
                            else:
                                policy_id = None
                        
                        if policy_id:
                            # Remove ID fields from update
                            update_dict = {k: v for k, v in update_params.items() if k not in ['_id', 'policy_number', 'transaction_id']}
                            supabase.table('policies').update(update_dict).eq('_id', policy_id).execute()
                            clear_policies_cache()
                    st.success("Policy details updated.")
                    st.rerun()

                st.markdown("### Policy Ledger (Editable)")                # Ensure Credit and Debit columns are numeric
                if "Credit (Commission Owed)" in ledger_df.columns:
                    ledger_df["Credit (Commission Owed)"] = pd.to_numeric(ledger_df["Credit (Commission Owed)"], errors="coerce").fillna(0.0)
                if "Debit (Paid to Agent)" in ledger_df.columns:
                    ledger_df["Debit (Paid to Agent)"] = pd.to_numeric(ledger_df["Debit (Paid to Agent)"], errors="coerce").fillna(0.0)

                # Lock formula columns
                formula_columns = []
                for col in ["Credit (Commission Owed)", "Debit (Paid to Agent)"]:
                    if col in ledger_df.columns:
                        formula_columns.append(col)
                column_config = {}
                for col in ledger_df.columns:
                    if col in formula_columns:
                        column_config[col] = {"disabled": True}

                # Prevent deletion of the first (opening) row
                ledger_df_display = ledger_df.copy()
                ledger_df_display["Delete"] = True
                ledger_df_display.loc[0, "Delete"] = False  # Opening row cannot be deleted
                display_cols = ledger_columns + ["Delete"]

                # --- Ensure all display columns exist in the DataFrame ---
                for col in display_cols:
                    if col not in ledger_df_display.columns:
                        # For Delete, default to True except first row; for others, empty string
                        if col == "Delete":
                            ledger_df_display[col] = [True] * len(ledger_df_display)
                            if len(ledger_df_display) > 0:
                                ledger_df_display.loc[0, "Delete"] = False
                        else:
                            ledger_df_display[col] = ""
                # Reorder columns safely
                ledger_df_display = ledger_df_display[display_cols]

                st.markdown("<b>Why are some columns locked?</b>", unsafe_allow_html=True)
                with st.expander("Locked Columns Explanation", expanded=False):
                    st.markdown("""
                    Some columns in the Policy Revenue Ledger are locked because their values are automatically calculated by the app based on other fields or business rules. These formula columns ensure the accuracy of your commission and revenue calculations. If you need to change a value in a locked column, you must update the underlying fields that drive the calculation, or contact your administrator if you believe the formula needs to be changed.
                    
                    **Locked columns in this ledger include:**
                    - Credit (Commission Owed)
                    - Debit (Paid to Agent)
                    """)

                # Show the table as editable, with locked columns and strictly no row add/delete
                # Lock formula columns and Transaction ID, and set num_rows to 'fixed' to prevent row add/delete
                column_config["Credit (Commission Owed)"] = {
                    "disabled": True,
                    "help": "This column is automatically calculated and cannot be edited. See 'Why are some columns locked?' above."
                }
                column_config["Debit (Paid to Agent)"] = {
                    "disabled": True,
                    "help": "This column is automatically calculated and cannot be edited. See 'Why are some columns locked?' above."
                }
                column_config["Transaction ID"] = {
                    "disabled": True,
                    "help": "Transaction ID is a unique identifier and cannot be changed."
                }
                edited_ledger_df = st.data_editor(
                    ledger_df_display,
                    use_container_width=True,
                    height=max(400, 40 + 40 * len(ledger_df_display)),
                    key="policy_ledger_editor",
                    num_rows="fixed",
                    column_config=column_config,
                    hide_index=True
                )

                # Strictly prevent row addition and deletion: only allow editing of existing rows
                # Remove the Delete column before saving (no row deletion allowed)
                edited_ledger_df = edited_ledger_df.drop(columns=["Delete"])
                # Only keep the original ledger columns (no extra columns)
                edited_ledger_df = edited_ledger_df[ledger_columns]                # --- Test Mapping (Preview Policy Ledger) Button and Expander ---
                if st.button("Test Mapping (Preview Policy Ledger)", key="test_mapping_policy_ledger_btn") or st.session_state.get("show_policy_ledger_mapping_preview", False):
                    st.session_state["show_policy_ledger_mapping_preview"] = True
                    def get_policy_ledger_column_mapping(col, val):
                        mapping = {
                            "Transaction ID": "transaction_id",
                            "Description": "description",
                            "Credit (Commission Owed)": "agent_estimated_comm",
                            "Debit (Paid to Agent)": "Agent Paid Amount (STMT)",
                            "Transaction Type": "transaction_type"
                        }
                        db_col = mapping.get(col, col)
                        return {
                            "UI Column Title": col,
                            "Database Column Name": db_col,
                            "Current Value": val
                        }
                    mapping_table = [get_policy_ledger_column_mapping(col, edited_ledger_df.iloc[0][col] if not edited_ledger_df.empty else "") for col in ledger_columns]
                    mapping_df = pd.DataFrame(mapping_table)
                    with st.expander("Preview Policy Ledger Mapping & Changes", expanded=True):
                        st.write("**Column Mapping (UI → Database):**")
                        st.dataframe(mapping_df, use_container_width=True)
                        st.write("**Proposed Policy Ledger Preview:**")
                        st.dataframe(edited_ledger_df, use_container_width=True)
                else:
                    st.session_state["show_policy_ledger_mapping_preview"] = False                # --- Ledger Totals Section ---
                total_credits = edited_ledger_df["Credit (Commission Owed)"].apply(pd.to_numeric, errors="coerce").sum() if "Credit (Commission Owed)" in edited_ledger_df.columns else 0.0
                total_debits = edited_ledger_df["Debit (Paid to Agent)"].apply(pd.to_numeric, errors="coerce").sum() if "Debit (Paid to Agent)" in edited_ledger_df.columns else 0.0
                balance_due = total_credits - total_debits

                st.markdown("#### Ledger Totals")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Credits", f"${total_credits:,.2f}")
                with col2:
                    st.metric("Total Debits", f"${total_debits:,.2f}")
                with col3:
                    st.metric("Balance Due", f"${balance_due:,.2f}")

                if st.button("Save Changes", key="save_policy_ledger_btn"):
                    for idx, row in edited_ledger_df.iterrows():
                        update_sql = "UPDATE policies SET " + ", ".join([f'"{col}" = :{col}' for col in ledger_columns if col != "Transaction ID"]) + " WHERE \"Transaction ID\" = :transaction_id"
                        update_params = {col: row[col] for col in ledger_columns if col != "Transaction ID"}
                        update_params["transaction_id"] = row["Transaction ID"]
                        
                        # Update via Supabase
                        try:
                            # Find the policy by Transaction ID
                            result = supabase.table('policies').select('_id').eq('Transaction ID', row["Transaction ID"]).execute()
                            if result.data:
                                policy_id = result.data[0]['_id']
                                # Remove ID fields from update
                                update_dict = {k: v for k, v in update_params.items() if k not in ['_id', 'transaction_id']}
                                supabase.table('policies').update(update_dict).eq('_id', policy_id).execute()
                                clear_policies_cache()
                            else:
                                st.error(f"Policy with Transaction ID {row['Transaction ID']} not found")
                        except Exception as e:
                            st.error(f"Error updating policy: {e}")
                    st.success("Policy ledger changes saved.")
                    st.rerun()    # --- Help ---
    elif page == "Help":
        st.title("📚 Help & Documentation")
        
        # Create tabs for organized help content
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["Getting Started", "Features Guide", "Troubleshooting", "Database Column Renaming", "About"])
        
        with tab1:
            st.subheader("🚀 Getting Started")
            st.write("""
            Welcome to the Commission Management Application! This powerful tool helps you manage insurance 
            commission data, track policies, and generate comprehensive reports.
            
            **Quick Start Steps:**
            1. **Add Policy Data**: Use "Add New Policy Transaction" to input your policy information
            2. **View Data**: Navigate to "All Policies in Database" to see your complete dataset
            3. **Generate Reports**: Visit the "Reports" section for analytics and summaries
            4. **Manage Accounting**: Use "Accounting" for commission reconciliation and payments
            """)
            
            st.info("💡 **Tip**: Start by adding a few sample policies to explore all features!")
        
        with tab2:
            st.subheader("📖 Features Guide")
            
            # Dashboard features
            st.write("**📊 Dashboard**")
            st.write("- View key metrics and policy summaries")
            st.write("- Search and edit policies quickly")
            st.write("- Monitor recent activity and statistics")
            
            st.divider()
            
            # Database features
            st.write("**💾 Database Management**")
            st.write("- **All Policies**: Browse paginated policy data with export options")
            st.write("- **Edit Policies**: Search and modify existing policy records")
            st.write("- **Add New**: Create new policy transactions with auto-generated IDs")
            
            st.divider()
            
            # Reporting features
            st.write("**📈 Reports & Analytics**")
            st.write("- **Reports**: Generate summary and commission analysis reports")
            st.write("- **Search & Filter**: Advanced filtering with multiple criteria")
            st.write("- **Policy Revenue Ledger**: Detailed revenue tracking and calculations")
            
            st.divider()
            
            # Administrative features
            st.write("**⚙️ Administrative Tools**")
            st.write("- **Admin Panel**: Database management and system tools")
            st.write("- **Tools**: Utilities for calculations and data formatting")
            st.write("- **Accounting**: Commission reconciliation and payment tracking")
        
        with tab3:
            st.subheader("🔧 Troubleshooting")
            
            st.write("**Common Issues and Solutions:**")
            
            st.write("**❓ No data showing in Dashboard**")
            st.write("- Solution: Add policy data using 'Add New Policy Transaction'")
            st.write("- Check that the database file 'commissions.db' exists")
            
            st.write("**❓ Error saving changes**")
            st.write("- Solution: Ensure all required fields are filled correctly")
            st.write("- Check database permissions and disk space")
            
            st.write("**❓ Slow performance**")
            st.write("- Solution: The app uses caching for better performance")
            st.write("- Large datasets may take longer to load initially")
            
            st.write("**❓ Export not working**")
            st.write("- Solution: Ensure you have data to export")
            st.write("- Check browser settings for file downloads")
            
            st.info("💭 **Need more help?** Contact your system administrator or developer.")
        
        with tab4:
            st.subheader("🔄 Database Column Renaming Guide")
            
            st.warning("""
            ⚠️ **Important**: Do NOT use the Admin Panel's column rename feature for database columns. 
            Instead, follow this procedure to ensure both the app code and database stay synchronized.
            """)
            
            st.write("### Why this procedure?")
            st.write("""
            Database column names are referenced throughout the application code. When you rename a column,
            both the application code AND the database must be updated together to avoid errors.
            """)
            
            st.write("### ✅ Correct Procedure for Renaming Database Columns:")
            
            st.write("**Step 1: Check with Claude**")
            st.code("""
Example request: "We need to update a database column title from: 
'Agency Estimated Comm/Revenue (AZ)' to 'Agency Estimated Comm/Revenue (CRM)'"
            """, language="text")
            
            st.write("**Step 2: Claude will investigate and provide options**")
            st.info("""
            Claude will:
            1. Search the codebase for all references to the column
            2. Identify where the column name appears
            3. Provide options (usually recommending direct database update)
            """)
            
            st.write("**Step 3: Execute the SQL command in Supabase**")
            st.write("1. Go to your Supabase project dashboard")
            st.write("2. Navigate to the SQL Editor")
            st.write("3. Paste and run the ALTER command provided by Claude:")
            
            st.code("""
ALTER TABLE policies 
RENAME COLUMN "Old Column Name" 
TO "New Column Name";
            """, language="sql")
            
            st.write("**Step 4: Claude updates the application code**")
            st.write("""
            After you confirm the database change succeeded, Claude will:
            - Update all code references to use the new column name
            - Update column mapping configurations
            - Update SQL schema documentation files
            - Create a backup of the application before changes
            """)
            
            st.write("### 📋 Recent Example:")
            st.success("""
            **Changed**: "Agency Estimated Comm/Revenue (AZ)" → "Agency Estimated Comm/Revenue (CRM)"
            
            1. Claude found the column in commission_app.py and schema files
            2. You ran the ALTER command in Supabase SQL Editor
            3. Claude updated all code references from (AZ) to (CRM)
            4. Application now works correctly with the new column name
            """)
            
            st.write("### ⚡ Quick Reference:")
            col1, col2 = st.columns(2)
            with col1:
                st.error("❌ **Wrong Way:**")
                st.write("Using Admin Panel → Column Mapping")
                st.write("(This only updates display names, not database)")
            
            with col2:
                st.success("✅ **Right Way:**")
                st.write("1. Ask Claude for help")
                st.write("2. Run SQL in Supabase")
                st.write("3. Let Claude update code")
            
            st.info("""
            💡 **Remember**: This procedure ensures your database and application code stay perfectly 
            synchronized, preventing errors like 'Could not find column in schema cache'.
            """)
        
        with tab5:
            st.subheader("ℹ️ About")
            st.write("""
            **Commission Management Application**
            
            This application is designed to help insurance agencies and professionals manage their 
            commission data efficiently. Built with Streamlit and Python, it provides a comprehensive 
            suite of tools for policy management, reporting, and financial tracking.
            
            **Key Technologies:**
            - **Frontend**: Streamlit
            - **Database**: Supabase (PostgreSQL)
            - **Data Processing**: Pandas
            - **File Formats**: CSV, Excel, PDF support
            
            **Features:**
            - ✅ Complete CRUD operations for policy data
            - ✅ Advanced search and filtering capabilities
            - ✅ Comprehensive reporting and analytics
            - ✅ Commission reconciliation tools
            - ✅ Data export and import functionality
            - ✅ User-friendly dashboard interface
            
            **Version**: 2.0
            **Last Updated**: 2025
            """)
            
            st.success("🎉 Thank you for using the Commission Management Application!")
    
    # --- Policy Revenue Ledger Reports ---
    elif page == "Policy Revenue Ledger Reports":
        st.subheader("Policy Revenue Ledger Reports")
        st.success("📊 Generate customizable reports for policy summaries with Balance Due calculations and export capabilities.")
        
        if all_data.empty:
            st.warning("No policy data loaded. Please check database connection or import data.")
        else:
            # Simple data processing - AGGREGATE BY POLICY NUMBER using mapped columns
            # This ensures one row per policy with totals from all transactions
            working_data = all_data.copy()
            
            # Get mapped column for Policy Number
            policy_number_col = get_mapped_column("Policy Number") or "Policy Number"
            
            # Group by Policy Number and aggregate the data
            if policy_number_col in working_data.columns and not working_data.empty:
                # Define aggregation rules using mapped column names
                agg_dict = {}
                
                # For descriptive fields, take the first value (they should be the same for all transactions of the same policy)
                descriptive_field_names = ["Customer", "Policy Type", "Carrier Name", "Effective Date", 
                                         "Policy Origination Date", "X-DATE", "Client ID"]
                for field_name in descriptive_field_names:
                    mapped_col = get_mapped_column(field_name)
                    target_col = mapped_col if mapped_col and mapped_col in working_data.columns else (field_name if field_name in working_data.columns else None)
                    if target_col:
                        agg_dict[target_col] = 'first'                  # For monetary fields, sum all transactions for each policy
                monetary_field_names = ["Agent Estimated Comm $", "Agent Paid Amount (STMT)", 
                                       "Agency Estimated Comm/Revenue (CRM)", "Premium Sold", "Policy Gross Comm %"]
                for field_name in monetary_field_names:
                    mapped_col = get_mapped_column(field_name)
                    target_col = mapped_col if mapped_col and mapped_col in working_data.columns else (field_name if field_name in working_data.columns else None)
                    if target_col:
                        # Convert to numeric first, then sum
                        working_data[target_col] = pd.to_numeric(working_data[target_col], errors='coerce').fillna(0)
                        agg_dict[target_col] = 'sum'
                
                # For percentage fields, take the first value (should be consistent per policy)
                percentage_fields = ["Agent Comm (NEW 50% RWL 25%)", "Policy Gross Comm %"]
                for field in percentage_fields:
                    if field in working_data.columns and field not in agg_dict:
                        agg_dict[field] = 'first'
                
                # Group by Policy Number and aggregate
                working_data = working_data.groupby('Policy Number', as_index=False).agg(agg_dict)
            
            # Calculate Policy Balance Due using the EXACT same formula as Policy Revenue Ledger page
            # This ensures consistency between the two pages
            
            # Calculate BALANCE DUE using the same logic as the Policy Revenue Ledger page
            # Force recalculation by removing existing Policy Balance Due column if it exists
            if "Policy Balance Due" in working_data.columns:
                working_data = working_data.drop(columns=["Policy Balance Due"])
            
            # Check for required columns and calculate
            # Formula: Policy Balance Due = "Agent Estimated Comm $" - "Agent Paid Amount (STMT)"
            if "Agent Paid Amount (STMT)" in working_data.columns and "Agent Estimated Comm $" in working_data.columns:
                paid_amounts = pd.to_numeric(working_data["Agent Paid Amount (STMT)"], errors="coerce").fillna(0)
                commission_amounts = pd.to_numeric(working_data["Agent Estimated Comm $"], errors="coerce").fillna(0)
                
                working_data["Policy Balance Due"] = commission_amounts - paid_amounts
            else:
                # If we don't have the required columns, create with zeros
                working_data["Policy Balance Due"] = 0
            
            # Ensure Policy Balance Due is numeric for calculations
            if "Policy Balance Due" in working_data.columns:
                working_data["Policy Balance Due"] = pd.to_numeric(working_data["Policy Balance Due"], errors="coerce").fillna(0)
              # Calculate metrics using the same logic as Policy Revenue Ledger page
            total_policies = len(working_data)
            
            # Count unique policies that have outstanding balance due (Policy Balance Due > 0)
            # This matches the UI formula used on the Policy Revenue Ledger page
            if "Policy Balance Due" in working_data.columns:
                # Get unique policies with Policy Balance Due > 0
                outstanding_policies = len(working_data[working_data["Policy Balance Due"] > 0])
                total_balance = working_data["Policy Balance Due"].sum()
            else:
                outstanding_policies = 0
                total_balance = 0
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Policies", f"{total_policies:,}")
            with col2:
                st.metric("Outstanding Balances", f"{outstanding_policies:,}")
            with col3:
                st.metric("Total Balance Due", f"${total_balance:,.2f}")
            
            # Enhanced Column Selection and Template Management
            st.markdown("### 🔧 Column Selection & Templates")
            
            # Initialize session state for templates
            if "prl_templates" not in st.session_state:
                st.session_state.prl_templates = {}
            
            # Get all available columns
            all_columns = list(working_data.columns)
              # Default column selection (updated to use correct column names)
            default_columns = ["Customer", "Policy Number", "Agent Estimated Comm $", "Agent Paid Amount (STMT)", "Policy Balance Due"]
            available_default_columns = [col for col in default_columns if col in all_columns]
            
            # Initialize selected columns in session state
            if "prl_selected_columns" not in st.session_state:
                st.session_state.prl_selected_columns = available_default_columns            # Column selection interface
            col_selection_col1, col_selection_col2 = st.columns([2, 1])
            
            with col_selection_col1:
                st.markdown("**Select Columns:**")
                
                # Available columns multiselect (restored to original style)
                selected_columns = st.multiselect(
                    "Choose columns to display in your report",
                    options=all_columns,
                    default=st.session_state.prl_selected_columns,
                    key="prl_column_multiselect"
                )
                
                # Update session state
                st.session_state.prl_selected_columns = selected_columns
                
                # Column reordering with streamlit_sortables (compact style like Edit Policies page)
                if selected_columns:
                    reorder_col1, reorder_col2 = st.columns([4, 1])
                    
                    with reorder_col1:
                        st.markdown("**Reorder columns by dragging the boxes below:**")
                    
                    with reorder_col2:
                        if st.button("🔄 Refresh", key="refresh_reorder", help="Refresh reorder section to sync with selected columns"):
                            # Force sync by updating the sortable key
                            if "prl_column_order_sortable_counter" not in st.session_state:
                                st.session_state.prl_column_order_sortable_counter = 0
                            st.session_state.prl_column_order_sortable_counter += 1
                            st.rerun()
                    
                    # Generate unique key for sortable component
                    sortable_key = f"prl_column_order_sortable_{st.session_state.get('prl_column_order_sortable_counter', 0)}"
                    
                    reordered_columns = streamlit_sortables.sort_items(
                        items=selected_columns,
                        direction="horizontal",
                        key=sortable_key
                    )
                    
                    # Update session state with reordered columns
                    st.session_state.prl_selected_columns = reordered_columns
                    selected_columns = reordered_columns
            
            with col_selection_col2:
                st.markdown("**Quick Presets:**")
                if st.button("💰 Financial Focus"):
                    financial_cols = ["Customer", "Policy Number", "Agency Estimated Comm/Revenue (CRM)", "Agent Estimated Comm $", "Agent Paid Amount (STMT)", "Policy Balance Due"]
                    st.session_state.prl_selected_columns = [col for col in financial_cols if col in all_columns]
                    st.rerun()
                
                if st.button("📋 Basic Info"):
                    basic_cols = ["Customer", "Policy Type", "Carrier Name", "Policy Number", "Effective Date"]
                    st.session_state.prl_selected_columns = [col for col in basic_cols if col in all_columns]
                    st.rerun()
            
            # Template Management Section
            st.markdown("### 💾 Template Management")
            template_col1, template_col2, template_col3 = st.columns(3)
            
            with template_col1:
                st.markdown("**Save New Template:**")
                new_template_name = st.text_input(
                    "Template Title",
                    placeholder="Enter custom report title",
                    help="Give your template a descriptive name"
                )
                
                if st.button("💾 Save Template", disabled=not new_template_name or not selected_columns):
                    if new_template_name in st.session_state.prl_templates:
                        st.error(f"Template '{new_template_name}' already exists!")
                    else:
                        st.session_state.prl_templates[new_template_name] = {
                            "columns": selected_columns.copy(),
                            "created": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        st.success(f"✅ Template '{new_template_name}' saved!")
                        st.rerun()
            
            with template_col2:
                st.markdown("**Load Template:**")
                if st.session_state.prl_templates:
                    template_to_load = st.selectbox(
                        "Select template to load",
                        options=list(st.session_state.prl_templates.keys()),
                        key="template_loader"
                    )
                    
                    if st.button("📂 Load Template"):
                        template_data = st.session_state.prl_templates[template_to_load]
                        # Only load columns that still exist in the data
                        valid_columns = [col for col in template_data["columns"] if col in all_columns]
                        st.session_state.prl_selected_columns = valid_columns
                        st.success(f"✅ Loaded template: {template_to_load}")
                        st.rerun()
                else:
                    st.info("No saved templates yet")
            
            with template_col3:
                st.markdown("**Manage Templates:**")
                if st.session_state.prl_templates:
                    template_to_manage = st.selectbox(
                        "Select template to manage",
                        options=list(st.session_state.prl_templates.keys()),
                        key="template_manager"
                    )
                    
                    manage_col1, manage_col2 = st.columns(2)
                    with manage_col1:
                        if st.button("✏️ Edit"):
                            # Load template for editing
                            template_data = st.session_state.prl_templates[template_to_manage]
                            valid_columns = [col for col in template_data["columns"] if col in all_columns]
                            st.session_state.prl_selected_columns = valid_columns
                            st.info(f"Loaded '{template_to_manage}' for editing. Modify columns above and save with a new name.")
                    
                    with manage_col2:
                        if st.button("🗑️ Delete"):
                            del st.session_state.prl_templates[template_to_manage]
                            st.success(f"✅ Deleted template: {template_to_manage}")
                            st.rerun()
                else:
                    st.info("No templates to manage")
              # Show current template status
            if st.session_state.prl_templates:
                with st.expander("📋 Saved Templates", expanded=False):
                    for name, data in st.session_state.prl_templates.items():
                        st.write(f"**{name}** - Created: {data['created']} - Columns: {len(data['columns'])}")
            
            # Data Preview with Selected Columns
            st.markdown("### 📊 Report Preview")
            if selected_columns and not working_data.empty:
                # Filter to only include columns that exist
                valid_columns = [col for col in selected_columns if col in working_data.columns]
                
                if valid_columns:
                    # Pagination controls
                    pagination_col1, pagination_col2, pagination_col3 = st.columns([2, 1, 1])
                    
                    with pagination_col1:
                        records_per_page = st.selectbox(
                            "Records per page:",
                            options=[20, 50, 100, 200, 500, "All"],
                            index=0,
                            key="prl_records_per_page"                        )
                    
                    with pagination_col2:
                        if records_per_page != "All":
                            # Ensure records_per_page is treated as integer for calculations
                            records_per_page_int = int(records_per_page)
                            total_pages = max(1, (len(working_data) + records_per_page_int - 1) // records_per_page_int)
                            current_page = st.number_input(
                                "Page:",
                                min_value=1,
                                max_value=total_pages,
                                value=1,
                                key="prl_current_page"
                            )
                        else:
                            current_page = 1
                            total_pages = 1
                    
                    with pagination_col3:
                        if records_per_page != "All":
                            st.write(f"Page {current_page} of {total_pages}")
                    
                    # Apply pagination
                    if records_per_page == "All":
                        display_data = working_data[valid_columns]
                        caption_text = f"Showing all {len(working_data):,} records with {len(valid_columns)} columns"
                    else:
                        # Ensure type safety for calculations
                        records_per_page_int = int(records_per_page)
                        current_page_int = int(current_page)
                        start_idx = (current_page_int - 1) * records_per_page_int
                        end_idx = start_idx + records_per_page_int
                        display_data = working_data[valid_columns].iloc[start_idx:end_idx]
                        caption_text = f"Showing records {start_idx + 1}-{min(end_idx, len(working_data))} of {len(working_data):,} total records with {len(valid_columns)} columns"
                    
                    st.dataframe(display_data, use_container_width=True)
                    st.caption(caption_text)                    # Create report metadata for export
                    export_metadata = {
                        "Report Generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Report Type": "Policy Revenue Ledger Report",
                        "Total Records": f"{len(working_data):,}",
                        "Selected Columns": f"{len(valid_columns)} columns: {', '.join(valid_columns)}",
                        "Data Aggregation": "Grouped by Policy Number (one row per unique policy)",
                        "Balance Due Formula": "Agent Estimated Comm $ - Agent Paid Amount (STMT)",
                        "Page Size": f"{records_per_page} records per page" if records_per_page != "All" else "All records displayed",
                        "Outstanding Policies": f"{outstanding_policies:,} policies with balance due > $0",
                        "Total Balance Due": f"${total_balance:,.2f}"
                    }
                    
                    # Show active report parameters
                    st.markdown("### 📋 Active Report Parameters")
                    st.markdown("*These parameters will be included in exported files:*")
                    metadata_df = pd.DataFrame(list(export_metadata.items()), columns=["Parameter", "Value"])
                    st.dataframe(metadata_df, use_container_width=True, height=min(300, 40 + 40 * len(metadata_df)))
                    
                    # Enhanced export with custom filename and metadata
                    export_col1, export_col2, export_col3 = st.columns(3)
                    with export_col1:
                        custom_filename = st.text_input(
                            "Custom Export Filename (optional)",
                            placeholder="Leave blank for auto-generated name"
                        )
                    
                    with export_col2:
                        filename = custom_filename if custom_filename else f"policy_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
                        csv_filename = filename if filename.endswith('.csv') else filename + '.csv'
                        
                        # Create CSV with metadata header
                        csv_lines = ["# Policy Revenue Ledger Report - Export Parameters"]
                        for param, value in export_metadata.items():
                            csv_lines.append(f"# {param}: {value}")
                        csv_lines.append("# " + "="*50)
                        csv_lines.append("")  # Empty line before data
                        
                        # Add the actual data
                        csv_data = working_data[valid_columns].to_csv(index=False)
                        csv_with_metadata = "\n".join(csv_lines) + "\n" + csv_data
                        
                        st.download_button(
                            "📄 Export as CSV (with Parameters)",
                            csv_with_metadata,
                            csv_filename,
                            "text/csv",
                            help="CSV file includes report parameters in header comments"
                        )
                    
                    with export_col3:
                        excel_filename = filename if filename.endswith('.xlsx') else filename + '.xlsx'
                        
                        # Create Excel file with metadata sheet
                        excel_buffer = io.BytesIO()
                        with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                            # Write metadata to first sheet
                            metadata_df.to_excel(writer, sheet_name='Report Parameters', index=False)
                            
                            # Write data to second sheet
                            working_data[valid_columns].to_excel(writer, sheet_name='Policy Revenue Report', index=False)
                            
                            # Get workbook and format metadata sheet
                            workbook = writer.book
                            metadata_sheet = writer.sheets['Report Parameters']                            # Format metadata sheet
                            header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC'})
                            metadata_sheet.set_column('A:A', 25)
                            metadata_sheet.set_column('B:B', 50)
                            metadata_sheet.write_row(0, 0, ['Parameter', 'Value'], header_format)
                        
                        excel_buffer.seek(0)
                        
                        st.download_button(
                            "📊 Export as Excel (with Parameters)",
                            excel_buffer.getvalue(),
                            excel_filename,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Excel file includes report parameters on separate sheet"
                        )
                else:
                    st.warning("Selected columns are not available in the current data.")
            else:
                if not selected_columns:
                    st.info("Please select columns to display the report.")
                else:
                    st.warning("No data available for report generation.")
            
            # PDF printing instructions
            st.markdown("### 🖨️ Print to PDF")
            st.info("💡 **Tip:** To save this report as PDF, use your browser's print feature (Ctrl+P or Cmd+P) and select 'Save as PDF'. The report parameters shown above will be included in the PDF.")
            
            st.success("✅ Enhanced Policy Revenue Ledger Reports with Templates & Export Parameters!")
    
    # --- Pending Policy Renewals ---
    elif page == "Pending Policy Renewals":
        st.subheader("Pending Policy Renewals")
        
        if 'deleted_renewals' not in st.session_state:
            st.session_state['deleted_renewals'] = []

        pending_renewals_df = get_pending_renewals(all_data)
        duplicated_renewals_df = duplicate_for_renewal(pending_renewals_df)
        
        # Filter out deleted renewals
        if not duplicated_renewals_df.empty:
            display_df = duplicated_renewals_df[~duplicated_renewals_df.index.isin(st.session_state['deleted_renewals'])]
        else:
            display_df = pd.DataFrame()

        if not display_df.empty:
            display_df.insert(0, "Select", False)
            edited_df = st.data_editor(display_df)
            
            selected_rows = edited_df[edited_df["Select"]]
            
            if st.button("Renew Selected"):
                renewed_rows = selected_rows.drop(columns=["Select"])
                
                try:
                    # Insert renewed policies via Supabase
                    for _, row in renewed_rows.iterrows():
                        renewal_data = row.to_dict()
                        # Handle NaN values
                        for key, value in renewal_data.items():
                            if pd.isna(value):
                                renewal_data[key] = None
                        try:
                            supabase.table('policies').insert(renewal_data).execute()
                        except Exception as e:
                            st.error(f"Error inserting renewal: {e}")
                    clear_policies_cache()
                    
                    # Log the renewal
                    if renewed_rows is not None and not renewed_rows.empty:
                        try:
                            # Get the transaction IDs  
                            renewed_ids = renewed_rows['Transaction ID'].tolist() if 'Transaction ID' in renewed_rows.columns else []
                            selected_transaction_ids = selected_rows['Transaction ID'].tolist() if 'Transaction ID' in selected_rows.columns else []
                            
                            # Save renewal history to Supabase
                            renewal_history_data = {
                                "renewal_timestamp": datetime.datetime.now().isoformat(),
                                "renewed_by": "User",
                                "original_transaction_id": ",".join(selected_transaction_ids) if selected_transaction_ids else "",
                                "new_transaction_id": ",".join(renewed_ids) if renewed_ids else "",
                                "details": json.dumps({"count": len(renewed_rows), "renewed_ids": renewed_ids})
                            }
                            supabase.table('renewal_history').insert(renewal_history_data).execute()
                        except Exception as e:
                            st.error(f"Error saving renewal history: {e}")
                    
                    st.success(f"{len(renewed_rows)} policies renewed successfully!")
                    # Remove renewed items from the pending list
                    st.session_state['deleted_renewals'].extend(selected_rows.index.tolist())
                    st.rerun()

                except Exception as e:
                    st.error(f"An error occurred during renewal: {e}")
            
            if st.button("Delete Selected"):
                st.session_state['deleted_renewals'].extend(selected_rows.index.tolist())
                st.rerun()
        else:
            st.info("No policies are pending renewal at this time.")
# Call main function
main()
